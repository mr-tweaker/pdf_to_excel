import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_financial_statements_excel(output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Create worksheets for each financial statement
    balance_sheet = wb.active
    balance_sheet.title = "Balance Sheet"
    manufacturing_account = wb.create_sheet("Manufacturing Account")
    trading_account = wb.create_sheet("Trading Account")
    profit_loss = wb.create_sheet("Profit and Loss")
    schedules = wb.create_sheet("Schedules")
    fixed_assets = wb.create_sheet("Fixed Assets")

    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="D9D9D9",
                             end_color="D9D9D9", fill_type="solid")

    # Create separate sheets for each schedule
    schedule_names = [
        "Schedule A - Capital Account",
        "Schedule B - Secured Loan",
        "Schedule C - Unsecured Loan",
        "Schedule D - Trade Payables",
        "Schedule E - Provisions",
        "Schedule F - Advance against Sale of Land",
        "Schedule G - Advance against Sale",
        "Schedule H - Investment",
        "Schedule I - Closing Stock",
        "Schedule J - Trade Receivable",
        "Schedule K - Loans and Advances",
        "Schedule L - Advances for Material Purchase",
        "Schedule M - Deposits",
        "Schedule N - Other Current Assets",
        "Schedule O - Cash and Bank"
    ]

    # Create each schedule sheet
    for schedule_name in schedule_names:
        sheet = wb.create_sheet(schedule_name)

        # Add header to each sheet
        sheet['A1'] = "M/S MO Infra"
        sheet['A1'].font = header_font

        sheet['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
        sheet['A2'].font = subheader_font

        # Extract schedule letter from the name (e.g., 'A' from 'Schedule A')
        schedule_letter = schedule_name.split(' ')[1]
        sheet['A4'] = f"Schedule : {schedule_letter}"
        sheet['A4'].font = bold_font

        # Set column widths for better readability
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15

    print(f"Created {len(schedule_names)} schedule sheets")

    # Function to set up header for each sheet
    def setup_header(sheet, title):
        sheet.merge_cells('A1:G1')
        sheet['A1'] = "M/S MO Infra"
        sheet['A1'].font = header_font
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('A2:G2')
        sheet['A2'] = "1, Wahid Nagar, Ratlam, Madhya Pradesh -457001"
        sheet['A2'].font = normal_font
        sheet['A2'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('A3:G3')
        sheet['A3'] = title
        sheet['A3'].font = subheader_font
        sheet['A3'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('A4:G4')
        sheet['A4'] = ""

    # Set up Balance Sheet
    setup_header(
        balance_sheet, "Provisional Balance Sheet As on 31st March 2024")

    # Balance Sheet Headers
    balance_sheet['A5'] = "Liabilities"
    balance_sheet['B5'] = "Sch.No"
    balance_sheet['C5'] = "Nature of assets"
    balance_sheet['D5'] = "Amount"
    balance_sheet['E5'] = "Assets"
    balance_sheet['F5'] = "Sch.No"
    balance_sheet['G5'] = "Nature of assets"
    balance_sheet['H5'] = "Amount"

    for col in range(1, 9):
        balance_sheet.cell(row=5, column=col).font = bold_font
        balance_sheet.cell(row=5, column=col).fill = header_fill
        balance_sheet.cell(row=5, column=col).border = thin_border
        balance_sheet.cell(row=5, column=col).alignment = Alignment(
            horizontal='center')

    # Balance Sheet Data
    balance_sheet_data = [
        ["Capital", "A", "", "37,96,52,885", "Fixed Assets", "",
            "some business and some personal", "21,12,33,071"],
        ["Loans (Liabilities)", "", "", "",
         "Long Term Loan and Advances", "", "", ""],
        ["CC Limit(Punjab & sind bank)90", "", "Business",
         "5,88,05,199", "(SMO Ferro Alloys)", "", "", ""],
        ["Secured Loan", "B", "some business and some personal",
            "10,48,32,447", "", "", "", ""],
        ["Unsecured Loans", "C", "Personal", "4,81,42,730",
            "Investments", "H", "Personal", "33,50,41,850"],
        ["HDFC KCC CA", "", "Personal", "4,06,78,904", "", "", "", ""],
        ["Payable for Fixed Assets CB", "", "Business", "47,71,422", "", "", "", ""],
        ["Current Liablities and Provisions", "",
            "", "", "Current Assets", "", "", ""],
        ["Trade Payable", "D", "business", "18,68,81,405",
            "Inventory", "I", "business", "10,49,87,149"],
        ["Provisions", "E", "business", "6,61,067",
            "Trade Receivables", "J", "business", "9,64,80,340"],
        ["Advance against sale of Land", "F", "personal", "6,00,000",
            "Loans & Advances", "K", "business", "5,26,23,269"],
        ["Advance against Sale", "G", "personal", "2,01,08,367",
            "Advances for Material Purchase", "L", "business", "2,43,07,860"],
        ["", "", "", "", "Deposits", "M", "business", "63,96,865"],
        ["", "", "", "", "Other Current Assets", "N", "business", "84,29,963"],
        ["", "", "", "", "Preliminary Expenses", "", "business", "11,87,821"],
        ["", "", "", "", "Cash and Bank", "O", "business", "44,46,238"],
        ["TOTAL", "", "", "84,51,34,426", "TOTAL", "", "", "84,51,34,426"]
    ]

    for row_idx, row_data in enumerate(balance_sheet_data, 6):
        for col_idx, cell_value in enumerate(row_data, 1):
            balance_sheet.cell(row=row_idx, column=col_idx).value = cell_value
            balance_sheet.cell(row=row_idx, column=col_idx).font = normal_font
            balance_sheet.cell(
                row=row_idx, column=col_idx).border = thin_border

            # Apply bold font and background to total row
            if row_idx == len(balance_sheet_data) + 5:  # Last row (TOTAL)
                balance_sheet.cell(
                    row=row_idx, column=col_idx).font = bold_font

    # Set column widths for balance sheet
    balance_sheet.column_dimensions['A'].width = 25
    balance_sheet.column_dimensions['B'].width = 10
    balance_sheet.column_dimensions['C'].width = 25
    balance_sheet.column_dimensions['D'].width = 15
    balance_sheet.column_dimensions['E'].width = 25
    balance_sheet.column_dimensions['F'].width = 10
    balance_sheet.column_dimensions['G'].width = 25
    balance_sheet.column_dimensions['H'].width = 15

    # Footer
    footer_row = len(balance_sheet_data) + 6
    balance_sheet.merge_cells(f'A{footer_row}:H{footer_row}')
    balance_sheet[f'A{footer_row}'] = 'Schedules "A" to "O" forms Integral Part of Accounts'
    balance_sheet[f'A{footer_row}'].font = normal_font
    balance_sheet[f'A{footer_row}'].alignment = Alignment(horizontal='center')

    footer_row += 2
    balance_sheet.merge_cells(f'A{footer_row}:D{footer_row}')
    balance_sheet[f'A{footer_row}'] = 'For: M/s MO Infra'
    balance_sheet[f'A{footer_row}'].font = normal_font

    footer_row += 2
    balance_sheet.merge_cells(f'A{footer_row}:D{footer_row}')
    balance_sheet[f'A{footer_row}'] = 'Mr. Saiyyed Akhtar Ali'
    balance_sheet[f'A{footer_row}'].font = normal_font

    footer_row += 1
    balance_sheet.merge_cells(f'A{footer_row}:D{footer_row}')
    balance_sheet[f'A{footer_row}'] = '(Proprietor)'
    balance_sheet[f'A{footer_row}'].font = normal_font

    footer_row += 2
    balance_sheet.merge_cells(f'A{footer_row}:D{footer_row}')
    balance_sheet[f'A{footer_row}'] = 'Place : Ratlam'
    balance_sheet[f'A{footer_row}'].font = normal_font

    footer_row += 1
    balance_sheet.merge_cells(f'A{footer_row}:D{footer_row}')
    balance_sheet[f'A{footer_row}'] = 'Date : 17-05-2024'
    balance_sheet[f'A{footer_row}'].font = normal_font

    # Set up Manufacturing Account
    setup_header(manufacturing_account,
                 "Provisional Manufacturing Account for the year ended 31st March, 2024")

    # Manufacturing Account Headers
    manufacturing_account['A5'] = "Particulars"
    manufacturing_account['C5'] = "Amount"
    manufacturing_account['E5'] = "Particulars"
    manufacturing_account['G5'] = "Amount"

    for col_idx in [1, 3, 5, 7]:
        manufacturing_account.cell(row=5, column=col_idx).font = bold_font
        manufacturing_account.cell(row=5, column=col_idx).fill = header_fill
        manufacturing_account.cell(row=5, column=col_idx).border = thin_border
        manufacturing_account.cell(
            row=5, column=col_idx).alignment = Alignment(horizontal='center')

    # Manufacturing Account Data
    manufacturing_account_data = [
        ["To Opening Stock", "", "-", "By Cost of Goods Produced-", "", "57,67,393"],
        ["", "", "", "(Transfer to Trdaing Account)", "", ""],
        ["To Direct Expenses", "", "", "", "", ""],
        ["Blasting Expenses", "", "4,95,000", "", "", "-"],
        ["District Mining Fund", "", "1,43,867", "", "", ""],
        ["Electricity Expenses", "", "16,82,221", "", "", ""],
        ["Mining Expenses", "", "-", "", "", ""],
        ["Royalty Expenses", "", "30,76,281", "53,97,369", "", ""],
        ["To Depreciation (Manufacturing)", "", "3,70,024", "", "", ""],
        ["Total", "", "57,67,393", "Total", "", "57,67,393"]
    ]

    for row_idx, row_data in enumerate(manufacturing_account_data, 6):
        for col_idx, col_pos in enumerate([0, 1, 2, 4, 5, 6]):
            if col_idx < len(row_data):
                cell_value = row_data[col_idx]
                actual_col = col_pos + 1
                manufacturing_account.cell(
                    row=row_idx, column=actual_col).value = cell_value
                manufacturing_account.cell(
                    row=row_idx, column=actual_col).font = normal_font
                manufacturing_account.cell(
                    row=row_idx, column=actual_col).border = Side(style='thin')

                # Apply bold font and background to total row
                if row_idx == len(manufacturing_account_data) + 5:  # Last row (Total)
                    manufacturing_account.cell(
                        row=row_idx, column=actual_col).font = bold_font

    # Set column widths for manufacturing account
    manufacturing_account.column_dimensions['A'].width = 25
    manufacturing_account.column_dimensions['B'].width = 10
    manufacturing_account.column_dimensions['C'].width = 15
    manufacturing_account.column_dimensions['D'].width = 10
    manufacturing_account.column_dimensions['E'].width = 25
    manufacturing_account.column_dimensions['F'].width = 10
    manufacturing_account.column_dimensions['G'].width = 15

    # Add footer to Manufacturing Account similar to Balance Sheet
    footer_row = len(manufacturing_account_data) + 7
    manufacturing_account.merge_cells(f'A{footer_row}:G{footer_row}')
    manufacturing_account[f'A{footer_row}'] = 'In Terms of Our Attached Report of Even Date'
    manufacturing_account[f'A{footer_row}'].font = normal_font
    manufacturing_account[f'A{footer_row}'].alignment = Alignment(
        horizontal='center')

    footer_row += 2
    manufacturing_account.merge_cells(f'A{footer_row}:D{footer_row}')
    manufacturing_account[f'A{footer_row}'] = 'For: M/s MO Infra'
    manufacturing_account[f'A{footer_row}'].font = normal_font

    # Set up Trading Account
    setup_header(trading_account,
                 "Provisional Trading Account for the year ended 31st March, 2024")

    # Set column widths for trading account
    trading_account.column_dimensions['A'].width = 25
    trading_account.column_dimensions['B'].width = 15
    trading_account.column_dimensions['C'].width = 15
    trading_account.column_dimensions['D'].width = 15
    trading_account.column_dimensions['E'].width = 25
    trading_account.column_dimensions['F'].width = 15
    trading_account.column_dimensions['G'].width = 15

    # Create the table structure
    for row in range(5, 23):  # Extend to row 23 to cover all content
        for col in range(1, 8):
            cell = trading_account.cell(row=row, column=col)
            cell.border = thin_border

    # Trading Account Headers
    trading_account['A5'] = "Particulars"
    trading_account['D5'] = "Amount"
    trading_account['E5'] = "Particulars"
    trading_account['G5'] = "Amount"

    for col_idx in [1, 4, 5, 7]:
        trading_account.cell(row=5, column=col_idx).font = bold_font
        trading_account.cell(row=5, column=col_idx).fill = header_fill
        trading_account.cell(row=5, column=col_idx).alignment = Alignment(
            horizontal='center')

    # Trading Account Data - matching the exact structure in the image
    row = 6
    trading_account.cell(row=row, column=1).value = "To Opening Stock"
    trading_account.cell(row=row, column=4).value = "8,57,40,793"
    trading_account.cell(row=row, column=5).value = "By Sales account"
    row += 1

    trading_account.cell(row=row, column=5).value = "Local Sales 5%"
    trading_account.cell(row=row, column=6).value = "2,09,90,859"
    row += 1

    trading_account.cell(row=row, column=1).value = "To Purchase account"
    trading_account.cell(row=row, column=5).value = "Local Sales 18%"
    trading_account.cell(row=row, column=6).value = "17,56,33,045"
    row += 1

    trading_account.cell(row=row, column=1).value = "Local Purchase @ 18%"
    trading_account.cell(row=row, column=2).value = "21,80,59,526"
    trading_account.cell(row=row, column=4).value = "21,80,59,526"
    trading_account.cell(row=row, column=5).value = "Igst Sales @ 18%"
    trading_account.cell(row=row, column=6).value = "4,09,27,000"
    row += 1

    trading_account.cell(row=row, column=5).value = "Work Contract"
    trading_account.cell(row=row, column=6).value = "76,92,308"
    row += 1

    trading_account.cell(row=row, column=1).value = "To Cost of Goods Produced"
    trading_account.cell(row=row, column=4).value = "57,67,393"
    trading_account.cell(row=row, column=5).value = "Sales Return"
    trading_account.cell(row=row, column=6).value = "(1,13,750)"
    trading_account.cell(row=row, column=7).value = "24,51,29,462"
    row += 1

    trading_account.cell(row=row, column=1).value = "(From Manufacturing A/c)"
    row += 1

    # Skip a row as in the image
    row += 1

    trading_account.cell(row=row, column=5).value = "To Direct incomes"
    row += 1

    trading_account.cell(row=row, column=5).value = "Land Rent Income"
    trading_account.cell(row=row, column=6).value = "3,00,000"
    row += 1

    trading_account.cell(row=row, column=1).value = "To Direct Expenses"
    trading_account.cell(row=row, column=5).value = "Royalty Income"
    trading_account.cell(row=row, column=6).value = "15,63,823"
    trading_account.cell(row=row, column=7).value = "18,63,823"
    row += 1

    trading_account.cell(row=row, column=1).value = "Frieght Charges"
    trading_account.cell(row=row, column=2).value = "38,500"
    trading_account.cell(row=row, column=4).value = "38,500"
    trading_account.cell(row=row, column=5).value = "To Closing Stock"
    row += 1

    trading_account.cell(row=row, column=5).value = "Raw Material"
    trading_account.cell(row=row, column=6).value = "11,26,525"
    row += 1

    trading_account.cell(row=row, column=5).value = "Stock in hand"
    trading_account.cell(row=row, column=6).value = "10,38,60,624"
    trading_account.cell(row=row, column=7).value = "10,49,87,149"
    row += 1

    trading_account.cell(row=row, column=1).value = "Gross Profit"
    trading_account.cell(row=row, column=4).value = "4,23,74,222"
    row += 1

    trading_account.cell(row=row, column=1).value = "Total"
    trading_account.cell(row=row, column=4).value = "35,19,80,434"
    trading_account.cell(row=row, column=5).value = "Total"
    trading_account.cell(row=row, column=7).value = "35,19,80,434"

    # Apply bold font to the "Total" row
    for col in [1, 4, 5, 7]:
        trading_account.cell(row=row, column=col).font = bold_font

    # Add footer
    row += 2
    trading_account.merge_cells(f'A{row}:G{row}')
    trading_account[f'A{row}'] = 'In Terms of Our Attached Report of Even Date'
    trading_account[f'A{row}'].font = normal_font
    trading_account[f'A{row}'].alignment = Alignment(horizontal='left')

    row += 2
    trading_account.merge_cells(f'A{row}:D{row}')
    trading_account[f'A{row}'] = 'For: M/s MO Infra'
    trading_account[f'A{row}'].font = normal_font

    row += 4
    trading_account.merge_cells(f'A{row}:D{row}')
    trading_account[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    trading_account[f'A{row}'].font = normal_font

    row += 1
    trading_account.merge_cells(f'A{row}:D{row}')
    trading_account[f'A{row}'] = '(Proprietor)'
    trading_account[f'A{row}'].font = normal_font

    row += 3
    trading_account.merge_cells(f'A{row}:D{row}')
    trading_account[f'A{row}'] = 'Place : Ratlam'
    trading_account[f'A{row}'].font = normal_font

    row += 1
    trading_account.merge_cells(f'A{row}:D{row}')
    trading_account[f'A{row}'] = 'Date : 17-05-2024'
    trading_account[f'A{row}'].font = normal_font

    # Set up Profit and Loss Account
    setup_header(
        profit_loss, "Provisional Profit and Loss a/c for the year ended 31st March, 2024")

    # Set column widths for profit and loss
    profit_loss.column_dimensions['A'].width = 25
    profit_loss.column_dimensions['B'].width = 15
    profit_loss.column_dimensions['C'].width = 15
    profit_loss.column_dimensions['D'].width = 15
    profit_loss.column_dimensions['E'].width = 25
    profit_loss.column_dimensions['F'].width = 15
    profit_loss.column_dimensions['G'].width = 15

    # Create the table structure
    for row in range(5, 38):  # Extend table to cover all content
        for col in range(1, 8):
            cell = profit_loss.cell(row=row, column=col)
            cell.border = thin_border

    # Profit and Loss Headers
    profit_loss['A5'] = "Particulars"
    profit_loss['C5'] = "Amount"
    profit_loss['E5'] = "Particulars"
    profit_loss['G5'] = "Amount"

    for col_idx in [1, 3, 5, 7]:
        profit_loss.cell(row=5, column=col_idx).font = bold_font
        profit_loss.cell(row=5, column=col_idx).fill = header_fill
        profit_loss.cell(row=5, column=col_idx).alignment = Alignment(
            horizontal='center')

    # Profit and Loss Data - matching the exact structure in the image
    row = 6
    profit_loss.cell(row=row, column=1).value = "To Indirect Expenses"
    profit_loss.cell(row=row, column=5).value = "By Gross Profit"
    profit_loss.cell(row=row, column=7).value = "4,23,74,222"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Audit Fees"
    profit_loss.cell(row=row, column=3).value = "25,000"
    row += 1

    profit_loss.cell(row=row, column=1).value = "CIBIL Expenses"
    profit_loss.cell(row=row, column=3).value = "8,685"
    profit_loss.cell(row=row, column=5).value = "By Indirect Income"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Depreciation"
    profit_loss.cell(row=row, column=3).value = "8,15,799"
    profit_loss.cell(row=row, column=5).value = "Bank Interest Received"
    profit_loss.cell(row=row, column=7).value = "8,091"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Fees & Penalty"
    profit_loss.cell(row=row, column=3).value = "32,560"
    profit_loss.cell(row=row, column=5).value = "Bank Interest on FD"
    profit_loss.cell(row=row, column=7).value = "2,37,488"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Insurance Expenses"
    profit_loss.cell(row=row, column=3).value = "1,08,913"
    profit_loss.cell(row=row, column=5).value = "Bank Interest on Gold Bond"
    profit_loss.cell(row=row, column=7).value = "2,904"
    row += 1

    profit_loss.cell(
        row=row, column=1).value = "Legal and Professional Expenses"
    profit_loss.cell(row=row, column=3).value = "10,91,711"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Miscelleneous Expenses"
    profit_loss.cell(row=row, column=3).value = "1,300"
    profit_loss.cell(row=row, column=5).value = "Director Sitting Fees"
    profit_loss.cell(row=row, column=7).value = "1,60,000"
    row += 1

    profit_loss.cell(row=row, column=1).value = "MSTC Registration Expenses"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Professional Tax"
    profit_loss.cell(row=row, column=3).value = "2,500"
    profit_loss.cell(row=row, column=5).value = "Gas Subsidies"
    profit_loss.cell(row=row, column=7).value = "219"
    row += 1

    profit_loss.cell(
        row=row, column=1).value = "Preliminary Expenses written off"
    profit_loss.cell(row=row, column=3).value = "2,96,956"
    profit_loss.cell(row=row, column=5).value = "Profit from Firm (M/s Saiyyed"
    row += 1

    profit_loss.cell(row=row, column=5).value = "Akhtar Ali)"
    profit_loss.cell(row=row, column=7).value = "2,69,88,723"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Repair & Maintenance Expenses"
    profit_loss.cell(row=row, column=5).value = "Interest on Capital of Firm"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Salary Expenses"
    profit_loss.cell(row=row, column=3).value = "5,95,000"
    profit_loss.cell(row=row, column=5).value = "(M/s Saiyyed Akhtar Ali)"
    profit_loss.cell(row=row, column=7).value = "1,80,83,759"
    profit_loss.cell(row=row, column=7).alignment = Alignment(
        horizontal='right')
    row += 1

    profit_loss.cell(row=row, column=1).value = "Staff Welfare Expenses"
    profit_loss.cell(row=row, column=3).value = "20,72,726"
    profit_loss.cell(row=row, column=7).value = "4,54,81,184"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Fastag Expenses"
    profit_loss.cell(row=row, column=3).value = "53,500"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Travelling Expenses"
    profit_loss.cell(row=row, column=3).value = "4,21,392"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Power & fuel Expneses"
    profit_loss.cell(row=row, column=3).value = "9,59,440"
    profit_loss.cell(row=row, column=3).alignment = Alignment(
        horizontal='right')
    profit_loss.cell(row=row, column=4).value = "64,85,482"
    row += 1

    profit_loss.cell(row=row, column=1).value = "To Finance Charges"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Due Diligence Expenses"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Bank Charges"
    profit_loss.cell(row=row, column=3).value = "4,59,550"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Bank Interest CC"
    profit_loss.cell(row=row, column=3).value = "53,21,939"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Bank Interest on Term Loan"
    profit_loss.cell(row=row, column=3).value = "1,59,93,741"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Brokerage Expenses"
    profit_loss.cell(row=row, column=3).value = "61,403"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Credit Card Charges"
    profit_loss.cell(row=row, column=3).value = "18,706"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Inspection Charges"
    profit_loss.cell(row=row, column=3).value = "40,503"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Interest on Unsecured Loan"
    profit_loss.cell(row=row, column=3).value = "17,76,102"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Loan Processing Charges"
    profit_loss.cell(row=row, column=3).value = "4,66,716"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Stamp Charges"
    profit_loss.cell(row=row, column=3).value = "5,02,075"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Valuation Charges"
    profit_loss.cell(row=row, column=3).value = "32,770"
    profit_loss.cell(row=row, column=3).alignment = Alignment(
        horizontal='right')
    profit_loss.cell(row=row, column=4).value = "2,46,73,505"
    row += 1

    profit_loss.cell(row=row, column=1).value = "To Net Profit"
    profit_loss.cell(row=row, column=4).value = "5,66,96,419"
    row += 1

    profit_loss.cell(row=row, column=1).value = "Total"
    profit_loss.cell(row=row, column=4).value = "8,78,55,405"
    profit_loss.cell(row=row, column=5).value = "Total"
    profit_loss.cell(row=row, column=7).value = "8,78,55,405"

    # Apply bold font to the "Total" row
    for col in [1, 4, 5, 7]:
        profit_loss.cell(row=row, column=col).font = bold_font

    # Draw a line under subtotals
    # Rows with subtotals (64,85,482 and 2,46,73,505)
    for row_with_subtotals in [22, 35]:
        for col in range(2, 5):
            profit_loss.cell(row=row_with_subtotals, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                # Darker bottom border
                bottom=Side(style='thin', color="000000")
            )

    # Add footer
    row += 2
    profit_loss.merge_cells(f'A{row}:G{row}')
    profit_loss[f'A{row}'] = 'In Terms of Our Attached Report of Even Date'
    profit_loss[f'A{row}'].font = normal_font
    profit_loss[f'A{row}'].alignment = Alignment(horizontal='left')

    row += 2
    profit_loss.merge_cells(f'A{row}:D{row}')
    profit_loss[f'A{row}'] = 'For: M/s MO Infra'
    profit_loss[f'A{row}'].font = normal_font

    row += 4
    profit_loss.merge_cells(f'A{row}:D{row}')
    profit_loss[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    profit_loss[f'A{row}'].font = normal_font

    row += 1
    profit_loss.merge_cells(f'A{row}:D{row}')
    profit_loss[f'A{row}'] = '(Proprietor)'
    profit_loss[f'A{row}'].font = normal_font

    row += 3
    profit_loss.merge_cells(f'A{row}:D{row}')
    profit_loss[f'A{row}'] = 'Place : Ratlam'
    profit_loss[f'A{row}'].font = normal_font

    row += 1
    profit_loss.merge_cells(f'A{row}:D{row}')
    profit_loss[f'A{row}'] = 'Date : 17-05-2024'
    profit_loss[f'A{row}'].font = normal_font

    # Set up Fixed Assets sheet for Schedules
    setup_header(
        fixed_assets, "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024")

    # Add a decorative line below the header
    fixed_assets.merge_cells('A2:I2')
    # Create a line of underscores for visual separation
    fixed_assets['A2'] = "_" * 120
    fixed_assets['A2'].font = Font(name='Arial', size=10, bold=True)
    fixed_assets['A2'].alignment = Alignment(horizontal='center')

    # Add the FIXED ASSETS title
    fixed_assets['A4'] = "FIXED ASSETS :-"
    fixed_assets['A4'].font = bold_font

    # Fixed Assets Headers - Row 6 for GROSS BLOCK
    fixed_assets.merge_cells('B6:E6')
    fixed_assets['B6'] = "GROSS BLOCK"
    fixed_assets['B6'].font = bold_font
    fixed_assets['B6'].alignment = Alignment(horizontal='center')
    fixed_assets['B6'].fill = header_fill
    fixed_assets['B6'].border = thin_border

    # NET BLOCK header
    fixed_assets['I6'] = "NET BLOCK"
    fixed_assets['I6'].font = bold_font
    fixed_assets['I6'].alignment = Alignment(horizontal='center')
    fixed_assets['I6'].fill = header_fill
    fixed_assets['I6'].border = thin_border

    # Fixed Assets Headers - Row 7 for column labels
    col_headers = [
        "ASSETS",
        "Balance\nas on\n01/04/2023",
        "Additions\nduring the\nyear (more than 6 months)",
        "Additions\nduring the\nyear (less than 6 months)",
        "Deduction /\nDisposal",
        "Balance\nas on\n31/03/2023",
        "Depreciation\nfor the year",
        "Rate of\nDepriciation",
        "NET BLOCK\nAs on\n31/3/2024"
    ]

    for idx, header in enumerate(col_headers, 1):
        cell = fixed_assets.cell(row=7, column=idx)
        cell.value = header
        cell.font = bold_font
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = thin_border

    # Set the column widths
    column_widths = [30, 15, 15, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        fixed_assets.column_dimensions[get_column_letter(i)].width = width

    # Fixed Assets Data
    fixed_assets_data = [
        ["Tangible Assets", "", "", "", "", "", "", "", ""],
        ["Land", "20,47,87,702", "1,24,664", "", "",
            "20,49,12,366", "-", "-", "20,49,12,366"],
        ["Plant & Machinery (Crusher)", "7,91,432", "", "",
         "", "7,91,432", "1,18,715", "15%", "6,72,717"],
        ["Plant & Machinery", "6,16,351", "", "", "",
            "6,16,351", "92,453", "15.0%", "5,23,898"],
        ["Plant & Machinery(Concentrating table)", "3,14,829",
         "", "", "", "3,14,829", "47,224", "15.0%", "2,67,605"],
        ["Weightbrige System", "3,07,984", "", "", "",
            "3,07,984", "46,198", "15%", "2,61,786"],
        ["Vehicle (JCB)", "10,52,584", "", "", "",
         "10,52,584", "1,57,888", "15%", "8,94,696"],
        ["Vehicle (Jeep)", "-", "", "", "", "-", "-", "15%", "-"],
        ["Vehicle(Fortuner BS-VI)", "40,76,541", "", "", "",
         "40,76,541", "6,11,481", "15%", "34,65,060"],
        ["Furniture & Fixture", "32,182", "", "",
            "-", "32,182", "3,218", "10%", "28,964"],
        ["Computers & mobile phone", "59,965", "1,68,644", "86,015",
            "-", "3,14,624", "1,08,647", "40%", "2,05,978"],
        ["Sub Total", "21,20,39,571", "2,93,308", "86,015", "-",
            "21,24,18,894", "11,85,823", "", "21,12,33,071"]
    ]

    # Starting row for data
    start_row = 8

    # Add data and apply formatting
    for row_idx, row_data in enumerate(fixed_assets_data, start_row):
        # Special formatting for the "Tangible Assets" row
        if row_idx == start_row:  # First row
            fixed_assets.cell(row=row_idx, column=1).font = bold_font

        for col_idx, cell_value in enumerate(row_data, 1):
            cell = fixed_assets.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.border = thin_border

            # Bold for the first column and totals row
            if col_idx == 1 or row_idx == len(fixed_assets_data) + start_row - 1:
                cell.font = bold_font
            else:
                cell.font = normal_font

            # Align numbers to the right
            if col_idx > 1:  # Columns with numbers
                cell.alignment = Alignment(horizontal='right')

    # Set column widths
    schedules.column_dimensions['A'].width = 35
    schedules.column_dimensions['B'].width = 10
    schedules.column_dimensions['C'].width = 15
    schedules.column_dimensions['D'].width = 35
    schedules.column_dimensions['E'].width = 10
    schedules.column_dimensions['F'].width = 15

    wb = create_schedule_a(wb, output_file)
    wb = create_schedule_b(wb, output_file)
    wb = create_schedule_c(wb, output_file) 
    wb = create_schedule_ca(wb, output_file)
    wb = create_schedule_d(wb, output_file)
    wb = create_schedule_e(wb, output_file)
    wb = create_schedule_f(wb, output_file)
    wb = create_schedule_g(wb, output_file)
    wb = create_schedule_h(wb, output_file)
    wb = create_schedule_i(wb, output_file)
    wb = create_schedule_j(wb, output_file)
    wb = create_schedule_k(wb, output_file)
    wb = create_schedule_l(wb, output_file)
    wb = create_schedule_m(wb, output_file)
    wb = create_schedule_n(wb, output_file)
    wb = create_schedule_o(wb, output_file)
    
    
    # Save the workbook
    wb.save(output_file)
    print(f"Financial statements Excel file created: {output_file}")


def create_schedule_a(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule A - Capital Account in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_a = wb.active
        schedule_a.title = "Schedule A - Capital Account"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule A - Capital Account" in wb.sheetnames:
            schedule_a = wb["Schedule A - Capital Account"]
        else:
            schedule_a = wb.create_sheet("Schedule A - Capital Account")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_a.column_dimensions['A'].width = 35
    schedule_a.column_dimensions['B'].width = 20
    schedule_a.column_dimensions['C'].width = 35
    schedule_a.column_dimensions['D'].width = 20
    
    # Add header
    schedule_a['A1'] = "M/S MO Infra"
    schedule_a['A1'].font = header_font
    
    schedule_a['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_a['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_a['A4'] = "Schedule : A"
    schedule_a['A4'].font = bold_font
    
    # Add Capital Account header
    schedule_a['A5'] = "Capital Account (Saiyyed Akhtar Ali)"
    schedule_a['A5'].font = bold_font
    schedule_a.merge_cells('A5:D5')
    schedule_a['A5'].alignment = Alignment(horizontal='center')
    
    # Create table headers
    headers = ['Particulars', 'Amount', 'Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_a.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["To Drawing", "21,93,261", "By Opening Balance", "32,09,74,715"],
        ["To Income Tax 20-21", "", "By Agriculture Income", "6,802"],
        ["To Saiyyed Akhtar Ali Firm", "", "By Saiyyed Akhtar Ali Firm", "25,32,812"],
        ["To Agriculture Expenses", "", "By VAT Appeal Refund", "1,17,250"],
        ["To Marriage Expenses", "", "By Land received through Gift Deed", ""],
        ["To Durga Prashad Kasera", "", "By Owais Metal and Mineral Process", "16,00,000"],
        ["To TDS (21-22)", "", "", ""],
        ["To TCS on Purchase (21-22)", "", "", ""],
        ["To Life General Insurance", "81,852", "", ""],
        ["To Property Tax", "", "", ""],
        ["To Land as gift deed", "", "", ""],
        ["", "", "", ""],
        ["To Closing Balance", "37,96,52,885", "By Profit and loss a/c", "5,66,96,419"],
        ["Total", "38,19,27,998", "", "38,19,27,998"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_a.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount columns
            if col_idx == 2 or col_idx == 4:  # Amount columns
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_a.merge_cells(f'A{row}:D{row}')
    schedule_a[f'A{row}'] = 'For: M/s MO Infra'
    schedule_a[f'A{row}'].font = normal_font
    
    row += 3
    schedule_a.merge_cells(f'A{row}:D{row}')
    schedule_a[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_a[f'A{row}'].font = normal_font
    
    row += 1
    schedule_a.merge_cells(f'A{row}:D{row}')
    schedule_a[f'A{row}'] = '(Proprietor)'
    schedule_a[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_b(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule B - Secured Loan in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_b = wb.active
        schedule_b.title = "Schedule B - Secured Loan"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule B - Secured Loan" in wb.sheetnames:
            schedule_b = wb["Schedule B - Secured Loan"]
        else:
            schedule_b = wb.create_sheet("Schedule B - Secured Loan")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_b.column_dimensions['A'].width = 50
    schedule_b.column_dimensions['B'].width = 20
    
    # Add header
    schedule_b['A1'] = "M/S MO Infra"
    schedule_b['A1'].font = header_font
    
    schedule_b['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_b['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_b['A4'] = "Schedule : B"
    schedule_b['A4'].font = bold_font
    
    # Add Secured Loan header
    schedule_b['A5'] = "2.Secured Loan"
    schedule_b['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_b.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Aditya Birla Finance", "24,61,455"],
        ["Axis Bank (A/c No BPR004409075975)", "22,97,890"],
        ["Bank of India Gold Loan 948077610000275", "5,50,000"],
        ["Bank of India Gold Loan 948077610000277", "2,99,946"],
        ["BOI Gold Loan 948077610000234", "2,14,000"],
        ["BOI Gold Loan  948077610000308", "4,42,000"],
        ["Central Bank Of India 5253986508", "3,47,67,479"],
        ["Central Bank Of India Loan 05276562822 (Rs.580000)", "4,11,08,145"],
        ["ICICI Bank LORAT0004355846 (Rs.115360)", "1,20,286"],
        ["ICICI Bank   UPRAT0004827513 (Rs.266374)", "61,51,874"],
        ["Indusind Bank Refinance IRR00466E(RS29452)", "7,71,835"],
        ["Kisetsu Saison Finance(India)Private Limited", "19,91,267"],
        ["Kotak Mahindra Bank Limited (Rs.220826/-)", "18,22,284"],
        ["Punjab & Sind Bank - 05941200001042", "4,67,928"],
        ["Punjab & Sind Bank  05941200001272", "29,73,404"],
        ["Punjab & Sind Bank  05941200001273", "80,40,469"],
        ["ICICI Credit card", "(47,814)"],
        ["Total", "10,48,32,447"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_b.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_b.merge_cells(f'A{row}:B{row}')
    schedule_b[f'A{row}'] = 'For: M/s MO Infra'
    schedule_b[f'A{row}'].font = normal_font
    
    row += 3
    schedule_b.merge_cells(f'A{row}:B{row}')
    schedule_b[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_b[f'A{row}'].font = normal_font
    
    row += 1
    schedule_b.merge_cells(f'A{row}:B{row}')
    schedule_b[f'A{row}'] = '(Proprietor)'
    schedule_b[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_c(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule C - Unsecured Loan in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_c = wb.active
        schedule_c.title = "Schedule C - Unsecured Loan"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule C - Unsecured Loan" in wb.sheetnames:
            schedule_c = wb["Schedule C - Unsecured Loan"]
        else:
            schedule_c = wb.create_sheet("Schedule C - Unsecured Loan")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_c.column_dimensions['A'].width = 50
    schedule_c.column_dimensions['B'].width = 20
    
    # Add header
    schedule_c['A1'] = "M/S MO Infra"
    schedule_c['A1'].font = header_font
    
    schedule_c['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_c['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_c['A4'] = "Schedule : C"
    schedule_c['A4'].font = bold_font
    
    # Add Unsecured Loan header
    schedule_c['A5'] = "3. Unsecured Loan"
    schedule_c['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_c.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Anita Jain", "5,00,000"],
        ["Ashok Bhandari HUF", "3,70,600"],
        ["Hemant Bhandari", "20,29,704"],
        ["Rajendra Kumar Jain", "9,34,248"],
        ["Rakesh Bhandari HUF", "3,70,600"],
        ["Sohanlal Jain", "12,00,000"],
        ["Durga Prashad Gabbu Lal Tank HUF", "96,760"],
        ["Ashok Kumar Anshul Kumar HUF", "10,00,000"],
        ["Ashok Kumar Katariya", "15,00,000"],
        ["Neha Jain D/o Santosh", "10,00,000"],
        ["Radheshyam Mundra", "10,00,000"],
        ["Akhil", "2,00,000"],
        ["Sohanlal Bee", "1,58,000"],
        ["Buruhauddin Sailana Wala", "34,00,000"],
        ["Darsh Marketing", "35,00,000"],
        ["Firoza Bee", "1,58,000"],
        ["Geetanjali Construction (Loan)", "15,00,000"],
        ["Mehroj Bee", "51,867"],
        ["Niyamat Ali", "29,815"],
        ["Owais Ali Overseas Proprietor", "14,65,000"],
        ["RG Jewellers", "47,55,371"],
        ["Saiyyed Asgar Ali", "36,56,230"],
        ["Saiyyed Neha Ali", "81,02,000"],
        ["Shamim Bee Khan", "8,00,000"],
        ["Shree Siddhi Vinayak", "10,00,000"],
        ["Sushil Kumar Moonat", "20,00,000"],
        ["Vijay Katariya", "13,64,535"],
        ["Acme Ferro Alloys Pvt Ltd", "60,00,000"],
        ["Total", "4,81,42,730"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_c.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_c.merge_cells(f'A{row}:B{row}')
    schedule_c[f'A{row}'] = 'For: M/s MO Infra'
    schedule_c[f'A{row}'].font = normal_font
    
    row += 3
    schedule_c.merge_cells(f'A{row}:B{row}')
    schedule_c[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_c[f'A{row}'].font = normal_font
    
    row += 1
    schedule_c.merge_cells(f'A{row}:B{row}')
    schedule_c[f'A{row}'] = '(Proprietor)'
    schedule_c[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_ca(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule CA - HDFC KCC and Payable for Fixed Assets in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_ca = wb.active
        schedule_ca.title = "Schedule CA - HDFC KCC"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule CA - HDFC KCC" in wb.sheetnames:
            schedule_ca = wb["Schedule CA - HDFC KCC"]
        else:
            schedule_ca = wb.create_sheet("Schedule CA - HDFC KCC")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_ca.column_dimensions['A'].width = 50
    schedule_ca.column_dimensions['B'].width = 20
    
    # Add header
    schedule_ca['A1'] = "M/S MO Infra"
    schedule_ca['A1'].font = header_font
    
    schedule_ca['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_ca['A2'].font = subheader_font
    
    # Add Schedule title for HDFC KCC
    schedule_ca['A4'] = "Schedule : CA"
    schedule_ca['A4'].font = bold_font
    
    # Add HDFC KCC header
    schedule_ca['A5'] = "4.HDFC KCC"
    schedule_ca['A5'].font = bold_font
    
    # Create table headers for HDFC KCC
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_ca.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the HDFC KCC data
    hdfc_data = [
        ["HDFC Bank -50200048190719", "3,24,64,965"],
        ["HDFC Bank -50200048205797", "77,81,281"],
        ["HDFC Bank -50200048192293", "4,32,658"],
        ["Total", "4,06,78,904"]
    ]
    
    # Populate the HDFC KCC data
    for row_idx, row_data in enumerate(hdfc_data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_ca.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(hdfc_data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add a gap between tables
    row = len(hdfc_data) + 8
    
    # Add Payable for Fixed Assets section
    schedule_ca.cell(row=row, column=1).value = "Schedule : CA"
    schedule_ca.cell(row=row, column=1).font = bold_font
    
    row += 1
    schedule_ca.cell(row=row, column=1).value = "Payble for Fixed Assets"
    schedule_ca.cell(row=row, column=1).font = bold_font
    
    # Create table headers for Payable for Fixed Assets
    row += 1
    for col, header in enumerate(headers, start=1):
        cell = schedule_ca.cell(row=row, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the Payable for Fixed Assets data
    fixed_assets_data = [
        ["Mewar Hitech Engineering", "47,71,422"],
        ["Total", "47,71,422"]
    ]
    
    # Populate the Payable for Fixed Assets data
    for row_idx, row_data in enumerate(fixed_assets_data, start=row+1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_ca.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == row + len(fixed_assets_data):  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = row + len(fixed_assets_data) + 2
    schedule_ca.merge_cells(f'A{row}:B{row}')
    schedule_ca[f'A{row}'] = 'For: M/s MO Infra'
    schedule_ca[f'A{row}'].font = normal_font
    
    row += 3
    schedule_ca.merge_cells(f'A{row}:B{row}')
    schedule_ca[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_ca[f'A{row}'].font = normal_font
    
    row += 1
    schedule_ca.merge_cells(f'A{row}:B{row}')
    schedule_ca[f'A{row}'] = '(Proprietor)'
    schedule_ca[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_d(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule D - Trade Payables in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_d = wb.active
        schedule_d.title = "Schedule D - Trade Payables"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule D - Trade Payables" in wb.sheetnames:
            schedule_d = wb["Schedule D - Trade Payables"]
        else:
            schedule_d = wb.create_sheet("Schedule D - Trade Payables")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_d.column_dimensions['A'].width = 50
    schedule_d.column_dimensions['B'].width = 20
    
    # Add header
    schedule_d['A1'] = "M/S MO Infra"
    schedule_d['A1'].font = header_font
    
    schedule_d['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_d['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_d['A4'] = "Schedule : D"
    schedule_d['A4'].font = bold_font
    
    # Add Trade Payables header
    schedule_d['A5'] = "4.Trade Payables"
    schedule_d['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_d.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Bade Baba Mining, Nagpur", "5,73,979"],
        ["EKTA TRADERS", "14,78,219"],
        ["Geetanjali Construction", "1,84,71,228"],
        ["GP 11 INDUSTRIES PRIVATE LIMITED", "8,16,25,413"],
        ["Hasim Khan", "1,60,749"],
        ["K.K. Shah", "1,25,000"],
        ["K.N. Developers and Buildcon", "50,000"],
        ["Madhya Bharat Machinery Stores", "1,700"],
        ["M O Industries", "6,34,26,246"],
        ["Net Planet Computers", "79,000"],
        ["Rk Enterprise", "1,58,151"],
        ["Sagar Automobiles", "1,48,620"],
        ["S.A. Infrastructure", "9,11,126"],
        ["S.B.M Traderes", "28,29,926"],
        ["Sensotech Weighting System Pvt Ltd", "51,200"],
        ["Shubham Tyres", "84,000"],
        ["SML Industries Private Limited", "1,21,37,804"],
        ["SMO Ferro Alloys Pvt Ltd", "9,85,211"],
        ["Tulsi Associates", "35,01,082"],
        ["V.R. Construction", "82,750"],
        ["Total", "18,68,81,405"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_d.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_d.merge_cells(f'A{row}:B{row}')
    schedule_d[f'A{row}'] = 'For: M/s MO Infra'
    schedule_d[f'A{row}'].font = normal_font
    
    row += 3
    schedule_d.merge_cells(f'A{row}:B{row}')
    schedule_d[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_d[f'A{row}'].font = normal_font
    
    row += 1
    schedule_d.merge_cells(f'A{row}:B{row}')
    schedule_d[f'A{row}'] = '(Proprietor)'
    schedule_d[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_e(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule E - Provisions in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_e = wb.active
        schedule_e.title = "Schedule E - Provisions"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule E - Provisions" in wb.sheetnames:
            schedule_e = wb["Schedule E - Provisions"]
        else:
            schedule_e = wb.create_sheet("Schedule E - Provisions")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_e.column_dimensions['A'].width = 50
    schedule_e.column_dimensions['B'].width = 20
    
    # Add header
    schedule_e['A1'] = "M/S MO Infra"
    schedule_e['A1'].font = header_font
    
    schedule_e['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_e['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_e['A4'] = "Schedule : E"
    schedule_e['A4'].font = bold_font
    
    # Add Provisions header
    schedule_e['A5'] = "5.Provisions"
    schedule_e['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_e.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Audit fees payble", "50,000"],
        ["Salary", "4,00,000"],
        ["TDS and TCS Payable", "2,11,067"],
        ["Total", "6,61,067"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_e.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_e.merge_cells(f'A{row}:B{row}')
    schedule_e[f'A{row}'] = 'For: M/s MO Infra'
    schedule_e[f'A{row}'].font = normal_font
    
    row += 3
    schedule_e.merge_cells(f'A{row}:B{row}')
    schedule_e[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_e[f'A{row}'].font = normal_font
    
    row += 1
    schedule_e.merge_cells(f'A{row}:B{row}')
    schedule_e[f'A{row}'] = '(Proprietor)'
    schedule_e[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_f(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule F - Advance against Sale of Land in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_f = wb.active
        schedule_f.title = "Schedule F - Advance against Sale of Land"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule F - Advance against Sale of Land" in wb.sheetnames:
            schedule_f = wb["Schedule F - Advance against Sale of Land"]
        else:
            schedule_f = wb.create_sheet("Schedule F - Advance against Sale of Land")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_f.column_dimensions['A'].width = 50
    schedule_f.column_dimensions['B'].width = 20
    
    # Add header
    schedule_f['A1'] = "M/S MO Infra"
    schedule_f['A1'].font = header_font
    
    schedule_f['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_f['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_f['A4'] = "Schedule : F"
    schedule_f['A4'].font = bold_font
    
    # Add Advance against Sale of Land header
    schedule_f['A5'] = "6.Advance against Sale of Land"
    schedule_f['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_f.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Vijay Katariya", "6,00,000"],
        ["Total", "6,00,000"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_f.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_f.merge_cells(f'A{row}:B{row}')
    schedule_f[f'A{row}'] = 'For: M/s MO Infra'
    schedule_f[f'A{row}'].font = normal_font
    
    row += 3
    schedule_f.merge_cells(f'A{row}:B{row}')
    schedule_f[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_f[f'A{row}'].font = normal_font
    
    row += 1
    schedule_f.merge_cells(f'A{row}:B{row}')
    schedule_f[f'A{row}'] = '(Proprietor)'
    schedule_f[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_g(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule G - Advance against Sale in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_g = wb.active
        schedule_g.title = "Schedule G - Advance against Sale"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule G - Advance against Sale" in wb.sheetnames:
            schedule_g = wb["Schedule G - Advance against Sale"]
        else:
            schedule_g = wb.create_sheet("Schedule G - Advance against Sale")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_g.column_dimensions['A'].width = 50
    schedule_g.column_dimensions['B'].width = 20
    
    # Add header
    schedule_g['A1'] = "M/S MO Infra"
    schedule_g['A1'].font = header_font
    
    schedule_g['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_g['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_g['A4'] = "Schedule : G"
    schedule_g['A4'].font = bold_font
    
    # Add Advance against Sale header
    schedule_g['A5'] = "7.Advance against Sale"
    schedule_g['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_g.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Raninga Ispat Pvt Ltd", "2,00,00,000"],
        ["Sarpanch Gram panchayat Rampuriya", "36,617"],
        ["Hanuman singh Mehta", "60,000"],
        ["Shree Ram Metals Unit-1", "11,750"],
        ["Total", "2,01,08,367"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_g.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_g.merge_cells(f'A{row}:B{row}')
    schedule_g[f'A{row}'] = 'For: M/s MO Infra'
    schedule_g[f'A{row}'].font = normal_font
    
    row += 3
    schedule_g.merge_cells(f'A{row}:B{row}')
    schedule_g[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_g[f'A{row}'].font = normal_font
    
    row += 1
    schedule_g.merge_cells(f'A{row}:B{row}')
    schedule_g[f'A{row}'] = '(Proprietor)'
    schedule_g[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_h(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule H - Investment in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_h = wb.active
        schedule_h.title = "Schedule H - Investment"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule H - Investment" in wb.sheetnames:
            schedule_h = wb["Schedule H - Investment"]
        else:
            schedule_h = wb.create_sheet("Schedule H - Investment")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_h.column_dimensions['A'].width = 50
    schedule_h.column_dimensions['B'].width = 20
    
    # Add header
    schedule_h['A1'] = "M/S MO Infra"
    schedule_h['A1'].font = header_font
    
    schedule_h['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_h['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_h['A4'] = "Schedule : H"
    schedule_h['A4'].font = bold_font
    
    # Add Investment header
    schedule_h['A5'] = "8.Investment"
    schedule_h['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_h.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["M/s Saiyyed Akhtar Ali (Firm)", "19,57,70,472"],
        ["SMO Aluminium Private Limited (Shares)", ""],
        ["SMO Copper smelting Private Limited (shares)", ""],
        ["SMO Ferro Alloys Pvt Ltd(274500 Shares)", "13,45,50,000"],
        ["SMO Fragrance Distillaery Private Limited (Shares)", ""],
        ["SMO Metals & Energy Private Limited (Shares)", ""],
        ["Owais Metal and Mineral Processing Ltd", "17,00,000"],
        ["Gold", "30,21,378"],
        ["Total", "33,50,41,850"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_h.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_h.merge_cells(f'A{row}:B{row}')
    schedule_h[f'A{row}'] = 'For: M/s MO Infra'
    schedule_h[f'A{row}'].font = normal_font
    
    row += 3
    schedule_h.merge_cells(f'A{row}:B{row}')
    schedule_h[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_h[f'A{row}'].font = normal_font
    
    row += 1
    schedule_h.merge_cells(f'A{row}:B{row}')
    schedule_h[f'A{row}'] = '(Proprietor)'
    schedule_h[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_i(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule I - Closing Stock in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_i = wb.active
        schedule_i.title = "Schedule I - Closing Stock"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule I - Closing Stock" in wb.sheetnames:
            schedule_i = wb["Schedule I - Closing Stock"]
        else:
            schedule_i = wb.create_sheet("Schedule I - Closing Stock")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_i.column_dimensions['A'].width = 50
    schedule_i.column_dimensions['B'].width = 20
    
    # Add header
    schedule_i['A1'] = "M/S MO Infra"
    schedule_i['A1'].font = header_font
    
    schedule_i['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_i['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_i['A4'] = "Schedule : I"
    schedule_i['A4'].font = bold_font
    
    # Add Closing Stock header
    schedule_i['A5'] = "9.Closing Stock"
    schedule_i['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_i.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Tin Slag", "11,26,526"],
        ["Gitty", "23,99,388"],
        ["Ferro Silicon", "1,96,34,694"],
        ["Ferro Silico Manganese", "5,34,19,516"],
        ["Ferro Silico Manganese Slag Touch", "2,21,628"],
        ["Ferro Silicon off Grade", "3,42,974"],
        ["Manganese Ore", "3,600"],
        ["Midium Carbon Ferro Manganese", "64,97,750"],
        ["Sand", "68,62,950"],
        ["Other Consumables", "1,28,125"],
        ["Total", "9,06,37,150"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_i.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_i.merge_cells(f'A{row}:B{row}')
    schedule_i[f'A{row}'] = 'For: M/s MO Infra'
    schedule_i[f'A{row}'].font = normal_font
    
    row += 3
    schedule_i.merge_cells(f'A{row}:B{row}')
    schedule_i[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_i[f'A{row}'].font = normal_font
    
    row += 1
    schedule_i.merge_cells(f'A{row}:B{row}')
    schedule_i[f'A{row}'] = '(Proprietor)'
    schedule_i[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_j(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule J - Trade Receivable in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_j = wb.active
        schedule_j.title = "Schedule J - Trade Receivable"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule J - Trade Receivable" in wb.sheetnames:
            schedule_j = wb["Schedule J - Trade Receivable"]
        else:
            schedule_j = wb.create_sheet("Schedule J - Trade Receivable")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_j.column_dimensions['A'].width = 50
    schedule_j.column_dimensions['B'].width = 20
    
    # Add header
    schedule_j['A1'] = "M/S MO Infra"
    schedule_j['A1'].font = header_font
    
    schedule_j['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_j['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_j['A4'] = "Schedule : J"
    schedule_j['A4'].font = bold_font
    
    # Add Trade Receivable header
    schedule_j['A5'] = "10.Trade Receivable"
    schedule_j['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_j.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Ishan Precast Compound Wall", "2,19,483"],
        ["JAI AMBE ALLOYS", "64,730"],
        ["Lohum Cleantech Private Limited", "24,78,000"],
        ["Shri Madhav International, Vadodra", "56,871"],
        ["Ar Ferro Alloys", "10,82,489"],
        ["Ekta Traders (Asgar Ali)", "1,96,330"],
        ["Gitanjali Construction Hub Private Limited", "9,06,88,907"],
        ["Krati Construction Co", "79,769"],
        ["Mahavideh Traders", "28,600"],
        ["Rehmat Ali & Sons (Debtors)", "15,84,732"],
        ["Sarpanch Grampanchayat Narela", "430"],
        ["Total", "9,64,80,340"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_j.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_j.merge_cells(f'A{row}:B{row}')
    schedule_j[f'A{row}'] = 'For: M/s MO Infra'
    schedule_j[f'A{row}'].font = normal_font
    
    row += 3
    schedule_j.merge_cells(f'A{row}:B{row}')
    schedule_j[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_j[f'A{row}'].font = normal_font
    
    row += 1
    schedule_j.merge_cells(f'A{row}:B{row}')
    schedule_j[f'A{row}'] = '(Proprietor)'
    schedule_j[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_k(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule K - Loans and Advances (Assets) in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_k = wb.active
        schedule_k.title = "Schedule K - Loans and Advances"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule K - Loans and Advances" in wb.sheetnames:
            schedule_k = wb["Schedule K - Loans and Advances"]
        else:
            schedule_k = wb.create_sheet("Schedule K - Loans and Advances")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_k.column_dimensions['A'].width = 50
    schedule_k.column_dimensions['B'].width = 20
    
    # Add header
    schedule_k['A1'] = "M/S MO Infra"
    schedule_k['A1'].font = header_font
    
    schedule_k['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_k['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_k['A4'] = "Schedule : K"
    schedule_k['A4'].font = bold_font
    
    # Add Loans and Advances header
    schedule_k['A5'] = "11. Loans and Advances (Assets)"
    schedule_k['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_k.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Abhishek Sharma", "32,28,435"],
        ["Anash Ali", "15,00,000"],
        ["Archana Singh", "1,50,000"],
        ["Jabir Hussain", "24,10,071"],
        ["Mirza", "5,00,000"],
        ["Mohammad Shafi Qureshi", "23,50,000"],
        ["Mr. Noor Ali", "86,800"],
        ["Murtuza Ali Loan", "5,33,000"],
        ["Nahid Ali", "1,03,134"],
        ["Owais Metal & Mineral Processing Ltd ( Un)", "14,000"],
        ["Saiyyed Murutaza Ali", "17,60,814"],
        ["SBM TRADERS(Unsecured Loan)", "2,16,076"],
        ["Shokhat Ansari", "55,00,000"],
        ["SMO FERRO ALLOYS PVT LTD(Loans & Advances)", "73,94,808"],
        ["SMO Gold and Refinary Pvt Ltd", "5,00,000"],
        ["SMO Metal & Energy Private Limited", "2,34,00,000"],
        ["S.R. Ferro Alloys", "46,007"],
        ["Tulsi Associates (Loan)", "1,89,074"],
        ["Padam Kumar Vora", "41,600"],
        ["Rajendra Singh Nayak", "25,54,750"],
        ["Sanjay Bareth", "87,700"],
        ["Smo Ferro Alloys Pvt Ltd (OXY)", "57,000"],
        ["Total", "5,26,23,269"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_k.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_k.merge_cells(f'A{row}:B{row}')
    schedule_k[f'A{row}'] = 'For: M/s MO Infra'
    schedule_k[f'A{row}'].font = normal_font
    
    row += 3
    schedule_k.merge_cells(f'A{row}:B{row}')
    schedule_k[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_k[f'A{row}'].font = normal_font
    
    row += 1
    schedule_k.merge_cells(f'A{row}:B{row}')
    schedule_k[f'A{row}'] = '(Proprietor)'
    schedule_k[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_l(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule L - Advances for Material Purchase in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_l = wb.active
        schedule_l.title = "Schedule L - Advances for Material Purchase"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule L - Advances for Material Purchase" in wb.sheetnames:
            schedule_l = wb["Schedule L - Advances for Material Purchase"]
        else:
            schedule_l = wb.create_sheet("Schedule L - Advances for Material Purchase")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_l.column_dimensions['A'].width = 50
    schedule_l.column_dimensions['B'].width = 20
    
    # Add header
    schedule_l['A1'] = "M/S MO Infra"
    schedule_l['A1'].font = header_font
    
    schedule_l['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_l['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_l['A4'] = "Schedule : L"
    schedule_l['A4'].font = bold_font
    
    # Add Advances for Material Purchase header
    schedule_l['A5'] = "12.Advances for Material Purchase"
    schedule_l['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_l.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["MO Infra Unit-1", "2,39,50,725"],
        ["M/s Saiyyed Akhtar Ali", "1,22,135"],
        ["Saundarya Sagar", "2,35,000"],
        ["Total", "2,43,07,860"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_l.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_l.merge_cells(f'A{row}:B{row}')
    schedule_l[f'A{row}'] = 'For: M/s MO Infra'
    schedule_l[f'A{row}'].font = normal_font
    
    row += 3
    schedule_l.merge_cells(f'A{row}:B{row}')
    schedule_l[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_l[f'A{row}'].font = normal_font
    
    row += 1
    schedule_l.merge_cells(f'A{row}:B{row}')
    schedule_l[f'A{row}'] = '(Proprietor)'
    schedule_l[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_m(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule M - Deposits in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_m = wb.active
        schedule_m.title = "Schedule M - Deposits"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule M - Deposits" in wb.sheetnames:
            schedule_m = wb["Schedule M - Deposits"]
        else:
            schedule_m = wb.create_sheet("Schedule M - Deposits")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_m.column_dimensions['A'].width = 50
    schedule_m.column_dimensions['B'].width = 20
    
    # Add header
    schedule_m['A1'] = "M/S MO Infra"
    schedule_m['A1'].font = header_font
    
    schedule_m['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_m['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_m['A4'] = "Schedule : M"
    schedule_m['A4'].font = bold_font
    
    # Add Deposits header
    schedule_m['A5'] = "13.Deposits"
    schedule_m['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_m.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["FD at HDFC Bank", "60,70,708"],
        ["FD at Punjab and Sind Bank", "60,309"],
        ["FDR", "1,62,862"],
        ["MPKVVCL Indore", "56,606"],
        ["Gold Bond", "46,380"],
        ["Total", "63,96,865"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_m.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_m.merge_cells(f'A{row}:B{row}')
    schedule_m[f'A{row}'] = 'For: M/s MO Infra'
    schedule_m[f'A{row}'].font = normal_font
    
    row += 3
    schedule_m.merge_cells(f'A{row}:B{row}')
    schedule_m[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_m[f'A{row}'].font = normal_font
    
    row += 1
    schedule_m.merge_cells(f'A{row}:B{row}')
    schedule_m[f'A{row}'] = '(Proprietor)'
    schedule_m[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_n(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule N - Other Current Assets in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_n = wb.active
        schedule_n.title = "Schedule N - Other Current Assets"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule N - Other Current Assets" in wb.sheetnames:
            schedule_n = wb["Schedule N - Other Current Assets"]
        else:
            schedule_n = wb.create_sheet("Schedule N - Other Current Assets")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_n.column_dimensions['A'].width = 50
    schedule_n.column_dimensions['B'].width = 20
    
    # Add header
    schedule_n['A1'] = "M/S MO Infra"
    schedule_n['A1'].font = header_font
    
    schedule_n['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_n['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_n['A4'] = "Schedule : N"
    schedule_n['A4'].font = bold_font
    
    # Add Other Current Assets header
    schedule_n['A5'] = "14.Other Current Assets"
    schedule_n['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_n.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["GST", "76,98,248"],
        ["TDS Receivable", "6,47,576"],
        ["TCS Receivable", "84,139"],
        ["Total", "84,29,963"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_n.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_n.merge_cells(f'A{row}:B{row}')
    schedule_n[f'A{row}'] = 'For: M/s MO Infra'
    schedule_n[f'A{row}'].font = normal_font
    
    row += 3
    schedule_n.merge_cells(f'A{row}:B{row}')
    schedule_n[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_n[f'A{row}'].font = normal_font
    
    row += 1
    schedule_n.merge_cells(f'A{row}:B{row}')
    schedule_n[f'A{row}'] = '(Proprietor)'
    schedule_n[f'A{row}'].font = normal_font
    
    return wb

def create_schedule_o(workbook=None, output_file="M_S_MO_Infra_Financial_Statements.xlsx"):
    """
    Creates Schedule O - Cash and Bank in the provided workbook
    or creates a new workbook if none is provided.
    """
    if workbook is None:
        # Create a new workbook if none is provided
        wb = openpyxl.Workbook()
        schedule_o = wb.active
        schedule_o.title = "Schedule O - Cash and Bank"
    else:
        # Use the existing workbook and create/get the sheet
        wb = workbook
        if "Schedule O - Cash and Bank" in wb.sheetnames:
            schedule_o = wb["Schedule O - Cash and Bank"]
        else:
            schedule_o = wb.create_sheet("Schedule O - Cash and Bank")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths for better readability
    schedule_o.column_dimensions['A'].width = 50
    schedule_o.column_dimensions['B'].width = 20
    
    # Add header
    schedule_o['A1'] = "M/S MO Infra"
    schedule_o['A1'].font = header_font
    
    schedule_o['A2'] = "SCHEDULES FORMING PART OF BALANCE SHEET AS AT 31.03.2024"
    schedule_o['A2'].font = subheader_font
    
    # Add Schedule title
    schedule_o['A4'] = "15.Schedule : O"
    schedule_o['A4'].font = bold_font
    
    # Add Cash and Bank header
    schedule_o['A5'] = "Cash and Bank"
    schedule_o['A5'].font = bold_font
    
    # Create table headers
    headers = ['Particulars', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = schedule_o.cell(row=6, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add the data
    data = [
        ["Bank of India Saving A/c- 6819", "77,088"],
        ["Central Bank of India -3436", "1,19,634"],
        ["Punjab & Sind Bank-6169", "11,568"],
        ["HDFC Bank", "282"],
        ["Cash", "42,37,667"],
        ["Total", "44,46,238"]
    ]
    
    # Populate the data
    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = schedule_o.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # Format the Total row with bold
            if row_idx == len(data) + 6:  # Last row (Total)
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Right-align amount column
            if col_idx == 2:  # Amount column
                cell.alignment = Alignment(horizontal='right')
    
    # Add footer
    row = len(data) + 8
    schedule_o.merge_cells(f'A{row}:B{row}')
    schedule_o[f'A{row}'] = 'For: M/s MO Infra'
    schedule_o[f'A{row}'].font = normal_font
    
    row += 3
    schedule_o.merge_cells(f'A{row}:B{row}')
    schedule_o[f'A{row}'] = 'Mr. Saiyyed Akhtar Ali'
    schedule_o[f'A{row}'].font = normal_font
    
    row += 1
    schedule_o.merge_cells(f'A{row}:B{row}')
    schedule_o[f'A{row}'] = '(Proprietor)'
    schedule_o[f'A{row}'].font = normal_font
    
    return wb

if __name__ == "__main__":
    create_financial_statements_excel()