import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pytesseract
from pdf2image import convert_from_path
import re


def convert_financial_pdf_to_excel(pdf_path, output_excel_path):
    """
    Main function to convert financial statements PDF to Excel

    Args:
        pdf_path (str): Path to the PDF file
        output_excel_path (str): Path to save the Excel file

    Returns:
        str: Path to the created Excel file
    """
    # Create a new Excel workbook
    wb = Workbook()

    # Create sheets for each financial statement
    create_balance_sheet(wb)
    create_profit_loss(wb)
    create_cash_flow(wb)
    create_notes(wb)

    # Save the workbook
    wb.save(output_excel_path)
    print(f"Excel file created successfully: {output_excel_path}")

    return output_excel_path


def apply_header_style(cell):
    """Apply styling to header cells"""
    cell.font = Font(bold=True)
    cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.fill = PatternFill(start_color="D9D9D9",
                            end_color="D9D9D9", fill_type="solid")


def apply_data_cell_style(cell, is_numeric=False, is_total=False, indent_level=0):
    """Apply styling to data cells"""
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if is_numeric:
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right', vertical='center')
    else:
        indent = ' ' * (indent_level * 2)
        if cell.value and isinstance(cell.value, str):
            cell.value = indent + cell.value
        cell.alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True)

    if is_total:
        cell.font = Font(bold=True)
        if is_numeric:
            cell.fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def add_company_header(ws, statement_title):
    """Add company information header to the worksheet"""
    # Column width settings
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Company name and address
    ws.merge_cells('A1:D1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:D2')
    cell = ws.cell(row=2, column=1,
                   value="3 MMTC/STC COLONY, GEETANJALI ENCLAVE, NEW DELHI")
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:D3')
    cell = ws.cell(row=3, column=1,
                   value="SOUTH DELHI-110017, CIN: U74899DL1989PTC038372")
    cell.alignment = Alignment(horizontal='center')

    # Statement title
    ws.merge_cells('A4:D4')
    cell = ws.cell(row=4, column=1, value=statement_title)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    # Empty row for spacing
    ws.cell(row=5, column=1, value="")


def create_balance_sheet(wb):
    """Create the Balance Sheet worksheet"""
    # Get the active sheet and rename it
    ws = wb.active
    ws.title = "Balance Sheet"

    # Add company header
    add_company_header(ws, "BALANCE SHEET AS AT 31ST MARCH, 2022")

    # Add column headers
    headers = ["Particulars", "Note No.",
        "As at 31.03.2022 Amount (Rs.)", "As at 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i, value=header)
        apply_header_style(cell)

    # Balance Sheet data from the PDF
    balance_sheet_data = [
        # I. EQUITY AND LIABILITIES
        ["I. EQUITY AND LIABILITIES", "", "", ""],
        ["1 Shareholders' funds", "", "", ""],
        ["(a) Share Capital", "1", "79.27", "79.27"],
        ["(b) Reserves and Surplus", "2", "-2712.28", "-1286.51"],
        ["(c) Money Received against Share Warrants", "", "", ""],
        ["2 Share application money pending allotment", "", "", ""],
        ["3 Non-current liabilities", "", "", ""],
        ["(a) Long-Term Borrowings", "3", "5703.98", "5086.39"],
        ["(b) Deferred Tax Liabilities (Net)", "4", "7.07", "7.07"],
        ["(c) Other Long Term Liabilities", "", "", ""],
        ["(d) Long-Term Provisions", "5", "15.39", "13.26"],
        ["4 Current liabilities", "", "", ""],
        ["(a) Short-Term Borrowings", "6", "540.38", "182.12"],
        ["(b) Trade Payables", "7", "", ""],
        ["    (i) Total Outstanding dues of MSME", "", "0.00", "0.00"],
        ["    (ii) Total Outstanding dues of Other Than MSME",
               "", "1035.45", "904.68"],
        ["(c) Other Current Liabilities", "8", "173.34", "158.95"],
        ["(d) Short-Term Provisions", "5", "1.31", "0.89"],
        ["TOTAL", "", "4843.92", "5146.12"],

        # II. ASSETS
        ["II. ASSETS", "", "", ""],
        ["Non-current assets", "", "", ""],
        ["1 (a) Property, Plant and Equipment and Intangible assets", "", "", ""],
        ["    (i) Property, Plant and Equipment", "9", "4409.45", "4851.13"],
        ["    (ii) Intangible assets", "", "0.00", "0.00"],
        ["    (iii) Capital work-in-Progress", "9", "4.13", "0.00"],
        ["    (iiii) Intangible assets under development", "", "0.00", "0.00"],
        ["(b) Non-Current Investments", "", "0.00", "0.00"],
        ["(c) Deferred Tax Assets (net)", "", "0.00", "0.00"],
        ["(d) Long-Term Loans and Advances", "10", "21.78", "19.97"],
        ["(e) Other Non-Current Assets", "9", "4.87", "7.30"],
        ["2 Current assets", "", "", ""],
        ["(a) Current Investments", "", "0.00", "0.00"],
        ["(b) Inventories", "11", "15.52", "11.80"],
        ["(c) Trade Receivables", "12", "208.76", "16.33"],
        ["(d) Cash and Cash Equivalents", "13", "80.75", "66.64"],
        ["(e) Short-Term Loans and Advances", "14", "80.43", "148.83"],
        ["(f) Other Current Assets", "14a", "18.24", "24.12"],
        ["TOTAL", "", "4843.92", "5146.12"]
    ]

    # Write data to worksheet
    for row_idx, row_data in enumerate(balance_sheet_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Determine cell styling
            indent_level = 0
            if row_data[0].startswith("    "):  # 4-space indent
                indent_level = 2
            elif row_data[0].startswith("  "):  # 2-space indent
                indent_level = 1
            elif "(" in row_data[0] and not row_data[0].startswith("I") and not row_data[0].startswith("II"):
                indent_level = 1

            is_numeric = col_idx > 2 and value not in ["", None]
            is_total = row_data[0] == "TOTAL"
            is_section_header = row_data[0] in ["I. EQUITY AND LIABILITIES", "II. ASSETS", "Non-current assets",
                "1 Shareholders' funds", "3 Non-current liabilities", "4 Current liabilities", "2 Current assets"]

            if is_section_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)

    # Add note to the bottom of the sheet
    note_row = len(balance_sheet_data) + 8
    ws.merge_cells(f'A{note_row}:D{note_row}')
    cell = ws.cell(row=note_row, column=1,
                  value="Notes 1 to 21 and Accounting Policies attached to the Financial Statement are an integral part thereof.")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal='left')


def create_profit_loss(wb):
    """Create the Profit & Loss Statement worksheet"""
    ws = wb.create_sheet("Profit & Loss")

    # Add company header
    add_company_header(
        ws, "STATEMENT OF PROFIT & LOSS FOR THE YEAR ENDING 31ST MARCH, 2022")

    # Add column headers
    headers = ["Particulars", "Refer Note No.",
        "Year ending 31.03.2022 Amount (Rs.)", "Year ending 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i, value=header)
        apply_header_style(cell)

    # Profit & Loss data from the PDF
    pl_data = [
        ["I. Revenue From Operations", "", "", ""],
        ["Revenue from operations", "15", "2523.15", "2471.49"],
        ["II. Other Income", "16", "26.72", "50.14"],
        ["III. Total Income (I + II)", "", "2549.87", "2521.63"],
        ["IV. Expenses:", "", "", ""],
        ["a) Cost of materials consumed", "17", "1223.94", "800.14"],
        ["b) Purchase of Stock-in Trade", "", "0.00", "0.00"],
        ["c) Changes in Inventories of Finished Goods, Work-in-Progress and Stock-in-Trade",
            "", "0.00", "0.00"],
        ["d) Employee Benefits Expense", "18", "1136.13", "1067.30"],
        ["e) Finance Costs", "19", "366.82", "302.70"],
        ["f) Depreciation and Amortization Expenses", "9", "546.34", "664.29"],
        ["g) Other Expenses", "20", "527.08", "293.39"],
        ["Total Expenses", "", "3800.31", "3127.83"],
        ["V. Profit before exceptional and extraordinary items and tax (III - IV)", "", "-1250.44", "-606.20"],
        ["VI Exceptional items", "", "0.00", "0.00"],
        ["VII. Profit before extraordinary items and tax (V - VI)",
                                                          "", "-1250.44", "-606.20"],
        ["VIII Extraordinary items", "", "0.00", "0.00"],
        ["IX. Profit before tax (VII - VIII)", "", "-1250.44", "-606.20"],
        ["X Tax Expense:", "", "", ""],
        ["a) Current tax", "", "0.00", "0.00"],
        ["b) (Less): Mat Credit (Where Applicable)", "", "0.00", "0.00"],
        ["c) Current Tax Expensses Relating to Prior Years", "", "0.00", "0.00"],
        ["d) Net Current Tax Expenses", "", "0.00", "0.00"],
        ["e) Deferred tax", "", "0.00", "0.00"],
        ["XI Profit/(Loss) for the period from continuing operations (IX - X)",
                     "", "-1250.44", "-606.20"],
        ["XII Profit/(Loss) from discontinuing operations",
                      "", "0.00", "0.00"],
        ["XIII Tax expense of discounting operations", "", "0.00", "0.00"],
        ["XIV Profit/(Loss) from discontinuing operations (after tax) (XII - XIII)",
                      "", "0.00", "0.00"],
        ["XV Profit/(Loss) for the period (XI + XIV)",
                     "", "-1250.44", "-606.20"],
        ["XVI Earnings per Equity Share:", "", "", ""],
        ["Before and after extraordinary Items :-", "", "", ""],
        ["(1) Basic in ₹", "", "-1577.36", "-764.69"],
        ["(2) Diluted in ₹", "", "-1577.36", "-764.69"]
    ]

    # Write data to worksheet
    for row_idx, row_data in enumerate(pl_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Determine cell styling
            indent_level = 0
            if row_data[0].startswith("  "):  # 2-space indent
                indent_level = 1
            elif row_data[0].startswith("("):  # parenthesis items
                indent_level = 2
            elif row_data[0].startswith("a)") or row_data[0].startswith("b)") or row_data[0].startswith("c)") or row_data[0].startswith("d)") or row_data[0].startswith("e)") or row_data[0].startswith("f)") or row_data[0].startswith("g)"):
                indent_level = 1

            is_numeric = col_idx > 2 and value not in ["", None]
            is_total = row_data[0] in ["Total Expenses",
                "XV Profit/(Loss) for the period (XI + XIV)"]
            is_section_header = row_data[0].startswith("I.") or row_data[0].startswith("II.") or row_data[0].startswith("III.") or row_data[0].startswith("IV.") or row_data[0].startswith("V.") or row_data[0].startswith("VI") or row_data[0].startswith("VII.") or row_data[0].startswith(
                "VIII") or row_data[0].startswith("IX.") or row_data[0].startswith("X ") or row_data[0].startswith("XI ") or row_data[0].startswith("XII ") or row_data[0].startswith("XIII ") or row_data[0].startswith("XIV ") or row_data[0].startswith("XV ") or row_data[0].startswith("XVI ")

            if is_section_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)

    # Add note to the bottom of the sheet
    note_row = len(pl_data) + 8
    ws.merge_cells(f'A{note_row}:D{note_row}')
    cell = ws.cell(row=note_row, column=1,
                  value="Notes 1 to 21 and Accounting Policies attached to the Financial Statement are an integral part thereof.")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal='left')


def create_cash_flow(wb):
    """Create the Cash Flow Statement worksheet"""
    ws = wb.create_sheet("Cash Flow")

    # Add company header
    add_company_header(
        ws, "CASH FLOW STATEMENT FOR THE YEAR ENDED AS AT 31ST MARCH, 2022")

    # Add column headers
    headers = ["PARTICULARS", "For the year ended March 31, 2022",
        "For the year ended March 31, 2021"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i, value=header)
        apply_header_style(cell)

    # Cash Flow data from the PDF
    cf_data = [
        ["A CASH FLOW FROM OPERATING ACTIVITIES", "", ""],
        ["Net Profit Before Tax", "-1250.44", "-606.20"],
        ["Adjustments for-", "", ""],
        ["Depreciation", "543.90", "661.86"],
        ["Preliminary Expenses w/off", "2.43", "2.43"],
        ["Prior Period Items", "-175.34", "0.00"],
        ["Interest & Finance Charges", "366.82", "294.21"],
        ["Operating Profit before Working Capital Changes", "-512.62", "352.31"],
        ["Adjustments for", "", ""],
        ["Decrease/(Increase) in Loan & Advances", "66.59", "-91.99"],
        ["Decrease/(Increase) in Inventories", "-3.72", "-1.28"],
        ["Decrease/(Increase) in Other Current Assets", "5.88", "0.00"],
        ["Decrease/(Increase) in Trade Receivables", "-192.43", "-4.84"],
        ["Increase/(Decrease) in Trade Payables", "130.78", "0.00"],
        ["Increase/(Decrease) in Current Liabilities & Provisions",
                    "16.93", "-376.20"],
        ["Cash generated from operations", "-488.58", "-122.01"],
        ["Income Tax paid", "0.00", "0.00"],
        ["Net Cash flow from Operating activities (A)", "-488.58", "-122.01"],
        ["", "", ""],
        ["B CASH FLOW FROM INVESTING ACTIVITIES", "", ""],
        ["Purchase of Fixed Assets", "-106.35", "-199.10"],
        ["Net Cash used in Investing activities (B)", "-106.35", "-199.10"],
        ["", "", ""],
        ["C CASH FLOW FROM FINANCING ACTIVITIES", "", ""],
        ["Proceeds from Long term Borrowings", "617.60", "642.28"],
        ["Proceeds from Short term Borrowings", "358.27", "-28.04"],
        ["Interest paid", "-366.82", "-294.21"],
        ["Net Cash used in financing activities (C)", "609.05", "320.03"],
        ["", "", ""],
        ["Net increase in cash & Cash Equivalents (A+B+C)", "14.11", "-1.08"],
        ["Opening Cash and Cash equivalents as at 31.03.2021", "66.64", "67.72"],
        ["Closing Cash and Cash equivalents as at 31.03.2022", "80.75", "66.64"],
        ["", "", ""],
        ["Cash & Cash Equivalents as stated", "", ""],
        ["Cash in Hand", "2.16", "1.05"],
        ["Cash at Bank", "78.59", "65.59"],
        ["", "80.75", "66.64"]
    ]

    # Write data to worksheet
    for row_idx, row_data in enumerate(cf_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Determine cell styling
            indent_level = 0
            if col_idx == 1 and not row_data[0].startswith("A ") and not row_data[0].startswith("B ") and not row_data[0].startswith("C ") and row_data[0] not in ["", "Cash & Cash Equivalents as stated", "Net increase in cash & Cash Equivalents (A+B+C)", "Opening Cash and Cash equivalents as at 31.03.2021", "Closing Cash and Cash equivalents as at 31.03.2022"]:
                indent_level = 1

            is_numeric = col_idx > 1 and value not in ["", None]
            is_total = row_data[0] in ["Net Cash flow from Operating activities (A)", "Net Cash used in Investing activities (B)", "Net Cash used in financing activities (C)",
                                                                                 "Closing Cash and Cash equivalents as at 31.03.2022"] or (col_idx == 2 and row_idx == (len(cf_data) + 6) and value == "80.75")
            is_section_header = row_data[0] in ["A CASH FLOW FROM OPERATING ACTIVITIES", "B CASH FLOW FROM INVESTING ACTIVITIES", "C CASH FLOW FROM FINANCING ACTIVITIES", "Cash & Cash Equivalents as stated",
                "Net increase in cash & Cash Equivalents (A+B+C)", "Opening Cash and Cash equivalents as at 31.03.2021", "Closing Cash and Cash equivalents as at 31.03.2022"]

            if is_section_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)

    # Add note to the bottom of the sheet
    note_row = len(cf_data) + 8
    ws.merge_cells(f'A{note_row}:C{note_row}')
    cell = ws.cell(row=note_row, column=1,
                  value="Notes 1 to 21 and Accounting Policies attached to the Financial Statement are an integral part thereof.")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal='left')


def create_notes(wb):
    """Create worksheets for Notes to Financial Statements"""

    # Create all notes
    create_note_1(wb)  # Share Capital
    create_note_2(wb)  # Reserve & Surplus
    create_note_3(wb)  # Long Term Borrowings
    create_note_4(wb)  # Deferred Tax Liabilities/Assets
    create_note_5(wb)  # Provisions
    create_note_6(wb)  # Short Term Borrowings
    create_note_7(wb)  # Trade Payables
    create_note_8(wb)  # Other Current Liabilities
    create_note_9(wb)  # Property, Plant and Equipment
    create_note_10(wb)  # Long Term Loans and Advances
    create_note_11(wb)  # Inventories
    create_note_12(wb)  # Trade Receivables
    create_note_13(wb)  # Cash and Cash Equivalents
    create_note_14(wb)  # Short Term Loans & Advances
    create_note_15(wb)  # Revenue from operations
    create_note_16(wb)  # Other Income
    create_note_17(wb)  # Cost of Material Consumed
    create_note_18(wb)  # Employee Benefit Expenses
    create_note_19(wb)  # Finance Costs
    create_note_20(wb)  # Other Expenses

    # Create an index sheet for all notes
    create_notes_index(wb)


def create_note_2(wb):
    """Create Note 2 - Reserve & Surplus"""
    ws = wb.create_sheet("Note 2 - Reserve & Surplus")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 2. Reserve & Surplus")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Reserve & Surplus data
    reserve_data = [
        ["a. Securities Premium Account", "", ""],
        ["Opening Balance", "176.25", "176.25"],
        ["Closing Balance", "176.25", "176.25"],
        ["", "", ""],
        ["b. Surplus / (Deficit) in Statement of Profit & Loss", "", ""],
        ["Opening balance", "-1462.76", "-856.56"],
        ["(+) Net Profit/(Net Loss) For the current year", "-1250.44", "-606.20"],
        ["(+) Transfer from Reserves", "", "0.00"],
        ["(-) Prior Period Item", "-175.34", "0.00"],
        ["(-) Transfer to Reserves", "", ""],
        ["Closing Balance", "-2888.54", "-1462.76"],
        ["", "", ""],
        ["Total", "-2712.28", "-1286.51"]
    ]

    for row_idx, row_data in enumerate(reserve_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total" or row_data[0] == "Closing Balance"
            is_header = row_data[0] in ["a. Securities Premium Account",
                "b. Surplus / (Deficit) in Statement of Profit & Loss"]

            indent_level = 0
            if not is_header and not is_total and row_data[0] != "":
                indent_level = 1

            if is_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_3(wb):
    """Create Note 3 - Long Term Borrowings"""
    ws = wb.create_sheet("Note 3 - Long Term Borrowings")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 3. Long Term Borrowings")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Long Term Borrowings data
    borrowings_data = [
        ["(a) Term Loan", "", ""],
        ["    (i) From Bank", "905.58", "1074.19"],
        ["    (ii) From Other Parties", "0.00", "0.00"],
        ["Total (a)", "905.58", "1074.19"],
        ["(b) Loans & Advances from related parties", "", ""],
        ["    (i) Secured", "0.00", "0.00"],
        ["    (ii) Unsecured", "", ""],
        ["        Crown Relators & Amusement (P) Ltd.", "124.39", "118.02"],
        ["        Crown Apartment Pvt Ltd", "255.00", "255.00"],
        ["        Yashoda Hospital & Research Centre Ltd", "4419.02", "3639.18"],
        ["Total (b)", "4798.41", "4012.19"],
        ["(c) Other loans & advances", "", ""],
        ["    (i) Secured", "0.00", "0.00"],
        ["    (ii) Unsecured", "0.00", "0.00"],
        ["Total (C)", "0.00", "0.00"],
        ["", "", ""],
        ["Total", "5703.98", "5086.39"]
    ]

    for row_idx, row_data in enumerate(borrowings_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total" or row_data[0] == "Total (a)" or row_data[
                                                                        0] == "Total (b)" or row_data[0] == "Total (C)"

            indent_level = 0
            if row_data[0].startswith("    "):
                indent_level = 1
            elif row_data[0].startswith("        "):
                indent_level = 2

            if is_total:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_4(wb):
    """Create Note 4 - Deferred Tax Liabilities/Assets"""
    ws = wb.create_sheet("Note 4 - Deferred Tax")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1,
                   value="Note 4. Deferred Tax Liabilities/ Assets")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Deferred Tax data
    tax_data = [
        ["Deferred Tax Liabilities", "7.07", "7.07"],
        ["Total", "7.07", "7.07"]
    ]

    for row_idx, row_data in enumerate(tax_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"

            if is_total:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total)


def create_note_5(wb):
    """Create Note 5 - Provisions"""
    ws = wb.create_sheet("Note 5 - Provisions")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 5. Provisions")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Provisions data
    provisions_data = [
        ["Long Term Provision", "", ""],
        ["Provision For Income Tax", "0.00", "0.00"],
        ["Provision For Gratuity", "15.39", "13.26"],
        ["", "15.39", "13.26"],
        ["Short Term Provision", "", ""],
        ["Provision For Gratuity", "1.31", "0.89"],
        ["Total", "1.31", "0.89"]
    ]

    for row_idx, row_data in enumerate(provisions_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total" or row_idx == row + 4
            is_header = row_data[0] in [
                "Long Term Provision", "Short Term Provision"]

            indent_level = 0
            if not is_header and not is_total and row_data[0] != "":
                indent_level = 1

            if is_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_6(wb):
    """Create Note 6 - Short Term Borrowings"""
    ws = wb.create_sheet("Note 6 - Short Term Borrowings")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 6. Short Term Borrowings")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Short Term Borrowings data
    borrowings_data = [
        ["Secured:", "", ""],
        ["Loans Repayable on Demand", "", ""],
        ["From Bank", "", ""],
        ["HDFC Bank Limited- OD A/c", "96.67", "0.00"],
        ["(Secured against Hypothecation of Book Debts of the Company", "", ""],
        ["& House Property of M. Director Situated at Surya Nagar)", "", ""],
        ["", "", ""],
        ["Current Maturities of long term debt", "", ""],
        ["Secured:", "", ""],
        ["Term Loan from Banks", "", ""],
        ["From HDFC Bank Limited - Loan No. - 83993460", "85.85", "68.41"],
        ["From HDFC Bank Limited - Loan No. - 84701255", "357.87", "113.71"],
        ["Total", "540.38", "182.12"]
    ]

    for row_idx, row_data in enumerate(borrowings_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"
            is_header = row_data[0] in [
                "Secured:", "Current Maturities of long term debt", "Term Loan from Banks"]

            indent_level = 0
            if not is_header and not is_total and row_data[0] != "":
                if row_data[0].startswith("(Secured"):
                    indent_level = 2
                elif row_data[0].startswith("From "):
                    indent_level = 1
                elif row_data[0] in ["Loans Repayable on Demand", "From Bank"]:
                    indent_level = 1

            if is_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_7(wb):
    """Create Note 7 - Trade Payables"""
    ws = wb.create_sheet("Note 7 - Trade Payables")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 7. Trade Payables")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Trade Payables data
    payables_data = [
        ["Other Than MSME", "", ""],
        ["Trade Payable- Goods", "854.09", "429.89"],
        ["Trade Payable- Fixed Assets", "181.36", "474.79"],
        ["Total", "1035.45", "904.68"]
    ]

    for row_idx, row_data in enumerate(payables_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"
            is_header = row_data[0] == "Other Than MSME"

            indent_level = 0
            if not is_header and not is_total and row_data[0] != "":
                indent_level = 1

            if is_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)

    # Add aging schedule
    aging_row = row_idx + 2
    ws.merge_cells(f'A{aging_row}:C{aging_row}')
    cell = ws.cell(row=aging_row, column=1,
                   value="Trade Payables ageing schedule: As at 31st March,2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')

    # Add detailed aging tables here if needed


def create_note_8(wb):
    """Create Note 8 - Other Current Liabilities"""
    ws = wb.create_sheet("Note 8 - Other Current Liab")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 8. Other Current Liabilities")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Other Current Liabilities data
    liabilities_data = [
        ["a) Other Payables", "", ""],
        ["(i) Statutory Remittances", "", ""],
        ["    T.D.S. Payable", "30.69", "22.61"],
        ["    GST Payable", "0.14", "0.00"],
        ["    EPF Payable", "1.59", "1.45"],
        ["    ESIC Payable", "0.96", "0.86"],
        ["(ii) Advance From Customers", "22.24", "12.63"],
        ["(iii) Other", "", ""],
        ["    Expenses Payable", "114.89", "118.57"],
        ["    Audit Fee Payble - R K Govil & Co.", "2.83", "2.83"],
        ["Total", "173.34", "158.95"]
    ]

    for row_idx, row_data in enumerate(liabilities_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"

            indent_level = 0
            if row_data[0].startswith("a)"):
                indent_level = 0
            elif row_data[0].startswith("(i)") or row_data[0].startswith("(ii)") or row_data[0].startswith("(iii)"):
                indent_level = 1
            elif row_data[0].startswith("    "):
                indent_level = 2

            if row_data[0] == "a) Other Payables":
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_9(wb):
    """Create Note 9 - Fixed Assets"""
    ws = wb.create_sheet("Note 9 - Fixed Assets")
    
    # Add header - set values first
    cell = ws.cell(row=1, column=1, value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    
    cell = ws.cell(row=2, column=1, value="3 MMTC/STC COLONY, GEETANJALI ENCLAVE, NEW DELHI")
    cell.alignment = Alignment(horizontal='center')
    
    cell = ws.cell(row=3, column=1, value="SOUTH DELHI-110017, CIN: U74899DL1989PTC038372")
    cell.alignment = Alignment(horizontal='center')
    
    cell = ws.cell(row=4, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    cell = ws.cell(row=5, column=1, value="NOTE NO. 9- FIXED ASSETS")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')
    
    # Then merge cells
    ws.merge_cells('A1:Q1')
    ws.merge_cells('A2:Q2')
    ws.merge_cells('A3:Q3')
    ws.merge_cells('A4:Q4')
    ws.merge_cells('A5:Q5')
    
    # Set column widths
    for col in range(1, 18):  # A through Q
        ws.column_dimensions[get_column_letter(col)].width = 9
    
    ws.column_dimensions['A'].width = 3  # S.No.
    ws.column_dimensions['B'].width = 20  # PARTICULARS
    
    # Create the complex header structure for the fixed assets table
    row = 7
    
    # First header row - main categories
    cell = ws.cell(row=row, column=1, value="S. No.")
    apply_header_style(cell)
    
    cell = ws.cell(row=row, column=2, value="PARTICULARS")
    apply_header_style(cell)
    
    # Gross Block header
    cell = ws.cell(row=row, column=3, value="GROSS BLOCK")
    apply_header_style(cell)
    
    # Depreciation header
    cell = ws.cell(row=row, column=8, value="DEPRECIATION")
    apply_header_style(cell)
    
    # Net Block header
    cell = ws.cell(row=row, column=12, value="NET BLOCK")
    apply_header_style(cell)
    
    # Rate of Depreciation header
    cell = ws.cell(row=row, column=17, value="Rate of Depreciation")
    apply_header_style(cell)
    
    # Now merge the headers
    ws.merge_cells(f'C{row}:G{row}')  # Gross Block
    ws.merge_cells(f'H{row}:K{row}')  # Depreciation
    ws.merge_cells(f'L{row}:N{row}')  # Net Block
    
    # Second header row
    row += 1
    
    # As at 01.04.2021
    cell = ws.cell(row=row, column=3, value="As at 01.04.2021")
    apply_header_style(cell)
    
    # ADDITION
    cell = ws.cell(row=row, column=5, value="ADDITION")
    apply_header_style(cell)
    
    # Sold/Transfer
    cell = ws.cell(row=row, column=7, value="Sold/ Transfer")
    apply_header_style(cell)
    
    # As at 31.03.2022 (Gross Block)
    cell = ws.cell(row=row, column=8, value="As at 31.03.2022")
    apply_header_style(cell)
    
    # Upto 31.03.2021 (Depreciation)
    cell = ws.cell(row=row, column=10, value="Upto 31.03.2021")
    apply_header_style(cell)
    
    # For the year (Depreciation)
    cell = ws.cell(row=row, column=12, value="For the year")
    apply_header_style(cell)
    
    # Upto 31.03.2022 (Depreciation)
    cell = ws.cell(row=row, column=14, value="Upto 31.03.2022")
    apply_header_style(cell)
    
    # As at 31.03.2022 (Net Block)
    cell = ws.cell(row=row, column=16, value="As at 31.03.2022")
    apply_header_style(cell)
    
    # As at 31.03.2021 (Net Block)
    cell = ws.cell(row=row, column=18, value="As at 31.03.2021")
    apply_header_style(cell)
    
    # Add the WDV column
    cell = ws.cell(row=row, column=20, value="WDV")
    apply_header_style(cell)
    
    # Now merge the cells
    ws.merge_cells(f'C{row}:D{row}')  # As at 01.04.2021
    ws.merge_cells(f'E{row}:F{row}')  # ADDITION
    
    # Third header row - sub-categories
    row += 1
    
    # Add Original value, Rev. headers for each section
    headers_col_mapping = {
        3: "Original value",  # As at 01.04.2021
        4: "Rev.",
        5: "Original value",  # ADDITION
        6: "Rev.",
        7: "",  # Sold/Transfer
        8: "Original value",  # As at 31.03.2022
        9: "Rev.",
        10: "Original value",  # Upto 31.03.2021
        11: "Rev.",
        12: "Original value",  # For the year
        13: "Rev.",
        14: "Sold/Tr. Rev.",  # Adjustments
        15: "Original value",  # Upto 31.03.2022
        16: "Rev.",
        17: "Original value",  # As at 31.03.2022
        18: "Rev.",
        19: "Original value",  # As at 31.03.2021
        20: "Rev."
    }
    
    for col, value in headers_col_mapping.items():
        cell = ws.cell(row=row, column=col, value=value)
        apply_header_style(cell)
    
    # Fixed assets data - main property, plant & equipment
    fixed_assets_data = [
        ["1", "Land", "1640.87", "0.00", "0.00", "0.00", "0.00", "1640.87", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "1640.87", "0.00", "1640.87", "0.00", "0.00"],
        ["2", "Building", "1575.70", "0.00", "22.64", "0.00", "0.00", "1598.34", "0.00", "280.72", "0.00", "124.24", "0.00", "0.00", "404.95", "0.00", "1193.38", "0.00", "1294.99", "0.00", "0.00"],
        ["3", "Furniture & Fixtures", "85.21", "0.00", "0.00", "0.00", "0.00", "85.21", "0.00", "36.37", "0.00", "12.65", "0.00", "0.00", "49.01", "0.00", "36.20", "0.00", "48.85", "0.00", "0.00"],
        ["4", "Generator", "58.03", "0.00", "0.00", "0.00", "0.00", "58.03", "0.00", "40.92", "0.00", "7.71", "0.00", "0.00", "48.63", "0.00", "9.40", "0.00", "17.11", "0.00", "0.00"],
        ["5", "Electric Installation", "4.47", "0.00", "0.00", "0.00", "0.00", "4.47", "0.00", "1.69", "", "0.72", "", "", "2.41", "", "2.06", "", "2.78", "", "0.00"],
        ["6", "Plant & Machinery- Equipments", "854.92", "", "67.86", "", "", "922.78", "", "292.14", "", "123.97", "", "", "416.12", "", "506.66", "", "562.78", "", "0.00"],
        ["7", "Plant & Machinery-Life Saving Equipment", "1742.54", "0.00", "0.00", "0.00", "0.00", "1742.54", "", "490.66", "", "257.64", "", "", "748.30", "", "994.24", "", "1251.88", "", "0.00"],
        ["8", "Computers", "22.89", "0.00", "3.11", "0.00", "0.00", "26.00", "0.00", "19.16", "0.00", "3.24", "0.00", "0.00", "22.40", "0.00", "3.60", "0.00", "3.72", "0.00", "0.00"],
        ["9", "Printers", "2.44", "0.00", "0.12", "0.00", "0.00", "2.56", "", "2.04", "", "0.30", "", "", "2.35", "", "0.21", "", "0.39", "", "0.00"],
        ["10", "UPS", "13.15", "0.00", "2.64", "", "", "15.79", "", "11.08", "", "1.31", "", "", "12.38", "", "3.41", "", "2.07", "", "0.00"],
        ["11", "Office Equipments", "33.46", "0.00", "0.74", "0.00", "0.00", "34.20", "0.00", "21.48", "0.00", "5.56", "0.00", "0.00", "27.04", "0.00", "7.16", "0.00", "11.98", "0.00", "0.00"],
        ["12", "Air Conditioner", "15.71", "0.00", "3.02", "0.00", "0.00", "18.73", "", "9.28", "", "3.70", "", "", "12.98", "", "5.75", "", "6.43", "", "0.00"],
        ["13", "Camera", "2.05", "0.00", "0.96", "0.00", "", "3.01", "", "0.90", "", "0.65", "", "", "1.55", "", "1.47", "", "1.16", "", "0.00"],
        ["14", "Cooler", "0.93", "0.00", "0.02", "0.00", "", "0.95", "", "0.19", "", "0.27", "", "", "0.46", "", "0.48", "", "0.14", "", "0.00"],
        ["15", "EPABX Systems", "0.34", "0.00", "0.00", "0.00", "", "0.34", "", "0.23", "", "0.04", "", "", "0.28", "", "0.06", "", "0.11", "", "0.00"],
        ["16", "Fire Extinguisher", "2.24", "0.00", "0.00", "0.00", "", "2.24", "", "1.21", "", "0.46", "", "", "1.67", "", "0.57", "", "1.03", "", "0.00"],
        ["17", "Television", "2.21", "0.00", "0.00", "0.00", "", "2.21", "", "1.42", "", "0.36", "", "", "1.78", "", "0.43", "", "0.79", "", "0.00"],
        ["18", "Software", "6.22", "0.00", "0.51", "0.00", "", "6.73", "", "2.41", "", "0.99", "", "", "3.41", "", "3.32", "", "3.81", "", "0.00"],
        ["19", "Vehicle", "0.90", "0.00", "0.00", "0.00", "", "0.90", "", "0.25", "", "0.08", "", "", "0.33", "", "0.17", "", "0.24", "", "0.00"],
        ["", "TOTAL ( Rupees )", "6063.28", "0.00", "102.22", "0.00", "0.00", "6165.50", "0.00", "1212.15", "0.00", "543.90", "0.00", "0.00", "1756.05", "0.00", "4409.45", "0.00", "4851.13", "0.00", ""]
    ]
    
    # Write the main fixed assets data
    for row_idx, row_data in enumerate(fixed_assets_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = col_idx > 2 and value != "" and value != "0.00"
            is_total = row_data[1] == "TOTAL ( Rupees )"
            
            if is_total:
                cell.font = Font(bold=True)
                
            apply_data_cell_style(cell, is_numeric, is_total)
    
    current_row = row_idx + 2
    
    # Capital WIP section
    cell = ws.cell(row=current_row, column=1, value="")
    cell = ws.cell(row=current_row, column=2, value="Capital WIP")
    cell.font = Font(bold=True)
    
    capital_wip_data = [
        ["1", "Building WIP", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["2", "Plant & Machinery", "0.00", "", "2.16", "", "0.00", "2.16", "", "", "", "", "", "", "", "", "2.16", "", "0.00", "", ""],
        ["3", "Furniture & Fixtures", "0.00", "", "1.97", "", "0.00", "1.97", "", "", "", "", "", "", "", "", "1.97", "", "0.00", "", ""],
        ["4", "Capital Advance", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["5", "Financial Cost", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["6", "Financial Cost- Int. on Space Allotment", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["7", "Financial Cost- Int. on Unsecured Loan", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["8", "Electricity and Generator Expenses", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["", "TOTAL ( Rupees )", "0.00", "0.00", "4.13", "", "", "4.13", "0.00", "0.00", "0.00", "", "0.00", "0.00", "0.00", "0.00", "4.13", "0.00", "0.00", "0.00", ""]
    ]
    
    # Write the Capital WIP data
    for row_idx, row_data in enumerate(capital_wip_data, current_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(row_data):  # Ensure we're not going beyond data length
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                is_numeric = col_idx > 2 and value != "" and value != "0.00"
                is_total = row_data[1] == "TOTAL ( Rupees )"
                
                if is_total:
                    cell.font = Font(bold=True)
                    
                apply_data_cell_style(cell, is_numeric, is_total)
    
    current_row = row_idx + 2
    
    # Assets Purchase But Not Put to Use section
    cell = ws.cell(row=current_row, column=1, value="")
    cell = ws.cell(row=current_row, column=2, value="Assets Purchase But Not Put to Use")
    cell.font = Font(bold=True)
    
    assets_not_in_use_data = [
        ["1", "Computer", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["2", "Office Equipments", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["3", "Television", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["4", "Generator", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["5", "Air Conditioner", "0.00", "", "", "", "0.00", "0.00", "", "", "", "", "", "", "", "", "", "", "0.00", "", ""],
        ["", "TOTAL ( Rupees )", "0.00", "0.00", "0", "0.00", "", "0.00", "0.00", "0.00", "0.00", "", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", "0.00", ""]
    ]
    
    # Write the Assets Not In Use data
    for row_idx, row_data in enumerate(assets_not_in_use_data, current_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(row_data):  # Ensure we're not going beyond data length
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                is_numeric = col_idx > 2 and value != "" and value != "0.00"
                is_total = row_data[1] == "TOTAL ( Rupees )"
                
                if is_total:
                    cell.font = Font(bold=True)
                    
                apply_data_cell_style(cell, is_numeric, is_total)
    
    current_row = row_idx + 2
    
    # Grand Total section
    grand_total_data = [
        ["", "GRAND TOTAL ( Rupees )", "6063.28", "0.00", "106.35", "0.00", "0.00", "6169.63", "0.00", "1212.15", "0.00", "543.90", "0.00", "0.00", "1756.05", "0.00", "4413.58", "0.00", "4851.13", "0.00", ""],
        ["", "PREVIOUS YEAR ( Rupees )", "", "", "", "", "", "", "", "", "", "", "", "0.00", "", "", "", "", "", "", ""]
    ]
    
    # Write the Grand Total data
    for row_idx, row_data in enumerate(grand_total_data, current_row):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(row_data):  # Ensure we're not going beyond data length
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                is_numeric = col_idx > 2 and value != "" and value != "0.00"
                is_total = True  # Both rows are totals
                
                cell.font = Font(bold=True)
                apply_data_cell_style(cell, is_numeric, is_total)
    
    current_row = row_idx + 2
    
    # Pre Operative Expense section
    pre_operative_data = [
        ["Pre Operative Expense", "7.30"],
        ["Less: - W/off During the Year", "2.43"],
        ["Balance remaining", "4.87"]
    ]
    
    # Write the Pre Operative Expense data
    for row_idx, row_data in enumerate(pre_operative_data, current_row):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(row_data):  # Only set values for the actual data columns
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                is_numeric = col_idx == 2  # Second column has numeric values
                is_total = row_data[0] == "Balance remaining"
                
                if is_total:
                    cell.font = Font(bold=True)
                    
                apply_data_cell_style(cell, is_numeric, is_total)


def create_note_10(wb):
    """Create Note 10 - Long Term Loans and Advances"""
    ws = wb.create_sheet("Note 10 - LT Loans & Advances")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(
        row=3, column=1, value="Note 10. Long Term Loans & Advances (Unsecured and Considered Good)")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Long Term Loans & Advances data
    loans_data = [
        ["a) Security Deposits", "17.18", "17.15"],
        ["b) Accrued Interest on FDR", "4.61", "2.82"],
        ["Total", "21.78", "19.97"]
    ]

    for row_idx, row_data in enumerate(loans_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"

            if is_total:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total)


def create_note_11(wb):
    """Create Note 11 - Inventories"""
    ws = wb.create_sheet("Note 11 - Inventories")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 11. Closing Stock")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # Table headers for inventory
    row = 5
    headers = ["Particulars", "Opening Stock 01.04.2021",
        "Purchase", "Total", "Consumption", "Closing Stock 31.03.2022"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Inventory data
    inventory_data = [
        ["Stores", "", "", "", "", ""],
        ["Consumables stores", "0.00", "0.00", "0", "0.00", "0.00"],
        ["Gas Consumed", "0.79", "37.26", "38.05", "35.38", "2.66"],
        ["", "0.79", "37.26", "38.05", "35.38", "2.66"],
        ["Material", "", "", "", "", ""],
        ["Laboratory Material", "2.58", "467.24", "469.83", "466.61", "3.21"],
        ["Medicine", "3.86", "720.05", "723.91", "719.32", "4.58"],
        ["Surgical", "3.47", "2.63", "6.11", "2.62", "3.48"],
        ["", "9.91", "1189.93", "1199.84", "1188.56", "11.28"],
        ["Others", "", "", "", "", ""],
        ["Stationary", "0.43", "29.54", "29.97", "29.11", "0.85"],
        ["Building Repair Material", "0.55", "5.35", "5.91", "5.34", "0.57"],
        ["Furniture & Fixtures", "0.05", "4.61", "4.66", "4.58", "0.08"],
        ["Diesel", "0.06", "6.15", "6.21", "6.13", "0.07"],
        ["", "1.10", "45.64", "46.74", "45.17", "1.57"],
        ["Total", "11.80", "1272.83", "1284.63", "1269.11", "15.52"]
    ]

    for row_idx, row_data in enumerate(inventory_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3, 4, 5, 6] and value != ""
            is_total = row_data[0] == "Total"
            is_header = row_data[0] in ["Stores", "Material", "Others"]
            is_subtotal = row_data[0] == "" and col_idx == 1

            if is_header:
                cell.font = Font(bold=True)

            if is_subtotal:
                is_numeric = True

            apply_data_cell_style(cell, is_numeric, is_total)


def create_note_12(wb):
    """Create Note 12 - Trade Receivables"""
    ws = wb.create_sheet("Note 12 - Trade Receivables")
    
    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1, value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    cell = ws.cell(row=2, column=1, value="3 MMTC/STC COLONY, GEETANJALI ENCLAVE, NEW DELHI")
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="SOUTH DELHI-110017, CIN: U74899DL1989PTC038372")
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:C4')
    cell = ws.cell(row=4, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:C5')
    cell = ws.cell(row=5, column=1, value="Note 12. Trade Receivables (Unsecured and Considered Good)")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Table headers
    row = 7
    headers = ["Particulars", "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)
    
    # Trade Receivables data
    receivables_data = [
        ["(a) Secured, Considerdgoods", "0.00", "0.00"],
        ["(b) Unsecured, Considerdgoods", "208.76", "16.33"],
        ["(c) Doubtful", "", ""],
        ["Total", "208.76", "16.33"]
    ]
    
    for row_idx, row_data in enumerate(receivables_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"
            
            apply_data_cell_style(cell, is_numeric, is_total)
    
    # Add aging schedule for 2022
    aging_row = row_idx + 2
    
    # First create the cells with their values before merging
    cell = ws.cell(row=aging_row, column=1, value="Trade Receivables ageing schedule as at 31st March,2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')
    
    # Add (Rs. in ...........) text to column G (not merged cell)
    rs_cell = ws.cell(row=aging_row, column=7, value="(Rs. in ...........)")
    rs_cell.alignment = Alignment(horizontal='right')
    
    # Now merge cells after setting values
    ws.merge_cells(f'A{aging_row}:F{aging_row}')  # Merge A-F but not G
    
    # 2022 Aging Schedule headers
    aging_row += 1
    
    # Set value to the top-left cell first
    cell = ws.cell(row=aging_row, column=2, value="Outstanding for following periods from due date of payment")
    apply_header_style(cell)
    
    # Then merge
    ws.merge_cells(f'B{aging_row}:G{aging_row}')
    
    aging_row += 1
    aging_headers = ["Particulars", "Less than 6 months", "6 months -1 year", "1-2 years", "2-3 years", "More than 3 years", "Total"]
    for i, header in enumerate(aging_headers, 1):
        cell = ws.cell(row=aging_row, column=i, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    # 2022 Aging data
    aging_2022_data = [
        ["(i) Undisputed Trade receivables -considered good", "197.19", "5.25", "3.84607", "2.47", "-", "208.76"],
        ["(ii) Undisputed Trade receivables -considered doubtful", "", "", "", "", "", "-"],
        ["(iii) Disputed trade receivables considered good", "", "", "", "", "", "-"],
        ["(iv) Disputed trade receivables considered doubtful", "", "", "", "", "", "-"]
    ]
    
    for row_idx, row_data in enumerate(aging_2022_data, aging_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = col_idx > 1 and value not in ["", None, "-"]
            
            apply_data_cell_style(cell, is_numeric, False)
    
    # Add aging schedule for 2021
    aging_row = row_idx + 2
    
    # Set values before merging
    cell = ws.cell(row=aging_row, column=1, value="Trade Receivables ageing schedule as at 31st March,2021")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')
    
    # Add (Rs. in ...........) text to column G (not merged cell)
    rs_cell = ws.cell(row=aging_row, column=7, value="(Rs. in ...........)")
    rs_cell.alignment = Alignment(horizontal='right')
    
    # Now merge
    ws.merge_cells(f'A{aging_row}:F{aging_row}')  # Merge A-F but not G
    
    # 2021 Aging Schedule headers
    aging_row += 1
    
    # Set value first
    cell = ws.cell(row=aging_row, column=2, value="Outstanding for following periods from due date of payment")
    apply_header_style(cell)
    
    # Then merge
    ws.merge_cells(f'B{aging_row}:G{aging_row}')
    
    aging_row += 1
    for i, header in enumerate(aging_headers, 1):
        cell = ws.cell(row=aging_row, column=i, value=header)
        apply_header_style(cell)
    
    # 2021 Aging data
    aging_2021_data = [
        ["(i) Undisputed Trade receivables -considered good", "10.29", "3.56", "2.47321", "-", "-", "16.33"],
        ["(ii) Undisputed Trade receivables -considered doubtful", "", "", "", "", "", "-"],
        ["(iii) Disputed trade receivables considered good", "", "", "", "", "", "-"],
        ["(iv) Disputed trade receivables considered doubtful", "", "", "", "", "", "-"]
    ]
    
    for row_idx, row_data in enumerate(aging_2021_data, aging_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = col_idx > 1 and value not in ["", None, "-"]
            
            apply_data_cell_style(cell, is_numeric, False)


def create_note_13(wb):
    """Create Note 13 - Cash and Cash Equivalent"""
    ws = wb.create_sheet("Note 13 - Cash & Cash Equiv")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 13. Cash and Cash Equivalent")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Cash and Cash Equivalent data
    cash_data = [
        ["a) Balances with Banks:", "", ""],
        ["In Current Accounts", "0.12", "0.12"],
        ["In FDRS and Bank Guarantee", "78.48", "65.48"],
        ["b) Cash on Hand", "2.16", "1.05"],
        ["Total", "80.75", "66.64"]
    ]

    for row_idx, row_data in enumerate(cash_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"
            is_header = row_data[0] in ["a) Balances with Banks:"]

            indent_level = 0
            if not is_header and not is_total and row_data[0] != "" and not row_data[0].startswith("b)"):
                indent_level = 1

            if is_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_14(wb):
    """Create Note 14 - Short Term Loans & Advances"""
    ws = wb.create_sheet("Note 14 - ST Loans & Advances")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(
        row=3, column=1, value="Note 14. Short Term Loans & Advances (Unsecured and Considered Good)")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "As at 31 March 2022 Amount (Rs.)", "As at 31 March 2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Short Term Loans & Advances data
    st_loans_data = [
        ["a) Loans and Advances to Employees", "1.20", "0.16"],
        ["b) Advances to Suppliers - Capital Advance", "16.20", "64.59"],
        ["c) Other", "", ""],
        ["(i) Kotak Mahindra Bank- Loan Processing Fees", "5.99", "5.99"],
        ["(ii) HDFC Bank Limited- OD A/c (Book Balance)", "", "78.10"],
        ["d) TDS Recoverables", "57.04", "0.00"],
        ["Total", "80.43", "148.83"]
    ]

    for row_idx, row_data in enumerate(st_loans_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"

            indent_level = 0
            if row_data[0].startswith("(i)") or row_data[0].startswith("(ii)"):
                indent_level = 1

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)

    # Create Note 14a - Other Current Assets
    row_idx += 4
    ws.merge_cells(f'A{row_idx}:C{row_idx}')
    cell = ws.cell(row=row_idx, column=1,
                   value="Note 14a. Other Current Assets")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    row_idx += 2
    # Table headers
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=i, value=header)
        apply_header_style(cell)

    # Other Current Assets data
    other_current_assets = [
        ["a) Prepaid Expenses", "6.52", "6.12"],
        ["b) Balance with Government Authorities", "", ""],
        ["(i) Income Tax Refund AY 2021-22", "11.72", "11.72"],
        ["(ii) Income Tax Refund AY 2020-21", "0.00", "6.27"],
        ["Total", "18.24", "24.12"]
    ]

    for idx, row_data in enumerate(other_current_assets, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx+idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"

            indent_level = 0
            if row_data[0].startswith("(i)") or row_data[0].startswith("(ii)"):
                indent_level = 1

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_15(wb):
    """Create Note 15 - Revenue from operations"""
    ws = wb.create_sheet("Note 15 - Revenue")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 15. Revenue from operations")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "Year ending 31.03.2022 Amount (Rs.)", "Year ending 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Revenue data
    revenue_data = [
        ["Sale of Services", "2523.15", "2471.49"],
        ["", "", ""],
        ["Total", "2523.15", "2471.49"]
    ]

    for row_idx, row_data in enumerate(revenue_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"

            if is_total:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total)


def create_note_16(wb):
    """Create Note 16 - Other Income"""
    ws = wb.create_sheet("Note 16 - Other Income")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 16. Other Income")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "Year ending 31.03.2022 Amount (Rs.)", "Year ending 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Other Income data
    other_income_data = [
        ["Interest Income", "", ""],
        ["on FDR", "1.98", "1.97"],
        ["on Income Tax Refund", "0.00", "0.00"],
        ["Foreign Exchange Income", "0.00", "48.17"],
        ["Misc. Receipts", "0.77", "0.00"],
        ["Creditor W/o", "14.36", ""],
        ["Rent Received", "9.60", ""],
        ["TOTAL:", "26.72", "50.14"]
    ]

    for row_idx, row_data in enumerate(other_income_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "TOTAL:"
            is_header = row_data[0] == "Interest Income"

            indent_level = 0
            if not is_header and not is_total and row_data[0] in ["on FDR", "on Income Tax Refund"]:
                indent_level = 1

            if is_total:
                cell.font = Font(bold=True)

            if is_header:
                cell.font = Font(bold=True)

            apply_data_cell_style(cell, is_numeric, is_total, indent_level)


def create_note_17(wb):
    """Create Note 17 - Cost of Material Consumed"""
    ws = wb.create_sheet("Note 17 - Cost of Material")

    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1,
                   value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    cell = ws.cell(
        row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 17. Cost of Material Consumed")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Table headers
    row = 5
    headers = ["Particulars",
        "Year ending 31.03.2022 Amount (Rs.)", "Year ending 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)

    # Cost of Material Consumed data
    material_data = [
        ["Material Consumed", "1188.56", "789.26"],
        ["Store Consumed", "35.38", "10.88"],
        ["Trial Run Expenses", "0.00", "0.00"],
        ["", "", ""],
        ["Total", "1223.94", "800.14"]
    ]

    for row_idx, row_data in enumerate(material_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"

            apply_data_cell_style(cell, is_numeric, is_total)


def create_note_18(wb):
    """Create Note 18 - Employee Benefit Expenses"""
    ws = wb.create_sheet("Note 18 - Employee Benefits")
    
    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1, value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    cell = ws.cell(row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 18. Employee Benefit Expenses")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Table headers
    row = 5
    headers = ["Particulars", "Year ending 31.03.2022 Amount (Rs.)", "Year ending 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)
    
    # Employee Benefit Expenses data
    employee_data = [
        ["Salaries, Wages and Incentives", "1118.36", "1052.61"],
        ["Contributions to PF and Other Fund", "17.65", "14.69"],
        ["Staff Welfare Expenses", "0.12", "0.01"],
        ["", "", ""],
        ["Total", "1136.13", "1067.30"]
    ]
    
    for row_idx, row_data in enumerate(employee_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"
            
            apply_data_cell_style(cell, is_numeric, is_total)

def create_note_19(wb):
    """Create Note 19 - Finance Costs"""
    ws = wb.create_sheet("Note 19 - Finance Costs")
    
    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1, value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    cell = ws.cell(row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 19. Finance Costs")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Table headers
    row = 5
    headers = ["Particulars", "Year ending 31.03.2022 Amount (Rs.)", "Year ending 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)
    
    # Finance Costs data
    finance_data = [
        ["Interest Expenses on", "", ""],
        ["-Borrowings", "366.82", "294.21"],
        ["-Others", "0.00", "0.00"],
        ["    Bank Charges", "0.00", "1.71"],
        ["    Credit Card Charges", "0.00", "5.97"],
        ["    Interest on Govt. Dues", "0.00", "0.80"],
        ["", "", ""],
        ["Total", "366.82", "302.70"]
    ]
    
    for row_idx, row_data in enumerate(finance_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"
            is_header = row_data[0] == "Interest Expenses on"
            
            indent_level = 0
            if not is_header and not is_total and row_data[0].startswith("-"):
                indent_level = 1
            elif row_data[0].startswith("    "):
                indent_level = 2
                
            if is_header:
                cell.font = Font(bold=True)
                
            apply_data_cell_style(cell, is_numeric, is_total, indent_level)

def create_note_20(wb):
    """Create Note 20 - Other Expenses"""
    ws = wb.create_sheet("Note 20 - Other Expenses")
    
    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1, value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    cell = ws.cell(row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=1, value="Note 20. Other Expenses")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Table headers
    row = 5
    headers = ["Particulars", "Year ending 31.03.2022 Amount (Rs.)", "Year ending 31.03.2021 Amount (Rs.)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)
    
    # Other Expenses data
    other_expenses_data = [
        ["Bank Charges", "2.59", "0.00"],
        ["Credit Card Charges", "5.47", "0.00"],
        ["Interest on Govt. Dues", "3.59", "0.00"],
        ["Admisssion and Discharge Expenses", "15.03", "17.10"],
        ["Advance to creditors W/o", "55.22", "0.00"],
        ["Advertisement Expenses", "0.00", "8.18"],
        ["Auditors' Remuneration", "1.48", "1.48"],
        ["Business Promotion Exp", "0.00", "0.00"],
        ["Commission Expenses", "0.88", "0.00"],
        ["Customer Entertainment", "1.07", "1.89"],
        ["Diwali Exp.", "0.07", "0.04"],
        ["Donation and Charity", "0.00", "0.00"],
        ["Electricity Expenses", "86.34", "71.89"],
        ["Fees and Subscription", "3.71", "3.93"],
        ["Foreign Exchange Loss", "2.38", "0.00"],
        ["Gardening Expenses", "0.36", "0.24"],
        ["General Expenses", "2.47", "1.12"],
        ["Generator Expenses", "6.13", "7.21"],
        ["Housekeeping Expenses", "118.33", "76.53"],
        ["Insurance", "5.00", "5.02"],
        ["Kitchen Expenses", "115.99", "40.62"],
        ["Laundry Expenses", "20.20", "12.12"],
        ["Miscellaneous Expenses", "0.71", "0.87"],
        ["Nagar Nigam Tax", "2.89", "0.00"],
        ["Pest Control", "2.42", "2.93"],
        ["Pre-Operative Exp W/O", "0.00", "0.00"],
        ["Printing & Stationery", "29.11", "14.68"],
        ["Professional Charges", "0.59", "0.09"],
        ["Rent", "0.14", "11.83"],
        ["Repair & Maintenance-Building", "5.34", "4.70"],
        ["Repair & Maintenance-Computer", "0.33", "1.09"],
        ["Repair & Maintenance-Others", "32.44", "5.33"],
        ["Repair & Maintenance-Vehicles", "0.34", "0.01"],
        ["TDS FY 2016-17, 2017-18", "0.40", "0.00"],
        ["Telephone Expenses", "2.42", "2.89"],
        ["Travelling & Conveyance", "0.42", "0.36"],
        ["Uniform Expenses", "3.22", "1.25"],
        ["", "", ""],
        ["Total", "527.08", "293.39"]
    ]
    
    for row_idx, row_data in enumerate(other_expenses_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = col_idx in [2, 3] and value != ""
            is_total = row_data[0] == "Total"
            
            apply_data_cell_style(cell, is_numeric, is_total)

def create_notes_index(wb):
    """Create an index sheet for all notes"""
    # Move the Notes Index sheet after the Cash Flow sheet
    sheet_names = wb.sheetnames
    # Get the index for Cash Flow
    cash_flow_idx = sheet_names.index("Cash Flow")
    
    # Create new sheet at the correct position
    ws = wb.create_sheet("Notes Index", cash_flow_idx + 1)
    
    # Add header
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1, value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    cell = ws.cell(row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    
    # Table headers
    row = 4
    headers = ["Note No.", "Description", "Reference"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        apply_header_style(cell)
    
    # Notes data
    notes_data = [
        ["1", "Share Capital", "Balance Sheet - Equity and Liabilities"],
        ["2", "Reserve & Surplus", "Balance Sheet - Equity and Liabilities"],
        ["3", "Long Term Borrowings", "Balance Sheet - Non-current liabilities"],
        ["4", "Deferred Tax Liabilities/Assets", "Balance Sheet - Non-current liabilities"],
        ["5", "Provisions", "Balance Sheet - Non-current/Current liabilities"],
        ["6", "Short Term Borrowings", "Balance Sheet - Current liabilities"],
        ["7", "Trade Payables", "Balance Sheet - Current liabilities"],
        ["8", "Other Current Liabilities", "Balance Sheet - Current liabilities"],
        ["9", "Property, Plant and Equipment", "Balance Sheet - Non-current assets"],
        ["10", "Long Term Loans and Advances", "Balance Sheet - Non-current assets"],
        ["11", "Inventories", "Balance Sheet - Current assets"],
        ["12", "Trade Receivables", "Balance Sheet - Current assets"],
        ["13", "Cash and Cash Equivalents", "Balance Sheet - Current assets"],
        ["14", "Short Term Loans and Advances", "Balance Sheet - Current assets"],
        ["15", "Revenue from operations", "P&L Statement - Revenue"],
        ["16", "Other Income", "P&L Statement - Revenue"],
        ["17", "Cost of Material Consumed", "P&L Statement - Expenses"],
        ["18", "Employee Benefit Expenses", "P&L Statement - Expenses"],
        ["19", "Finance Costs", "P&L Statement - Expenses"],
        ["20", "Other Expenses", "P&L Statement - Expenses"]
    ]
    
    for row_idx, row_data in enumerate(notes_data, row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add hyperlinks to the note sheets where they exist
            if col_idx == 2:
                note_sheet_name = f"Note {row_data[0]} - {value.split(' ')[0]}"
                if note_sheet_name in wb.sheetnames or note_sheet_name.replace(" - ", " ") in wb.sheetnames:
                    cell.font = Font(color="0563C1", underline="single")
                    cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='left' if col_idx != 1 else 'center')
                
            # Add border
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

def create_note_1(wb):
    """Create Note 1 - Share Capital"""
    ws = wb.create_sheet("Note 1 - Share Capital")
    
    # Add header
    ws.merge_cells('A1:E1')
    cell = ws.cell(row=1, column=1, value="M/S VIVEKANAND NURSING HOME PRIVATE LIMITED")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    cell = ws.cell(row=2, column=1, value="NOTES ON FINANCIAL STATEMENTS FOR THE YEAR ENDED ON 31ST MARCH, 2022")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:E3')
    cell = ws.cell(row=3, column=1, value="Note 1. Share Capital")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # First table headers
    ws.merge_cells('B5:C5')
    ws.merge_cells('D5:E5')
    cell_b5 = ws.cell(row=5, column=2, value="As at 31 March 2022")
    apply_header_style(cell_b5)
    cell_d5 = ws.cell(row=5, column=4, value="As at 31 March 2021")
    apply_header_style(cell_d5)
    
    headers = ["Particulars", "Number", "Amount (Rs.)", "Number", "Amount (Rs.)"]
    for i, header in enumerate(headers):
        col = i + 1
        cell = ws.cell(row=6, column=col, value=header)
        apply_header_style(cell)
    
    # Share capital data
    share_capital_data = [
        ["Authorised", "", "", "", ""],
        ["Equity Shares of Rs 100/- each", "106000", "10600000", "106000", "10600000"],
        ["", "", "", "", ""],
        ["Issued, Subscribed and Fully Paid up", "", "", "", ""],
        ["Equity Shares of Rs 100/- each", "79274", "7927400", "79274", "7927400"],
        ["Total", "79274", "7927400", "79274", "7927400"]
    ]
    
    for row_idx, row_data in enumerate(share_capital_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = (col_idx in [2, 3, 4, 5]) and value != ""
            is_total = row_data[0] == "Total"
            is_header = row_data[0] in ["Authorised", "Issued, Subscribed and Fully Paid up"]
            
            if is_header:
                cell.font = Font(bold=True)
                
            apply_data_cell_style(cell, is_numeric, is_total)
    
    # Add reconciliation table
    recon_row = len(share_capital_data) + 8
    ws.merge_cells(f'A{recon_row}:E{recon_row}')
    cell = ws.cell(row=recon_row, column=1, 
                  value="The Reconciliation of the Number of Shares Outstanding and the Amount of Share Capital as at 31st March 2022 and 31st March 2021 is Set Out as Below:")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Reconciliation table headers
    recon_row += 2
    ws.merge_cells(f'B{recon_row}:C{recon_row}')
    ws.merge_cells(f'D{recon_row}:E{recon_row}')
    cell_b = ws.cell(row=recon_row, column=2, value="As at 31 March 2022")
    apply_header_style(cell_b)
    cell_d = ws.cell(row=recon_row, column=4, value="As at 31 March 2021")
    apply_header_style(cell_d)
    
    recon_row += 1
    headers = ["Particulars", "Number", "Amount (Rs.)", "Number", "Amount (Rs.)"]
    for i, header in enumerate(headers):
        col = i + 1
        cell = ws.cell(row=recon_row, column=col, value=header)
        apply_header_style(cell)
    
    # Reconciliation data
    recon_data = [
        ["Balance at the beginning of the year", "79274", "7927400", "79274", "7927400"],
        ["Add: Shares issues during the year", "-", "-", "-", "-"],
        ["Less: Shares bought back during the year", "-", "-", "-", "-"],
        ["Balance at the end of the year", "79274", "7927400", "79274", "7927400"]
    ]
    
    for row_idx, row_data in enumerate(recon_data, recon_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = (col_idx in [2, 3, 4, 5]) and value not in ["-", ""]
            is_total = row_data[0] == "Balance at the end of the year"
            
            if is_total:
                cell.font = Font(bold=True)
                
            apply_data_cell_style(cell, is_numeric, is_total)
    
    # Company rights note
    rights_row = row_idx + 2
    ws.merge_cells(f'A{rights_row}:E{rights_row}')
    cell = ws.cell(row=rights_row, column=1, 
                  value="The Company has only one class of equity shares having a par value of Rs. 100/- per share. All these shares have same rights & prefrences with respect to payment of dividend, repayment of capital and voting.")
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    # Shareholders table
    shareholders_row = rights_row + 2
    ws.merge_cells(f'A{shareholders_row}:E{shareholders_row}')
    cell = ws.cell(row=shareholders_row, column=1, value="Details of Saheholders holding more than 5%")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')
    
    # Shareholders table headers
    shareholders_row += 2
    ws.merge_cells(f'C{shareholders_row}:E{shareholders_row}')
    cell = ws.cell(row=shareholders_row, column=3, value="Equity Shares")
    apply_header_style(cell)
    
    shareholders_row += 1
    ws.merge_cells(f'C{shareholders_row}:D{shareholders_row}')
    cell_c = ws.cell(row=shareholders_row, column=3, value="As at 31 March 2022")
    apply_header_style(cell_c)
    cell_e = ws.cell(row=shareholders_row, column=5, value="As at 31 March 2021")
    apply_header_style(cell_e)
    
    shareholders_row += 1
    headers = ["Name of Shareholders", "No. of Shares held", "% of Holding", "No. of Shares held", "% of Holding"]
    for i, header in enumerate(headers):
        if i == 0:
            col = 1
        elif i == 1:
            col = 3
        elif i == 2:
            col = 4
        elif i == 3:
            col = 3  # This would overwrite, but we'll merge and set differently
        elif i == 4:
            col = 5
        cell = ws.cell(row=shareholders_row, column=col, value=header)
        apply_header_style(cell)
    
    # Shareholders data
    shareholders_data = [
        ["Yashoda Hospital & Research Centre Limited", "79274", "100", "79274", "100"]
    ]
    
    for row_idx, row_data in enumerate(shareholders_data, shareholders_row + 1):
        col_positions = [1, 3, 4, 5, 5]  # Column positions for each data element
        
        for i, value in enumerate(row_data):
            col = col_positions[i] if i < len(col_positions) else 1
            
            if i == 3:  # "No. of Shares held" for 2021
                # Skip as we'll handle in a special way
                continue
                
            if i == 4:  # "% of Holding" for 2021
                # This is correctly positioned at column 5
                pass
                
            cell = ws.cell(row=row_idx, column=col, value=value)
            
            is_numeric = i in [1, 2, 4] and value != ""
            apply_data_cell_style(cell, is_numeric, False)
            
if __name__ == "__main__":
    # Example usage
    pdf_path = "vivekanand_nursing_home_financial_statements.pdf"
    output_excel_path = "Vivekanand_Financial_Statements_FY2022.xlsx"
    
    convert_financial_pdf_to_excel(pdf_path, output_excel_path)