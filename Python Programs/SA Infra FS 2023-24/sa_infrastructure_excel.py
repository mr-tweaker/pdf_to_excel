import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_financial_statements_excel():
    """
    Creates a complete Excel workbook with all financial statements
    from the SA Infrastructure PDF.
    """
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create main financial statement sheets
    create_balance_sheet(wb)
    create_manufacturing_account(wb)
    create_trading_account(wb)
    create_profit_loss_account(wb)
    
    # Create schedules
    create_capital_account(wb)  # Schedule 1
    create_secured_loans(wb)    # Schedule 2
    create_unsecured_loans(wb)  # Schedule 3
    create_working_capital(wb)  # Schedule 4
    create_agriculture_credit(wb)  # Schedule 4A
    create_trade_payables(wb)   # Schedule 5
    create_provisions(wb)       # Schedule 6
    create_advance_against_land(wb)  # Schedule 6A
    create_advance_against_sales(wb)  # Schedule 6B
    create_fixed_assets(wb)     # Schedule 7
    create_fixed_assets_mfg(wb) # Schedule 7A
    create_investments(wb)      # Schedule 8
    create_long_term_loans(wb)  # Schedule 9
    create_deposits(wb)         # Schedule 10
    create_trade_receivables(wb)  # Schedule 11
    create_trade_receivables_assets(wb)  # Schedule 12
    create_duties_and_taxes(wb)  # Schedule 13
    create_short_term_loans(wb)  # Schedule 14
    create_advances_against_purchase(wb)  # Schedule 14A
    create_cash_and_bank(wb)    # Schedule 15
    create_direct_expenses(wb)  # Schedule 16
    create_gross_receipt(wb)    # Schedule 17
    create_purchase(wb)         # Schedule 18
    create_indirect_income(wb)  # Schedule 19
    create_indirect_expenses(wb)  # Schedule 20
    create_finance_expenses(wb)  # Schedule 21
    
    # Create annexures
    create_agriculture_account(wb)  # Annexure 1
    create_land_account(wb)     # Annexure 2
    
    # Save the workbook
    wb.save('SA_Infrastructure_Financial_Statements_March_2024.xlsx')
    print("Excel file created successfully!")
    return wb

def apply_header_style(ws, title):
    """Apply standard header styling to a worksheet."""
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    # Add title
    ws.merge_cells('A1:F1')
    ws['A1'] = "M/s SA Infrastructure Prop. Saiyyed Afsar Ali (PAN No. AZGPS7572J)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "16/10, Rehmat Nagar, Ratlam (M.P.)-457001"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:F3')
    ws['A3'] = title
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')

def apply_footer(ws, start_row):
    """Apply standard footer to a worksheet."""
    # Add footer
    ws.merge_cells(f'A{start_row}:C{start_row}')
    ws[f'A{start_row}'] = "M/s SA Infrastructure"
    
    ws.merge_cells(f'D{start_row}:F{start_row}')
    ws[f'D{start_row}'] = "In terms of our attached report of even date"
    
    ws.merge_cells(f'D{start_row+1}:F{start_row+1}')
    ws[f'D{start_row+1}'] = "For- Saloni Asawa and Associates"
    
    ws.merge_cells(f'D{start_row+2}:F{start_row+2}')
    ws[f'D{start_row+2}'] = "Chartered Accountant"
    
    ws.merge_cells(f'D{start_row+3}:F{start_row+3}')
    ws[f'D{start_row+3}'] = "FRN.-022450C"
    
    ws.merge_cells(f'A{start_row+5}:C{start_row+5}')
    ws[f'A{start_row+5}'] = "Saiyyed Afsar Ali"
    
    ws.merge_cells(f'D{start_row+5}:F{start_row+5}')
    ws[f'D{start_row+5}'] = "Saloni Asawa"
    
    ws.merge_cells(f'A{start_row+6}:C{start_row+6}')
    ws[f'A{start_row+6}'] = "(Proprietor)"
    
    ws.merge_cells(f'D{start_row+6}:F{start_row+6}')
    ws[f'D{start_row+6}'] = "(Membership No.-430411)"
    
    ws.merge_cells(f'A{start_row+8}:C{start_row+8}')
    ws[f'A{start_row+8}'] = "UDIN-"
    
    ws.merge_cells(f'A{start_row+10}:C{start_row+10}')
    ws[f'A{start_row+10}'] = "Date-"
    
    ws.merge_cells(f'A{start_row+11}:C{start_row+11}')
    ws[f'A{start_row+11}'] = "Place- Ratlam"

def create_balance_sheet(wb):
    """Create Balance Sheet worksheet."""
    ws = wb.create_sheet("Balance Sheet")
    apply_header_style(ws, "Balance Sheet as at 31st March 2024")
    
    # Add headers
    ws['A5'] = "Liabilities"
    ws['B5'] = "Sch. No."
    ws['C5'] = "Amount (Rs.)"
    ws['D5'] = "Assets"
    ws['E5'] = "Sch. No."
    ws['F5'] = "Amount (Rs.)"
    for cell in ['A5', 'B5', 'C5', 'D5', 'E5', 'F5']:
        ws[cell].font = Font(bold=True)
    
    # Add Liabilities data
    liabilities_data = [
        ("Proprietor's Fund", "", ""),
        ("Proprietor's Capital", "1", "166,087,985"),
        ("", "", ""),
        ("Non Current Liabilities", "", ""),
        ("Secured Loans", "2", "61,188,596"),
        ("Unsecured Loans", "3", "25,650,048"),
        ("Cash Credit A/c- Bank of India", "4", "39,684,622"),
        ("Agriculture Credit Overdraft - HDFC", "4A", "-"),
        ("", "", ""),
        ("Current Liabilites", "", ""),
        ("Trade Payables", "5", "137,092,207"),
        ("Provisions", "6", "1,845,423"),
        ("Advance against land", "6A", "600,000"),
        ("Advance against Sales", "6B", "30,710,480"),
        ("", "", ""),
        ("Total", "", "462,859,361")
    ]
    
    # Add Assets data
    assets_data = [
        ("Non Current Assets", "", ""),
        ("Fixed Assets", "7", "99,791,905"),
        ("Non Current Investment", "8", "112,066,713"),
        ("Long Term Loan and Advances", "9", "29,196,432"),
        ("Government Deposits", "10", "1,736,600"),
        ("", "", ""),
        ("", "", ""),
        ("", "", ""),
        ("", "", ""),
        ("Current Assets", "", ""),
        ("Closing Stock", "", "123,375,824"),
        ("Trade Receivables", "11", "83,645,859"),
        ("Trade Receivables for Assets", "12", "50,000"),
        ("Duties and Taxes", "13", "5,398,899"),
        ("Short Term Loan and Advances", "14", "749,126"),
        ("Advance Against Purchase", "14A", "1,828,845"),
        ("Cash and Bank Balance", "15", "5,019,157"),
        ("Total", "", "462,859,361")
    ]
    
    # Write Liabilities data
    for i, (name, sch_no, amount) in enumerate(liabilities_data, start=6):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = sch_no
        ws[f'C{i}'] = amount
        if "Total" in name or name in ["Proprietor's Fund", "Non Current Liabilities", "Current Liabilites"]:
            ws[f'A{i}'].font = Font(bold=True)
        if "Total" in name:
            ws[f'C{i}'].font = Font(bold=True)
    
    # Write Assets data
    for i, (name, sch_no, amount) in enumerate(assets_data, start=6):
        ws[f'D{i}'] = name
        ws[f'E{i}'] = sch_no
        ws[f'F{i}'] = amount
        if "Total" in name or name in ["Non Current Assets", "Current Assets"]:
            ws[f'D{i}'].font = Font(bold=True)
        if "Total" in name:
            ws[f'F{i}'].font = Font(bold=True)
    
    # Add footer
    apply_footer(ws, 23)

def create_manufacturing_account(wb):
    """Create Manufacturing Account worksheet."""
    ws = wb.create_sheet("Manufacturing Account")
    apply_header_style(ws, "Manufacturing Account For the Year 1st April 2023 to 31st March 2024")
    
    # Add headers
    ws['A5'] = "Dr"
    ws['D5'] = "Cr"
    ws['A5'].alignment = Alignment(horizontal='center')
    ws['D5'].alignment = Alignment(horizontal='center')
    ws['A5'].font = Font(bold=True)
    ws['D5'].font = Font(bold=True)
    
    ws['A6'] = "Particulars"
    ws['B6'] = "Sch. No."
    ws['C6'] = "Amount (Rs.)"
    ws['D6'] = "Particulars"
    ws['E6'] = "Sch. No."
    ws['F6'] = "Amount (Rs.)"
    for cell in ['A6', 'B6', 'C6', 'D6', 'E6', 'F6']:
        ws[cell].font = Font(bold=True)
    
    # Add data
    dr_data = [
        ("To Opening Stock (Work in Process)", "", "-"),
        ("To Direct Expenses", "16", "44,476,112"),
        ("To Depreciation (Manufacturing)", "7(A)", "2,873,169"),
        ("", "", ""),
        ("Total", "", "47,349,281")
    ]
    
    cr_data = [
        ("By Cost of Goods Produced- (Transfer to Trdaing Account)", "", "47,349,281"),
        ("", "", ""),
        ("By Closing Stock (Work in Process)", "", "-"),
        ("", "", ""),
        ("Total", "", "47,349,281")
    ]
    
    # Write Dr data
    for i, (name, sch_no, amount) in enumerate(dr_data, start=7):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = sch_no
        ws[f'C{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'].font = Font(bold=True)
    
    # Write Cr data
    for i, (name, sch_no, amount) in enumerate(cr_data, start=7):
        ws[f'D{i}'] = name
        ws[f'E{i}'] = sch_no
        ws[f'F{i}'] = amount
        if "Total" in name:
            ws[f'D{i}'].font = Font(bold=True)
            ws[f'F{i}'].font = Font(bold=True)
    
    # Add footer
    apply_footer(ws, 13)

def create_trading_account(wb):
    """Create Trading Account worksheet."""
    ws = wb.create_sheet("Trading Account")
    apply_header_style(ws, "Trading Account For the Year 1st April 2023 to 31st March 2024")
    
    # Add headers
    ws['A5'] = "Dr"
    ws['D5'] = "Cr"
    ws['A5'].alignment = Alignment(horizontal='center')
    ws['D5'].alignment = Alignment(horizontal='center')
    ws['A5'].font = Font(bold=True)
    ws['D5'].font = Font(bold=True)
    
    ws['A6'] = "Particulars"
    ws['B6'] = "Sch. No."
    ws['C6'] = "Amount (Rs.)"
    ws['D6'] = "Particulars"
    ws['E6'] = "Sch. No."
    ws['F6'] = "Amount (Rs.)"
    for cell in ['A6', 'B6', 'C6', 'D6', 'E6', 'F6']:
        ws[cell].font = Font(bold=True)
    
    # Add data
    dr_data = [
        ("To Opening Stock", "", "59,877,197"),
        ("To Cost of Goods Produced (From Manufacturing A/c)", "", "47,349,281"),
        ("To Purchase", "18", "192,730,086"),
        ("To Gross Profit", "", "38,646,331"),
        ("", "", ""),
        ("Total", "", "338,602,895")
    ]
    
    cr_data = [
        ("By Gross Receipt", "17", "215,227,071"),
        ("By Closing Stock (Finished Goods)", "", "123,375,824"),
        ("", "", ""),
        ("", "", ""),
        ("", "", ""),
        ("Total", "", "338,602,895")
    ]
    
    # Write Dr data
    for i, (name, sch_no, amount) in enumerate(dr_data, start=7):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = sch_no
        ws[f'C{i}'] = amount
        if "Total" in name or "Gross Profit" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'].font = Font(bold=True)
    
    # Write Cr data
    for i, (name, sch_no, amount) in enumerate(cr_data, start=7):
        ws[f'D{i}'] = name
        ws[f'E{i}'] = sch_no
        ws[f'F{i}'] = amount
        if "Total" in name:
            ws[f'D{i}'].font = Font(bold=True)
            ws[f'F{i}'].font = Font(bold=True)
    
    # Add footer
    apply_footer(ws, 14)

def create_profit_loss_account(wb):
    """Create Profit and Loss Account worksheet."""
    ws = wb.create_sheet("Profit and Loss Account")
    apply_header_style(ws, "Profit and Loss Account For the Year 1st April 2023 to 31st March 2024")
    
    # Add headers
    ws['A5'] = "Dr"
    ws['D5'] = "Cr"
    ws['A5'].alignment = Alignment(horizontal='center')
    ws['D5'].alignment = Alignment(horizontal='center')
    ws['A5'].font = Font(bold=True)
    ws['D5'].font = Font(bold=True)
    
    ws['A6'] = "Particulars"
    ws['B6'] = "Sch. No."
    ws['C6'] = "Amount (Rs.)"
    ws['D6'] = "Particulars"
    ws['E6'] = "Sch. No."
    ws['F6'] = "Amount (Rs.)"
    for cell in ['A6', 'B6', 'C6', 'D6', 'E6', 'F6']:
        ws[cell].font = Font(bold=True)
    
    # Add data
    dr_data = [
        ("To Indirect Expenses", "", ""),
        ("General and Administrative Exp.", "20", "8,768,927"),
        ("Depreciation", "7", "2,545,411"),
        ("Finance and Bank Expenses", "21", "14,657,301"),
        ("To Net Profit", "", "21,639,494"),
        ("", "", ""),
        ("", "", "47,611,133")
    ]
    
    cr_data = [
        ("By Gross Profit", "", "38,646,331"),
        ("By Indirect Incomes", "19", "8,964,802"),
        ("", "", ""),
        ("", "", ""),
        ("", "", ""),
        ("", "", ""),
        ("", "", "47,611,133")
    ]
    
    # Write Dr data
    for i, (name, sch_no, amount) in enumerate(dr_data, start=7):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = sch_no
        ws[f'C{i}'] = amount
        if "To Indirect Expenses" in name:
            ws[f'A{i}'].font = Font(bold=True)
        if "Net Profit" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'].font = Font(bold=True)
        if i == 13:  # Total row
            ws[f'C{i}'].font = Font(bold=True)
    
    # Write Cr data
    for i, (name, sch_no, amount) in enumerate(cr_data, start=7):
        ws[f'D{i}'] = name
        ws[f'E{i}'] = sch_no
        ws[f'F{i}'] = amount
        if "Gross Profit" in name:
            ws[f'D{i}'].font = Font(bold=True)
            ws[f'F{i}'].font = Font(bold=True)
        if i == 13:  # Total row
            ws[f'F{i}'].font = Font(bold=True)
    
    # Add footer
    apply_footer(ws, 15)

def create_schedule_header(ws, schedule_num, title):
    """Create standard header for schedule worksheets."""
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    
    # Add title
    ws.merge_cells('A1:B1')
    ws['A1'] = "M/s SA Infrastructure Prop. Saiyyed Afsar Ali (PAN No. AZGPS7572J)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:B2')
    ws['A2'] = "16/10, Rehmat Nagar, Ratlam (M.P.)-457001"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:B3')
    ws['A3'] = "Schedule Forming Part of Balance Sheet as at 31st March 2024"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add schedule number and title
    ws['A5'] = f"Schedule No. {schedule_num}"
    ws['A5'].font = Font(bold=True)
    
    ws['A6'] = title
    ws['A6'].font = Font(bold=True)

def create_capital_account(wb):
    """Create Schedule 1 - Capital Account."""
    ws = wb.create_sheet("Schedule 1 - Capital")
    
    # Adjust for 4-column layout
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    
    # Add title
    ws.merge_cells('A1:D1')
    ws['A1'] = "M/s SA Infrastructure Prop. Saiyyed Afsar Ali (PAN No. AZGPS7572J)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = "16/10, Rehmat Nagar, Ratlam (M.P.)-457001"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:D3')
    ws['A3'] = "Schedule Forming Part of Balance Sheet as at 31st March 2024"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add schedule number and title
    ws.merge_cells('A5:D5')
    ws['A5'] = "Schedule No. 1"
    ws['A5'].font = Font(bold=True)
    
    ws.merge_cells('A6:D6')
    ws['A6'] = "Capital Account of Proprietor Saiyyed Afsar Ali"
    ws['A6'].font = Font(bold=True)
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['C8'] = "Particulars" 
    ws['D8'] = "Amount (Rs.)"
    
    for cell in ['A8', 'B8', 'C8', 'D8']:
        ws[cell].font = Font(bold=True)
    
    # Add data
    dr_data = [
        ("To Drawing", "3,755,286"),
        ("To Diversion Fees", "147,900"),
        ("To Sudlife", "510,000"),
        ("", ""),
        ("To Income Tax and TDS", ""),
        ("M/s Saiyyed Akhtar Ali TDS and Income Tax (20%)", "2,082,624"),
        ("", ""),
        ("", "")
    ]
    
    cr_data = [
        ("By Opening Balance", "149,335,711"),
        ("", ""),
        ("By Agriculture Income (As Per Agriculture A/c Annexure)", "1,370,260"),
        ("By star Union Daiichi Lif", "238,330"),
        ("By Profit for the year", "21,639,494"),
        ("", ""),
        ("By Closing Balance", "166,087,985"),
        ("", "")
    ]
    
    # Write Dr data
    for i, (name, amount) in enumerate(dr_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
    
    # Write Cr data
    for i, (name, amount) in enumerate(cr_data, start=9):
        ws[f'C{i}'] = name
        ws[f'D{i}'] = amount
    
    # Add total row
    ws['B17'] = "172,583,795"
    ws['D17'] = "172,583,795"
    ws['B17'].font = Font(bold=True)
    ws['D17'].font = Font(bold=True)

def create_secured_loans(wb):
    """Create Schedule 2 - Secured Loans."""
    ws = wb.create_sheet("Schedule 2 - Secured Loans")
    create_schedule_header(ws, 2, "Secured Loans")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add secured loans data
    loans_data = [
        ("948077610000216", "195,000"),
        ("948077610000217", "197,500"),
        ("948077610000243", "220,000"),
        ("Axis Bank", "1,893,672"),
        ("Bank Of India 948070410000068", "86,866"),
        ("Bank Of India 948070410000098 GECL Loan", "1,249,546"),
        ("HDFC BANK 108503 (Legender)", "3,348,210"),
        ("Hdfc Bank Loan Covid", "19,935"),
        ("HDFC credit card", "-1,143,695"),
        ("Icici Bank 8561", "1,182,266"),
        ("Indusind Bank 45517", "1,195,358"),
        ("Indusind Bank 45517 Refinance 2", "1,242,103"),
        ("New crusher 71000", "1,987,322"),
        ("Sundaram Finance Limited - Covid", "148,864"),
        ("Tata Capital Financial Service", "1,813,384"),
        ("PSB Lap (05941200001322)", "46,627,702"),
        ("Yes Bank Tata Hitachi", "924,563"),
        ("Total", "61,188,596")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(loans_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_unsecured_loans(wb):
    """Create Schedule 3 - Unsecured Loans."""
    ws = wb.create_sheet("Schedule 3 - Unsecured Loans")
    create_schedule_header(ws, 3, "Unsecured Loan")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add unsecured loans data
    loans_data = [
        ("Acme Ferro Alloys(Loan)", "4,000,000"),
        ("Darsh Marketing", "3,000,000"),
        ("GEETANJALI CONSTRUCTION", "1,033,750"),
        ("IQRAR AHMED ANSARI", "125,000"),
        ("LAVISH BHANDARI", "2,500,010"),
        ("MEHRUNISA ANSARI", "400,000"),
        ("Mohammad Ansari", "500,000"),
        ("Nahid Ali", "1,315,964"),
        ("Rakesh Bhandari", "659,435"),
        ("SARITA RAKESH BHANDARI", "131,887"),
        ("Shahid Ali", "650,000"),
        ("SHAHID ANSARI", "700,000"),
        ("Sharif Uddin Ansari", "400,000"),
        ("SULTANA BEE", "1,100,000"),
        ("Suresh Kataria", "648,600"),
        ("TABASUM MIRZA", "500,000"),
        ("VIJAY KATARIA", "7,336,802"),
        ("Vinod Kataria", "648,600"),
        ("TOTAL", "25,650,048")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(loans_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "TOTAL" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_working_capital(wb):
    """Create Schedule 4 - Working Capital Loan."""
    ws = wb.create_sheet("Schedule 4 - Working Capital")
    create_schedule_header(ws, 4, "Working Capital Loan")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    loans_data = [
        ("Bank of India 948030110000086", "9,727,108"),
        ("Bank of India 948030110000122", "29,957,514"),
        ("TOTAL", "39,684,622")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(loans_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "TOTAL" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_agriculture_credit(wb):
    """Create Schedule 4A - Agriculture Credit Overdraft."""
    ws = wb.create_sheet("Schedule 4A - Agriculture Credit")
    create_schedule_header(ws, "4A", "Agriculture Credit Overdraft")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    loans_data = [
        ("HDFC Bank CC-50200047918597", ""),
        ("HDFC Bank CC Account-50200047920894", ""),
        ("HDFC Bank DOD-50200047920204", ""),
        ("Total", "-")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(loans_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_trade_payables(wb):
    """Create Schedule 5 - Trade Payables."""
    ws = wb.create_sheet("Schedule 5 - Trade Payables")
    create_schedule_header(ws, 5, "Trade Payables")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    payables_data = [
        ("ACME FERRO ALLOYS PRIVATE LIMITED", "5,281,096"),
        ("AUTO TRADE RATLAM", "5,198"),
        ("BAJAJ ALLIANZ GENERAL INSURANCE CO. LTD", "22,317"),
        ("Ekta Traders", "51,594"),
        ("E S INFRA SERVE PVT LTD", "23,378"),
        ("Garima Edibles Pvt. Ltd", "38,930,580"),
        ("Geo Infrbuild Company LLP", "75,777"),
        ("GMMCO Limited", "5,379"),
        ("Hindustan Automobiles", "12,352"),
        ("Hindustan Power Solutions", "64,225,610"),
        ("MADHYA BHARAT MACHINERY", "9,995"),
        ("Mahaveer Gases Ratlam", "8,500"),
        ("MEWAR HITECH ENGINEERING LTD", "194,029"),
        ("MPPKVVCL", "764,284"),
        ("Nakoda Steels", "395,035"),
        ("Narayan Diesal", "16,468"),
        ("NAVKAR INFRA", "627,505"),
        ("NEW AMRIT TYRE HOUSE", "24,741"),
        ("Pankaj Auto Spares", "161,000"),
        ("P.CHOURDIYA & COMPANY", "12,514"),
        ("Raninga Ispat Pvt. Ltd.", "20,000,000"),
        ("RC KING INFRATECH", "479,269"),
        ("R C King Traders", "29,187"),
        ("S.B.M. Traders", "226,950"),
        ("SMO FERRO ALLOYS PVT LTD(OXY)", "1,680,726"),
        ("STALCO", "74,984"),
        ("Tulsi Associates", "446,611"),
        ("TVS Mobility Private Limited", "126,025"),
        ("Vatsalya Minerals and Power Pvt.Ltd.", "3,181,104"),
        ("Total", "137,092,207")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(payables_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_provisions(wb):
    """Create Schedule 6 - Provisions."""
    ws = wb.create_sheet("Schedule 6 - Provisions")
    create_schedule_header(ws, 6, "Provisions")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    provisions_data = [
        ("Audit Fees Payable", "50,000"),
        ("Electricity Bill Payable", ""),
        ("Provision for Salary and Labour expenses", "976,800"),
        ("Tds Payable", "633,132"),
        ("TCS payable", "183,741"),
        ("RCM payable", "1,750"),
        ("Total", "1,845,423")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(provisions_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_advance_against_land(wb):
    """Create Schedule 6A - Advance against land."""
    ws = wb.create_sheet("Schedule 6A - Advance against land")
    create_schedule_header(ws, "6A", "Advance against land")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Vijay Kataria", "600,000"),
        ("Total", "600,000")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_advance_against_sales(wb):
    """Create Schedule 6B - Advance against Sales."""
    ws = wb.create_sheet("Schedule 6B - Advance against Sales")
    create_schedule_header(ws, "6B", "Advance against Sales")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Metro Craft Industries", "13,940,000"),
        ("Nu Tech", "11,306,000"),
        ("UPI Sales", "1,678,158"),
        ("PN Alloys", "127,265"),
        ("V.R. Construction", "3,659,057"),
        ("Total", "30,710,480")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_fixed_assets(wb):
    """Create Schedule 7 - Fixed Assets."""
    ws = wb.create_sheet("Schedule 7 - Fixed Assets")
    
    # Set column widths for all columns
    for col in range(1, 10):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15
    
    # Column A needs more width for descriptions
    ws.column_dimensions['A'].width = 30
    
    # Add title
    ws.merge_cells('A1:I1')
    ws['A1'] = "M/s SA Infrastructure Prop. Saiyyed Afsar Ali (PAN No. AZGPS7572J)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "16/10, Rehmat Nagar, Ratlam (M.P.)-457001"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:I3')
    ws['A3'] = "Schedule Forming Part of Balance Sheet as at 31st March 2024"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add schedule title
    ws['A4'] = "Schedule No. 7"
    ws['A4'].font = Font(bold=True)
    
    ws['A5'] = "Fixed Assets"
    ws['A5'].font = Font(bold=True)
    
    # Add complex headers for fixed assets table
    headers = [
        "Particulars",
        "Opening Balance (A)",
        "Assets purchase (B) Not Put To use",
        "Assets purchase (B) Use for < 6 Month",
        "Assets purchase (B) Use for > 6 Month",
        "Assets Sold/ Transfered (C)",
        "Closing Balance (A+B-C) (D)",
        "Depreciation (E)",
        "Net Block (D-E)"
    ]
    
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws[f'{col_letter}7'] = header
        ws[f'{col_letter}7'].font = Font(bold=True)
        ws[f'{col_letter}7'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    
    # Set row height for header row
    ws.row_dimensions[7].height = 40
    
    # Add fixed asset categories and sample data
    asset_data = [
        ["Land and Property", "", "", "", "", "", "", ""],
        ["Land", "67,801,518", "-", "-", "-", "-", "67,801,518", "-", "67,801,518"],
        ["Land", "-", "-", "-", "-", "-", "-", "-", "-"],
        ["", "", "", "", "", "", "", ""],
        ["Plant and Machinery @ 15 %", "", "", "", "", "", "", "", ""],
        ["Mobile", "11,416", "-", "-", "-", "-", "11,416", "1,712", "9,704"],
        ["", "", "", "", "", "", "", ""],
        ["Vehical- Depreciation @ 15%", "", "", "", "", "", "", "", ""],
        ["Dumper U-2523", "1,411,223", "-", "-", "-", "-", "1,411,223", "211,683", "1,199,540"],
        ["Dumper U-2523", "1,411,223", "-", "-", "-", "-", "1,411,223", "211,683", "1,199,540"],
        ["Pockland", "1,636,799", "-", "-", "-", "-", "1,636,799", "245,520", "1,391,280"],
        ["Tata Hitachi Hydraulic", "3,476,745", "-", "-", "-", "512,157", "2,964,588", "521,512", "2,443,076"],
        ["Activa Honda", "14,712", "-", "-", "-", "-", "14,712", "2,207", "12,505"],
        ["Bullet", "55,121", "-", "-", "-", "-", "55,121", "8,268", "46,853"],
        ["Hero HF Deluxe", "26,622", "-", "-", "-", "-", "26,622", "3,993", "22,629"],
        ["Bullet New", "150,370", "-", "-", "-", "-", "150,370", "22,556", "127,815"],
        ["Bike", "10,133", "-", "-", "-", "-", "10,133", "1,520", "8,613"],
        ["TH Backhoe Loader", "1,993,840", "-", "-", "-", "-", "1,993,840", "299,076", "1,694,764"],
        ["Wheel loader", "2,464,549", "-", "-", "-", "-", "2,464,549", "369,682", "2,094,867"],
        ["Verna Car", "985,697.40", "-", "-", "-", "837,843", "147,855", "(0)"],
        ["Fortuner Legender (Car)", "-", "-", "5,134,783.00", "-", "-", "5,134,783", "385,109", "4,749,674"],
        ["", "", "", "", "", "", "", "", ""],
        ["Furniture and Fixtures @ 10%", "", "", "", "", "", "", "", ""],
        ["Furniture", "1,007,033", "-", "-", "-", "-", "1,007,033", "100,703", "906,330"],
        ["LG LED55 MODEL NO.55UN7300PTC", "42,082", "-", "-", "-", "-", "42,082", "4,208", "37,874"],
        ["LG Split AC 1.5 TON", "69,135", "-", "-", "-", "-", "69,135", "6,913", "62,221"],
        ["Computer @ 40%", "", "", "", "", "", "", "", ""],
        ["Computer", "3,024", "-", "-", "-", "-", "3,024", "1,210", "1,814"],
        ["", "", "", "", "", "", "", "", ""],
        ["Total", "82,571,244", "-", "5,134,783", "-", "512,157", "87,046,016", "2,545,411", "83,810,616"]
    ]
    
    # Write fixed asset data
    for i, row_data in enumerate(asset_data, start=8):
        ws[f'A{i}'] = row_data[0]
        
        # Add values for remaining columns if present
        for j, value in enumerate(row_data[1:], start=2):
            col_letter = get_column_letter(j)
            ws[f'{col_letter}{i}'] = value
        
        # Apply bold formatting to category headers and total row
        if row_data[0] in ["Land and Property", "Plant and Machinery @ 15 %", "Vehical- Depreciation @ 15%", 
                           "Furniture and Fixtures @ 10%", "Computer @ 40%", "Total"]:
            ws[f'A{i}'].font = Font(bold=True)
            if row_data[0] == "Total":
                for j in range(2, 10):
                    col_letter = get_column_letter(j)
                    if i < len(asset_data) + 8 and j-2 < len(row_data):
                        ws[f'{col_letter}{i}'].font = Font(bold=True)
    
    # Add footer
    row_num = len(asset_data) + 10
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws[f'A{row_num}'] = "M/s SA Infrastructure"
    
    ws.merge_cells(f'E{row_num}:I{row_num}')
    ws[f'E{row_num}'] = "In terms of our attached report of even date"
    
    ws.merge_cells(f'E{row_num+1}:I{row_num+1}')
    ws[f'E{row_num+1}'] = "For- Saloni Asawa and Associates"
    
    ws.merge_cells(f'E{row_num+2}:I{row_num+2}')
    ws[f'E{row_num+2}'] = "Chartered Accountant"
    
    ws.merge_cells(f'E{row_num+3}:I{row_num+3}')
    ws[f'E{row_num+3}'] = "FRN.-022450C"
    
    ws.merge_cells(f'A{row_num+5}:D{row_num+5}')
    ws[f'A{row_num+5}'] = "Saiyyed Afsar Ali"
    
    ws.merge_cells(f'E{row_num+5}:I{row_num+5}')
    ws[f'E{row_num+5}'] = "Saloni Asawa"
    
    ws.merge_cells(f'A{row_num+6}:D{row_num+6}')
    ws[f'A{row_num+6}'] = "(Proprietor)"
    
    ws.merge_cells(f'E{row_num+6}:I{row_num+6}')
    ws[f'E{row_num+6}'] = "(Membership No.-430411)"
    
    ws.merge_cells(f'A{row_num+8}:D{row_num+8}')
    ws[f'A{row_num+8}'] = "UDIN-"
    
    ws.merge_cells(f'A{row_num+10}:D{row_num+10}')
    ws[f'A{row_num+10}'] = "Date-"
    
    ws.merge_cells(f'A{row_num+11}:D{row_num+11}')
    ws[f'A{row_num+11}'] = "Place- Ratlam"

def create_fixed_assets_mfg(wb):
    """Create Schedule 7A - Fixed Assets (Manufacturing)."""
    ws = wb.create_sheet("Schedule 7A - Fixed Assets (Mfg)")
    
    # Set column widths for all columns
    for col in range(1, 10):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15
    
    # Column A needs more width for descriptions
    ws.column_dimensions['A'].width = 30
    
    # Add title
    ws.merge_cells('A1:I1')
    ws['A1'] = "M/s SA Infrastructure Prop. Saiyyed Afsar Ali (PAN No. AZGPS7572J)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "16/10, Rehmat Nagar, Ratlam (M.P.)-457001"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:I3')
    ws['A3'] = "Schedule Forming Part of Balance Sheet as at 31st March 2024"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add schedule title
    ws['A4'] = "Schedule No. 7(A)"
    ws['A4'].font = Font(bold=True)
    
    ws['A5'] = "Fixed Assets (Manufacturing)"
    ws['A5'].font = Font(bold=True)
    
    # Add complex headers for fixed assets table
    headers = [
        "Particulars",
        "Opening Balance (A)",
        "Assets purchase (B) Not Put To use",
        "Assets purchase (B) Use for < 6 Month",
        "Assets purchase (B) Use for > 6 Month",
        "Assets Sold (C)",
        "Closing Balance (A+B-C) (D)",
        "Depreciation (E)",
        "Net Block (D-E)"
    ]
    
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws[f'{col_letter}7'] = header
        ws[f'{col_letter}7'].font = Font(bold=True)
        ws[f'{col_letter}7'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    
    # Set row height for header row
    ws.row_dimensions[7].height = 40
    
    # Add manufacturing fixed asset data
    asset_data = [
        ["Plant and Machinery @ 15 %", "", "", "", "", "", "", "", ""],
        ["Crusher Machine -2", "123,169.82", "-", "-", "-", "-", "123,170", "18,475", "104,694"],
        ["Crusher Machine -3", "187,831.65", "-", "-", "-", "-", "187,832", "28,175", "159,657"],
        ["Stone Crusher Plant", "4,782,256.33", "-", "-", "1,200,000", "300,000", "5,682,256", "897,338.45", "4,784,918"],
        ["Plant and Machinery", "2,316,750.92", "-", "-", "186,014", "-", "2,502,765", "375,414.74", "2,127,350"],
        ["Steel", "8,528,554.31", "-", "515,510", "1,314,371", "-", "10,358,435", "1,553,765.30", "8,804,670"],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["Total", "15,938,563", "-", "515,510", "2,700,385", "300,000", "18,854,458", "2,873,169", "15,981,289"]
    ]
    
    # Write manufacturing fixed asset data
    for i, row_data in enumerate(asset_data, start=8):
        ws[f'A{i}'] = row_data[0]
        
        # Add values for remaining columns
        for j, value in enumerate(row_data[1:], start=2):
            col_letter = get_column_letter(j)
            ws[f'{col_letter}{i}'] = value
        
        # Apply bold formatting to category headers and total row
        if row_data[0] in ["Plant and Machinery @ 15 %", "Total"]:
            ws[f'A{i}'].font = Font(bold=True)
            if row_data[0] == "Total":
                for j in range(2, 10):
                    col_letter = get_column_letter(j)
                    if j-2 < len(row_data):
                        ws[f'{col_letter}{i}'].font = Font(bold=True)
    
    # Add footer
    row_num = len(asset_data) + 10
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws[f'A{row_num}'] = "M/s SA Infrastructure"
    
    ws.merge_cells(f'E{row_num}:I{row_num}')
    ws[f'E{row_num}'] = "In terms of our attached report of even date"
    
    ws.merge_cells(f'E{row_num+1}:I{row_num+1}')
    ws[f'E{row_num+1}'] = "For- Saloni Asawa and Associates"
    
    ws.merge_cells(f'E{row_num+2}:I{row_num+2}')
    ws[f'E{row_num+2}'] = "Chartered Accountant"
    
    ws.merge_cells(f'E{row_num+3}:I{row_num+3}')
    ws[f'E{row_num+3}'] = "FRN.-022450C"
    
    ws.merge_cells(f'A{row_num+5}:D{row_num+5}')
    ws[f'A{row_num+5}'] = "Saiyyed Afsar Ali"
    
    ws.merge_cells(f'E{row_num+5}:I{row_num+5}')
    ws[f'E{row_num+5}'] = "Saloni Asawa"
    
    ws.merge_cells(f'A{row_num+6}:D{row_num+6}')
    ws[f'A{row_num+6}'] = "(Proprietor)"
    
    ws.merge_cells(f'E{row_num+6}:I{row_num+6}')
    ws[f'E{row_num+6}'] = "(Membership No.-430411)"
    
    ws.merge_cells(f'A{row_num+8}:D{row_num+8}')
    ws[f'A{row_num+8}'] = "UDIN-"
    
    ws.merge_cells(f'A{row_num+10}:D{row_num+10}')
    ws[f'A{row_num+10}'] = "Date-"
    
    ws.merge_cells(f'A{row_num+11}:D{row_num+11}')
    ws[f'A{row_num+11}'] = "Place- Ratlam"

def create_investments(wb):
    """Create Schedule 8 - Investments."""
    ws = wb.create_sheet("Schedule 8 - Investments")
    create_schedule_header(ws, 8, "Investments")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Gold", "-"),
        ("Investment in Partnership Firm- M/s Saiyyed Akhter Ali (20% )", "66,136,713"),
        ("Investment in Partnership Firm- M/s Gharib Nawaj Infra (12% )", "-"),
        ("Investment in Private Limited Company -SMO Ferro Alloys Private Limited", "45,930,000"),
        ("(9300 Shares @ 100 Rs. Per share Non Traded)", ""),
        ("Total", "112,066,713")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_long_term_loans(wb):
    """Create Schedule 9 - Long Term Loan and Advances."""
    ws = wb.create_sheet("Schedule 9 - Long Term Loans")
    create_schedule_header(ws, 9, "Long Term Loans and Advances")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Rakhi", "500,000"),
        ("Smo Ferro Alloys Pvt Ltd (Adv)", "27,913,932"),
        ("V S M Trade Link", "782,500"),
        ("Total", "29,196,432")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_deposits(wb):
    """Create Schedule 10 - Deposits."""
    ws = wb.create_sheet("Schedule 10 - Deposits")
    create_schedule_header(ws, 10, "Deposits")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Security Deposits (Long Term)", ""),
        ("Khanij Adhikardi Khanij Shaka Jila Ratlam", "300,000"),
        ("MPEB Deposit", "1,436,600"),
        ("Total", "1,736,600")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name or "Security Deposits (Long Term)" in name:
            ws[f'A{i}'].font = Font(bold=True)
        if "Total" in name:
            ws[f'B{i}'].font = Font(bold=True)

def create_trade_receivables(wb):
    """Create Schedule 11 - Trade Receivables."""
    ws = wb.create_sheet("Schedule 11 - Trade Receivables")
    create_schedule_header(ws, 11, "Trade Receivables")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("AHA Alloybreeze Industries Pvt. Ltd.", "7,000,267"),
        ("A.M. Enterprises", "36,799"),
        ("Balaji Agro Foods", "879,465"),
        ("Bhimji Velji Sorathia Construction Pvt Ltd", "148,394"),
        ("CREDIT INFORMATION BUREAU INDIA LTD", "2,400"),
        ("Mohammaed Juber", "415,200"),
        ("M/s R.P. Rathore", "3,709,263"),
        ("Prathvi Infrastructure Pvt Ltd", "22,077"),
        ("Raamesh Construction", "267,662"),
        ("Samdariya Builders Pvt Ltd", "12,585"),
        ("Samdariya Builders (Ratlam) Pvt Ltd", "283,997"),
        ("Shri Madhav International", "3,309"),
        ("SML Industries Pvt Ltd", "63,434,911"),
        ("M/s Saiyyed Akhtar Ali", "7,429,531"),
        ("Total", "83,645,859")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_trade_receivables_assets(wb):
    """Create Schedule 12 - Trade Receivables for Assets."""
    ws = wb.create_sheet("Schedule 12 - Assets Receivables")
    create_schedule_header(ws, 12, "Trade Receivables for Assets")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Anjum Bee", "25,000"),
        ("Kiranbala Parihar W/o Arjan Parihar", "25,000"),
        ("Total", "50,000")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_duties_and_taxes(wb):
    """Create Schedule 13 - Duties and Taxes."""
    ws = wb.create_sheet("Schedule 13 - Duties and Taxes")
    create_schedule_header(ws, 13, "Duties and Taxes")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Goods and Service Tax (Input Tax Credit)", "3,658,741"),
        ("TCS Receivable", "109,718"),
        ("TDS Receivable", "1,630,440"),
        ("Total", "5,398,899")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_short_term_loans(wb):
    """Create Schedule 14 - Short Term Loan and Advances."""
    ws = wb.create_sheet("Schedule 14 - Short Term Loans")
    create_schedule_header(ws, 14, "Short Term Loan and Advances")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("MO Infra", "749,126"),
        ("Total", "749,126")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_advances_against_purchase(wb):
    """Create Schedule 14A - Advances against purchase."""
    ws = wb.create_sheet("Schedule 14A - Purchase Advances")
    create_schedule_header(ws, "14A", "Advances against purchase")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("AKASHGANGA CONSTRUCTIONAL MACHINES PVT LTD", "1,079"),
        ("GITANJALI CONSTRUCTION", "15,570"),
        ("M.A.TRADERS", "13,465"),
        ("MT Enterprises", "65,500"),
        ("Praveen Sulphochem", "4,340"),
        ("Saroj Bala Maheshwari", "79,912"),
        ("Shri Ram Fuel (Gharot)", "1,420,000"),
        ("Siddhi Vinayak Traders", "228,979"),
        ("Total", "1,828,845")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_cash_and_bank(wb):
    """Create Schedule 15 - Cash and Bank Balance."""
    ws = wb.create_sheet("Schedule 15 - Cash and Bank")
    create_schedule_header(ws, 15, "Cash and Bank Balance")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Cash In Hand", "4,529,122"),
        ("Bank Accounts", ""),
        ("Bank of India - Saving Account (948010110009116)", "77,692"),
        ("Hdfc Saving Account50100343676346", "6,232"),
        ("Axis Bank", "406,111"),
        ("Total", "5,019,157")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name or "Bank Accounts" in name:
            ws[f'A{i}'].font = Font(bold=True)
        if "Total" in name:
            ws[f'B{i}'].font = Font(bold=True)

def create_direct_expenses(wb):
    """Create Schedule 16 - Direct Expenses."""
    ws = wb.create_sheet("Schedule 16 - Direct Expenses")
    create_schedule_header(ws, 16, "Direct Expenses")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Consumables, Spare Parts and Others", "2,335,230"),
        ("Diesel & Petrol Exp.", "25,162,656"),
        ("District Mining Fund", "302,132"),
        ("Electricity Exp.", "7,468,029"),
        ("Freight / Cartgae / Hammali", "30,454"),
        ("Labour Exp.", "1,000,575"),
        ("Machinery Running Expenses", "1,092,561"),
        ("Vehicle spares", "2,831,017"),
        ("Lease Machinery / Plant Rental", "1,444,916"),
        ("Royalty Exp. ( RCM/ITC)", "2,597,840"),
        ("Contract Expenses", "210,102"),
        ("Regional Transport Office Expenses", "600"),
        ("Total", "44,476,112")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_gross_receipt(wb):
    """Create Schedule 17 - Gross Receipt."""
    ws = wb.create_sheet("Schedule 17 - Gross Receipt")
    create_schedule_header(ws, 17, "Gross Receipt")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Gitti", "27,337,187"),
        ("Sales 18%(silico manganesse)", "50,190,000"),
        ("ferro silicon sales", "72,055,000"),
        ("Sales Medium Carbon Ferro Manganese", "33,465,000"),
        ("Work Contract 18%", "32,179,884"),
        ("Total", "215,227,071")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_purchase(wb):
    """Create Schedule 18 - Purchase."""
    ws = wb.create_sheet("Schedule 18 - Purchase")
    create_schedule_header(ws, 18, "Purchase")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Cement", "4,540,086"),
        ("GST Purchase 18%", "150,265,000"),
        ("GST Purchase 18% (Silico Manganese)", "37,925,000"),
        ("Work contract purchase", "-"),
        ("Total", "192,730,086")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_indirect_income(wb):
    """Create Schedule 19 - Indirect Income."""
    ws = wb.create_sheet("Schedule 19 - Indirect Income")
    create_schedule_header(ws, 19, "Indirect Income")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Interest from M/s Saiyyed Akhtar Ali", "3,816,738"),
        ("Interest Received for Bank", "19,755"),
        ("Interest Recevied from SMO Ferro Alloys", "4,133,700"),
        ("Interest Recevied For MPPKVVCL", "109,206"),
        ("Refund", "1,005"),
        ("Discount A/c received", "94,098"),
        ("Share of profit from firm M/s Saiyyed Akhtar Ali (1% Partner)", "790,300"),
        ("Total", "8,964,802")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_indirect_expenses(wb):
    """Create Schedule 20 - Indirect Expenses."""
    ws = wb.create_sheet("Schedule 20 - Indirect Expenses")
    create_schedule_header(ws, 20, "Indirect Expenses")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Audit Fees", "25,000"),
        ("Cibil Expenses", "6,000"),
        ("lease /machinery rent", ""),
        ("Outstanding demand", "41,000"),
        ("Late fees", "500"),
        ("Insurance Expenses", ""),
        ("Interest on RCM", ""),
        ("Interest on TCS", ""),
        ("Interest on GST", "2,526"),
        ("Interest On TDS", ""),
        ("Road Tax", ""),
        ("Mining Expenses", "389,960"),
        ("Office Exp.", ""),
        ("Professional expenses", "508,078"),
        ("Professional Tax", "5,000"),
        ("Repair and Maintenance Exp.", "4566673"),
        ("Rating", ""),
        ("Salary & Staff Wellfare", "700,266"),
        ("Stationary Expenses", "100,000"),
        ("Stock audit fees", "17,700"),
        ("Round off", "-18"),
        ("Vehical Running and Maintenance Exp.", "274,929"),
        ("Telephone & mobile exp", ""),
        ("Travelling and hospitility expenses", "2,131,313"),
        ("Total", "8768927.48")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

def create_finance_expenses(wb):
    """Create Schedule 21 - Finance and Bank expenses."""
    ws = wb.create_sheet("Schedule 21 - Finance Expenses")
    create_schedule_header(ws, 21, "Finance and Bank expenses")
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    
    # Add data
    data = [
        ("Bank expenses", "5"),
        ("Bank commission & other charges", "358,253"),
        ("Interest on CC", "5,074,122"),
        ("Interest on term loan", "7,642,509"),
        ("Interest on unsecured loan", "1,079,258"),
        ("Interest on working capital limit", ""),
        ("Interest on vehicle loan", ""),
        ("Interest on KCC", ""),
        ("Brokerage Expenses", "36,750"),
        ("Other interest", "4,420"),
        ("Loan Processing Expneses", "461,984"),
        ("Total", "14,657,301")
    ]
    
    # Write data
    for i, (name, amount) in enumerate(data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
        if "Total" in name:
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)
    
    # Add standard footer
    ws.merge_cells('A22:B22')
    ws['A22'] = "M/s SA Infrastructure"
    
    ws.merge_cells('A23:B23')
    ws['A23'] = "In terms of our attached report of even date"
    
    ws.merge_cells('A24:B24')
    ws['A24'] = "For- Saloni Asawa and Associates"
    
    ws.merge_cells('A25:B25')
    ws['A25'] = "Chartered Accountant"
    
    ws.merge_cells('A26:B26')
    ws['A26'] = "FRN.-022450C"
    
    ws.merge_cells('A28:B28')
    ws['A28'] = "Saiyyed Afsar Ali"
    
    ws.merge_cells('A29:B29')
    ws['A29'] = "(Proprietor)"
    
    ws.merge_cells('A30:B30')
    ws['A30'] = "Saloni Asawa"
    
    ws.merge_cells('A31:B31')
    ws['A31'] = "(Membership No.-430411)"
    
    ws.merge_cells('A33:B33')
    ws['A33'] = "UDIN-"
    
    ws.merge_cells('A35:B35')
    ws['A35'] = "Date-"
    
    ws.merge_cells('A36:B36')
    ws['A36'] = "Place- Ratlam"

def create_agriculture_account(wb):
    """Create Annexure 1 - Agriculture Account."""
    ws = wb.create_sheet("Annexure 1 - Agriculture")
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    
    # Add title
    ws.merge_cells('A1:D1')
    ws['A1'] = "M/s SA Infrastructure Prop. Saiyyed Afsar Ali (PAN No. AZGPS7572J)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = "16/10, Rehmat Nagar, Ratlam (M.P.)-457001"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:D3')
    ws['A3'] = "Annexures Forming Part of Balance Sheet as at 31st March 2024"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add annexure title
    ws.merge_cells('A5:D5')
    ws['A5'] = "Annexure-1"
    ws['A5'].font = Font(bold=True)
    
    ws.merge_cells('A6:D6')
    ws['A6'] = "Agriculture A/c"
    ws['A6'].font = Font(bold=True)
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['C8'] = "Particulars"
    ws['D8'] = "Amount (Rs.)"
    
    for cell in ['A8', 'B8', 'C8', 'D8']:
        ws[cell].font = Font(bold=True)
    
    # Add data
    dr_data = [
        ("To Agriculture Expenses", "210,700"),
        ("", ""),
        ("By Closing Balance", "1,370,260"),
        ("(Transfer to Capital Account)", ""),
        ("", "")
    ]
    
    cr_data = [
        ("By Opening Balance", "-"),
        ("By Agriculture Income", "1,580,960"),
        ("", ""),
        ("", ""),
        ("", "")
    ]
    
    # Write Dr data
    for i, (name, amount) in enumerate(dr_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
    
    # Write Cr data
    for i, (name, amount) in enumerate(cr_data, start=9):
        ws[f'C{i}'] = name
        ws[f'D{i}'] = amount
    
    # Add total row
    ws['B14'] = "1,580,960"
    ws['D14'] = "1,580,960"
    ws['B14'].font = Font(bold=True)
    ws['D14'].font = Font(bold=True)
    
    # Add standard footer
    ws.merge_cells('A16:B16')
    ws['A16'] = "M/s SA Infrastructure"
    
    ws.merge_cells('C16:D16')
    ws['C16'] = "In terms of our attached report of even date"
    
    ws.merge_cells('C17:D17')
    ws['C17'] = "For- Saloni Asawa and Associates"
    
    ws.merge_cells('C18:D18')
    ws['C18'] = "Chartered Accountant"
    
    ws.merge_cells('C19:D19')
    ws['C19'] = "FRN.-022450C"
    
    ws.merge_cells('A21:B21')
    ws['A21'] = "Saiyyed Afsar Ali"
    
    ws.merge_cells('C21:D21')
    ws['C21'] = "Saloni Asawa"
    
    ws.merge_cells('A22:B22')
    ws['A22'] = "(Proprietor)"
    
    ws.merge_cells('C22:D22')
    ws['C22'] = "(Membership No.-430411)"
    
    ws.merge_cells('A24:B24')
    ws['A24'] = "UDIN-"
    
    ws.merge_cells('A26:B26')
    ws['A26'] = "Date-"
    
    ws.merge_cells('A27:B27')
    ws['A27'] = "Place- Ratlam"

def create_land_account(wb):
    """Create Annexure 2 - Land Account."""
    ws = wb.create_sheet("Annexure 2 - Land Account")
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    
    # Add title
    ws.merge_cells('A1:D1')
    ws['A1'] = "M/s SA Infrastructure Prop. Saiyyed Afsar Ali (PAN No. AZGPS7572J)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = "16/10, Rehmat Nagar, Ratlam (M.P.)-457001"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:D3')
    ws['A3'] = "Annexures Forming Part of Balance Sheet as at 31st March 2024"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add annexure title
    ws.merge_cells('A5:D5')
    ws['A5'] = "Annexure-2"
    ws['A5'].font = Font(bold=True)
    
    ws.merge_cells('A6:D6')
    ws['A6'] = "Land (168/31) Account A/c"
    ws['A6'].font = Font(bold=True)
    
    # Add column headers
    ws['A8'] = "Particulars"
    ws['B8'] = "Amount (Rs.)"
    ws['C8'] = "Particulars"
    ws['D8'] = "Amount (Rs.)"
    
    for cell in ['A8', 'B8', 'C8', 'D8']:
        ws[cell].font = Font(bold=True)
    
    # Add data
    dr_data = [
        ("To Opening Stock", "2,736,198"),
        ("(11724 Sq Ft.)", ""),
        ("", ""),
        ("", ""),
        ("", "")
    ]
    
    cr_data = [
        ("By Sales", ""),
        ("", ""),
        ("By Closing Stock", "2,736,198"),
        ("(11724 Sq ft)", ""),
        ("", "")
    ]
    
    # Write Dr data
    for i, (name, amount) in enumerate(dr_data, start=9):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = amount
    
    # Write Cr data
    for i, (name, amount) in enumerate(cr_data, start=9):
        ws[f'C{i}'] = name
        ws[f'D{i}'] = amount
    
    # Add profit and total rows
    ws['A14'] = "By Profit on sale of land"
    ws['A15'] = "(Transfer to Capital Account)"
    ws['B14'] = "-"
    
    ws['B16'] = "2,736,198"
    ws['D16'] = "2,736,198"
    ws['B16'].font = Font(bold=True)
    ws['D16'].font = Font(bold=True)
    
    # Add standard footer
    ws.merge_cells('A18:B18')
    ws['A18'] = "M/s SA Infrastructure"
    
    ws.merge_cells('C18:D18')
    ws['C18'] = "In terms of our attached report of even date"
    
    ws.merge_cells('C19:D19')
    ws['C19'] = "For- Saloni Asawa and Associates"
    
    ws.merge_cells('C20:D20')
    ws['C20'] = "Chartered Accountant"
    
    ws.merge_cells('C21:D21')
    ws['C21'] = "FRN.-022450C"
    
    ws.merge_cells('A23:B23')
    ws['A23'] = "Saiyyed Afsar Ali"
    
    ws.merge_cells('C23:D23')
    ws['C23'] = "Saloni Asawa"
    
    ws.merge_cells('A24:B24')
    ws['A24'] = "(Proprietor)"
    
    ws.merge_cells('C24:D24')
    ws['C24'] = "(Membership No.-430411)"
    
    ws.merge_cells('A26:B26')
    ws['A26'] = "UDIN-"
    
    ws.merge_cells('A28:B28')
    ws['A28'] = "Date-"
    
    ws.merge_cells('A29:B29')
    ws['A29'] = "Place- Ratlam"

# Main function to run the code
if __name__ == "__main__":
    create_financial_statements_excel()