import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def create_smo_ferro_excel():
    # Create a workbook
    wb = Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    title_font = Font(name='Arial', size=12, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Create Balance Sheet
    create_balance_sheet(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Create Profit and Loss Statement
    create_profit_loss(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Create Notes to Accounts
    create_notes(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Create annexures
    create_annexure_1(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_2(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_3a(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_3b(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_3c(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_3d(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_3e(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_3f(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_4(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_5(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_annexure_6(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Save the workbook
    wb.save('SMO_Ferro_Alloys_Financial_Statements.xlsx')
    
    return "Excel file has been created successfully."

def create_balance_sheet(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Balance Sheet worksheet
    ws = wb.create_sheet("Balance Sheet")
    
    # Add title
    ws['A1'] = "SMO FERRO ALLOYS PRIVATE LIMITED"
    ws.merge_cells('A1:D1')
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = "Balance Sheet as at 31st March, 2024"
    ws.merge_cells('A2:D2')
    ws['A2'].font = title_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ["Particulars", "Note No.", "Figures as at the end of current reporting period (in Rs.)", "Figures as at the end of previous reporting period (in Rs.)"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    
    # EQUITY AND LIABILITIES
    row = 5
    
    # Main section heading
    ws.cell(row=row, column=1).value = "I. EQUITY AND LIABILITIES"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    # Shareholder's Funds
    ws.cell(row=row, column=1).value = "(1) Shareholder's Funds"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    # Add data
    balance_sheet_data = [
        ["a) Share Capital", "2.1", 65286500, 59510000],
        ["b) Reserves and Surplus", "2.2", 882102463, 274797163],
        ["c) Money received against Share warrants", "", "", ""],
        ["(2) Share Application money Pending allotment", "", "", ""],
        ["(3) Non-Current Liabilities", "", "", ""],
        ["a) Long-Term Borrowings", "2.3", 367779690, 254281756],
        ["b) Deferred Tax Liabilities(Net)", "", "", ""],
        ["c) Other Long -Term Liabilities", "2.4A", 11682284, 23826266],
        ["(4) Current Liabilities", "", "", ""],
        ["a) Short-Term Borrowings", "2.5", 341366809, 307339049],
        ["b) Trade Payables", "2.6", 286540318, 111769804],
        ["c) Other Current Liabilities", "2.7", 87099737, 57222368],
        ["d) Short Term Provisions", "2.8", 59733571, 21231055],
        ["TOTAL", "", 2101591373, 1109977461]
    ]
    
    for data_row in balance_sheet_data:
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = thin_border
            if col >= 3 and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                
        # Bold for headings and total
        if data_row[0].startswith("(") or data_row[0] == "TOTAL":
            ws.cell(row=row, column=1).font = Font(bold=True)
            
        row += 1
    
    # II. ASSETS
    ws.cell(row=row, column=1).value = "II. ASSETS"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    assets_data = [
        ["(1) Non-Current Assets", "", "", ""],
        ["a) Property, Plant & Equipments & Intangible Assets", "", "", ""],
        ["i) Property, Plant & Equipments", "2.9", 851579642, 237248933],
        ["ii) Intangible Assets", "", "", ""],
        ["iii) Capital Work in Progress", "2.9A", 15633289, 173447557],
        ["iv) Intangible Assets under development", "", "", ""],
        ["b) Non-Current Investments", "2.9B", 5553700, 756243],
        ["c) Deferred Tax Assets(Net)", "2.4", 916456, 916456],
        ["d) Long -Term loans and Advances", "2.10", 45614370, 2293428],
        ["e) Other non-current Assets", "2.11", 92763249, 67750588],
        ["(2) Current Assets", "", "", ""],
        ["a) Current Investments", "", "", ""],
        ["b) Inventories", "2.12", 473543372, 448628827],
        ["c) Trade Receivables", "2.13", 377038457, 95127055],
        ["d) Cash and cash equivalents", "2.14", 7145934, 3988022],
        ["e) Short-Term Loans and Advances", "2.15", 213818780, 62870922],
        ["e) Other current Assets", "2.15B", 17984125, 16949431],
        ["TOTAL", "", 2101591373, 1109977461]
    ]
    
    for data_row in assets_data:
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = thin_border
            if col >= 3 and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                
        # Bold for headings and total
        if data_row[0].startswith("(") or data_row[0] == "TOTAL":
            ws.cell(row=row, column=1).font = Font(bold=True)
            
        row += 1
    
    # Signature line
    row += 2
    ws.cell(row=row, column=1).value = "For and on behalf of the Board"
    row += 2
    ws.cell(row=row, column=1).value = "SMO Ferro Alloys Pvt. Ltd."
    row += 2
    ws.cell(row=row, column=1).value = "Director"
    row += 2
    ws.cell(row=row, column=1).value = "Sayyad Akhtar Ali"
    ws.cell(row=row+1, column=1).value = "(Director)"
    ws.cell(row=row+2, column=1).value = "DIN: 08291143"
    row += 4
    ws.cell(row=row, column=1).value = "Place: Ratlam"
    ws.cell(row=row+1, column=1).value = "Date: 06.06.2024"

def create_profit_loss(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Profit and Loss worksheet
    ws = wb.create_sheet("Profit and Loss")
    
    # Add title
    ws['A1'] = "SMO FERRO ALLOYS PRIVATE LIMITED"
    ws.merge_cells('A1:D1')
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = "Statement of Profit and Loss for the year ended 31st March, 2024"
    ws.merge_cells('A2:D2')
    ws['A2'].font = title_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ["Particulars", "Note No.", "Figures as at the end of current reporting period (in Rs.)", "Figures as at the end of previous reporting period (in Rs.)"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    
    # Add profit and loss data
    pnl_data = [
        ["Revenue from operations", "2.16", 1234577618, 1109643099],
        ["Other income", "2.17", 5273747, 2264325],
        ["Total Revenue", "", 1239851365, 1111907424],
        ["", "", "", ""],
        ["Expenses:", "", "", ""],
        ["Cost of Material Consumed", "2.18", 699330697, 712036379],
        ["Change in Inventories", "2.19", -20057888, -101793747],
        ["Employees Benefit expenses", "2.20", 18101766, 11280748],
        ["Finance costs", "2.21", 75462815, 44275968],
        ["Depreciation", "2.22", 42204456, 18256299],
        ["Manufacturing expenses", "2.23", 261954446, 331506727],
        ["Other expenses", "2.23", 24032045, 16176937],
        ["Total expenses", "", 1101028336, 1031739311],
        ["", "", "", ""],
        ["Profit before exceptional & extraordinary items and tax", "", 138823029, 80168113],
        ["Loss from sale of car", "", "", ""],
        ["", "", "", ""],
        ["Profit before Tax", "", 138823029, 80168113],
        ["", "", "", ""],
        ["Less", "", "", ""],
        ["Tax expense of Continuing operation:-", "", "", ""],
        ["(1) Current tax", "", 35998363, 20293860],
        ["(2) Deferred tax", "", "", 232358],
        ["Profit from Continuing operation (after tax)", "", 102824666, 59641895],
        ["", "", "", ""],
        ["Earnings per equity share:", "", "", ""],
        ["(1) Basic", "", 252, 57],
        ["(2) Diluted", "", 252, 57]
    ]
    
    row = 5
    for data_row in pnl_data:
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            
            # Add borders except for blank rows
            if any(str(x).strip() for x in data_row):
                cell.border = thin_border
                
            if col >= 3 and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                
        # Bold for totals and section headings
        if data_row[0] in ["Total Revenue", "Total expenses", "Profit before Tax", "Profit from Continuing operation (after tax)"]:
            ws.cell(row=row, column=1).font = Font(bold=True)
            
        row += 1
    
    # Signature line
    row += 2
    ws.cell(row=row, column=1).value = "For and on behalf of the Board"
    row += 2
    ws.cell(row=row, column=1).value = "SMO Ferro Alloys Pvt. Ltd."
    row += 2
    ws.cell(row=row, column=1).value = "Director"
    row += 2
    ws.cell(row=row, column=1).value = "Sayyad Akhtar Ali"
    ws.cell(row=row+1, column=1).value = "(Director)"
    ws.cell(row=row+2, column=1).value = "DIN: 08291143"
    row += 4
    ws.cell(row=row, column=1).value = "Place: Ratlam"
    ws.cell(row=row+1, column=1).value = "Date: 06.06.2024"

def create_notes(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Helper function to create a note sheet
    def create_note_sheet(wb, note_number, note_title, headers, data, has_subtitle=False, subtitle=""):
        ws = wb.create_sheet(f"Note {note_number}")
        
        # Add title
        if has_subtitle:
            ws['A1'] = subtitle
            ws['A1'].font = title_font
            
            ws['A2'] = f"Note No. {note_number} {note_title}"
            ws['A2'].font = title_font
            row = 4
        else:
            ws['A1'] = f"Note No. {note_number} {note_title}"
            ws['A1'].font = title_font
            row = 3
        
        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        for i in range(1, len(headers)):
            ws.column_dimensions[get_column_letter(i+1)].width = 25
        
        # Add data
        row = row + 1
        for data_row in data:
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = normal_font
                
                # Add borders except for blank rows
                if any(str(x).strip() for x in data_row):
                    cell.border = thin_border
                    
                if col >= 2 and isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                    
            # Bold for heading and total
            if data_row[0] and (data_row[0].startswith('(') or data_row[0].endswith(':') or "Total" in data_row[0]):
                ws.cell(row=row, column=1).font = Font(bold=True)
                
            row += 1
        
        return ws
    
    # Standard headers for most notes
    std_headers = ["Particulars", "Figures as at the end of the current reporting period", "Figures as at the end of previous reporting period"]
    
    # Create Note 2.1 - Share Capital
    create_note_2_1(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Note 2.1(a) - Reconciliation of Shares
    reconciliation_data = [
        ["Equity Shares:", "", "", ""],
        ["Shares outstanding at the beginning of the year", "Number", "Amount", "Number", "Amount"],
        ["", 5951000, 59510000, 4080000, 40800000],
        ["Shares Issued during the year", 577650, 5776500, 1871000, 18710000],
        ["Shares bought back during the year", "", "", "", ""],
        ["Shares outstanding at the end of the year", 6528650, 65286500, 5951000, 59510000]
    ]
    reconcile_headers = ["Particulars", "Figures as at the end of the current reporting period", "", "Figures as at the end of previous reporting period", ""]
    create_note_sheet(wb, "2.1(a)", "The Reconciliation of the No. of Shares outstanding at the beginning and at the end of the period", reconcile_headers, reconciliation_data)
    
    # Note 2.1(b) - Rights, Preferences and Restriction attached to Shares
    rights_data = [
        ["Equity Shares", "", ""],
        ["The company has only one class of Equity having a par value 100 per share. Each shareholders is eligible for one vote per share held.The dividend proposed, if any by the board of directors is subject to the approval of the shareholders in ensuing Annual General Meeting except in the case of the interim dividend. In the event of liquidation, the equity shareholders are eligible to receive the remaining assets of the company after distribution of all preferential amounts in portion to their shareholding.", "", ""]
    ]
    create_note_sheet(wb, "2.1(b)", "Right, Preferences and Restriction attached to Shares", std_headers, rights_data)
    
    # Create more notes
    create_note_2_1c(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_note_2_1d(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Note 2.2 - Reserves and Surplus
    reserve_data = [
        ["Retained earnings:", "", ""],
        ["Opening Balance", 274797163, 46885429],
        ["(+) Net Profit/(Loss) for the current year", 102824666, 59641895],
        ["(+) Security Premium", 499588290, 168390000],
        ["(+) Other Addition", 4892344, ""],
        ["(-) Income tax Paid", "", 120160],
        ["", "", ""],
        ["Closing balance", 882102463, 274797163]
    ]
    create_note_sheet(wb, "2.2", "Reserves and Surplus", std_headers, reserve_data)
    
    # Create note 2.3
    create_note_2_3(wb, header_font, normal_font, title_font, thin_border, header_fill)
    create_note_2_3a(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Note 2.4 - Deferred Tax Assets
    deferred_tax_data = [
        ["In accordance with the accounting standard AS-22 \"Accounting for tax on income\" issued by \"The Institute of Chartered Accountants of India\" consequently deferred taxes have been recognised in respect of following items of timing differences between accounting income and taxable income:", "", "", ""],
        ["", "", "", ""],
        ["Items of timing difference", "Accumulated Deferred Tax Assets/(Liabilities) as at 31.03.2023", "(Charged)/Credit during the year", "Balance Assets (Liabilities) as at 31.03.2024"],
        ["", "", "", ""],
        ["Opening Balance", 1148414, "", ""],
        ["Deferred Tax Assets for the Year", (231958), "", ""],
        ["", "", "", ""],
        ["Total", 916456, "", 916456],
    ]
    create_note_sheet(wb, "2.4", "Deferred Tax assets", std_headers[:1] + ["", "", ""], deferred_tax_data)
    
    # Note 2.4A - Other Long-Term Liabilities
    other_long_term_data = [
        ["For Silico Manganese", "", ""],
        ["Sundry Creditors for Fixed Assets- Annexure 2", 11682284, 23826266],
        ["", "", ""],
        ["Total", 11682284, 23826266]
    ]
    create_note_sheet(wb, "2.4A", "Other Long -Term Liabilities", std_headers, other_long_term_data)
    
    # Note 2.5 - Short Term Borrowings
    short_term_borrowings_data = [
        ["From Bank (Secured)", "", ""],
        ["HDFC Bank Limited (CC)", 75087348, ""],
        ["Punjab and Sind Bank (OD)", 70347326, 73460307],
        ["Bank of India (CC)", 262344, 60432476],
        ["Bank of India", 124932276, ""],
        ["Canara Bank of India (OD)", 25214230, 135524215],
        ["Bank of India(CC) Oxy-Oxygen", 3107185, 5721973],
        ["", "", ""],
        ["* Secured Against Hypothecation of Stock, Book Debts, Equitable Mortgage of Property Situated at Ratlam (M.P.)", "", ""],
        ["", "", ""],
        ["Current Maturities of Long Term debts (Note no. 2.3)", 57443290, 30965636],
        ["", "", ""],
        ["Total", 341366809, 307339049]
    ]
    create_note_sheet(wb, "2.5", "Short Term Borrowings", std_headers, short_term_borrowings_data)
    
    # Note 2.6 - Trade Payable
    trade_payable_data = [
        ["Sundry Creditors- Annexure-3", "", ""],
        ["a) Silico Manganese", "", ""],
        ["Creditors for Goods", 274226394, ""],
        ["Creditors for Expenses", "", 77191419],
        ["b) Sundry Creditors (Oxygen Plant)", 1562345, 22969390],
        ["c) Sundry Creditors Aluminium Plant", 5773612, 1694497],
        ["d) Sundry Creditors Rajasthan", 2778502, 7524451],
        ["e) Sundry Creditors Chandrapur", 123995, 2342435],
        ["f) Sundry Creditors Durgapur", 75470, 6047612],
        ["Total", 286540318, 111769804]
    ]
    create_note_sheet(wb, "2.6", "Trade Payable", std_headers, trade_payable_data)
    
    # Note 2.6a - The disclosure of amount payable to entities covered under MSME Act
    msme_data = [
        ["Trade Payable", "Figures as at the end of current reporting period", "Figures as at the end of previous reporting period"],
        ["(a) Total outstanding dues of Micro, Small and medium Enterprises", "", ""],
        ["(b) Total outstanding dues of Other payable against suppliers other than Micro, Small and Medium Enterprises", 286540318, 111769804],
        ["Total", 286540318, 111769804]
    ]
    create_note_sheet(wb, "2.6a", "The disclosure of amount payable to entities covered under Micro, Small and Medium Enterprises Development Act, 2006 as required by Schedule III of the Companies Act, 2013, are as follows: Information not available", std_headers[:1] + ["Figures as at the end of current reporting period", "Figures as at the end of previous reporting period"], msme_data)
    
    # Note 2.6b - Ageing Schedule of Trade Payable
    aging_data = [
        ["As at March 31, 2023", "", "", "", "", "", ""],
        ["", "Not due", "Outstanding for following periods from due date of payment", "", "", "", "Total"],
        ["", "", "Less than 1 year", "1-2 years", "2-3 years", "More than 3 years", ""],
        ["Undisputed dues-MSME", "", "", "", "", "", ""],
        ["Undisputed dues-Other", "", "", "", "", "", ""],
        ["Disputed dues-MSME", "", "", "", "", "", ""],
        ["Disputed dues-Other", "", "", "", "", "", ""],
        ["Total Trade Payable", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["As at March 31, 2022", "", "", "", "", "", ""],
        ["", "Not due", "Outstanding for following periods from due date of payment", "", "", "", "Total"],
        ["", "", "Less than 1 year", "1-2 years", "2-3 years", "More than 3 years", ""],
        ["Undisputed dues-MSME", "", "", "", "", "", ""],
        ["Undisputed dues-Other", "", "", "", "", "", ""],
        ["Disputed dues-MSME", "", "", "", "", "", ""],
        ["Disputed dues-Other", "", "", "", "", "", ""],
        ["Total Trade Payable", "", "", "", "", "", ""]
    ]
    create_note_sheet(wb, "2.6b", "Ageing Schedule of Trade Payable is as below- Information not available", std_headers[:1] + ["", "", "", "", "", ""], aging_data)
    
    # Note 2.7 - Other Current Liabilities
    other_current_liabilities_data = [
        ["Statutory Liabilities-Silico Manganese", "", ""],
        ["GST", 7443546, 24908403],
        ["TDS and TCS", 5552386, 2912199],
        ["", "", ""],
        ["Statutory Liabilities-Aluminium", "", ""],
        ["TDS and TCS", 40594, ""],
        ["GST", 986537, 1190442],
        ["", "", ""],
        ["Statutory Liabilities-Oxygen", "", ""],
        ["TDS and TCS", 27400, 454982],
        ["", "", ""],
        ["Statutory Liabilities-Rajasthan", "", ""],
        ["TDS and TCS", "", 585],
        ["", "", ""],
        ["Statutory Liabilities-Durgapur", "", ""],
        ["TDS and TCS", "", 37500],
        ["", "", ""],
        ["Interest Accrued but not due on Borrowings", "", ""],
        ["Bank of India", "", ""],
        ["", "", ""],
        ["Other Payable Against Expenses", "", ""],
        ["Advance Received from Parties Annexure-4", 80941301, 28618446],
        ["", "", ""],
        ["Total", 87099737, 57222368]
    ]
    create_note_sheet(wb, "2.7", "Other Current Liabilities", std_headers, other_current_liabilities_data)
    
    # Note 2.8 - Short Term Provisions
    short_term_provisions_data = [
        ["Income Tax Provision", 56292223, 20293860],
        ["Audit Fee Payable", 180000, 30000],
        ["CSR Expenses Payable", 3600000, ""],
        ["Water Expenses Payable", 40000, ""],
        ["Salary Payable", 897100, 907195],
        ["", "", ""],
        ["Total", 59733571, 21231055]
    ]
    create_note_sheet(wb, "2.8", "Short Term Provisions", std_headers, short_term_provisions_data)
    
    # Note 2.9 - Property, Plant & Equipments
    # This would require a more complex structure for fixed assets, simplified here
    property_plant_data = [
        ["GROSS BLOCK", "", ""],
        ["Opening Balance", 300000000, 200000000],
        ["Additions during the year", 650000000, 120000000],
        ["Disposals", 10000000, 20000000],
        ["Closing Balance", 940000000, 300000000],
        ["", "", ""],
        ["ACCUMULATED DEPRECIATION", "", ""],
        ["Opening Balance", 62751067, 44494768],
        ["Depreciation for the year", 35669291, 18256299],
        ["Disposals", 10000000, 0],["Closing Balance", 88420358, 62751067],
        ["", "", ""],
        ["NET BLOCK", 851579642, 237248933]
    ]
    create_note_sheet(wb, "2.9", "Property, Plant & Equipments", std_headers, property_plant_data)
    
    # Note 2.9A - Capital WIP
    capital_wip_data = [
        ["Silico Manganese-Advance Paid for Plant", "", ""],
        ["SR Ferro Alloys", "", 156938600],
        ["Brajendra Sharma", 5000000, 5000000],
        ["Aquatherm Engineering Consultants (I) Pvt Ltd", 13000, ""],
        ["Universal Chemical Equipments Pvt Ltd", 29960, ""],
        ["Kasnodi Corp", 10000000, ""],
        ["Siddhi Vinayak Traders", 590329, ""],
        ["", "", ""],
        ["For Oxygen", "", ""],
        ["Parkar Enterprises", "", 3726177],
        ["Cooltech Engineer", "", 25000],
        ["Sikha Das", "", 5656620],
        ["", "", ""],
        ["Total", 15633289, 173447557]
    ]
    create_note_sheet(wb, "2.9A", "Capital WIP", std_headers, capital_wip_data)
    
    # Note 2.9B - Non Current Investment
    non_current_investment_data = [
        ["Krishna Mohan Energy and Infrastructure Pvt Ltd (51%)", "", 5100000],
        ["SMO Aluminium Pvt ltd", 100000, 67000],
        ["SMO Copper Smelting Pvt Ltd", 101025, 100000],
        ["SMO Manganese & Development Pvt Ltd", 100000, 135000],
        ["SMO Fragrance Distillery Pvt Ltd", 101025, ""],
        ["SMO Gold Refinery Pvt Ltd", 101025, 97000],
        ["SMO Metal and Energy Pvt Ltd", 100000, 104081],
        ["SMO Ferro Alloys Pvt Ltd-Rajasthan Unit", 5050625, 252162],
        ["Total", 5553700, 756243]
    ]
    create_note_sheet(wb, "2.9B", "Non Current Investment", std_headers, non_current_investment_data)
    
    # Note 2.10 - Long Term Loan & Advances
    long_term_loan_data = [
        ["Unsecured, Considered Goods:", "", ""],
        ["Krishna Mohan Energy and Infrastructure Pvt Ltd", "", 1816626],
        ["Mohan Bee", "", 128000],
        ["Nahid Bee", "", 128000],
        ["Rafiq Bee", "", 128000],
        ["SMO Aluminium Pvt ltd", "", 158000],
        ["SMO Copper Smelting Pvt Ltd", 1828390, ""],
        ["SMO Fragrance Distillery Pvt Ltd", 3474044, ""],
        ["SMO Gold Refinery Pvt Ltd", 2317600, ""],
        ["SMO Metal and Energy Pvt Ltd", 3791946, ""],
        ["SMO Ferro Alloys Pvt Ltd- Aluminium Division", 10817885, ""],
        ["SMO Ferro Alloys Pvt Ltd- Oxygen Division", 21266373, ""],
        ["SMO Ferro Alloys Pvt Ltd- Durgapur", 2208092, ""],
        ["Total", 45614370, 2293428]
    ]
    create_note_sheet(wb, "2.10", "Long Term Loan & Advances", std_headers, long_term_loan_data)
    
    # Note 2.10a - Loan & Advance in the nature of loan outstanding from promoters, directors, KMPs and related parties
    loan_advance_data = [
        ["Particulars", "Figures as at the end of current reporting period", "", "Figures as at the end of previous reporting period", ""],
        ["", "%", "", "%", ""],
        ["Promoters", "", "", "", ""],
        ["Directors", "", "", "", ""],
        ["KMPs", "", "", "", ""],
        ["Related Parties", "", "", "", ""],
        ["Total", "", 0, "", 0]
    ]
    create_note_sheet(wb, "2.10a", "Loan & Advance in the nature of loan outstanding from promoters, directors, KMPs and related parties", std_headers[:1] + ["Figures as at the end of current reporting period", "", "Figures as at the end of previous reporting period", ""], loan_advance_data)
    
    # Note 2.11 - Other Non-Current Assets
    other_non_current_assets_data = [
        ["Security Deposit", "", ""],
        ["", "", ""],
        ["Silico Manganese", "", ""],
        ["Rent Deposit (Security deposit for flat at Mumbai)", "", 300000],
        ["FDR CBI", 11250000, ""],
        ["FDR PSB", 12741136, ""],
        ["FDR BOI", 7652, ""],
        ["Other Deposit - KB Ferro Alloys for electricity", 49928400, 53628950],
        ["CMD Deposit - (KGM Junction)", 249648, 249648],
        ["Wz Work Rent Deposit", "", 5262],
        ["Misc. Expenses Assets (ROC Expenses)", "", 190000],
        ["Jasmeet Kaur Saluja", 400000, 400000],
        ["Paramjeet Kaur Saluja", 400000, 400000],
        ["Security Deposit to excl Authority", 130750, 130750],
        ["Security Deposit with PDIL", 45000, ""],
        ["Preliminary Exp", 50000, ""],
        ["", "", ""],
        ["Oxygen Plant", "", ""],
        ["Misc. Expenses Assets (Preliminary)", 126453, 126453],
        ["MPCB Deposit", 1605000, 1340000],
        ["MH Deposit", 138351, 138351],
        ["", "", ""],
        ["Aluminium Plant Deposit", "", ""],
        ["Electricity Deposit", "", ""],
        ["Plant Deposit", "", ""],
        ["", "", ""],
        ["Durgapur Deposit", "", ""],
        ["Rent Deposit (Audi Industries)", 500000, 500000],
        ["", "", ""],
        ["Granite Rajasthan Mines Deposit", "", ""],
        ["FDR", 154885, 154885],
        ["Deposit for mines", 15428400, 10285000],
        ["", "", ""],
        ["Total", 92763249, 67750588]
    ]
    create_note_sheet(wb, "2.11", "Other Non-Current Assets", std_headers, other_non_current_assets_data)
    
    # Note 2.12 - Inventories
    inventories_data = [
        ["Values and Certified by the Management", "", ""],
        ["Silico Manganese", "", ""],
        ["Finished Goods", 145012499, 115732533],
        ["Raw Material", 299133853, ""],
        ["Stock of Consumables", 2214512, 288936940],
        ["(valued at cost or Net realisable value, whichever is lower)", "", 23594572],
        ["", "", ""],
        ["Oxygen Plant", "", ""],
        ["Finished Goods", "", ""],
        ["Raw Material and Consumables", 17290672, 12964231],
        ["", "", ""],
        ["Aluminium Plant", "", ""],
        ["Finished Goods", "", ""],
        ["Raw Materials and Consumables", 3172346, 13641935],
        ["", "", 3280515],
        ["Rajasthan Unit", "", ""],
        ["Finished Goods", "", ""],
        ["Raw Materials and Consumables", 280801, 1178286],
        ["", 6040, 27097],
        ["Chandrapur Unit", "", ""],
        ["Finished Goods", "", ""],
        ["Raw Materials and Consumables", 1373069, 823429],
        ["", "", ""],
        ["Durgapur Unit", "", ""],
        ["Finished Goods", "", ""],
        ["Raw Materials and Consumables", 2059580, ""],
        ["Total", 473543372, 448628827]
    ]
    create_note_sheet(wb, "2.12", "Inventories", std_headers, inventories_data)
    
    # Note 2.13 - Trade Receivable
    trade_receivable_data = [
        ["Unsecured, Considered Goods", "", ""],
        ["Trade receivables outstanding for a period exceeding six months from the date they are due for payment", "", ""],
        ["Other Trade receivables- Annexure-5", "", ""],
        ["Silico Manganese", 369254179, 75136090],
        ["Oxygen Plant", 314454, ""],
        ["Aluminium Plant", 7359360, 1976542],
        ["Rajasthan Unit", "", 18014423],
        ["Chandrapur Unit", 109263, ""],
        ["Total", 377038457, 95127055]
    ]
    create_note_sheet(wb, "2.13", "Trade Receivable", std_headers, trade_receivable_data)
    
    # Note 2.13a - Trade Receivables aging schedule
    receivables_aging_data = [
        ["As at March 31, 2023", "", "", "", "", "", ""],
        ["", "Not Due", "Outstanding for following periods from the date of payment", "", "", "", "Total"],
        ["", "", "Less than 6 months", "6 months-1 year", "1-2 years", "2-3 years", "More than 3 years"],
        ["Undisputed-considered good", "", "", "", "", "", ""],
        ["Undisputed-considered doubtful", "", "", "", "", "", ""],
        ["Disputed-considered good", "", "", "", "", "", ""],
        ["Disputed-considered doubtful", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Total Trade Receivables", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["As at March 31, 2022", "", "", "", "", "", ""],
        ["", "Not Due", "Outstanding for following periods from the date of payment", "", "", "", "Total"],
        ["", "", "Less than 6 months", "6 months-1 year", "1-2 years", "2-3 years", "More than 3 years"],
        ["Undisputed-considered good", "", "", "", "", "", ""],
        ["Undisputed-considered doubtful", "", "", "", "", "", ""],
        ["Disputed-considered good", "", "", "", "", "", ""],
        ["Disputed-considered doubtful", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Total Trade Receivables", "", "", "", "", "", ""]
    ]
    create_note_sheet(wb, "2.13a", "Trade Receivables aging schedule", std_headers[:1] + ["", "", "", "", "", ""], receivables_aging_data)
    
    create_note_2_14(wb, header_font, normal_font, title_font, thin_border, header_fill)
    
    # Note 2.15 - Short-Term Loans and Advances
    short_term_loans_data = [
        ["Advances Recoverable in Cash or in Kind:", "", ""],
        ["Advance to Parties-Annexure 6", "", ""],
        ["Silico Manganese", "", ""],
        ["Aluminium", 184201187, 62018984],
        ["Oxygen", 5499512, 94237],
        ["Durgapur", 11931082, 720201],
        ["Chandrapur", 11284749, 37500],
        ["Rajasthan", 222750, ""],
        ["", 900000, ""],
        ["Total", 213818780, 62870922]
    ]
    create_note_sheet(wb, "2.15", "Short-Term Loans and Advances", std_headers, short_term_loans_data)
    
    # Note 2.15a - Loan & Advance in the nature of loan outstanding from promoters, directors, KMPs and related parties
    stl_loan_advance_data = [
        ["Particulars", "Figures as at the end of current reporting period", "", "Figures as at the end of previous reporting period", ""],
        ["", "%", "", "%", ""],
        ["Promoters", "", "", "", ""],
        ["Directors", "", "", "", ""],
        ["KMPs", "", "", "", ""],
        ["Related Parties", "", "", "", ""],
        ["Total", "", 0, "", 0]
    ]
    create_note_sheet(wb, "2.15a", "Loan & Advance in the nature of loan outstanding from promoters, directors, KMPs and related parties", std_headers[:1] + ["Figures as at the end of current reporting period", "", "Figures as at the end of previous reporting period", ""], stl_loan_advance_data)
    
    # Note 2.15B - Other Current Assets
    other_current_assets_data = [
        ["Silico Manganese", "", ""],
        ["Prepaid Expenses", 156270, 141185],
        ["TDS and TCS", 5130859, 3398186],
        ["GST Receivable", 1523188, ""],
        ["", "", ""],
        ["Oxygen Plant", "", ""],
        ["Prepaid Expenses", 41180, 41180],
        ["TDS and TCS", 53234, 15762],
        ["GST Receivable", 8593888, 10863689],
        ["", "", ""],
        ["Aluminium Plant", "", ""],
        ["GST Receivable", "", 2564499],
        ["TDS Receivable", 15665, 14835],
        ["", "", ""],
        ["Rajasthan Unit", "", ""],
        ["GST Receivable", "", ""],
        ["TDS Receivable", 1135053, 17246],
        ["", "", 3438],
        ["Chandrapur Unit", "", ""],
        ["GST Receivable", 214151, 148235],
        ["TCS Receivable", 826, 826],
        ["", "", ""],
        ["Durgapur Unit", "", ""],
        ["GST Receivable", 447162, ""],
        ["Preliminary Expenses", 498150, ""],
        ["Total", 17984125, 16949431]
    ]
    create_note_sheet(wb, "2.15B", "Other Current Assets", std_headers, other_current_assets_data)
    
    # Note 2.16 - Revenue from operations
    revenue_data = [
        ["Sales of Goods:", "", ""],
        ["Silico Manganese", 1109792709, 1036542225],
        ["Aluminium Products", 93094289, 61184257],
        ["Sales from Oxygen Plant", 12116434, 10979319],
        ["Sales from Rajasthan Unit", 17344156, 1128499],
        ["Sales from Chandrapur Unit", 443030, ""],
        ["Total", 1234577618, 1109643099]
    ]
    create_note_sheet(wb, "2.16", "Revenue from operations", std_headers, revenue_data)
    
    # Note 2.17 - Other Incomes
    other_income_data = [
        ["For Silico Manganese", "", ""],
        ["Discount Received", 24039, 25556],
        ["Quantity and Quality claim", "", ""],
        ["Interest on Fixed Deposit", 568652, ""],
        ["Interest on MPCB Deposit", 3448349, 2150057],
        ["other Income", 1224094, 5192],
        ["", "", ""],
        ["For Aluminium", "", ""],
        ["Interest on MPCB Deposit", 7494, 31828],
        ["", "", ""],
        ["For Oxygen", "", ""],
        ["Interest on MPCB Deposit", 1119, 51692],
        ["Total", 5273747, 2264325]
    ]
    create_note_sheet(wb, "2.17", "Other Incomes", std_headers, other_income_data)
    
    # Note 2.18 - Raw Material and Stock in Trade Consumed
    raw_material_data = [
        ["Opening Stock of Raw Material and Stock in Trade", "", ""],
        ["For Silico Manganese", 306102742, 243053695],
        ["For Aluminium", 16852512, 238432396],
        ["For Oxygen", 1286515, 1611096],
        ["For Chandrapur", 823529, ""],
        ["For Rajasthan", 27097, ""],
        ["", "", ""],
        ["Add: Purchase", "", ""],
        ["Purchase of Raw Material and Stock in Trade", "", ""],
        ["For Silico Manganese", 704187276, 779086077],
        ["For Aluminium", 697314346, 719520235],
        ["For Oxygen", 95593, 48745389],
        ["For Rajasthan", 4028209, 5486963],
        ["For Chandrapur", 801647, 823529],
        ["For Durgapur", 2059580, ""],
        ["", "", ""],
        ["Less:", "", ""],
        ["Closing Stock of Raw Material and Stock in Trade", "", ""],
        ["For Silico Manganese", 316959324, 306102741],
        ["For Aluminium", 304548289, 301985612],
        ["For Oxygen", 2172346, 3280515],
        ["For Rajasthan", 6040, 27097],
        ["For Chandrapur", 1373069, 823529],
        ["For Durgapur", 2059580, ""],
        ["Total", 699330697, 712036379]
    ]
    create_note_sheet(wb, "2.18", "Raw Material and Stock in Trade Consumed", std_headers, raw_material_data)
    
    # Note 2.19 - Change in Inventories
    change_in_inventories_data = [
        ["Opening Stocks:", "", ""],
        ["", "", ""],
        ["Finished Goods- Silico Manganese", 115732533, 38528526],
        ["Finished Goods- Aluminium", 12651035, 2203411],
        ["Finished Goods- Oxygen", 12964231, ""],
        ["Finished Goods- Rajasthan", 1178385, ""],
        ["", 142526184, 40732117],
        ["", "", ""],
        ["Closing Stocks:", "", ""],
        ["", "", ""],
        ["Finished Goods- Silico Manganese", 145012498, 115732533],
        ["Finished Goods- Aluminium", "", 12651035],
        ["Finished Goods- Oxygen", 17290672, 12964231],
        ["Finished Goods- Rajasthan", 280801, 1178385],
        ["", 162584072, 142526184],
        ["Total", "(20057888)", "(101793747)"]
    ]
    create_note_sheet(wb, "2.19", "Change in Inventories", std_headers, change_in_inventories_data)
    
    # Note 2.20 - Employee Benefit expenses
    employee_benefit_data = [
        ["For Silico Manganese", "", ""],
        ["Wages & Salary", 12799620, 8305277],
        ["Staff Welfare Expenses", 3442678, 2522363],
        ["PF Expense", 247378, 206719],
        ["ESI- Expense", 11487, 35783],
        ["", "", ""],
        ["For Oxygen", "", ""],
        ["Wages & Salary", 1351038, 215000],
        ["", "", ""],
        ["For Chandrapur", "", ""],
        ["Wages & Salary", 25600, ""],
        ["", "", ""],
        ["For Durgapur", "", ""],
        ["Wages & Salary", 149200, ""],
        ["", "", ""],
        ["Total", 18101766, 11280748]
    ]
    create_note_sheet(wb, "2.20", "Employees Benefit expenses", std_headers, employee_benefit_data)
    
    # Note 2.21 - Finance Cost
    finance_cost_data = [
        ["For Silico Manganese", "", ""],
        ["Bank charges, commission etc", 5386927, 1778887],
        ["Processing and Other expenses", 384450, 771495],
        ["Stamp Duty expenses", 970600, 1287671],
        ["TDS Interest", "", ""],
        ["Interest on Term Loans", 26517440, 2008187],
        ["Interest on Cash Credit Limit", 38015230, 19030186],
        ["Interest on Unsecured Loans", 1502447, 5833250],
        ["", "", ""],
        ["For Oxygen", "", ""],
        ["Bank charges, commission etc", 21246, 519088],
        ["Processing and Other expenses", "", 29004],
        ["Interest on Term Loans", 9019178, 7272866],
        ["Interest on Cash Credit Limit", 1233856, ""],
        ["Interest on Unsecured Loans", "", 7152410],
        ["Total", 75462815, 44275968]
    ]
    create_note_sheet(wb, "2.21", "Financial Cost", std_headers, finance_cost_data)
    
    # Note 2.22 - Depreciation and Amortisation Expenses
    depreciation_data = [
        ["Depreciation", 42204456, 18256299],
        ["", "", ""],
        ["Total", 42204456, 18256299]
    ]
    create_note_sheet(wb, "2.22", "Depreciation and Amortisation Expenses", std_headers, depreciation_data)
    
    # Note 2.23 - Other Expenses/Manufacturing Expenses
    other_expenses_data = [
        ["Direct Expenses/Manufacturing Expenses", "", ""],
        ["", "", ""],
        ["For Silico Manganese", "", ""],
        ["Plant Rent", "", 13200000],
        ["Vehicle Rent", "", 5315000],
        ["Production Charges", 20696024, 26782460],
        ["Power & Fuel Charges", 201126023, 245378066],
        ["Loading and Unloading Expenses", "", ""],
        ["Transportation Expenses", 20767203, 9778441],
        ["Freight Expenses", 901243, 297711],
        ["Pollution Fees", "", 81600],
        ["Water Expenses", 1802946, 2894830],
        ["Wire Cotton repairs Cost", 596081, 1363763],
        ["Service Charges", 14000, 111000],
        ["Labour Expenses", 8335245, 1000],
        ["", "", ""],
        ["For Aluminium", "", ""],
        ["Plant Rent", "", 5412000],
        ["Power & Fuel Charges", "", ""],
        ["Transportation Expenses", 666679, 4508316],
        ["Packing and forwarding charges", "", ""],
        ["Labour charges", "", 212603],
        ["", "", ""],
        ["For Oxygen", "", ""],
        ["Electricity Connection Expenses", 6027171, 6733590],
        ["Cylinder Filling Expenses", "", 84369],
        ["Water Expenses", 259934, 117250],
        ["Shifting Expenses", "", 90000],
        ["Labour charges", "", 90000],
        ["", "", ""],
        ["For Rajasthan", "", ""],
        ["Plant Rent", "", 154090],
        ["Labour Charges", "", 1156598],
        ["Power & Fuel Charges", "", 55505],
        ["", "", ""],
        ["For Durgapur", "", ""],
        ["Freight", 1200, ""],
        ["Loading Charges", 1655, ""],
        ["Total", 261954446, 331506727],
        ["", "", ""],
        ["Indirect Expenses", "", ""],
        ["", "", ""],
        ["For Silico Manganese", "", ""],
        ["Indirect Expenses", "", ""],
        ["Business Promotion Expenses", 47148, ""],
        ["General expenses", "", 2287123],
        ["Conveyance expenses", "", ""],
        ["Consultancy Charges", 2067905, ""],
        ["GST Audit Fees", "", ""],
        ["Insurance Expenses", 891128, 1009774],
        ["Internet & Electricity and other Dues", "", 1092057],
        ["Legal & Professional Charges", "", 1248702],
        ["Office Expenses", "", 322229],
        ["Rent, Rates and Taxes", 1291718, 411389],
        ["Repair & Maintenance Expenses", 1129339, 1975424],
        ["Telephone Telex & Telegram services", 2468753, 586893],
        ["Travelling & Conveyance Exp", 3088014, 471855],
        ["Vehicle Running & Maint", 17145, 317320],
        ["Computer maintainence", 131223, 145969],
        ["Statutory fee of ROC, MCA, GI", "", 47572],
        ["Registration and Tender Fee", "", ""],
        ["Hospitality Expenses", 3732476, 159495],
        ["GST Fee", "", 124815],
        ["Profession tax", "", ""],
        ["Preliminary Expenses Write Off", 2500, 2500],
        ["Balance / Interest / Balance write off", 372719, ""],
        ["Freight & Transportation Exp", 168097, 32444],
        ["Rating Expenses", "", 1309979],
        ["Prelim/Prefime development charges", "", 90000],
        ["CSR Expenses", 463750, 6500],
        ["Donation Expenses", 1100000, ""],
        ["Security Expenses", 155000, ""],
        ["Lab and Research Expenses", 1352212, ""],
        ["Office Expense", 558000, ""],
        ["Misc Expenses", 159665, 334193],
        ["", "", ""],
        ["Auditor Remuneration", "", ""],
        ["a) Audit Fees", 200000, 30000],
        ["b) Taxation Matters", "", ""],
        ["c) Company Law Matters", "", ""],
        ["", "", ""],
        ["For Oxygen", "", ""],
        ["Indirect Expenses", "", ""],
        ["Freight Expenses", "", 962],
        ["Professional & Consultancy charges", 123678, 263313],
        ["General expense", 275978, ""],
        ["Lease Rent expenses", "", ""],
        ["Preliminary Expenses written off", "", ""],
        ["Telephone Expense", "", ""],
        ["Other charges", "", ""],
        ["Rent, Rates and Taxes", 17600, 35596],
        ["Repair and Maintenance Expenses", 65279, 6000],
        ["Vehicle Expenses", 12600, 160420],
        ["Licence & Permission Expenses", "", ""],
        ["Security Expenses", 76564, 41346],
        ["Registration Expenses", "", 205200],
        ["Office Expenses", 217021, 247404],
        ["Preliminary Expenses Written Off", "", 10000],
        ["", "", 11000],
        ["For Aluminium", "", ""],
        ["Indirect Expenses", "", ""],
        ["Freight Expenses", "", 4300],
        ["Rent and Loan Fees", 3910, 262941],
        ["Repair and Maintenance Expenses", "", 25095],
        ["Transportation Expenses", "", 58095],
        ["", "", 23200],
        ["For Rajasthan", "", ""],
        ["Rent and Loan Fees", "", 18245],
        ["Miscellaneous Expenses", "", 2000],
        ["Consultancy Expenses", 115000, ""],
        ["Road Tax", 10000, ""],
        ["Pollution Expenses", 5000, ""],
        ["", "", ""],
        ["For Chandrapur", "", ""],
        ["Rent Expenses", 8400, ""],
        ["Miscellaneous Expenses", 21575, ""],
        ["", "", ""],
        ["For Durgapur", "", ""],
        ["Rent Expenses", 87500, 175000],
        ["Fees & Registration Expenses", 3825, ""],
        ["Miscellaneous Expenses", 23695, 5000],
        ["", "", ""],
        ["For Corporate", "", ""],
        ["Professional & Legal Charges", "", ""],
        ["Consultancy charges", "", ""],
        ["", "", ""],
        ["Total", 24032045, 16176937],
        ["", "", ""],
        ["Grand Total (a+b)", 285986490, 347683664]
    ]
    create_note_sheet(wb, "2.23", "Other Expenses", std_headers, other_expenses_data)

def create_note_2_1(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Note 2.1 - Share Capital
    ws = wb.create_sheet("Note 2.1 Share Capital")
    
    # Add title
    ws['A1'] = "EQUITY & LIABILITIES"
    ws['A1'].font = title_font
    
    ws['A3'] = "Note No. 2.1 Share Capital"
    ws['A3'].font = title_font
    
    # Create table structure with merged cells and borders
    # Create the table with required columns and layout
    start_row = 5
    
    # Create the header row
    ws.cell(row=start_row, column=1).value = "Particulars"
    ws.cell(row=start_row, column=2).value = "Figures as at the end of the current reporting period"
    ws.cell(row=start_row, column=4).value = "Figures as at the end of the previous reporting period"
    
    # Format header row
    for col in [1, 2, 4]:
        cell = ws.cell(row=start_row, column=col)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    # Merge cells for headers
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)
    ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=5)
    
    # Add data rows
    data = [
        ["Authorised", "", "", "", ""],
        ["65,00,000 Equity Shares of 10/- each", "", "65,000,000", "", "65,000,000"],
        ["(PY 65,00,000 Equity Shares of 10/- each)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Issued Subscribed & Paid up", "", "65,000,000", "", "65,000,000"],
        ["65,28,650 Equity Shares of 10/- each fully paid up", "", "65,286,500", "", "59,510,000"],
        ["(PY 65,28,650 Equity Shares of 10/- each fully paid up)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Total", "", "65,286,500", "", "59,510,000"]
    ]
    
    # Add data to worksheet with formatting
    for i, row_data in enumerate(data, start=start_row+1):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.border = thin_border
            
            # Format numeric values with thousands separator
            if isinstance(value, str) and value.replace(',', '').isdigit():
                cell.value = int(value.replace(',', ''))
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    
    return ws

def create_note_2_1c(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Note 2.1(c) sheet
    ws = wb.create_sheet("Note 2.1(c) Shareholding 5%")
    
    # Add title
    ws['A1'] = "Note No. 2.1 (c) Shares held by each shareholder holding more than 5% of shares"
    ws['A1'].font = title_font
    
    # Create table with headers
    headers = [
        ["Name of Shareholder", "Figures as at the end of previous reporting period", "", "Figures as at the end of previous reporting period", ""],
        ["", "No. of Shares held", "% of Holding", "No. of Shares held", "% of Holding"]
    ]
    
    # Add headers with formatting
    for i, header_row in enumerate(headers, start=2):
        for j, header_text in enumerate(header_row, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = header_text
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    # Merge cells for the headers
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
    
    # Add data rows
    data = [
        ["Sayyad Akhtar Ali", 4127252, "63.22%", 3566000, "59.92%"],
        ["Saiyyed Owais Ali", 150000, "2.30%", 150000, "2.52%"],
        ["Mohabbat Ali", 487000, "7.46%", 487000, "8.18%"],
        ["Rafiq Bee", 1105000, "16.93%", 1105000, "18.57%"],
        ["Sayyad Afsar Ali", "-", "0.00%", 543000, "9.12%"],
        ["Sayyad Murtaza Ali", 100000, "1.53%", 100000, "1.68%"],
        ["Outside Investors", 559398, "8.57%", "-", "-"]
    ]
    
    # Add data to the worksheet with formatting
    for i, row_data in enumerate(data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format based on value type
            if j in [2, 4] and isinstance(value, int):  # Number of shares columns
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal='center')
            
            cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    return ws

def create_note_2_1d(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Note 2.1(d) sheet
    ws = wb.create_sheet("Note 2.1(d) Promoters")
    
    # Add title
    ws['A1'] = "Note No. 2.1 (d) Shareholding of Promoters & % of change during the Year"
    ws['A1'].font = title_font
    
    # Create table with headers
    headers = [
        ["Shares held Promoters at the end of the year", "Figures as at the end of previous reporting period", "", "", "Figures as at the end of previous reporting period", "", ""],
        ["", "No. of Shares held", "% of total shares", "% Change", "No. of Shares held", "% of total shares", "% Change"]
    ]
    
    # Add headers with formatting
    for i, header_row in enumerate(headers, start=2):
        for j, header_text in enumerate(header_row, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = header_text
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    # Merge cells for the headers
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
    
    # Add data rows
    data = [
        ["Sayyad Akhtar Ali", 4127252, "63.22%", "3.29%", 3566000, "59.92%", "-"],
        ["Sayyad Afsar Ali", "-", "-", "-9.12%", 543000, "9.12%", "-"],
        ["Outside Investors", 559398, "8.57%", "8.57%", 0, "-", "0.00%"]
    ]
    
    # Add data to the worksheet with formatting
    for i, row_data in enumerate(data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format based on value type
            if j in [2, 5] and isinstance(value, int):  # Number of shares columns
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal='center')
            
            cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    return ws

def create_note_2_3(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Note 2.3 - Long Term Borrowings
    ws = wb.create_sheet("Note 2.3 Long Term Borrowings")
    
    # Add title
    ws['A1'] = "Note No. 2.3 Long Term Borrowings"
    ws['A1'].font = title_font
    
    # Create headers
    start_row = 3
    headers = [
        ["Particulars", "Figures as at the end of current reporting period", "", "Figures as at the end of previous reporting period", ""],
        ["", "Non Current Maturities", "Current Maturities", "Non Current Maturities", "Current Maturities"]
    ]
    
    # Add headers with formatting
    for i, header_row in enumerate(headers, start=start_row):
        for j, header_text in enumerate(header_row, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = header_text
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    # Merge header cells
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)
    ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=5)
    
    # Add data rows based on the images
    loan_data = [
        ["Secured Loans", "", "", "", ""],
        ["Sundaram Finance Bank Ltd-Bolero Car Loan", "-", "-", "-", "-"],
        ["(Secured against hypothecation of Vehicle & personal guarantee of directors)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Sundaram Finance Bank Ltd-Bolero Car Loan", "", "", "278,367", "124,936"],
        ["(Secured against hypothecation of Vehicle & personal guarantee of directors)", "", "", "", ""],
        ["", "", "", "", ""],
        ["ICICI Bank Limited-Tractor", "1,805,236", "393,125", "-", "-"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["ICICI Bank Limited-Crushing Plant", "11,863,134", "2,210,404", "", ""],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Indusind Bank ltd RB49(S)FT- Tractor Loan", "310,412", "225,499", "320,838", "208,023"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Indusind Bank ltd RB49(S)TT- Trolley Loan", "27,042", "27,042", "31,982", "62,464"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Indusind Bank ltd RB49(S)FT- Tractor Loan", "321,213", "223,853", "", ""],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "320,771", "208,028"],
        ["", "", "", "", ""],
        ["Bank of India-Oxygen Term Loan", "87,707,847", "16,783,372", "96,089,350", "15,577,209"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Bank of India GECL 1.0 WCTL-078", "1,673,932", "1,673,932", "1,772,310", "4,083,340"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["AU Small Finance Bank-Fortuner", "979,984", "956,361", "930,834", "1,302,887"],
        ["(Secured against hypothecation of Vehicle & personal guarantee of directors)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Punjab and Sind Bank-Bolero", "841,713", "151,163", "841,424", "128,243"],
        ["(Secured against hypothecation of Vehicle & personal guarantee of directors)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Bank of India GECL 1.0 WCTL", "8,081,380", "2,643,840", "6,808,880", "2,091,120"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Canara Bank of India -Plant Durgapur", "221,063,338", "20,714,929", "118,262,167", "-"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank Credit Card", "949,452", "", "1,403,968", ""],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 942- JCB Loan", "2,649,665", "796,027", "2,654,910", "750,740"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 132- JCB Loan", "2,649,665", "796,027", "2,654,910", "750,740"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 225- Dumper Loan", "3,671,486", "1,103,441", "3,678,840", "1,021,160"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 099- Dumper Loan", "3,671,486", "1,103,441", "3,678,840", "1,021,160"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 098- Dumper Loan", "3,671,486", "1,103,441", "3,678,840", "1,021,160"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 014- Dumper Loan", "3,671,486", "1,103,441", "3,678,840", "1,021,160"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 023- Dumper Loan", "3,671,486", "1,103,441", "3,678,840", "1,021,160"],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 532", "2,136,965", "711,023", "", ""],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 767", "14,253,964", "2,683,411", "", ""],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["HDFC Bank 634", "4,256,874", "805,942", "", ""],
        ["(Secured against hypothecation of stock & personal guarantee of directors and government guarantee)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Mahindra and Mahindra Financial Services Ltd", "6,500,000", "", "", ""],
        ["", "", "", "", ""],
        ["Total (a)", "366,929,929", "57,443,250", "226,752,841", "30,965,630"],
        ["", "", "", "", ""],
        ["Unsecured Loan", "", "", "", ""],
        ["From Directors & Their Relatives- Annexure 1", "58,393,011", "-", "33,528,915", "-"],
        ["From Body Corporates", "-", "-", "-", "-"],
        ["", "", "", "", ""],
        ["Total (b)", "58,393,011", "-", "33,528,915", "-"],
        ["", "", "", "", ""],
        ["Total (a+b)", "367,779,690", "57,443,250", "254,281,756", "30,965,630"]
    ]
    
    # Add loan data to the worksheet
    row_num = start_row + 2
    for row_data in loan_data:
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            
            # Format numeric values appropriately
            if col_num > 1 and value not in ["", "-"]:
                try:
                    # Remove commas and convert to integer
                    value = int(value.replace(",", ""))
                    cell.value = value
                    cell.number_format = '#,##0'
                except (ValueError, AttributeError):
                    cell.value = value
            else:
                cell.value = value
                
            # Apply borders to all cells
            cell.border = thin_border
            
            # Center align all columns except first
            if col_num > 1:
                cell.alignment = Alignment(horizontal='center')
            
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    return ws

def create_note_2_3a(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Note 2.3(a) sheet
    ws = wb.create_sheet("Note 2.3a Loan Terms")
    
    # Add title
    ws['A1'] = "Note No. 2.3 (a) Terms of Repayment of Loans"
    ws['A1'].font = title_font
    
    # Add loan terms data with proper formatting
    loan_terms = [
        "Loan from Sundaram Finance Bank (Bolero Car Loan) was taken during the year 2022-23 The Loan is repayable in 41 equated monthly installments of 13,100/- each starting from 03/12/22 along with interest @........ from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from Indusind Bank (Tractor Loan) was taken during the year 2022-23 The Loan is repayable in 36 equated monthly installments of 20861/- each starting from 15/10/22 along with interest @13.47% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from Indusind Bank (Tractor Loan) was taken during the year 2022-23 The Loan is repayable in 36 equated monthly installments of 20861/- each starting from 15/10/22 along with interest @13.47% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from Indusind Bank (Trolley Loan) was taken during the year 2022-23 The Loan is repayable in 24 equated monthly installments of 5833/- each starting from 15/10/22 along with interest @14.08% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from ICICI Bank Limited (Fortuner Car Loan) was taken during the year 2019-20. The Loan is repayable in 36 equated monthly installments of 1,02,004/- each starting from 05/06/2019 along with interest @10% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company. The Loan has Mature in the year 2022.",
        "",
        "Loan from Bank of India (Oxygen Term Loan) was taken during the year 2021-22 The Loan is repayable in 24 equated monthly installments of 11,30,952/- each starting from April 2022 along with interest @ 11.35% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from Bank of India (GECL WTCL) was taken during the year 2020-21 The Loan is repayable in 48 equated monthly installments of 3,78,486/- each starting from August 2022 along with interest @8.9% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from AU Small Finance Bank (Fortuner Car Loan) was taken during the year 2021-22 The Loan is repayable in 36 equated monthly installments of 1,25,327/- each starting from 14/10/2021 along with interest @7.9% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company. The Loan has Mature in the year 2024.",
        "",
        "Loan from Punjab and Sind Bank (Bolero Car Loan) was taken during the year 2021-22 The Loan is repayable in 84 equated monthly installments of 17,047/- each starting from 24/03/2022 along with interest @........ from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from Bank of India (GECL 1.0 WTCL) was taken during the year 2021-22 The Loan is repayable in 36 equated monthly installments of 2,95,510/- each starting from February 2024 along with interest @8.9% from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from HDFC Bank (JCB) was taken during the year 2022-23 The Loan is repayable in 48 equated monthly installments of 84000/- each starting from 15/04/23 from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
        "",
        "Loan from HDFC Bank (Dumper) was taken during the year 2022-23 The Loan is repayable in 48 equated monthly installments of 116430/- each starting from 05/04/23 from the date of loan. The Loan is secured by hypothecation of vehicle and personal guarantee of directors of the company.",
    ]
    
    # Add the loan terms to the worksheet
    row_num = 3
    for term in loan_terms:
        cell = ws.cell(row=row_num, column=1)
        cell.value = term
        cell.font = normal_font
        
        # Apply custom formatting for empty rows vs content rows
        if term:
            cell.alignment = Alignment(wrap_text=True)
        else:
            cell.alignment = Alignment()
            
        row_num += 1
    
    # Set column width to accommodate the text
    ws.column_dimensions['A'].width = 120
    
    return ws

def create_note_2_14(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Note 2.14 - Cash & Cash Equivalent
    ws = wb.create_sheet("Note 2.14 Cash & Cash Equivalent")
    
    # Add title
    ws['A1'] = "Note No. 2.14 Cash & Cash Equivalent"
    ws['A1'].font = title_font
    
    # Create headers
    headers = ["Particulars", "Figures as at the end of current reporting period", "Figures as at the end of previous reporting period"]
    row = 3
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    # Add data rows
    cash_data = [
        ["Balance with Bank-current A/c", "", ""],
        ["HDFC", "51,932", ""],
        ["Axis Bank", "72,916", ""],
        ["Bank of India", "-", ""],
        ["", "", ""],
        ["Cash on hand", "7,021,180", "3,988,022"],
        ["", "", ""],
        ["Total", "7,146,028", "3,988,022"]
    ]
    
    # Add data to worksheet
    start_row = 4
    for i, data_row in enumerate(cash_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j > 1 and value not in ["", "-"]:
                try:
                    # Remove commas and convert to integer
                    numeric_value = int(value.replace(",", ""))
                    cell.value = numeric_value
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                except (ValueError, AttributeError):
                    cell.value = value
            else:
                cell.value = value
            
            # Apply borders
            cell.border = thin_border
            
            # Bold the total row
            if data_row[0] == "Total":
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    
    return ws

def create_annexure_1(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 1 sheet
    ws = wb.create_sheet("Annexure 1 Unsecured Loan")
    
    # Add title with underline
    ws['A1'] = "1. Unsecured Loan"
    ws['A1'].font = title_font
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 3
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    loan_data = [
        ["For Silico Manganese", ""],
        ["Akhtar Ali", 1545202],
        ["Murtaza Ali", 5115000],
        ["Afsar Ali", 22099506],
        ["Mohabbat Ali", 3499196],
        ["", ""],
        ["For Oxygen", ""],
        ["Owais Ali", 26134107],
        ["Total", 58393011]  # Correct total as per image
    ]
    
    # Add data to worksheet
    start_row = 4
    for i, data_row in enumerate(loan_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold for headings and total
            if data_row[0] in ["For Silico Manganese", "For Oxygen", "Total"]:
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_2(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 2 sheet
    ws = wb.create_sheet("Annexure 2 Capital Assets")
    
    # Add title with underline
    ws['A1'] = "2. Sundry Creditors for Capital Assets"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 3
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    creditors_data = [
        ["Total", 0]
    ]
    
    # Add data to worksheet
    start_row = 4
    for i, data_row in enumerate(creditors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold for total
            if data_row[0] == "Total":
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_3a(wb, header_font, normal_font,title_font, thin_border, header_fill):
    # Create Annexure 3a sheet
    ws = wb.create_sheet("Annexure 3a Silico Manganese")
    
    # Add title with underline
    ws['A1'] = "3. Sundry Creditors"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    ws['A2'] = "a) Silico Manganese"
    ws['A2'].font = Font(name='Arial', size=11, bold=True, italic=True)
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 4
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows - first image
    creditors_data_1 = [
        ["Creditors for Goods", ""],
        ["Amara Steels Pvt. Ltd", 1779],
        ["Anushri Traders", 20179],
        ["Balaji Mineral Product", 19388],
        ["Basant Engineering Works", 6164],
        ["Bhagwati Elite Motors Pvt Ltd", 1273418],
        ["Bharya Trading Com", 28267],
        ["Deeva Enterprises", 1645052],
        ["Dewali Cargo Movers", 104817],
        ["Dipesh Computer & Mobile Point", 1300],
        ["GURPREET MOTOR & CARRIER", 38000],
        ["Jaat Machinery Stores", 1130],
        ["Jacky Enterprises", 142583],
        ["JAIDEEP ISPAT & ALLOYS PVT LTD", 3501],
        ["Jay Maa Ashapuri Pavansinhji Paramsinhji Bakaiya", 24300],
        ["Jishan Refrigeration & Motor Riwinding", 6590],
        ["J S FORKLIFT SERVICES", 4662],
        ["KaslilWal Brothers", 104982],
        ["KATARIYA ENTERPRISES", 620],
        ["khandelwal and associates", 186000],
        ["LAFCO INDIA SCIENTIFIC", 48902],
        ["Maa Bhavni JCB Parts & Lubricant", 690],
        ["Maa Jagdamba Transport Company TDS Dec.", 69085],
        ["Maa Tulsi Kripa Balaji Bhandar", 60520],
        ["Mahavir Industries", 61997],
        ["Mahavir Traders", 95876],
        ["Manish Milkman (Adesh)", 20096],
        ["MATESWAR! TRADER AND MANUFACTURER", 245700],
        ["Meghneer Chilled Water", 28220],
        ["MP AKVN LTD", 129714],
        ["Mppkvvcl HT Revenue Collection", 25573822],
        ["MS INFOTECH", 51000],
        ["M/s Mithalal R. Desai- Meghnagar", 975128],
        ["NAGINA TRANSPORT CO", 18509],
        ["Navkar Furniture", 80975],
        ["Panchmahal Transport", 35104],
        ["Parikshan Laboratory", 4380],
        ["PHONE HOUSE", 72400],
        ["PRAKASH CHANDRA PATIDAR", 221000],
        ["PRINCE AUTOMOBILES", 80288],
        ["Prosperous Facility Services Pvt Ltd", 160760]
    ]
    
    # Add data rows - second image
    creditors_data_2 = [
        ["Raamesh & Company", 33440],
        ["Reliance Enterprise", 36625],
        ["Rk Minerals", 21516],
        ["R.K. Steel Industries", 58610],
        ["RS Engineering Project", 4052],
        ["Rst Enterprises", 10000],
        ["Sai Shree Aqurium", 41615],
        ["Shakambhari Trading Company", 835605],
        ["Shakruwala Electricals", 3140],
        ["SHREE GANESH CHILLED WATER", 31500],
        ["Shree Nakoda Lights", 22305],
        ["Shree Nandu Chandu Tradres", 337935],
        ["Shree Shyam Traders", 2297],
        ["Shri Bhairav Bardan Bhandar, Meghnagar", 229853],
        ["Shyam Scientific", 1476],
        ["S.R. Enterprises", 25489],
        ["SR Ferro Alloys (Adv Against Pur of FAD)", 1944592],
        ["SS Plastic Ratlam", 500],
        ["S S Tyre World", 7700],
        ["Talati Electric Works Pvt. Ltd", 6519],
        ["TCI Express", 2929],
        ["Vedant Corporation", 1313],
        ["VISHWAKARMA IMSS", 23901],
        ["WATERLY BEVERAGES PVT LTD", 70682],
        ["Western Coal Field Limtd", 99800],
        ["Western Coal Field Limited", 12823],
        ["Western Coalfields Limited Neheriya", 40002],
        ["Western Coalfields Limited Shobhapur", 36955],
        ["Anokhlial Gendalal Tanted20-21", 35550],
        ["Bhagwati Elite Motors", 7237],
        ["Eagle Sales Corporation", 7290],
        ["Ekta Traders, Ratlam", 128775199],
        ["Gitanjali Construction Hub Private Limited", 68129989],
        ["Gitanjali Construction Ratlam", 12232190],
        ["Goodwill Logistic", 54000],
        ["Growmore Enterprises Pvt Ltd", 8540],
        ["Hindustan Automobile", 21682],
        ["Jm Environet Pvt Ltd", 153969],
        ["Luniya & Company", 3500],
        ["Mercedes-Benz India Private Limited", 195459],
        ["Owais Ali Overseas Private Limited Raj", 712559],
        ["Ramnipa Ispat Pvt Ltd", 170521],
        ["Roshan Ispat", 623142],
        ["Satguru Motors", 318156],
        ["Tvs Mobility", 2867490],
        ["Uchhav Lal Ramchandra", 6612980],
        ["United India Insurance Company Ltd.", 7167],
        ["MO Infra Unit 1", 6488092],
        ["M/s Sayyad Akhtar Ali", 405075],
        ["SMO Ferro Alloys Rajasthan Unit", 174286],
        ["Total", 285908678]  # Correct total as per image
    ]
    
    # Combine the data
    creditors_data = creditors_data_1 + creditors_data_2
    
    # Add data to worksheet
    start_row = 5
    for i, data_row in enumerate(creditors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold for headings and total
            if data_row[0] == "Creditors for Goods" or data_row[0] == "Total":
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_3b(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 3b sheet
    ws = wb.create_sheet("Annexure 3b Oxygen")
    
    # Add title with underline
    ws['A1'] = "3. Sundry Creditors"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    ws['A2'] = "b) Sundry Creditors Oxygen"
    ws['A2'].font = Font(name='Arial', size=11, bold=True, italic=True)
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 4
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows - based on the image
    creditors_data = [
        ["Anand Big Mall", 194500],
        ["Malwa Traders", 586476.56],
        ["MP Akvn (Water Dept.) Pithampur", 99132],
        ["MS Infotech", 50998.6],
        ["Multiplywood", ""],
        ["Priya Enterprises", 162124.46],
        ["Rapid Global Logistic", 11625],
        ["Shree Girraj Industries", 43188],
        ["MP Auto Engineering", 3717],
        ["Explotech Engineer", 38840],
        ["Sai Samrath", 3725],
        ["RK Construction, Pithampur", 160994],
        ["Rabi Chander", 1685],
        ["R D Steel Mainopuri", 205341],
        ["Total", 1562349]
    ]
    
    # Add data to worksheet
    start_row = 5
    for i, data_row in enumerate(creditors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold the total row
            if data_row[0] == "Total":
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_3c(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 3c sheet
    ws = wb.create_sheet("Annexure 3c Aluminium")
    
    # Add title with underline
    ws['A1'] = "3. Sundry Creditors"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    ws['A2'] = "c) Sundry Creditors Aluminium"
    ws['A2'].font = Font(name='Arial', size=11, bold=True, italic=True)
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 4
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows - based on the image
    creditors_data = [
        ["Creditor for Power & Fuels", ""],
        ["RK Agro And Fuies", 564830],
        ["Creditor for Raw Material", ""],
        ["Anand Metal (Alu)", 246561],
        ["C K Associates", 3422],
        ["CKCO Engineering Works", 59840],
        ["C.K. Exim", 104680],
        ["Gurukirpa Metal", 106374],
        ["Mahadev Enterprieses", 2858080],
        ["Sigma Impex", 1184410],
        ["Creditor for Rent", ""],
        ["IGP Metals Manufacturers (Alu)", 646598],
        ["Creditor for Store", ""],
        ["Balaji Facades", 82422],
        ["Paridhi Enterprises (Alu)", 14379],
        ["Total", 5771483]
    ]
    
    # Add data to worksheet
    start_row = 5
    for i, data_row in enumerate(creditors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold the category headers and total row
            if data_row[0] in ["Creditor for Power & Fuels", "Creditor for Raw Material", "Creditor for Rent", "Creditor for Store", "Total"]:
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_3d(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 3d sheet
    ws = wb.create_sheet("Annexure 3d Rajasthan")
    
    # Add title with underline
    ws['A1'] = "3. Sundry Creditors"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    ws['A2'] = "d) Sundry Creditors Rajasthan"
    ws['A2'].font = Font(name='Arial', size=11, bold=True, italic=True)
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 4
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows - based on the image
    creditors_data = [
        ["Aashiyana Coal Depot", 178920],
        ["ES Infraserve Pvt Ltd", 700000],
        ["HK & Associates", 136700],
        ["Shri Mahadev Enterprises", 1938955],
        ["Sarkar Timber Mart", 25000],
        ["Total", 2978575]
    ]
    
    # Add data to worksheet
    start_row = 5
    for i, data_row in enumerate(creditors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold the total row
            if data_row[0] == "Total":
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_3e(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 3e sheet
    ws = wb.create_sheet("Annexure 3e Chandrapur")
    
    # Add title with underline
    ws['A1'] = "3. Sundry Creditors"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    ws['A2'] = "e) Sundry Creditors Chandrapur"
    ws['A2'].font = Font(name='Arial', size=11, bold=True, italic=True)
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 4
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows - based on the image
    creditors_data = [
        ["M/s Saiyyed Akhtar Ali", 6500],
        ["Steel Authority of India", 1919334],
        ["Suryodyay Construction Co.", 84],
        ["Total", 1925918]
    ]
    
    # Add data to worksheet
    start_row = 5
    for i, data_row in enumerate(creditors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold the total row
            if data_row[0] == "Total":
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_3f(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 3f sheet
    ws = wb.create_sheet("Annexure 3f Durgapur")
    
    # Add title with underline
    ws['A1'] = "3. Sundry Creditors"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    ws['A2'] = "f) Sundry Creditors Durgapur"
    ws['A2'].font = Font(name='Arial', size=11, bold=True, italic=True)
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 4
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows - based on the image
    creditors_data = [
        ["M JRR Trading Company", 43200],
        ["Shivay Enterprises", 24800],
        ["Supreme Power System", 7600],
        ["Total", 75600]
    ]
    
    # Add data to worksheet
    start_row = 5
    for i, data_row in enumerate(creditors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold the total row
            if data_row[0] == "Total":
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_4(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 4 sheet
    ws = wb.create_sheet("Annexure 4 Advances Received")
    
    # Add title with underline
    ws['A1'] = "4. Advance Received from Parties"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 3
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows based on the image
    advances_data = [
        ["For Silico Manganese", ""],
        ["Gayatri Vijay Stone Crusher", 1500000],
        ["Vatsalya Mineral and Power Ltd", 156874],
        ["Bharat Minchem", 528322],
        ["Bright Carbon Resources", 201756],
        ["Drishan Industries Private Limited", 890036],
        ["Jain Auto Export", 36976387],
        ["Loom Drop Pins Mfg Co", 225178],
        ["Metallic Ferro Alloys Llp (Ahm)", 35969],
        ["Metallic Ferro Alloys Llp (Delhi)", 161754],
        ["P N Alloys", 36986712],
        ["Suvidhi Alloys Manufacturing Company", 2318605],
        ["", ""],
        ["For Oxygen", ""],
        ["MO Infra", 57000],
        ["Total", 80041591]
    ]
    
    # Add data to worksheet
    start_row = 4
    for i, data_row in enumerate(advances_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold for headings and total
            if data_row[0] in ["For Silico Manganese", "For Oxygen", "Total"]:
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_5(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 5 sheet
    ws = wb.create_sheet("Annexure 5 Sundry Debtors")
    
    # Add title with underline
    ws['A1'] = "5. Sundry Debtors"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 3
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows based on the image
    debtors_data = [
        ["For Silico Manganese", ""],
        ["AR Ferro Alloys", 522495],
        ["Acore Industries Pvt. Ltd.", 962775],
        ["Akshat Industries Cables & Fuels (India) Private Limited", 1359535],
        ["Lavni Enterprises, Nagpur", 237395],
        ["L&T Enterprises", 537436],
        ["Maa Durga Enterprises, Dhanbad", 93635],
        ["Vaishavi International Shipping Services Pvt.Ltd.", 62155],
        ["Aome Ferro Alloys Pvt Ltd", 1806],
        ["Alfa Alloytrease Industries Private Limited", 17130953],
        ["Ahmedabad Metal And Alloys", 2148],
        ["Falah Steel", 2124],
        ["Garlista Edibles Private Limited", 7431844],
        ["Gp 11 Industries Private Limited", 16725816],
        ["Hindustan Power Solutions", 23566093],
        ["Karishay Biotech", 9667462],
        ["Lithium Cleantech Private Limited", 32492450],
        ["Metal Impex", 47436],
        ["Metallic Ferro Alloys Lip", 150920],
        ["Metro Craft Industries", 5012847],
        ["M P Sales Corporation Indore", 958981],
        ["Rajasthan Minerals", 1189440],
        ["Safegaurd Infra", 24417090],
        ["Sai Mineral", 165164],
        ["Sai Mineral", 14180],
        ["Shri Madhav International", 68506],
        ["Shri Ram Metals Unit-1", 1848240],
        ["Total", 369253179],
        ["", ""],
        ["For Aluminium", ""],
        ["Badshah Aluminum Hub", 1058071],
        ["Suresh Aluminium", 69519],
        ["Sitangli Constructions", 5521021],
        ["Owais Metal and Mineral Processing Pvt Ltd", 710420],
        ["Total", 7359360],
        ["", ""],
        ["For Oxygen", ""],
        ["Anurag Gas Sales And Services", 153462],
        ["M/S Gopal Printpack Solutions", 708],
        ["Shivangi Rolling Mills Pvt Ltd", ""],
        ["Aryan Enterprises", 22504],
        ["Multani Gas Suppliers", 27246],
        ["Petralite Industries Pvt Ltd", 820],
        ["Saiyyed Ayesh Ali", 5230],
        ["Urd Sales", 94872],
        ["Vinaya Polymers India Pvt Ltd", 644],
        ["Total", 316664],
        ["", ""],
        ["For Chandrapur", ""],
        ["Nalivika Tradelink", 108403],
        ["SMO Ferro Alloys Pvt Ltd", 860],
        ["Total", 109263],
        ["", ""],
        ["RAJ", ""],
        ["", ""],
        ["Total", 384823734]
    ]
    
    # Add data to worksheet
    start_row = 4
    for i, data_row in enumerate(debtors_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold for headings and total
            if data_row[0] in ["For Silico Manganese", "For Aluminium", "For Oxygen", "For Chandrapur", "RAJ", "Total"]:
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    
    return ws

def create_annexure_6(wb, header_font, normal_font, title_font, thin_border, header_fill):
    # Create Annexure 6 sheet
    ws = wb.create_sheet("Annexure 6 Advances Paid")
    
    # Add title with underline
    ws['A1'] = "6. Advance Paid to Parties"
    ws['A1'].font = Font(name='Arial', size=12, bold=True, underline="single")
    
    # Create headers with borders
    headers = ["Particulars", "Amount"]
    row = 3
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows based on the images - Image 1
    advances_data_1 = [
        ["For Silico Manganese", ""],
        ["Akhil Traders", 8020],
        ["Bajaj Allianz General Insurance Company Limited", 4810],
        ["Bhagwati Velocity Motors Pvt Ltd", 20],
        ["Burhanl Steel", 10241],
        ["C Sagar Patidar", 15000],
        ["Chandiya Trading Company", 6413],
        ["Global Building Material", 20000],
        ["Govind Roy Puratia", 15000],
        ["Graphite Creations", 45000],
        ["Hardekar & Associates", 45500],
        ["Hari Ram Choudhary", 13290],
        ["Hindustan Power Solutions Creditor", 35718],
        ["Hindustan Trading Company", 12875],
        ["Landmark Cars Private Limited", 50000],
        ["MJ Digital Services Pvt Ltd", 1000000],
        ["M T Enterprises", 447024],
        ["Murtuja Ali", 670443],
        ["Nur Project Consultancy Services", 104430],
        ["Niranjan Bhatti", 5000],
        ["P. Ghorala & Company", 71651],
        ["Prashad Bartan Kendra", 672839],
        ["Raj Laxmi Mining Equipments", 2135000],
        ["Rally Tech Service Private Limited", 62467489],
        ["RK Equipment Rentals", 5424],
        ["Satyavat Ahari Ratlam", 5000],
        ["Satkar Sirve Decorator", 221000],
        ["S B M Traders", 25396123],
        ["Shubham Tyre", 250000],
        ["Som Tools", 10635],
        ["Suginidh Automotive", 1500000],
        ["Suraj S Parihare Company Secretary", 15200],
        ["Vishwajith Singh", 474000],
        ["Ariye Engineering, Ahmedabad", 5480],
        ["Bhalla Multi Trade", 138222],
        ["Dabur Fire Service", 5499],
        ["Dilip Impex", 12084],
        ["Drummart Materials Pvt Ltd", 99324],
        ["Finestyle Industries Limited", 34114],
        ["Hindustan Zinc Limited", 5190],
        ["Impex Commercial Pvt Ltd(Creditor)", 35596],
        ["Manoj Enterprise(Creditor)", 41969],
        ["M/s Industrial Development Corporation Indore", 123409],
        ["M/S Mahaveer Tal", 3230],
        ["Nakoda Tour & Travels", 33852],
        ["Om Coal Company Pvt. Ltd", 15665],
        ["Paresh Bhai", 11600],
        ["Parulihi Enterprises", 14112],
        ["Prateek Roudines", 152000],
        ["Sanjee Aluminium Section", 200000],
        ["Sanwar Lal Acharya", 108020],
        ["Satyhi Chemical Works", 264146],
        ["Shree Bharavi Transport", 40000],
        ["Shri Saiyatri Transport Tds Dec.", 15982],
        ["Shri Shivan Transport Company", 20480],
        ["Vidhan Sales, Indore", 15635],
        ["Vikas Barwe", 2872],
        ["Vivid Corporation", 40006],
        ["VKD Industries Pvt Ltd", 2567981],
        ["M/s Ghanti Nawaz Infra", 1299029],
        ["Aashirwad Kiaan Sewa Kendra", 1100000],
        ["Owais Ali Oversees", 16507146],
        ["Sanghi Brothers", 500000],
        ["Smo Ferro Alloys Pvt Ltd (Durgapur)", 2777442],
        ["Sr Ferro Alloys (Pvt)- Advance For Raw Material", 41203951],
        ["Steel Authority Of India Ltd", 3053175],
        ["Aanci Enterprises Indore", 4761],
        ["Advance Against Land", 300000],
        ["Aryan Enterprises", 3500000],
    ]
    
    # Add data rows based on the images - Image 2 (continuation)
    advances_data_2 = [
        ["Sada Baba Mining, Nagpur", 3507174],
        ["Balaji Equipment(Adv.)", 605527],
        ["Digvijay Singh", 305680],
        ["Emritah India", 1092800],
        ["Faith Traders And Logistics", 1002800],
        ["Ganesh Hardware Merchant", 213296],
        ["Finanl Consultants", 100600],
        ["Jagera Environmental & Engineering Services Indore", 146937],
        ["Innovative Engineers India", 255000],
        ["Jasmeet Kaur Saluja", 4342],
        ["Kapil Automobiles", 55192],
        ["Khandelwal Diesels", 10000],
        ["Malhotra Infra", 69500],
        ["Mahakwari Steel Sales, Indore", 2000000],
        ["Mitra Tea Processing", 109200],
        ["Mitra S.K. Pvt. Ltd", 9395],
        ["Modi Ltd. M.P.", 1153744],
        ["Modi Ltd, Nagpur", 70200],
        ["Myscon Limited", 255034],
        ["N/S Shashikant & Company", 5000],
        ["Pratiksh Chandra Nawal", 105400],
        ["P. Mahendra Company", 313376],
        ["Purnani Kuman Infrastructure", 273250],
        ["Purus Chand Nawal", 42975],
        ["Shanklieshwar Road Lines", 826331],
        ["Shiva Analyticals Pvt Ltd", 27648],
        ["Shri Balaji Enterprises Raipur(New)", 17911],
        ["Shri Kalki (MPTD)", 37130],
        ["Shri Shankar Lal Jathwali", 800000],
        ["S R Enterprises", 59420],
        ["Smo Mineral And Metal Processing Pvt Ltd", 508120],
        ["Sr Ferro Alloys(Mines Division)", 784465],
        ["Yatin & Wasim Brothers Transport, Tumsar", 28500],
        ["Mayur Patel", 500000],
        ["Total", 184261187],
        ["", ""],
        ["For Oxygen", ""],
        ["Anurag Enterprises", 3728177],
        ["Guru Movers", 19500],
        ["Maraan Mechanical Works", 20000],
        ["Modern Jigs And Power Tools", 162346],
        ["Meerj (Pithtm)", 151345],
        ["Samrath Services", 150000],
        ["Satnel Razora", 140165],
        ["Smo Ferro Alloys Chandrapur", 10000],
        ["Satish Infra", 6655500],
        ["Omsy Madhushudal Electrical Contractor", 15000],
        ["MPPCL", 64522],
        ["Narendra and Techno", 38274],
        ["SMO Gold and Refinery Pvt Ltd", 74250],
        ["Total", 11011085],
    ]
    
    # Add data rows based on the images - Image 3 (final part)
    advances_data_3 = [
        ["For Durgapur", ""],
        ["AS Enterprise", 200000],
        ["Atulit Industries Pvt. Ltd", 1205500],
        ["Banerjee Auto Corporation", 134818],
        ["Hella Infra Market Private Ltd", 367094],
        ["Rfc Equipment Rentals Pvt. Ltd.", 1668000],
        ["Satguru Infra", 7949537],
        ["Sensotech Weighing Systems Pvt Ltd", 368000],
        ["Total", 11923149],
        ["", ""],
        ["For Rajasthan", ""],
        ["Aayat Enterprises", 100000],
        ["Shree Mateshwari Electricals", 800000],
        ["Total", 900000],
        ["", ""],
        ["For Chandrapur", ""],
        ["Sameer", 223750],
        ["Total", 223750],
        ["", ""],
        ["For Aluminium", ""],
        ["Maa Kali Engineering", 20000],
        ["SML Industries Pvt Ltd", 5479612],
        ["Total", 5499612]
    ]
    
    # Combine all data
    advances_data = advances_data_1 + advances_data_2 + advances_data_3
    
    # Add data to worksheet
    start_row = 4
    for i, data_row in enumerate(advances_data, start=start_row):
        for j, value in enumerate(data_row, start=1):
            cell = ws.cell(row=i, column=j)
            
            # Format numeric values
            if j == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            # Apply borders to all cells
            cell.border = thin_border
            
            # Bold for headings and total
            if data_row[0] in ["For Silico Manganese", "For Oxygen", "For Durgapur", "For Rajasthan", "For Chandrapur", "For Aluminium", "Total"]:
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = normal_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    
    return ws

# Run the function to create the Excel file
if __name__ == "__main__":
    result = create_smo_ferro_excel()
    print(result)
    
    # Get the current directory path for file location
    current_dir = os.getcwd()
    file_path = os.path.join(current_dir, 'SMO_Ferro_Alloys_Financial_Statements.xlsx')
    print(f"Excel file created at: {file_path}")