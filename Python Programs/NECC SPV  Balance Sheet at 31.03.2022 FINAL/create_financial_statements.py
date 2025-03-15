import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()

# Create sheets
balance_sheet = wb.active
balance_sheet.title = "Balance Sheet"

profit_loss = wb.create_sheet("Profit and Loss")
cash_flow = wb.create_sheet("Cash Flow Statement")
fixed_assets = wb.create_sheet("Fixed Assets")
share_capital = wb.create_sheet("Share Capital")
notes = wb.create_sheet("Notes to Accounts")
sundry_creditors = wb.create_sheet("Sundry Creditors")

# Create Notes sheets
for i in range(2, 25):
    wb.create_sheet(f"Note {i}")

# Define styles
header_font = Font(bold=True, size=12)
subheader_font = Font(bold=True, size=11)
normal_font = Font(size=10)

thin_border = Side(border_style="thin")
border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
total_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")

# Balance Sheet data
balance_sheet_data = [
    ["", "", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", ""],
    ["BALANCE SHEET AS ON 31ST MARCH, 2022", "", "", ""],
    ["(Amt in INR)", "", "", ""],
    ["", "", "", ""],
    ["Particulars", "Note No", "As on 31st March 2022", "As on 31st March 2021"],
    ["", "", "", ""],
    ["EQUITY AND LIABILITIES", "", "", ""],
    ["Shareholder's Funds", "", "", ""],
    ["Share Capital", "2", "1,00,000.00", "1,00,000.00"],
    ["Reserves and Surplus", "3", "6,04,03,413.40", "4,64,63,916.96"],
    ["", "", "", ""],
    ["Non-Current Liabilities", "", "", ""],
    ["Long-Term Borrowings", "4", "81,52,99,928.08", "24,05,85,315.08"],
    ["Other Long Term Liabilities", "5", "42,34,846.00", "26,24,979.00"],
    ["Deferred Tax", "", "4,70,899.75", "-"],
    ["", "", "", ""],
    ["Current Liabilities", "", "", ""],
    ["Short-Term Borrowings", "6", "-", "-"],
    ["Trade Payables", "7", "8,94,35,064.65", "8,56,01,385.61"],
    ["Other Current Liabilities", "8", "61,31,89,014.00", "11,18,25,105.00"],
    ["Short Term Provisions", "9", "42,21,897.28", "93,89,122.00"],
    ["", "", "", ""],
    ["TOTAL", "", "1,58,73,55,063.16", "49,65,89,823.65"],
    ["", "", "", ""],
    ["ASSETS", "", "", ""],
    ["Non-Current Assets", "", "", ""],
    ["Fixed Assets", "", "", ""],
    ["Tangible Assets", "10", "8,01,62,449.98", "3,07,73,937.96"],
    ["Intangible Assets", "10", "-", "-"],
    ["Long Term Loans & Advances", "11", "38,66,100.00", "48,06,100.00"],
    ["Deferred Tax", "12", "-", "2,24,731.53"],
    ["Capital work in Progress", "", "21,53,22,037.68", "13,11,65,800.00"],
    ["", "", "", ""],
    ["Current Assets", "", "", ""],
    ["Inventories & Work in Progress", "13", "65,38,400.00", "46,28,900.00"],
    ["Trade Receivables", "14", "36,59,03,669.00", "1,21,02,833.00"],
    ["Cash and Cash Equivalents", "15", "49,55,61,553.82", "5,26,47,684.58"],
    ["Short Term Loans & Advances", "16", "41,99,80,185.98", "26,02,16,873.58"],
    ["Other Current Assets", "17", "20,666.70", "22,963.00"],
    ["", "", "", ""],
    ["TOTAL", "", "1,58,73,55,063.16", "49,65,89,823.65"]
]

# Profit and Loss data
profit_loss_data = [
    ["", "", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", ""],
    ["Statement of Profit and Loss for the year ended March 31, 2022", "", "", ""],
    ["(Amt in INR)", "", "", ""],
    ["", "", "", ""],
    ["Particulars", "Note No", "For the year Ended 31.03.2022", "For the year Ended 31.03.2021"],
    ["", "", "", ""],
    ["INCOME", "", "", ""],
    ["Revenue From operation", "18", "40,92,51,547.00", "38,13,90,243.00"],
    ["Other Income", "19", "1,56,683.92", "2,10,146.93"],
    ["TOTAL", "", "40,94,08,230.92", "38,16,00,389.93"],
    ["", "", "", ""],
    ["EXPENSES", "", "", ""],
    ["Purchase of Traded Goods", "20", "12,95,02,435.98", "6,43,65,699.38"],
    ["Changes in inventories of finished goods and Stock-in-Trade", "21", "(19,09,500.00)", "15,86,840.00"],
    ["Employee benefit expense", "22", "10,73,61,047.69", "11,97,28,974.87"],
    ["Financial costs", "23", "3,29,32,572.49", "6,34,530.65"],
    ["Depreciation and amortization expense", "10", "65,28,330.45", "61,99,000.00"],
    ["Other expenses", "24", "11,61,36,319.31", "15,60,08,028.91"],
    ["TOTAL", "", "39,05,51,205.92", "34,85,23,073.81"],
    ["", "", "", ""],
    ["Profit before exceptional and extraordinary items and tax", "", "1,88,57,025.00", "3,30,77,316.12"],
    ["Exceptional Items", "", "-", "-"],
    ["Extraordinary Items", "", "-", "-"],
    ["Profit before tax", "", "1,88,57,025.00", "3,30,77,316.12"],
    ["", "", "", ""],
    ["Tax expense:", "", "", ""],
    ["Current tax", "", "(42,21,897.28)", "(93,89,122.00)"],
    ["Deferred tax", "", "(6,95,631.28)", "1,89,480.97"],
    ["MAT Credit", "", "-", "-"],
    ["Prior Period Income", "", "-", "-"],
    ["Profit/(Loss) for the period", "", "1,39,39,496.44", "2,38,77,675.09"],
    ["", "", "", ""],
    ["Earning per Equity Share", "", "", ""],
    ["(1) Basic", "", "17.42", "29.85"],
    ["(2) Diluted", "", "17.42", "29.85"]
]

# Cash Flow data
cash_flow_data = [
    ["", "", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", ""],
    ["Cash Flow Statement For The Year Ended March 31, 2022", "", "", ""],
    ["", "", "", ""],
    ["", "", "For the Year ended 31st March, 2022", "For the Year ended 31st March, 2021"],
    ["", "", "", ""],
    ["I) Cash Flows from Operating Activities", "", "", ""],
    ["Net profit before tax,and extraordinary item", "", "1,88,57,025.00", "3,30,77,316.12"],
    ["Adjustments For Non cash & Non Operating Activities", "", "", ""],
    ["Loss on sale of fixed assest", "", "-", "-"],
    ["Depreciation", "", "65,28,330.45", "61,99,000.00"],
    ["Rental Income", "", "-", "-"],
    ["Interest Expenses", "", "3,29,32,572.49", "6,34,530.65"],
    ["Prior Period Adjustments", "", "-", "23,936.00"],
    ["MAT Credit", "", "-", "-"],
    ["", "", "5,83,17,927.94", "3,99,34,782.77"],
    ["Operating Profit Before Working Capital Changes", "", "5,83,17,927.94", "3,99,34,782.77"],
    ["Change in Trade Receivables", "", "(35,38,00,836.00)", "83,21,167.00"],
    ["Change in Inventory", "", "(19,09,500.00)", "15,86,840.00"],
    ["Change in Short Term loans & Advances", "", "(15,97,63,312.40)", "(19,15,19,764.80)"],
    ["Change in Long Term loans & Advances", "", "9,40,000.00", "(10,44,000.00)"],
    ["Change in Other Current Assets", "", "2,296.30", "2,552.00"],
    ["Change in Capitcal Work in Progress", "", "(8,41,56,237.68)", "(11,31,08,660.00)"],
    ["Change in Short Term Borrowings", "", "-", "-"],
    ["Change in Trade Payables", "", "38,33,679.04", "4,89,16,426.49"],
    ["Change in Other Current Liabilities", "", "50,13,63,909.00", "(8,38,55,453.20)"],
    ["Change in Other Long Term Liabilities", "", "16,09,867.00", "26,24,979.00"],
    ["", "", "(9,18,80,134.74)", "(32,80,75,913.51)"],
    ["Cash Generated from Operations", "", "(3,35,62,206.80)", "(28,81,41,130.74)"],
    ["Cash Flow from Operating Activities before Taxes", "", "(3,35,62,206.80)", "(28,81,41,130.74)"],
    ["Income Tax paid", "", "(93,89,122.00)", "(87,67,909.00)"],
    ["Net Cash From Operating Activity (A)", "", "(4,29,51,328.80)", "(29,69,09,039.74)"],
    ["", "", "", ""],
    ["II) Cash Flows from Investing Activities", "", "", ""],
    ["Purchase of Fixed Assets", "", "(5,59,16,842.48)", "(3,17,65,011.47)"],
    ["Decrease in WIP", "", "-", "-"],
    ["Sale Proceeds from Fixed Assets", "", "-", "-"],
    ["Rental Income", "", "-", "-"],
    ["Net Cash From Investing Activity (B)", "", "(5,59,16,842.48)", "(3,17,65,011.47)"],
    ["", "", "", ""],
    ["III) Cash Flows From Financing Activities", "", "", ""],
    ["Increase / (decrease) in Long Term Borrowings", "", "57,47,14,613.00", "21,42,92,192.00"],
    ["Interest paid", "", "(3,29,32,572.49)", "(6,34,530.65)"],
    ["", "", "54,17,82,040.51", "21,36,57,661.35"],
    ["Net Cash From Financing Activity (C)", "", "54,17,82,040.51", "21,36,57,661.35"],
    ["", "", "", ""],
    ["Net Increase in Cash And Cash equivalents(A+B+C)", "", "44,29,13,869.23", "(11,50,16,389.86)"],
    ["Cash & Cash Equivalents at the beginning of the year", "", "5,26,47,683.94", "16,76,64,073.80"],
    ["Cash & Cash Equivalents at the end of the year", "", "49,55,61,553.17", "5,26,47,683.94"]
]

# Fixed Assets data
fixed_assets_data = [
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", "", "", "", "", "", "", ""],
    ["SCHEDULE OF FIXED ASSETS AS PER COMPANIES ACT, 2013 AS ON 31ST MARCH 2022", "", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["", "GROSS BLOCK", "", "", "", "DEPRECIATION", "", "", "NET BLOCK", "", ""],
    ["PARTICULARS", "COST AS ON 01.04.2021", "ADDITION", "DEDUCTION", "COST AS ON 31.03.2022", "UP TO 01.04.2021", "DURING THE YEAR", "ADJUSTMENTS", "TOTAL UP TO 31.03.2022", "W.D.V. AS ON 31.03.2022", "W.D.V. AS ON 31.03.2021"],
    ["", "", "", "", "", "", "", "ON A/C OF SALE or Other Adjustment", "", "", ""],
    ["CWIP-TOOLS & EQUIPMENTS", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["TANGIBLE ASSETS", "", "", "", "", "", "", "", "", "", ""],
    ["BUILDING", "-", "-", "-", "-", "", "", "", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["PLANT & MACHINERY", "", "", "", "", "", "", "", "", "", ""],
    ["Plant & Machinery", "2,46,73,469.93", "-", "-", "2,46,73,469.93", "43,03,528.00", "44,65,898.06", "-", "87,69,426.06", "2,02,07,571.87", "2,46,73,469.93"],
    ["Generator", "0.00", "-", "-", "0.00", "-", "-", "-", "-", "-", "-"],
    ["Lab Equipments", "14,30,132.60", "45,140.00", "", "14,75,272.60", "7,88,924.00", "3,73,111.33", "", "11,62,035.33", "11,02,161.27", "14,30,132.60"],
    ["Tools & Equipments", "93,542.00", "-", "-", "93,542.00", "23,216.00", "16,931.10", "-", "40,147.10", "76,610.90", "93,542.00"],
    ["Electrical Assets", "1,74,263.83", "-", "-", "1,74,263.83", "35,487.00", "31,541.75", "-", "67,028.75", "1,42,722.08", "1,74,263.83"],
    ["Crusher", "", "5,55,84,745.00", "", "5,55,84,745.00", "", "", "", "-", "5,55,84,745.00", ""],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["OFFICE EQUIPMENTS", "", "", "", "", "", "", "", "", "", ""],
    ["Mobile and Mac Book", "67,975.00", "24,500.00", "-", "92,475.00", "1,69,737.00", "54,094.01", "-", "2,23,831.01", "38,380.99", "67,975.00"],
    ["Computer and Laptops", "4,53,196.00", "2,62,457.48", "-", "7,15,653.48", "8,04,719.00", "3,93,187.30", "-", "11,97,906.30", "3,22,466.18", "4,53,196.00"],
    ["Office Equipments", "2,19,695.00", "-", "-", "2,19,695.00", "2,00,920.00", "99,016.54", "-", "2,99,936.54", "1,20,678.46", "2,19,695.00"],
    ["Air Conditioner", "3,15,349.00", "-", "-", "3,15,349.00", "1,02,073.00", "81,643.86", "-", "1,83,716.86", "2,33,705.14", "3,15,349.00"],
    ["Refrigerator", "36,089.00", "-", "", "36,089.00", "9,504.00", "6,532.11", "", "16,036.11", "29,556.89", "36,089.00"],
    ["Camera", "1,06,327.00", "-", "-", "1,06,327.00", "41,158.00", "47,921.58", "-", "89,079.58", "58,405.42", "1,06,327.00"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["VEHICLES", "24,15,046.19", "-", "-", "24,15,046.19", "3,40,596.00", "7,54,218.93", "-", "10,94,814.93", "16,60,827.26", "24,15,046.19"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["FURNITURE & FIXTURES", "7,88,852.41", "-", "-", "7,88,852.41", "4,13,655.00", "2,04,233.89", "-", "6,17,888.89", "5,84,618.52", "7,88,852.41"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["INTANGIBLE ASSETS", "", "", "", "", "", "", "", "", "", ""],
    ["Computer Softwares", "-", "-", "-", "-", "", "", "", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["TOTAL", "3,07,73,937.96", "5,59,16,842.48", "0.00", "8,66,90,780.44", "72,33,517.00", "65,28,330.45", "0.00", "1,37,61,847.45", "8,01,62,449.98", "3,07,73,937.96"]
]

# Share Capital data
share_capital_data = [
    ["", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", ""],
    ["NOTES ANNEXED TO AND FORMING THE PART OF BALANCE SHEET FOR THE YEAR ENDED 31ST MARCH 2022", "", ""],
    ["", "", ""],
    ["", "(Amt in INR)", ""],
    ["Particulars", "As on 31st March 2022", "As on 31st March 2021"],
    ["", "", ""],
    ["NOTE 2 : SHARE CAPITAL", "", ""],
    ["Authorised Capital", "", ""],
    ["10,000 Equity Shares of Rs. 10/- each", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["Issued, Subscribed & Paid up", "", ""],
    ["10,000 Equity Shares of 10/- each fully Paid up", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["TOTAL", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["The Company has only one class of shares referred to as equity shares having a par value of Rs. 10/- each.", "", ""],
    ["Each holder of Equity share is entitled to one vote per share. The company declares & pay dividend", "", ""],
    ["when the same is approved by the shareholders in the ensuring Annual General Meeting.", "", ""],
    ["", "", ""],
    ["Reconciliation of the number of shares outstanding as on the date of Balance Sheet:", "", ""],
    ["", "", ""],
    ["Particulars", "Equity Shares", ""],
    ["", "Number", "Rs."],
    ["Shares outstanding at the beginning of the year", "1,00,000.00", "1,00,000.00"],
    ["Add/Less: Shares issued/redeem during the year", "-", "-"],
    ["Shares outstanding at the end of the year", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["Details of Shareholder holding more than 5% share as on the date of Balance Sheet :", "", ""],
    ["", "", ""],
    ["Name of Shareholders", "As on 31-03-2022", "As on 31-03-2021"],
    ["", "%", "%"],
    ["Nandhra Eng. & Construction(India) Pvt. Ltd.", "51.99", "51.99"],
    ["Santosh Ankush Mohite", "24.00", "24.00"],
    ["Kishor Ankush Mohite", "24.00", "24.00"],
    ["Rahat Virk", "0.01", "0.01"],
    ["", "100.00", "100.00"],
    ["", "", ""],
    ["NOTE 3 : RESERVES & SURPLUS", "", ""],
    ["Surplus", "", ""],
    ["Opening balance", "4,64,63,916.96", "2,25,62,305.87"],
    ["(+) Net Profit/(Net Loss) For the current year", "1,39,39,496.44", "2,38,77,675.09"],
    ["(+) Adjustment for earlier years", "-", "23,936.00"],
    ["Total", "6,04,03,413.40", "4,64,63,916.96"]
]

# Notes data
notes_data = [
    ["", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", ""],
    ["NOTES TO FINANCIAL STATEMENTS", ""],
    ["For the year ended March 31, 2022", ""],
    ["", ""],
    ["Note 1: Significant Accounting Policies", ""],
    ["", ""],
    ["a) Basis of Preparation:", "The financial statements have been prepared in accordance with the Generally Accepted Accounting Principles in India (Indian GAAP) to comply with the Accounting Standards specified under Section 133 of the Companies Act, 2013."],
    ["", ""],
    ["b) Use of Estimates:", "The preparation of financial statements in conformity with Indian GAAP requires the management to make judgments, estimates and assumptions that affect the reported amounts of revenues, expenses, assets and liabilities and the disclosure of contingent liabilities, at the end of the reporting period."],
    ["", ""],
    ["c) Fixed Assets:", "Fixed Assets are stated at cost less accumulated depreciation and impairment losses if any. Cost comprises the purchase price and any attributable cost of bringing the asset to its working condition for its intended use."],
    ["", ""],
    ["d) Depreciation:", "Depreciation on Fixed Assets is provided on Straight Line Method as per the useful life prescribed in Schedule II to the Companies Act, 2013."],
    ["", ""],
    ["e) Revenue Recognition:", "Revenue is recognized to the extent that it is probable that the economic benefits will flow to the Company and the revenue can be reliably measured."],
    ["", ""],
    ["f) Taxes on Income:", "Current tax is the amount of tax payable on the taxable income for the year as determined in accordance with the provisions of the Income Tax Act, 1961."],
    ["", "Deferred tax is recognized on timing differences, being the differences between the taxable income and the accounting income that originate in one period and are capable of reversal in one or more subsequent periods."],
    ["", ""],
    ["g) Provisions and Contingent Liabilities:", "A provision is recognized when the Company has a present obligation as a result of past events and it is probable that an outflow of resources will be required to settle the obligation in respect of which a reliable estimate can be made."],
    ["", "Contingent liabilities are disclosed when there is a possible obligation arising from past events, the existence of which will be confirmed only by the occurrence or non-occurrence of one or more uncertain future events not wholly within the control of the Company."]
]

# Additional Notes data
note5_data = [
    ["NOTE 5 : Other Long Term Liabilities", "", ""],
    ["Security Deposits- Labour", "42,34,846.00", "26,24,979.00"],
    ["Total", "42,34,846.00", "26,24,979.00"]
]

note6_data = [
    ["NOTE 6 : SHORT TERM BORROWINGS", "", ""],
    ["Loans repayable on demand", "", ""],
    ["From Banks", "-", ""],
    ["TOTAL", "-", "-"]
]

note7_data = [
    ["NOTE 7 : TRADE PAYABLES", "", ""],
    ["For Expenses", "3,06,87,283.18", "3,84,69,885.00"],
    ["For Goods Purchase", "2,72,97,063.40", "1,40,22,539.54"],
    ["For Subcontractor", "3,14,50,718.07", "3,31,08,961.07"],
    ["TOTAL", "8,94,35,064.65", "8,56,01,385.61"]
]

note8_data = [
    ["NOTE 8 : OTHER CURRENT LIABILITIES", "", ""],
    ["Current maturities of Long term Debt", "4,12,288.00", ""],
    ["Others", "", ""],
    [" - Advances From Customers", "55,63,00,000.00", "-"],
    [" - Cheque Issue but not Clear", "-", ""],
    [" - Expenses Payables", "3,50,000.00", "8,90,000.00"],
    [" - Statutory Liabilities", "30,00,604.00", "1,08,50,046.00"],
    [" - Payable to employees", "62,88,410.00", "51,72,771.00"],
    [" - Other Payable", "4,72,50,000.00", "9,45,00,000.00"],
    ["TOTAL", "61,31,89,014.00", "11,18,25,105.00"]
]

note9_data = [
    ["NOTE 9 : SHORT TERM PROVISIONS", "", ""],
    ["Provision for Income Tax", "42,21,897.28", "93,89,122.00"],
    ["TOTAL", "42,21,897.28", "93,89,122.00"]
]

note11_data = [
    ["NOTE 11 : LONG TERM LOANS AND ADVANCES", "", ""],
    ["UnSecured Considered Good", "", ""],
    ["Security Deposits", "38,66,100.00", "48,06,100.00"],
    ["TOTAL", "38,66,100.00", "48,06,100.00"]
]

note12_data = [
    ["NOTE 12 : DEFERRED TAX ASSET", "", ""],
    ["Deferred Tax Asset", "-", "2,24,731.53"],
    ["TOTAL", "-", "2,24,731.53"]
]

note13_data = [
    ["NOTE 13 : INVENTORIES", "", ""],
    ["General Consumable & Stock", "65,38,400.00", "46,28,900.00"],
    ["TOTAL", "65,38,400.00", "46,28,900.00"]
]

note14_data = [
    ["NOTE 14 : TRADE RECEIVABLES", "", ""],
    ["Trade Receivables", "", ""],
    [" -UnSecured Considered Good", "", ""],
    ["Executive Engineer, PWD, Khamgaon", "", ""],
    [" -Less than Six Months from Due Date", "36,59,03,669.00", "1,21,02,833.00"],
    [" -Others", "-", ""],
    ["TOTAL", "36,59,03,669.00", "1,21,02,833.00"]
]

note15_data = [
    ["NOTE 15 : CASH & CASH EQUIVALENTS", "", ""],
    ["Balances with Banks :", "", ""],
    ["In Current and Escrow Accounts", "49,55,44,363.81", "5,25,21,709.58"],
    ["Cheque received but not clear", "-", ""],
    ["Cash in hand & Imprest Balances", "17,190.01", "1,25,975.00"],
    ["TOTAL", "49,55,61,553.82", "5,26,47,684.58"]
]

note16_data = [
    ["NOTE 16 : SHORT TERM LOANS & ADVANCES", "", ""],
    ["Others", "", ""],
    ["Unsecured Considered Good", "", ""],
    [" - Advances to vendors", "39,35,29,989.28", "24,10,12,701.00"],
    [" - Loans & advances to Employees", "1,43,46,262.40", "67,39,109.40"],
    [" - Prepaid Expenses", "14,83,235.00", "16,22,683.00"],
    [" - GST & Service Tax", "49,25,850.55", "84,62,059.26"],
    [" - Income Tax & TDS", "56,94,848.75", "23,80,320.92"],
    ["TOTAL", "41,99,80,185.98", "26,02,16,873.58"]
]

note17_data = [
    ["NOTE 17 : OTHER CURRENT ASSETS", "", ""],
    ["Others", "", ""],
    [" - Preoperating Expenses", "20,666.70", "22,963.00"],
    [" - MAT Credit Entitlement", "-", ""],
    ["TOTAL", "20,666.70", "22,963.00"]
]

note18_data = [
    ["NOTE 18 : REVENUE FROM OPERATIONS", "", ""],
    ["Sale of products:", "", ""],
    [" Works Contract PWD", "40,92,51,547.00", "38,13,90,243.00"],
    ["", "40,92,51,547.00", "38,13,90,243.00"],
    ["Less: Excise Duty", "-", ""],
    ["Total", "40,92,51,547.00", "38,13,90,243.00"]
]

note19_data = [
    ["NOTE 19 : OTHER INCOME", "", ""],
    ["Discount Received", "1,56,683.92", "2,10,146.93"],
    ["Round Off", "-", ""],
    ["Total", "1,56,683.92", "2,10,146.93"]
]

note20_data = [
    ["NOTE 20 : PURCHASE OF STOCK IN TRADE", "", ""],
    ["Electrical and Hardware Material Purchase", "2,28,671.65", "18,17,492.26"],
    ["Safety Material Purchase", "25,78,054.38", "4,61,637.69"],
    ["Other Material Purchase", "6,53,55,646.20", "4,15,14,939.81"],
    ["RCC Pipe and Material", "16,83,200.27", "20,68,300.00"],
    ["Steel Purchase", "4,06,51,236.71", "2,77,32,575.40"],
    ["Cement Purchase", "1,09,30,551.79", "2,15,75,050.00"],
    ["Diesel Purchase", "9,22,31,312.66", "8,23,04,364.22"],
    ["", "21,36,58,673.66", "17,74,74,359.38"],
    ["Add:- Material transferred to work in progress for previous period", "13,11,65,800.00", "1,80,57,140.00"],
    ["Less:- Material transferred to work in progress", "21,53,22,037.68", "13,11,65,800.00"],
    ["Total", "12,95,02,435.98", "6,43,65,699.38"]
]

note21_data = [
    ["NOTE 21 : CHANGE IN INVENTORIES", "", ""],
    ["CLOSING STOCK OF", "", ""],
    ["General Consumable", "65,38,400.00", "46,28,900.00"],
    ["Total", "65,38,400.00", "46,28,900.00"],
    ["Less : OPENING STOCK OF", "", ""],
    ["General Consumable", "46,28,900.00", "62,15,740.00"],
    ["Total", "46,28,900.00", "62,15,740.00"],
    ["Change in Inventories", "(19,09,500.00)", "15,86,840.00"],
    ["Change in Inventories", "(19,09,500.00)", "15,86,840.00"]
]

note22_data = [
    ["NOTE 22 : EMPLOYEE BENEFIT EXPENSES", "", ""],
    ["Salaries, Wages & other Allowances", "10,18,33,392.73", "11,43,61,744.90"],
    ["Contribution to PF, ESI & Labour Welfare Fund", "8,03,420.00", "11,02,937.00"],
    ["Staff & Labour Welfare", "47,24,234.96", "42,64,292.97"],
    ["Total", "10,73,61,047.69", "11,97,28,974.87"]
]

note23_data = [
    ["NOTE 23 : FINANCIAL EXPENSES", "", ""],
    ["Interest Expenses", "", ""],
    ["Interest to Banks", "", ""],
    [" - Interest on Car Loan", "1,40,792.00", "60,489.00"],
    [" - Interest on TDS", "18,01,129.00", "4,27,504.00"],
    [" - Interest on IT", "15,74,929.00", "64,258.00"],
    [" '- Interest on Term Loan", "2,93,20,942.00", ""],
    ["Bank Charges", "94,780.49", "82,279.65"],
    ["Bank Guarantee Expenses", "-", ""],
    ["Total", "3,29,32,572.49", "6,34,530.65"]
]

note24_data = [
    ["NOTE 24 : OTHER EXPENSES", "", ""],
    ["INDIRECT EXPENDITURE", "", ""],
    ["(A) ADMINISTRATIVE EXPENSES", "", ""],
    ["Audit Fees", "3,50,000.00", "3,50,000.00"],
    ["Cgst 2.5% (Rcm) (Itc Ineligible)", "-", ""],
    ["Conveyance Charges", "-", ""],
    ["Crop Compensation Paid", "11,32,821.00", "9,08,185.00"],
    ["Crusher Lease", "-", ""],
    ["Director'S Salary", "40,00,000.00", "2,06,40,000.00"],
    ["Directors Accomadation Expenses", "3,21,590.00", "3,67,412.00"],
    ["Total (A + Other expenses...)", "11,61,36,319.31", "15,60,08,028.91"]
]

# Note 4 data - Long Term Borrowings
note4_data = [
    ["", "", ""],
    ["NOTE 4 : LONG TERM BORROWINGS", "", ""],
    ["Secured", "", ""],
    ["Term Loans", "", ""],
    [" - From Banks", "12,63,341.00", "16,75,629.00"],
    [" - Union Bank of India", "34,35,10,078.00", ""],
    [" - Axis Bank Ltd.", "22,81,770.00", ""],
    ["Less : Repayable in next 1 year", "4,12,288.00", ""],
    ["Un-Secured", "", ""],
    ["Loans from Directors & Others", "46,82,44,739.08", "23,93,21,974.08"],
    ["(Interest Bearing, Payable after 1 year)", "", ""],
    ["Total", "81,52,99,928.08", "24,05,85,315.08"]
]

# Sundry Creditors Group Summary
sundry_creditors_data = [
    ["Nandhra Engineering & Construction (SPV) Pvt. Ltd.", "", ""],
    ["2nd Floor, Nathkrupa Complex,", "", ""],
    ["Near Hinjwadi Police Chowki,", "", ""],
    ["Shivaji Chowk, Hinjawadi, Pune.", "", ""],
    ["CIN: U45203PN2018PTC180980", "", ""],
    ["", "", ""],
    ["Sundry Creditors", "", ""],
    ["Group Summary", "", ""],
    ["1-Apr-2021 to 31-Mar-2022", "", ""],
    ["", "", ""],
    ["", "Sundry Creditors", ""],
    ["", "Nandhra Engineering &", ""],
    ["Particulars", "1-Apr-2021 to 31-Mar-2022", ""],
    ["", "Closing Balance", ""],
    ["", "Debit", "Credit"],
    ["", "", ""],
    ["Sundry Creditors - Expenses", "", ""],
    ["Aaadnath Trade-Wings Private Limited", "7,69,494.24", ""],
    ["A & A Consulting Engineer", "4,00,000.00", ""],
    ["Accurate Valuers And Engineers", "29,500.00", ""],
    ["Ajabrao Gulabrao Ambhore", "", "3,46,217.00"],
    ["Akul Enterprises", "", "29,441.00"],
    ["Amit Vikas Mahajan", "5,00,000.00", ""],
    ["Ananta Namdeo Unde", "", "2,95,644.00"],
    ["Anil Kisan Raut", "", "1,05,458.00"],
    ["Anirudha Baban Dahake", "", "11,46,720.00"],
    ["Ankush Shriram Jagtap", "", "1,92,555.00"]
]

# Function to safely populate the notes sheet
def populate_notes_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Second pass: Apply merges for headers
    for row_idx in [2, 3]:  # Headers that need to be merged
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    
    # Third pass: Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            try:
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Header formatting
                if row_idx in [2, 3]:  # Main headers
                    if col_idx == 1:  # First cell of merged range
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                elif col_idx == 1 and value and isinstance(value, str) and value.startswith("Note"):
                    # Note headers
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                
                # Add borders to cells with content
                if value:
                    cell.border = border
            except:
                # Skip any cells that can't be formatted
                pass
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 75

# Function to safely populate a sheet with data and apply formatting
def populate_sheet(sheet, data, has_note_column=False):
    # Set the column offset based on note column presence
    column_offset = 1 if has_note_column else 0
    
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Second pass: Apply merges
    for row_idx, row in enumerate(data, 1):
        if row_idx in [2, 3]:  # Company name and report title rows
            if len(row) > 1:
                sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(row))
    
    # Third pass: Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            try:
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Apply styles
                if row_idx in [2, 3]:  # Company name and report title
                    if col_idx == 1:  # Only format the first cell of merged cells
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                elif row_idx in [5, 6]:  # Column headers
                    cell.font = subheader_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                elif row_idx > 6 and col_idx > (1 + column_offset):  # Numeric values
                    # Right-align numeric values
                    cell.alignment = Alignment(horizontal='right')
                
                # Special formatting for section headers
                if col_idx == 1 and row_idx > 6 and isinstance(value, str) and value and not value.startswith(" "):
                    if value in ["EQUITY AND LIABILITIES", "ASSETS", "INCOME", "EXPENSES", 
                               "I) Cash Flows from Operating Activities", 
                               "II) Cash Flows from Investing Activities",
                               "III) Cash Flows From Financing Activities",
                               "TANGIBLE ASSETS", "INTANGIBLE ASSETS"]:
                        cell.font = subheader_font
                        cell.fill = subheader_fill
                
                # Special formatting for totals
                if isinstance(value, str) and value == "TOTAL" or (row_idx > 6 and isinstance(value, str) and "total" in value.lower()):
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
            except:
                # Skip cells that can't be formatted (like merged cells)
                pass
                
    # Add borders to all non-empty cells
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            if value:
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = border
    
    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max(max_length + 2, 12), 50)  # Min width 12, max 50
        sheet.column_dimensions[column_letter].width = adjusted_width