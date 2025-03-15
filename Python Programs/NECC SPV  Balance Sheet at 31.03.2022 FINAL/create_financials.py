import os
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()

# Create main sheets
balance_sheet = wb.active
balance_sheet.title = "Balance Sheet"

profit_loss = wb.create_sheet("Profit and Loss")
cash_flow = wb.create_sheet("Cash Flow Statement")
fixed_assets = wb.create_sheet("Fixed Assets")
share_capital = wb.create_sheet("Share Capital")

# Create Notes sheets
for i in range(2, 25):
    wb.create_sheet(f"Note {i}")

# Create Annexure sheets
annexure_b = wb.create_sheet("Annexure B")
annexure_note4 = wb.create_sheet("Annexure to Note 4")
annexure_note8 = wb.create_sheet("Annexure to Note 8")
annexure_note11 = wb.create_sheet("Annexure to Note 11")
annexure_note12 = wb.create_sheet("Annexure to Note 12")
annexure_note15 = wb.create_sheet("Annexure to Note 15")
annexure_note16 = wb.create_sheet("Annexure to Note 16")

# Define styles
header_font = Font(bold=True, size=12)
subheader_font = Font(bold=True, size=11)
normal_font = Font(size=10)

thin_border = Side(border_style="thin")
border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
all_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
total_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
light_blue_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")

# Balance Sheet data
balance_sheet_data = [
    ["", "", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", ""],
    ["BALANCE SHEET AS ON 31ST MARCH, 2022", "", "", ""],
    ["(Amt in INR)", "", "", ""],
    ["", "", "", ""],
    ["Particulars", "Note No", "As on 31st March 2022", "As on 31st March 2021"],
    ["", "", "", ""],
    ["EQUITY AND LIABILITIES", "", "", ""],
    ["Shareholder's Funds", "", "", ""],
    ["Share Capital", "2", "1,00,000.00", "1,00,000.00"],
    ["Reserves and Surplus", "3", "6,04,03,413.40", "4,64,63,916.96"],
    ["", "", "", ""],
    ["Non-Current Liabilities", "", "", ""],
    ["Long-Term Borrowings", "4", "81,52,99,928.08", "24,05,85,315.08"],
    ["Other Long Term Liabilities", "5", "42,34,846.00", "26,24,979.00"],
    ["Deferred Tax", "", "4,70,899.75", "-"],
    ["", "", "", ""],
    ["Current Liabilities", "", "", ""],
    ["Short-Term Borrowings", "6", "-", "-"],
    ["Trade Payables", "7", "8,94,35,064.65", "8,56,01,385.61"],
    ["Other Current Liabilities", "8", "61,31,89,014.00", "11,18,25,105.00"],
    ["Short Term Provisions", "9", "42,21,897.28", "93,89,122.00"],
    ["", "", "", ""],
    ["TOTAL", "", "1,58,73,55,063.16", "49,65,89,823.65"],
    ["", "", "", ""],
    ["ASSETS", "", "", ""],
    ["Non-Current Assets", "", "", ""],
    ["Fixed Assets", "", "", ""],
    ["Tangible Assets", "10", "8,01,62,449.98", "3,07,73,937.96"],
    ["Intangible Assets", "10", "-", "-"],
    ["Long Term Loans & Advances", "11", "38,66,100.00", "48,06,100.00"],
    ["Deferred Tax", "12", "-", "2,24,731.53"],
    ["Capital work in Progress", "", "21,53,22,037.68", "13,11,65,800.00"],
    ["", "", "", ""],
    ["Current Assets", "", "", ""],
    ["Inventories & Work in Progress", "13", "65,38,400.00", "46,28,900.00"],
    ["Trade Receivables", "14", "36,59,03,669.00", "1,21,02,833.00"],
    ["Cash and Cash Equivalents", "15", "49,55,61,553.82", "5,26,47,684.58"],
    ["Short Term Loans & Advances", "16", "41,99,80,185.98", "26,02,16,873.58"],
    ["Other Current Assets", "17", "20,666.70", "22,963.00"],
    ["", "", "", ""],
    ["TOTAL", "", "1,58,73,55,063.16", "49,65,89,823.65"]
]

# Profit and Loss data
profit_loss_data = [
    ["", "", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", ""],
    ["Statement of Profit and Loss for the year ended March 31, 2022", "", "", ""],
    ["(Amt in INR)", "", "", ""],
    ["", "", "", ""],
    ["Particulars", "Note No", "For the year Ended 31.03.2022", "For the year Ended 31.03.2021"],
    ["", "", "", ""],
    ["INCOME", "", "", ""],
    ["Revenue From operation", "18", "40,92,51,547.00", "38,13,90,243.00"],
    ["Other Income", "19", "1,56,683.92", "2,10,146.93"],
    ["TOTAL", "", "40,94,08,230.92", "38,16,00,389.93"],
    ["", "", "", ""],
    ["EXPENSES", "", "", ""],
    ["Purchase of Traded Goods", "20", "12,95,02,435.98", "6,43,65,699.38"],
    ["Changes in inventories of finished goods and Stock-in-Trade", "21", "(19,09,500.00)", "15,86,840.00"],
    ["Employee benefit expense", "22", "10,73,61,047.69", "11,97,28,974.87"],
    ["Financial costs", "23", "3,29,32,572.49", "6,34,530.65"],
    ["Depreciation and amortization expense", "10", "65,28,330.45", "61,99,000.00"],
    ["Other expenses", "24", "11,61,36,319.31", "15,60,08,028.91"],
    ["TOTAL", "", "39,05,51,205.92", "34,85,23,073.81"],
    ["", "", "", ""],
    ["Profit before exceptional and extraordinary items and tax", "", "1,88,57,025.00", "3,30,77,316.12"],
    ["Exceptional Items", "", "-", "-"],
    ["Extraordinary Items", "", "-", "-"],
    ["Profit before tax", "", "1,88,57,025.00", "3,30,77,316.12"],
    ["", "", "", ""],
    ["Tax expense:", "", "", ""],
    ["Current tax", "", "(42,21,897.28)", "(93,89,122.00)"],
    ["Deferred tax", "", "(6,95,631.28)", "1,89,480.97"],
    ["MAT Credit", "", "-", "-"],
    ["Prior Period Income", "", "-", "-"],
    ["Profit/(Loss) for the period", "", "1,39,39,496.44", "2,38,77,675.09"],
    ["", "", "", ""],
    ["Earning per Equity Share", "", "", ""],
    ["(1) Basic", "", "17.42", "29.85"],
    ["(2) Diluted", "", "17.42", "29.85"]
]

# Cash Flow data based on the image
cash_flow_data = [
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", "", ""],
    ["Cash Flow Statement For The Year Ended March 31, 2022", "", "", "", ""],
    ["", "", "", "", ""],
    ["", "For the Year ended 31st March, 2022", "", "For the Year ended 31st March, 2021", ""],
    ["", "", "", "", ""],
    ["I) Cash Flows from Operating Activities", "", "", "", ""],
    ["", "", "", "", ""],
    ["Net profit before tax,and extraordinary item", "1,88,57,025.00", "", "3,30,77,316.12", ""],
    ["Adjustments For Non cash & Non Operating Activities", "", "", "", ""],
    ["Loss on sale of fixed assest", "-", "", "-", ""],
    ["Depreciation", "65,28,330.45", "", "61,99,000.00", ""],
    ["Rental Income", "-", "", "-", ""],
    ["Interest Expenses", "3,29,32,572.49", "", "6,34,530.65", ""],
    ["Prior Period Adjustments", "-", "", "23,936.00", ""],
    ["MAT Credit", "-", "5,83,17,927.94", "-", "3,99,34,782.77"],
    ["", "", "5,83,17,927.94", "", "3,99,34,782.77"],
    ["Operating Profit Before Working Capital Changes", "", "", "", ""],
    ["Change in Trade Receivables", "(35,38,00,836.00)", "", "83,21,167.00", ""],
    ["Change in Inventory", "(19,09,500.00)", "", "15,86,840.00", ""],
    ["Change in Short Term loans & Advances", "(15,97,63,312.40)", "", "(19,15,19,764.80)", ""],
    ["Change in Long Term loans & Advances", "9,40,000.00", "", "(10,44,000.00)", ""],
    ["Change in Other Current Assets", "2,296.30", "", "2,552.00", ""],
    ["Change in Capital Work in Progress", "(8,41,56,237.68)", "", "(11,31,08,660.00)", ""],
    ["Change in Short Term Borrowings", "-", "", "-", ""],
    ["Change in Trade Payables", "38,33,679.04", "", "4,89,16,426.49", ""],
    ["Change in Other Current Liabilities", "50,13,63,909.00", "", "(8,38,55,453.20)", ""],
    ["Change in Other Long Term Liabilities", "16,09,867.00", "(9,18,80,134.74)", "26,24,979.00", "(32,80,75,913.51)"],
    ["Cash Generated from Operations", "", "(3,35,62,206.80)", "", "(28,81,41,130.74)"],
    ["Cash Flow from Operating Activities before Taxes", "", "(3,35,62,206.80)", "", "(28,81,41,130.74)"],
    ["Income Tax paid", "", "(93,89,122.00)", "", "(87,67,909.00)"],
    ["Net Cash From Operating Activity (A)", "", "(4,29,51,328.80)", "", "(29,69,09,039.74)"],
    ["", "", "", "", ""],
    ["II) Cash Flows from Investing Activities", "", "", "", ""],
    ["Purchase of Fixed Assets", "(5,59,16,842.48)", "", "(3,17,65,011.47)", ""],
    ["Decrease in WIP", "-", "", "-", ""],
    ["Sale Proceeds from Fixed Assets", "-", "", "-", ""],
    ["Rental Income", "-", "", "-", ""],
    ["Net Cash From Investing Activity (B)", "", "(5,59,16,842.48)", "", "(3,17,65,011.47)"],
    ["", "", "", "", ""],
    ["III) Cash Flows From Financing Activities", "", "", "", ""],
    ["Increase / (decrease) in Long Term Borrowings", "57,47,14,613.00", "", "21,42,92,192.00", ""],
    ["Interest paid", "(3,29,32,572.49)", "54,17,82,040.51", "(6,34,530.65)", "21,36,57,661.35"],
    ["Net Cash From Financing Activity (C)", "", "54,17,82,040.51", "", "21,36,57,661.35"],
    ["", "", "", "", ""],
    ["Net Increase in Cash And Cash equivalents(A+B+C)", "", "44,29,13,869.23", "", "(11,50,16,389.86)"],
    ["Cash & Cash Equivalents at the beginning of the year", "", "5,26,47,683.94", "", "16,76,64,073.80"],
    ["Cash & Cash Equivalents at the end of the year", "", "49,55,61,553.17", "", "5,26,47,683.94"]
]

# Function to add data to the sheet and format it
def populate_cash_flow_sheet(sheet, data):
    # First pass: Set all cell values and merge cells
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Merge title cells
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    
    # Merge the column header cells
    sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
    sheet.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)
    
    # Second pass: Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Apply borders to all cells
            cell.border = border
            
            # Title cell formatting
            if row_idx == 1:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            elif row_idx == 2:
                cell.font = subheader_font
                cell.alignment = Alignment(horizontal='center')
            
            # Header row formatting
            elif row_idx == 4:
                cell.font = subheader_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Format section headers
            elif col_idx == 1 and (row_idx == 6 or row_idx == 33 or row_idx == 40):
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Format the "Operating Profit Before Working Capital Changes" header
            elif col_idx == 1 and row_idx == 17:
                cell.font = subheader_font
            
            # Format numeric cells with right alignment
            elif col_idx in [2, 3, 4, 5] and value and isinstance(value, str) and (value.startswith('(') or value[0].isdigit()):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total rows
            elif row_idx in [15, 16, 27, 28, 29, 30, 31, 36, 38, 39, 43, 44, 45]:
                if col_idx in [3, 5] and value:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
                    cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 45  # For descriptions
    sheet.column_dimensions['B'].width = 20  # For values
    sheet.column_dimensions['C'].width = 20  # For subtotals
    sheet.column_dimensions['D'].width = 20  # For previous year values
    sheet.column_dimensions['E'].width = 20  # For previous year subtotals
    
    # Apply light blue background to header rows
    for row_idx in [4]:
        for col_idx in range(1, 6):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = light_blue_fill

# Populate the sheet
populate_cash_flow_sheet(cash_flow, cash_flow_data)

# Note 2 - Share Capital data based on the image
note2_data = [
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", ""],
    ["NOTES ANNEXED TO AND FORMING THE PART OF BALANCE SHEET FOR THE YEAR ENDED 31ST MARCH 2022", "", ""],
    ["", "", ""],
    ["", "( Amt in INR)", ""],
    ["", "As on 31st March 2022", "As on 31st March 2021"],
    ["", "", ""],
    ["NOTE 2 : SHARE CAPITAL", "", ""],
    ["Authorised Capital", "", ""],
    ["10,000 Equity Shares of Rs. 10/- each", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["Issued, Subscribed & Paid up", "", ""],
    ["10,000 Equity Shares of 10/- each fully Paid up", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["TOTAL", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["The Company has only one class of shares referred to as equity shares having a par value of Rs. 10/- each.", "", ""],
    ["Each holder of Equity share is entitled to one vote per share. The company declares & pay dividend", "", ""],
    ["when the same is approved by the shareholders in the ensuring Annual General Meeting.", "", ""],
    ["", "", ""],
    ["Reconciliation of the number of shares outstanding as on the date of Balance Sheet:", "", ""],
    ["", "", ""],
    ["Particulars", "Equity Shares", ""],
    ["", "Number", "Rs."],
    ["Shares outstanding at the beginning of the year", "1,00,000.00", "1,00,000.00"],
    ["Add/Less: Shares issued/redeem during the year", "-", "-"],
    ["Shares outstanding at the end of the year", "1,00,000.00", "1,00,000.00"],
    ["", "", ""],
    ["Details of Shareholder holding more than 5% share as on the date of Balance Sheet :", "", ""],
    ["", "", ""],
    ["Name of Shareholders", "As on 31-03-2022", "As on 31-03-2021"],
    ["", "%", "%"],
    ["Nandhra Eng. & Construction(India) Pvt. Ltd.", "51.99", "51.99"],
    ["Santosh Ankush Mohite", "24.00", "24.00"],
    ["Kishor Ankush Mohite", "24.00", "24.00"],
    ["Rahat Virk", "0.01", "0.01"],
    ["", "", ""],
    ["", "100.00", "100.00"]
]

# Function to populate Note 2 - Share Capital sheet
note2_sheet = wb['Note 2']
def populate_note2_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Merge title cells
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
    
    # Merge the Equity Shares header
    sheet.merge_cells(start_row=22, start_column=2, end_row=22, end_column=3)
    
    # Second pass: Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Company title formatting
            if row_idx == 1:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Balance sheet title formatting
            elif row_idx == 2:
                cell.font = subheader_font
                cell.alignment = Alignment(horizontal='center')
            
            # Header rows formatting
            elif row_idx == 5:
                cell.font = subheader_font
                cell.fill = header_fill
                cell.border = all_border
                cell.alignment = Alignment(horizontal='center')
            
            # Note title formatting
            elif row_idx == 7:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Subheader formatting
            elif row_idx in [8, 11]:
                cell.font = Font(bold=True)
            
            # Format numeric cells with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value == "-"):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row
            elif row_idx == 14 and col_idx == 1:
                cell.font = Font(bold=True)
            elif row_idx == 14 and col_idx in [2, 3]:
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
            
            # Double underline above total row
            elif row_idx == 13 and col_idx in [2, 3]:
                cell.border = Border(bottom=Side(border_style="double"))
            
            # Table headers
            elif row_idx in [22, 23, 30, 31]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = all_border
                cell.alignment = Alignment(horizontal='center')
            
            # Table data cells
            elif row_idx in [24, 25, 26, 32, 33, 34, 35, 37]:
                cell.border = all_border
                if col_idx in [2, 3]:
                    cell.alignment = Alignment(horizontal='center')
    
    # Add double underline for totals
    for col_idx in [2, 3]:
        for row_idx in [8, 11]:
            cell = sheet.cell(row=row_idx+1, column=col_idx)
            cell.border = Border(bottom=Side(border_style="double"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    
    # Add special double underlines below totals
    cell = sheet.cell(row=14, column=2)
    cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
    
    cell = sheet.cell(row=14, column=3)
    cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))

# Populate the sheet
populate_note2_sheet(note2_sheet, note2_data)

# Note 3 data - Reserves and Surplus
note3_data = [
    ["NOTE 3 : RESERVES & SURPLUS", "", ""],
    ["Surplus", "", ""],
    ["Opening balance", "4,64,63,916.96", "2,25,62,305.87"],
    ["(+) Net Profit/(Net Loss) For the current year", "1,39,39,496.44", "2,38,77,675.09"],
    ["(+) Adjustment for earlier years", "", "23,936.00"],
    ["", "", ""],
    ["Total", "6,04,03,413.40", "4,64,63,916.96"]
]

# Function to populate Note 3 - Reserves and Surplus sheet
def populate_note3_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Subheader formatting
            elif row_idx == 2 and col_idx == 1:
                cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value == "-"):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row
            elif row_idx == 7 and col_idx == 1:
                cell.font = Font(bold=True)
            elif row_idx == 7 and col_idx in [2, 3]:
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
                # Add double border
                cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
    
    # Add double underline before total row
    for col_idx in [2, 3]:
        cell = sheet.cell(row=6, column=col_idx)
        cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    
    # Add signature image (placeholder)
    # Note: In a real scenario, you would insert an actual image
    signature_row = 12
    signature_cell = sheet.cell(row=signature_row, column=2)
    signature_cell.value = "(Signature placeholder)"
    signature_cell.font = Font(italic=True, color="808080")
    signature_cell.alignment = Alignment(horizontal='center')

# Get the Note 3 sheet and populate it
note3_sheet = wb['Note 3']
populate_note3_sheet(note3_sheet, note3_data)

# Note 4 data - Long Term Borrowings (updated to match image)
note4_data = [
    ["NOTE 4 : LONG TERM BORROWINGS", "", ""],
    ["Secured", "", ""],
    ["Term Loans", "", ""],
    [" - From Banks", "12,63,341.00", "16,75,629.00"],
    [" - Union Bank of India", "34,35,10,078.00", ""],
    [" - Axis Bank Ltd.", "22,81,770.00", ""],
    ["Less : Repayable in next 1 year", "", "4,12,288.00"],
    ["", "", ""],
    ["Un-Secured", "", ""],
    ["Loans from Directors & Others", "46,82,44,739.08", "23,93,21,974.08"],
    ["(Interest Bearing, Payable after 1 year)", "", ""],
    ["", "", ""],
    ["Total", "81,52,99,928.08", "24,05,85,315.08"]
]

# Function to populate Note 4 - Long Term Borrowings sheet
def populate_note4_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Section headers formatting
            elif (row_idx == 2 or row_idx == 9) and col_idx == 1:
                cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value == "-"):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row
            elif row_idx == 13 and col_idx == 1:
                cell.font = Font(bold=True)
            elif row_idx == 13 and col_idx in [2, 3]:
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
                # Add double border
                cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
    
    # Add underline before total row
    for col_idx in [2, 3]:
        cell = sheet.cell(row=12, column=col_idx)
        cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

# Replace the existing Note 4 population code with this updated function
note4_sheet = wb['Note 4']
populate_note4_sheet(note4_sheet, note4_data)

# Note 5 data - Other Long Term Liabilities
note5_data = [
    ["NOTE 5 : Other Long Term Liabilities", "", ""],
    ["Security Deposits- Labour", "42,34,846.00", "26,24,979.00"],
    ["", "", ""],
    ["Total", "42,34,846.00", "26,24,979.00"]
]

# Note 6 data - Short Term Borrowings
note6_data = [
    ["NOTE 6 : SHORT TERM BORROWINGS", "", ""],
    ["Loans repayable on demand", "", ""],
    ["From Banks", "-", "-"],
    ["", "", ""],
    ["TOTAL", "-", "-"]
]

# Note 7 data - Trade Payables
note7_data = [
    ["NOTE 7 : TRADE PAYABLES", "", ""],
    ["For Expenses", "3,06,87,283.18", "3,84,69,885.00"],
    ["For Goods Purchase", "2,72,97,063.40", "1,40,22,539.54"],
    ["For Subcontractor", "3,14,50,718.07", "3,31,08,961.07"],
    ["", "", ""],
    ["TOTAL", "8,94,35,064.65", "8,56,01,385.61"]
]

# Note 8 data - Other Current Liabilities
note8_data = [
    ["NOTE 8 : OTHER CURRENT LIABILITIES", "", ""],
    ["Current maturities of Long term Debt", "", "4,12,288.00"],
    ["Others", "", ""],
    [" - Advances From Customers", "55,63,00,000.00", "-"],
    [" - Cheque Issue but not Clear", "-", "-"],
    [" - Expenses Payables", "3,50,000.00", "8,90,000.00"],
    [" - Statutory Liabilities", "30,00,604.00", "1,08,50,046.00"],
    [" - Payable to employees", "62,88,410.00", "51,72,771.00"],
    [" - Other Payable", "4,72,50,000.00", "9,45,00,000.00"],
    ["", "", ""],
    ["TOTAL", "61,31,89,014.00", "11,18,25,105.00"]
]

# Note 9 data - Short Term Provisions
note9_data = [
    ["NOTE 9 : SHORT TERM PROVISIONS", "", ""],
    ["Provision for Income Tax", "42,21,897.28", "93,89,122.00"],
    ["", "", ""],
    ["TOTAL", "42,21,897.28", "93,89,122.00"]
]

# Generic function to populate note sheets with consistent formatting
def populate_note_sheet_generic(sheet, data, title_in_caps=False):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Subheader formatting
            elif row_idx > 1 and col_idx == 1 and not value.startswith(" ") and value not in ["", "TOTAL", "Total"]:
                cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row - find the row with TOTAL or Total
            if isinstance(value, str) and (value == "TOTAL" or value == "Total"):
                cell.font = Font(bold=True)
                
                # Format total amount cells
                total_row = row_idx
                for total_col in [2, 3]:
                    total_cell = sheet.cell(row=total_row, column=total_col)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                
                # Add underline before total row
                for underline_col in [2, 3]:
                    underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                    underline_cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

# Populate Notes 5-9
note5_sheet = wb['Note 5']
note6_sheet = wb['Note 6']
note7_sheet = wb['Note 7']
note8_sheet = wb['Note 8']
note9_sheet = wb['Note 9']

populate_note_sheet_generic(note5_sheet, note5_data)
populate_note_sheet_generic(note6_sheet, note6_data)
populate_note_sheet_generic(note7_sheet, note7_data)
populate_note_sheet_generic(note8_sheet, note8_data)
populate_note_sheet_generic(note9_sheet, note9_data)

# Note 11 data - Long Term Loans and Advances
note11_data = [
    ["NOTE 11 : LONG TERM LOANS AND ADVANCES", "", ""],
    ["UnSecured Considered Good", "", ""],
    ["Security Deposits", "38,66,100.00", "48,06,100.00"],
    ["", "", ""],
    ["TOTAL", "38,66,100.00", "48,06,100.00"]
]

# Note 12 data - Deferred Tax Asset
note12_data = [
    ["NOTE 12 : DEFERRED TAX ASSET", "", ""],
    ["Deferred Tax Asset", "-", "2,24,731.53"],
    ["", "", ""],
    ["TOTAL", "-", "2,24,731.53"]
]

# Note 13 data - Inventories
note13_data = [
    ["NOTE 13 : INVENTORIES", "", ""],
    ["General Consumable & Stock", "65,38,400.00", "46,28,900.00"],
    ["", "", ""],
    ["TOTAL", "65,38,400.00", "46,28,900.00"]
]

# Note 14 data - Trade Receivables
note14_data = [
    ["NOTE 14 : TRADE RECEIVABLES", "", ""],
    ["Trade Receivables", "", ""],
    ["-UnSecured Considered Good", "", ""],
    ["Executive Engineer, PWD, Khamgaon", "", ""],
    ["-Less than Six Months from Due Date", "36,59,03,669.00", "1,21,02,833.00"],
    ["-Others", "-", "-"],
    ["", "", ""],
    ["TOTAL", "36,59,03,669.00", "1,21,02,833.00"]
]

# Note 15 data - Cash & Cash Equivalents
note15_data = [
    ["NOTE 15 : CASH & CASH EQUIVALENTS", "", ""],
    ["Balances with Banks :", "", ""],
    ["In Current and Escrow Accounts", "49,55,44,363.81", "5,25,21,709.58"],
    ["Cheque received but not clear", "-", "-"],
    ["Cash in hand & Imprest Balances", "17,190.01", "1,25,975.00"],
    ["", "", ""],
    ["TOTAL", "49,55,61,553.82", "5,26,47,684.58"]
]

# Enhanced function to populate note sheets with style variations
def populate_note_sheet_enhanced(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Special formatting for italic text like "UnSecured Considered Good"
            elif row_idx == 2 and col_idx == 1 and "UnSecured Considered Good" in value:
                cell.font = Font(italic=True)
            
            # Section headers like "Balances with Banks :"
            elif col_idx == 1 and (":") in value and row_idx > 1:
                cell.font = Font(bold=True, underline="single")
            
            # Bold but not section headers
            elif row_idx > 1 and col_idx == 1 and not value.startswith("-") and value and value != "TOTAL":
                if "UnSecured" not in value:  # Skip items that should be in italic
                    cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row
            if isinstance(value, str) and value == "TOTAL":
                cell.font = Font(bold=True)
                
                # Get the total row
                total_row = row_idx
                for total_col in [2, 3]:
                    total_cell = sheet.cell(row=total_row, column=total_col)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                
                # Add underline before total row
                for underline_col in [2, 3]:
                    underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                    underline_cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

# Populate Notes 11-15
note11_sheet = wb['Note 11']
note12_sheet = wb['Note 12']
note13_sheet = wb['Note 13']
note14_sheet = wb['Note 14']
note15_sheet = wb['Note 15']

populate_note_sheet_enhanced(note11_sheet, note11_data)
populate_note_sheet_enhanced(note12_sheet, note12_data)
populate_note_sheet_enhanced(note13_sheet, note13_data)
populate_note_sheet_enhanced(note14_sheet, note14_data)
populate_note_sheet_enhanced(note15_sheet, note15_data)

# Note 16 data - Short Term Loans & Advances
note16_data = [
    ["NOTE 16 : SHORT TERM LOANS & ADVANCES", "", ""],
    ["Others", "", ""],
    ["Unsecured Considered Good", "", ""],
    ["  - Advances to vendors", "39,35,29,989.28", "24,10,12,701.00"],
    ["  - Loans & advances to Employees", "1,43,46,262.40", "67,39,109.40"],
    ["  - Prepaid Expenses", "14,83,235.00", "16,22,683.00"],
    ["  - GST & Service Tax", "49,25,850.55", "84,62,059.26"],
    ["  - Income Tax & TDS", "56,94,848.75", "23,80,320.92"],
    ["", "", ""],
    ["TOTAL", "41,99,80,185.98", "26,02,16,873.58"]
]

# Note 17 data - Other Current Assets
note17_data = [
    ["NOTE 17 : OTHER CURRENT ASSETS", "", ""],
    ["Others", "", ""],
    ["  - Preoperating Expenses", "20,666.70", "22,963.00"],
    ["  - MAT Credit Entitlement", "-", "-"],
    ["", "", ""],
    ["TOTAL", "20,666.70", "22,963.00"]
]

# Function to populate Notes 16 & 17 with proper formatting
def populate_notes_16_17(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Section headers - "Others"
            elif row_idx == 2 and col_idx == 1:
                cell.font = Font(bold=True)
            
            # Italicized "Unsecured Considered Good"
            elif row_idx == 3 and col_idx == 1 and "Unsecured Considered Good" in value:
                cell.font = Font(italic=True)
            
            # Format indented items
            elif col_idx == 1 and value.startswith("  - "):
                # No special formatting for indented items, just keep the indentation
                pass
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row
            if isinstance(value, str) and value == "TOTAL":
                cell.font = Font(bold=True)
                
                # Get the total row
                total_row = row_idx
                for total_col in [2, 3]:
                    total_cell = sheet.cell(row=total_row, column=total_col)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                
                # Add underline before total row
                for underline_col in [2, 3]:
                    underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                    underline_cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

# Populate Notes 16-17
note16_sheet = wb['Note 16']
note17_sheet = wb['Note 17']

populate_notes_16_17(note16_sheet, note16_data)
populate_notes_16_17(note17_sheet, note17_data)

# Note 18 data - Revenue from Operations
note18_data = [
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", ""],
    ["NOTES ANNEXED TO AND FORMING THE PART OF STATEMENT OF PROFIT & LOSS FOR THE YEAR ENDED", "", ""],
    ["31ST MARCH 2022", "( Amt in INR)", ""],
    ["", "For the year Ended 31.03.2022", "For the year Ended 31.03.2021"],
    ["", "", ""],
    ["NOTE 18 : REVENUE FROM OPERATIONS", "", ""],
    ["Sale of products:", "", ""],
    ["  Works Contract PWD", "40,92,51,547.00", "38,13,90,243.00"],
    ["", "", ""],
    ["", "40,92,51,547.00", "38,13,90,243.00"],
    ["Less: Excise Duty", "-", "-"],
    ["", "", ""],
    ["Total", "40,92,51,547.00", "38,13,90,243.00"]
]

# Note 19 data - Other Income
note19_data = [
    ["NOTE 19 : OTHER INCOME", "", ""],
    ["Discount Received", "1,56,683.92", "2,10,146.93"],
    ["Round Off", "-", "-"],
    ["", "", ""],
    ["Total", "1,56,683.92", "2,10,146.93"]
]

# Note 20 data - Purchase of Stock in Trade
note20_data = [
    ["NOTE 20 : PURCHASE OF STOCK IN TRADE", "", ""],
    ["Electrical and Hardware Material Purchase", "2,28,671.65", "18,17,492.26"],
    ["Safety Material Purchase", "25,78,054.38", "4,61,637.69"],
    ["Other Material Purchase", "6,53,55,646.20", "4,15,14,939.81"],
    ["RCC Pipe and Material", "16,83,200.27", "20,68,300.00"],
    ["Steel Purchase", "4,06,51,236.71", "2,77,32,575.40"],
    ["Cement Purchase", "1,09,30,551.79", "2,15,75,050.00"],
    ["Diesel Purchase", "9,22,31,312.66", "8,23,04,364.22"],
    ["", "", ""],
    ["", "21,36,58,673.66", "17,74,74,359.38"],
    ["", "", ""],
    ["Add:- Material transferred to work in progress for previous period", "13,11,65,800.00", "1,80,57,140.00"],
    ["", "", ""],
    ["Less:- Material transferred to work in progress", "21,53,22,037.68", "13,11,65,800.00"],
    ["", "", ""],
    ["Total", "12,95,02,435.98", "6,43,65,699.38"]
]

# Function to populate Notes 18-20 with proper formatting, including header
def populate_notes_18_to_20(sheet, data, is_header_note=True):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply special formatting for header (only for Note 18 or if specifically requested)
    if is_header_note:
        # Merge company name across columns
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        cell = sheet.cell(row=1, column=1)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        # Merge statement title across columns
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        cell = sheet.cell(row=2, column=1)
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center')
        
        # Format column headers (row 4)
        for col_idx in range(1, 4):
            cell = sheet.cell(row=4, column=col_idx)
            cell.font = subheader_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Set special formatting for amt in INR text
        cell = sheet.cell(row=3, column=2)
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='right')
        
        # Starting row for the note title depends on whether we have a header
        note_title_row = 6
    else:
        # If no header, the note title is in the first row
        note_title_row = 1
    
    # Apply formatting for all cells
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            # Skip already formatted header cells
            if is_header_note and row_idx <= 4:
                continue
                
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == note_title_row and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value == "-"):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row
            if isinstance(value, str) and value == "Total":
                cell.font = Font(bold=True)
                
                # Format total amount cells
                total_row = row_idx
                for total_col in [2, 3]:
                    total_cell = sheet.cell(row=total_row, column=total_col)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                
                # Add underline before total row
                for underline_col in [2, 3]:
                    underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                    underline_cell.border = Border(bottom=Side(border_style="thin"))
    
    # Format subtotal rows in Note 20
    if "PURCHASE OF STOCK IN TRADE" in data[0][0] or (is_header_note and "PURCHASE OF STOCK IN TRADE" in data[5][0]):
        # Find the subtotal row (row with first substantial sum)
        subtotal_row = 0
        for row_idx, row in enumerate(data, 1):
            if len(row) > 1 and isinstance(row[1], str) and "21,36,58,673.66" in row[1]:
                subtotal_row = row_idx
                break
        
        if subtotal_row > 0:
            for col_idx in [2, 3]:
                subtotal_cell = sheet.cell(row=subtotal_row, column=col_idx)
                subtotal_cell.font = Font(bold=True)
                subtotal_cell.border = Border(bottom=Side(border_style="thin"))
                subtotal_cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25

# Populate Notes 18-20
note18_sheet = wb['Note 18']
note19_sheet = wb['Note 19']
note20_sheet = wb['Note 20']

# Apply formatting with headers only for Note 18
populate_notes_18_to_20(note18_sheet, note18_data, True)

# For Note 19 and 20, we don't repeat the header
note19_data_adjusted = note19_data  # No header needed
note20_data_adjusted = note20_data  # No header needed

populate_notes_18_to_20(note19_sheet, note19_data_adjusted, False)
populate_notes_18_to_20(note20_sheet, note20_data_adjusted, False)

# Note 21 data - Change in Inventories
note21_data = [
    ["NOTE 21 : CHANGE IN INVENTORIES", "", ""],
    ["CLOSING STOCK OF", "", ""],
    ["General Consumable", "65,38,400.00", "46,28,900.00"],
    ["", "", ""],
    ["Total", "65,38,400.00", "46,28,900.00"],
    ["", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["Less : OPENING STOCK OF", "", ""],
    ["General Consumable", "46,28,900.00", "62,15,740.00"],
    ["", "", ""],
    ["Total", "46,28,900.00", "62,15,740.00"],
    ["", "", ""],
    ["Change in Inventories", "(19,09,500.00)", "15,86,840.00"],
    ["", "", ""],
    ["Change in Inventories", "(19,09,500.00)", "15,86,840.00"]
]

# Note 22 data - Employee Benefit Expenses
note22_data = [
    ["NOTE 22 : EMPLOYEE BENEFIT EXPENSES", "", ""],
    ["Salaries, Wages & other Allowances", "10,18,33,392.73", "11,43,61,744.90"],
    ["Contribution to PF, ESI & Labour Welfare Fund", "8,03,420.00", "11,02,937.00"],
    ["Staff & Labour Welfare", "47,24,234.96", "42,64,292.97"],
    ["", "", ""],
    ["Total", "10,73,61,047.69", "11,97,28,974.87"]
]

# Note 23 data - Financial Expenses
note23_data = [
    ["NOTE 23 : FINANCIAL EXPENSES", "", ""],
    ["Interest Expenses", "", ""],
    ["Interest to Banks", "", ""],
    ["- Interest on Car Loan", "1,40,792.00", "60,489.00"],
    ["- Interest on TDS", "18,01,129.00", "4,27,504.00"],
    ["- Interest on IT", "15,74,929.00", "64,258.00"],
    ["- Interest on Term Loan", "2,93,20,942.00", ""],
    ["Bank Charges", "94,780.49", "82,279.65"],
    ["Bank Guarantee Expenses", "", ""],
    ["", "", ""],
    ["Total", "3,29,32,572.49", "6,34,530.65"]
]

# Function to populate Notes 21-23 with proper formatting
def populate_notes_21_to_23(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Subheader formatting (all caps items like "CLOSING STOCK OF")
            elif col_idx == 1 and value and value.isupper() and row_idx > 1:
                cell.font = Font(bold=True)
            
            # Second level headers (like "Interest Expenses", "Interest to Banks")
            elif col_idx == 1 and value and not value.startswith("-") and not "Total" in value and not "Change" in value and row_idx > 1:
                if not value.isupper():  # Skip if already handled as uppercase header
                    cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str):
                cell.alignment = Alignment(horizontal='right')
            
            # Format "Total" rows
            if isinstance(value, str) and value == "Total":
                cell.font = Font(bold=True)
                
                # Format total amount cells
                total_row = row_idx
                for total_col in [2, 3]:
                    total_cell = sheet.cell(row=total_row, column=total_col)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                
                # Add underline before total row
                for underline_col in [2, 3]:
                    underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                    underline_cell.border = Border(bottom=Side(border_style="thin"))
            
            # Special formatting for "Change in Inventories" lines in Note 21
            if "Change in Inventories" in str(value):
                cell.font = Font(bold=True)
                
                if col_idx in [2, 3]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='right')
                    if row_idx == 21:  # The final Change in Inventories row
                        cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
    
    # Add horizontal line after Note 21 first total row to match the image
    if "NOTE 21" in data[0][0]:
        for col_idx in range(1, 4):
            cell = sheet.cell(row=11, column=col_idx)
            cell.border = Border(bottom=Side(border_style="medium"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25

# Populate Notes 21-23
note21_sheet = wb['Note 21']
note22_sheet = wb['Note 22']
note23_sheet = wb['Note 23']

populate_notes_21_to_23(note21_sheet, note21_data)
populate_notes_21_to_23(note22_sheet, note22_data)
populate_notes_21_to_23(note23_sheet, note23_data)

# Note 24 data - Other Expenses (combines all three images)
note24_data = [
    ["NOTE 24 : OTHER EXPENSES", "", ""],
    ["INDIRECT EXPENDITURE", "", ""],
    ["(A) ADMINISTRATIVE EXPENSES", "", ""],
    ["Audit Fees", "3,50,000.00", "3,50,000.00"],
    ["Cgst 2.5% (Rcm) (Itc Ineligible)", "", ""],
    ["Conveyance Charges", "", "-"],
    ["Crop Compensation Paid", "11,32,821.00", "9,08,185.00"],
    ["Crusher Lease", "", "-"],
    ["Director'S Salary", "40,00,000.00", "2,06,40,000.00"],
    ["Directors Accomadation Expenses", "3,21,590.00", "3,67,412.00"],
    ["Documanation Charges", "", "-"],
    ["Drinking Water Charges", "3,71,763.00", "4,08,858.00"],
    ["Electricity Charges", "6,75,738.00", "54,820.00"],
    ["Fees  Paid", "2,60,525.00", "4,16,630.00"],
    ["Govt. Tax Paid", "65,875.00", "2,49,720.00"],
    ["Gst Late Filling Fee", "3,050.00", "29,380.00"],
    ["Hotel And Loading Charges", "6,44,645.21", "11,07,341.63"],
    ["Insurance Paid", "4,83,321.00", "28,35,605.00"],
    ["Interest On Gst", "-", ""],
    ["Interest On Mobilisation Advance", "45,31,833.00", "44,77,189.00"],
    ["Internet Charges", "24,387.00", "37,444.00"],
    ["Labour Charges Expenses", "", ""],
    ["Labour Welfare", "", "-"],
    ["Land Rent", "21,81,309.00", "51,27,182.00"],
    ["Land Rent (Murum)", "24,85,000.00", "10,35,000.00"],
    ["Loading & Unloading Exp.", "845.00", "8,513.00"],
    ["Legal Charges", "3,850.00", ""],
    ["Loan Processing Fees", "47,25,605.00", ""],
    ["Machinery Hire Charges", "5,65,55,229.11", "6,84,33,260.97"],
    ["Mess Expenses", "", "-"],
    ["Mobile Expenses", "", "2,929.14"],
    ["N.A Tax Paid", "", "60,900.00"],
    ["Office Rent Paid", "64,000.00", "7,25,000.00"],
    ["Oil & Grease - 18%", "(5,75,464.29)", "8,37,883.71"],
    ["Pooja & Festival Exp.", "4,280.00", "28,325.00"],
    ["Postage & Courier", "41.00", "608.00"],
    ["Preliminary Expense Write/Off", "2,296.30", "2,552.00"],
    ["Printing And Stationery", "5,57,803.20", "4,92,805.66"],
    ["Professional And Technical Fees Paid", "17,65,795.00", "32,85,088.00"],
    ["Professional Charges", "12,00,000.00", ""],
    ["Rent Paid", "15,92,396.00", "14,57,959.00"],
    ["Repair & Maintenance", "17,73,216.18", "26,53,035.75"],
    ["Round Off", "(58.77)", "9,823.15"],
    ["Royalty Paid", "95,60,000.00", "72,06,400.00"],
    ["Security Charges", "12,48,216.00", "46,20,151.00"],
    ["Sensor Paper (Appolo ANP-550) Transportation", "", "85,000.00"],
    ["Sgst 2.5% (Rcm) (Itc Ineligible)", "", ""],
    ["Site Expenses", "9,53,350.77", "6,16,762.20"],
    ["Stamp Paper & Legal Exp.", "20,860.00", "65,800.00"],
    ["Sub Contract Work Charges", "", ""],
    ["Telephone/Internet Charges", "16,757.00", "93,891.00"],
    ["Testing Charges 18 %", "3,24,000.00", "20,500.00"],
    ["Tour And Trevelling Expenses", "5,10,632.00", "1,56,57,293.54"],
    ["Transportation Charges (URP-RCM)", "1,19,09,357.60", "56,28,538.16"],
    ["Transport Charges", "36,409.00", "4,12,096.00"],
    ["Tree Cutting Expenses", "", "-"],
    ["Transit Insurance Paid", "", "290.00"],
    ["TDS Late Fees", "3,27,600.00", ""],
    ["Tally Subscription Fees", "10,800.00", ""],
    ["Vehicle Expenses", "", "53.00"],
    ["Vehicle Hire Charges (Urp)", "48,02,666.00", "41,53,467.00"],
    ["Water Charges", "10,58,980.00", "14,04,390.00"],
    ["", "", ""],
    ["(B) SELLING & DISTRIBUTION EXPENSES", "", ""],
    ["Business Promotion Exp.", "1,55,000.00", "-"],
    ["", "", ""],
    ["", "11,61,36,319.31", "15,60,08,028.91"],
    ["", "", ""],
    ["(A+B)", "11,61,36,319.31", "15,60,08,028.91"]
]

# Create a separate sheet for the Details of Employee Benefit Expenses
employee_benefit_details_data = [
    ["Details of Employee Benefit Expenses", "", ""],
    ["Salaries, wages & Other Allowances", "", ""],
    ["Salaries", "1,52,47,020.00", "1,83,71,700.00"],
    ["Labour Charges", "8,03,14,909.73", "8,46,14,119.90"],
    ["Conveyance", "9,55,119.00", "14,93,858.00"],
    ["HRA", "28,22,154.00", "44,68,666.00"],
    ["Other Allowance", "11,07,111.00", "17,57,991.00"],
    ["Special Allowance", "13,87,079.00", "22,12,555.00"],
    ["Workers Welfare Cess 1%", "", "14,42,855.00"],
    ["", "", ""],
    ["", "10,18,33,392.73", "11,43,61,744.90"],
    ["", "", ""],
    ["Contribution to PF, ESI, Labour Welfare Fund", "", ""],
    ["PF Admn Charges", "50,829.00", "78,806.00"],
    ["PF contribution", "6,49,440.00", "9,54,023.00"],
    ["ESI Expenses", "61,816.00", "70,108.00"],
    ["Interest on PF/ Demages", "41,335.00", ""],
    ["", "", ""],
    ["", "8,03,420.00", "11,02,937.00"]
]

# Function to populate Note 24 with proper formatting
def populate_note24(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Note title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Section headers formatting
            elif col_idx == 1 and any(header in str(value) for header in ["INDIRECT EXPENDITURE", "(A) ADMINISTRATIVE EXPENSES", "(B) SELLING & DISTRIBUTION EXPENSES"]):
                cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value.startswith("-") or value.startswith("(")):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total rows
            if row_idx in [64, 66] and col_idx in [2, 3]:
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
                cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
            
            # Add underline before total rows
            if row_idx == 63 and col_idx in [2, 3]:
                cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25

# Function to populate Employee Benefit Details sheet
def populate_employee_benefit_details(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Main title and section headers
            if row_idx in [1, 2, 12] and col_idx == 1:
                cell.font = Font(bold=True, underline="single")
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value.startswith("-")):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total rows
            if row_idx in [10, 18] and col_idx in [2, 3]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right')
                cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
            
            # Add underline before total rows
            if row_idx in [9, 17] and col_idx in [2, 3]:
                cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25

# Populate Note 24
note24_sheet = wb['Note 24']
populate_note24(note24_sheet, note24_data)

# Create a new sheet for Employee Benefit Details
employee_details_sheet = wb.create_sheet("Employee Benefit Details")
populate_employee_benefit_details(employee_details_sheet, employee_benefit_details_data)

# Annexure to Note 4 data - Long Term Borrowings
annexure_to_note4_data = [
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", ""],
    ["ANNEXURE TO NOTES FORMING THE PART OF BALANCE SHEET FOR THE YEAR ENDED 31ST MARCH 2022", "", ""],
    ["", "( Amt in INR)", ""],
    ["", "As at 31st March 2022", "As at 31st March 2021"],
    ["", "", ""],
    ["ANNEXURE TO NOTE- 4", "", ""],
    ["LONG TERM BORROWINGS", "", ""],
    ["SECURED", "", ""],
    ["TERM LOANS", "", ""],
    ["Union Bank of India ( T.L.) 620306390000017", "34,35,10,078.00", ""],
    ["", "", ""],
    ["Total (A)", "34,35,10,078.00", "-"],
    ["", "", ""],
    ["VEHICLES LOANS", "", ""],
    ["Axis Bank Ltd. Loan A/c No.BPR004106898586", "22,81,770.00", ""],
    ["ICICI Bank (Tata Harrier) Loan A/C", "12,63,341.00", "16,75,629.00"],
    ["(Note :-vehicle loans availed from banks are secured by an exclusive charge on the", "", ""],
    ["vehicles financed by the said bank)", "", ""],
    ["", "", ""],
    ["Total (B)", "35,45,111.00", "16,75,629.00"],
    ["", "", ""],
    ["Grand Total (A+B)", "34,70,55,189.00", "16,75,629.00"],
    ["", "", ""],
    ["Unsecured Loans from Director / Shareholders and others", "", ""],
    ["Gurtej singh Nandhra", "-", "3,26,130.00"],
    ["Santosh Ankush Mohite", "10,28,60,232.00", "6,52,38,917.00"],
    ["Rahat virk", "2,000.00", "2,000.00"],
    ["Iqbal Singh", "28,46,63,973.08", "17,37,54,927.08"],
    ["Rajput Abhijeet Narayan Singh", "1,00,00,000.00", ""],
    ["Kishor Mohite ( U. Loan)", "7,35,00,000.00", ""],
    ["", "", ""],
    ["Total", "47,10,26,205.08", "23,93,21,974.08"],
    ["", "", ""],
    ["OTHER LONG TERM LIABILITIES", "", ""],
    ["SECURITY DEPOSITS - LABOUR", "", ""],
    ["Amol Ramesh Gawai (SD)", "24,750.00", "5,625.00"],
    ["Bhumi Constructions (SD)", "25,789.00", "25,789.00"],
    ["B V Parmar (S.D.)", "18,733.00", "18,733.00"],
    ["Chandraprakash Dipchand Tailor (SD)", "6,56,557.00", "1,22,745.00"],
    ["Dilip Chavan (SD)", "15,000.00", ""],
    ["Dnyaneshwar Tulashiram Pandhari (SD)", "6,950.00", "6,950.00"],
    ["Ganesh B Pandhare (SD)", "2,654.00", "2,654.00"],
    ["Itewar Rameshwar Ramdas (SD)", "1,512.00", "1,512.00"],
    ["Kawaljit Singh Chhabra ( SD)", "6,08,612.00", "3,71,713.00"],
    ["Khalil Kalandar Khan (SD)", "65,782.00", "59,267.00"],
    ["Matoshri Construction (SD)", "1,51,620.00", "93,301.00"],
    ["Nawal Ramdas Pandhare (SD)", "2,625.00", "2,625.00"],
    ["N.V. Parmar (SD)", "5,366.00", "5,366.00"],
    ["Om Sai Construction (SD)", "28,842.00", "16,744.00"],
    ["Paritosh S. Mandal (SD)", "19,042.00", "19,042.00"],
    ["Poonam Subhash Sawale (SD)", "42,009.00", "9,880.00"],
    ["Rajmudra Construction (SD)", "1,20,388.00", "75,334.00"],
    ["Sandip Suresh Sheogkar (SD)", "10,500.00", "10,500.00"],
    ["Sanjeev Shibu Sardar (SD)", "48,050.00", "48,050.00"],
    ["Shabbir Khan Gulsher Khan (SD)", "2,791.00", "2,791.00"],
    ["Shahbaz Khan Subhan Khan (SD)", "14,401.00", "14,401.00"],
    ["Shaikh Gani Shaikh Juneed (SD)", "21,485.00", "21,485.00"],
    ["Shripada Constructions (S.D.)", "22,98,639.00", "16,55,223.00"],
    ["Shuddhodan Laxman Wankhede (SD)", "7,500.00", ""],
    ["Ukharda Haribhau Dharamail (SD)", "1,520.00", "1,520.00"],
    ["Vaibhav Zingraji Parale (SD)", "33,729.00", "33,729.00"],
    ["", "", ""],
    ["Total", "42,34,846.00", "26,24,979.00"],
    ["TRADE PAYABLES - SUB CONTRACTOR", "", ""],
    ["Amol Ramesh Gawai (Lc)", "", "1,05,750.00"],
    ["Ashok Arjun Dhore (LC)", "28,808.00", "63,063.00"],
    ["Chandraprakash Dipchand Tailor", "", "18,04,488.00"],
    ["Creditor Contractor", "2,32,20,621.07", "2,32,20,621.07"],
    ["Department Labour Charges Payable", "59,68,030.00", "40,28,592.00"],
    ["Gajanan Bhanudas Sonalkar ( LC)", "75,718.00", "1,16,900.00"],
    ["Ganesh Narayan Karale (LC)", "", "1,91,862.00"],
    ["Ghugha Rahim Mamadbhai (LC)", "", "1,54,782.00"],
    ["Jitendra Kumar Singh(LC)", "7,56,813.00", ""],
    ["Kailash Suryabhan Wakode (LC)", "", "81,378.00"],
    ["Kawaljit Singh Chhabra (LC)", "", "11,03,027.00"],
    ["Matoshri Construction (LC)", "", "5,69,806.00"],
    ["Nawal Ramdas Pandhare (Lc)", "", "49,350.00"],
    ["Om Sai Construction (Lc)", "", "3,14,782.00"],
    ["Poonam Subhash Sawale (LC)", "", "1,85,748.00"],
    ["Rajmudra Constructions (LC)", "", "6,25,312.00"],
    ["Sanjay Punjaji Wankhede (Lc)", "", "19,350.00"],
    ["Vaibhav Zingaraji Parale (LC)", "52,667.00", ""],
    ["Vilas Motiram Shirsath ( LC)", "13,48,061.00", "4,74,150.00"],
    ["", "", ""],
    ["Total", "3,14,50,718.07", "3,31,08,961.07"],
    ["", "", ""],
]

# Annexure to Note 8 data - Other Current Liabilities
annexure_to_note8_data = [
    ["ANNEXURE TO NOTE-8", "", ""],
    ["OTHER CURRENT LIABILITIES", "", ""],
    ["EXPENSES PAYABLES", "", ""],
    ["Audit Fee Payable", "3,50,000.00", "3,50,000.00"],
    ["Neha Devi - Professinal Charges", "-", "5,40,000.00"],
    ["", "", ""],
    ["Total", "3,50,000.00", "8,90,000.00"],
    ["", "", ""],
    ["PAYABLE TO EMPLOYEES", "", ""],
    ["Salary Payable", "62,88,410.00", "51,72,771.00"],
    ["", "", ""],
    ["Total", "62,88,410.00", "51,72,771.00"],
    ["", "", ""],
    ["STATUTORY LIABILITIES", "", ""],
    ["PF Payable", "2,38,210.00", "2,95,528.00"],
    ["Admin Charges PF Payable", "8,477.00", "12,448.00"],
    ["ESI Payable", "18,526.00", "24,578.00"],
    ["Employee's Labour Welfare Fund", "2,976.00", "2,162.00"],
    ["Employee's Professional Tax", "1,00,200.00", "1,28,400.00"],
    ["TDS Payable", "26,32,215.00", "1,03,86,930.00"],
    ["", "", ""],
    ["Total", "30,00,604.00", "1,08,50,046.00"],
    ["", "", ""],
    ["OTHER PAYABLES", "", ""],
    ["Mobilization Advance", "4,72,50,000.00", "9,45,00,000.00"],
    ["", "", ""],
    ["", "4,72,50,000.00", "9,45,00,000.00"],
    ["", "", ""],
    ["Grand Total", "5,68,89,014.00", "11,14,12,817.00"]
]

# Function to populate the Annexure sheets with headers and proper formatting
def populate_annexure_sheet_with_header(sheet, data, has_header=True):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply header formatting if needed
    if has_header:
        # Merge company name across columns
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        cell = sheet.cell(row=1, column=1)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        # Merge statement title across columns
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        cell = sheet.cell(row=2, column=1)
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center')
        
        # Format column headers (row 4)
        for col_idx in [2, 3]:
            sheet.cell(row=4, column=col_idx).font = subheader_font
            sheet.cell(row=4, column=col_idx).fill = header_fill
            sheet.cell(row=4, column=col_idx).border = border
            sheet.cell(row=4, column=col_idx).alignment = Alignment(horizontal='center')
        
        # Set special formatting for amt in INR text
        cell = sheet.cell(row=3, column=2)
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='right')
        
        # Set the start row for formatting after the header
        format_start_row = 6
    else:
        format_start_row = 1
    
    # Apply formatting for the rest of the cells
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            # Skip already formatted header cells
            if has_header and row_idx <= 4:
                continue
                
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Format section headers
            if col_idx == 1 and value and value == value.upper() and row_idx >= format_start_row:
                cell.font = subheader_font
                if "ANNEXURE TO NOTE" in str(value):
                    cell.fill = subheader_fill
            
            # Format sub-headers (like "SECURED", "VEHICLES LOANS", etc.)
            elif col_idx == 1 and value and not value.startswith(" ") and not "Total" in str(value) and row_idx > format_start_row:
                if not any(char.islower() for char in str(value)):
                    cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value == "-" or value.startswith("(")):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total rows
            if isinstance(value, str) and ("Total" in value or "Grand Total" in value):
                cell.font = Font(bold=True)
                
                # Find total row
                total_row = row_idx
                for total_col in [2, 3]:
                    try:
                        total_cell = sheet.cell(row=total_row, column=total_col)
                        total_cell.font = Font(bold=True)
                        # Only apply fill to the final totals, not subtotals
                        if "Grand Total" in value:
                            total_cell.fill = total_fill
                        total_cell.alignment = Alignment(horizontal='right')
                        
                        # Add double border for Grand Total
                        if "Grand Total" in value:
                            total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                        else:
                            total_cell.border = Border(bottom=Side(border_style="thin"))
                    except:
                        pass
                
                # Add underline before total rows
                try:
                    for underline_col in [2, 3]:
                        underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                        underline_cell.border = Border(bottom=Side(border_style="thin"))
                except:
                    pass
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    
    # Set blue background for headers in Annexures
    if has_header:
        for col_idx in range(1, 4):
            cell = sheet.cell(row=4, column=col_idx)
            cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

# Populate annexure sheets
annexure_note4_sheet = wb['Annexure to Note 4']
annexure_note8_sheet = wb['Annexure to Note 8']

populate_annexure_sheet_with_header(annexure_note4_sheet, annexure_to_note4_data, True)
populate_annexure_sheet_with_header(annexure_note8_sheet, annexure_to_note8_data, False)

# Annexure to Note 11 data - Long Term Loans and Advances
annexure_to_note11_data = [
    ["ANNEXURE TO NOTE 11", "", ""],
    ["LONG TERM LOANS AND ADVANCES", "", ""],
    ["Security Deposits", "", ""],
    ["Anand Gas Agency (Deposit)", "1,700.00", "1,700.00"],
    ["Madhav Manikrao Deshmukh (Deposite)", "-", "10,00,000.00"],
    ["MSEDCL ( Deposit)", "50,000.00", "-"],
    ["Pratik Nandkishor Purohit (Room Deposite)", "24,000.00", "24,000.00"],
    ["Poonam Construction & Co.) - Security Deposit", "25,00,000.00", "25,00,000.00"],
    ["Rajesh Shankarlal Kothari (Deposit)", "10,00,000.00", "10,00,000.00"],
    ["Ramdas Tukaram Dhole (Deposit)", "8,000.00", "-"],
    ["Sarita Manoj Keshlani (Deposite)", "20,000.00", "20,000.00"],
    ["Shaikh Jafar Gafur ( Oxygen Gas Deposit)", "6,000.00", "6,000.00"],
    ["Shri Gajanan Gramin LPG Vitrak (Gas Deposit)", "3,400.00", "3,400.00"],
    ["Shyam Shelake ( Deposit)", "2,000.00", "-"],
    ["Vinode Vishal Borde (Deposit)", "2,51,000.00", "2,51,000.00"],
    ["", "", ""],
    ["Total", "38,66,100.00", "48,06,100.00"]
]

# Annexure to Note 15 data - Cash & Cash Equivalents
annexure_to_note15_data = [
    ["ANNEXURE TO NOTE 15", "", ""],
    ["CASH & CASH EQUIVALENTS", "", ""],
    ["Balances with Banks in Current Accounts", "", ""],
    [" - Union Bank of India A/C No.511201010030756", "11,61,793.87", "34,57,585.82"],
    [" - Union Bank of India A/C No.511201010030785", "3,856.09", "91,106.41"],
    [" - Union Bank of India A/C No.620301010050215", "49,43,78,713.85", "4,89,73,017.35"],
    ["", "", ""],
    ["", "49,55,44,363.81", "5,25,21,709.58"]
]

# Function to populate the Annexure to Note sheets
def populate_annexure_to_note_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Annexure title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Section header formatting (LONG TERM LOANS AND ADVANCES, CASH & CASH EQUIVALENTS)
            elif row_idx == 2 and col_idx == 1:
                cell.font = Font(bold=True)
            
            # Subsection header formatting (Security Deposits, Balances with Banks)
            elif col_idx == 1 and row_idx > 2 and not value.startswith(" ") and value and value != "Total":
                cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value == "-"):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total row
            if isinstance(value, str) and "Total" in value:
                # Format total label
                cell.font = Font(bold=True)
                
                # Format total amount cells
                total_row = row_idx
                for total_col in [2, 3]:
                    total_cell = sheet.cell(row=total_row, column=total_col)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                
                # Add underline before total row
                for underline_col in [2, 3]:
                    underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                    underline_cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    
# Annexure to Note 12 data - Deferred Tax Asset/Liability
annexure_to_note12_data = [
    ["", "", "ANNEXURE TO NOTE-12", "", "", "", ""],
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", "", "", "", ""],
    ["COMPUTATION OF DEFERRED TAX ASSET/LIABILITY", "", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["PARTICULARS", "COMPANIES ACT", "INCOME TAX", "DIFFERENCE", "ASSET/LIABILITY", "AMOUNT", "TAX 26%"],
    ["", "", "", "", "", "", ""],
    ["CLOSING WDV", "8,01,62,449.98", "7,83,51,297.08", "(18,11,152.90)", "Liability", "18,11,152.90", "4,70,899.75"],
    ["Preliminary Expenses", "", "", "", "Asset", "-", ""],
    ["", "", "", "", "", "", ""],
    ["DEFERRED TAX ASSET", "801624449.98", "783512970.8", "(18,11,152.90)", "Liability", "18,11,152.90", "4,70,899.75"],
    ["", "", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["", "DEFERRED TAX ASSET/ LIABILITY A/C", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["PARTICULARS", "AMOUNT", "", "PARTICULARS", "AMOUNT", "", ""],
    ["", "", "", "", "", "", ""],
    ["To Bal b/d", "2,24,731.53", "", "", "", "", ""],
    ["", "", "", "To P&L A/C", "695631.28", "", ""],
    ["To P&L A/C", "", "", "", "", "", ""],
    ["TO BAL. C/D", "470899.75", "", "TO BAL. C/D", "-", "", ""],
    ["", "", "", "", "", "", ""],
    ["TOTAL", "6,95,631.28", "", "TOTAL", "6,95,631.28", "", ""],
    ["", "", "", "", "", "", ""],
    ["", "", "", "", "P&L DEFERRED TAX Liability", "6,95,631.28", ""],
    ["", "", "", "", "B&S DEFERRED TAX Liability", "4,70,899.75", ""],
    ["", "", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["PARTICULARS", "COMPANIES ACT", "INCOME TAX", "DIFFERENCE", "ASSET/LIABILITY", "AMOUNT", "TAX 27.82%"],
    ["", "", "", "", "", "", ""],
    ["CLOSING WDV.", "30773937.96", "31581743.60", "807805.64", "Asset", "807805.64", "2,24,731.53"],
    ["Preliminary Expenses", "0.00", "0.00", "0.00", "Asset", "0.00", "0.000"],
    ["", "", "", "", "", "", ""],
    ["DEFERRED TAX ASSET", "30773937.96", "31581743.60", "807805.64", "ASSET", "807805.64", "224731.53"],
    ["", "", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["", "DEFERRED TAX ASSET/ LIABILITY A/C", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["PARTICULARS", "AMOUNT", "", "PARTICULARS", "AMOUNT", "", ""],
    ["", "", "", "", "", "", ""],
    ["To Bal b/d", "35250.56", "", "", "", "", ""],
    ["To P&L A/C", "189480.97", "", "", "", "", ""],
    ["", "", "", "TO BAL. C/D", "224731.53", "", ""],
    ["", "", "", "", "", "", ""],
    ["TOTAL", "224731.53", "", "TOTAL", "224731.53", "", ""],
    ["", "", "", "", "", "", ""],
    ["", "", "", "", "P&L DEFERRED TAX ASSET", "189480.97", ""],
    ["", "", "", "", "B&S DEFERRED TAX ASSET", "224731.53", ""],
]

# Function to populate the Annexure to Note 12 sheet
def populate_annexure_to_note12_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Define merged cell ranges
    merge_ranges = [
        # Title merges
        (1, 3, 1, 5),  # ANNEXURE TO NOTE-12
        (2, 1, 2, 7),  # NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.
        (3, 1, 3, 7),  # COMPUTATION OF DEFERRED TAX ASSET/LIABILITY
        
        # A/C statement header merges
        (13, 2, 13, 6),  # DEFERRED TAX ASSET/ LIABILITY A/C
        (35, 2, 35, 6),  # Second DEFERRED TAX ASSET/ LIABILITY A/C
        
        # Summary information merges
        (24, 5, 24, 6),  # P&L DEFERRED TAX Liability
        (25, 5, 25, 6),  # B&S DEFERRED TAX Liability
        (44, 5, 44, 6),  # P&L DEFERRED TAX ASSET
        (45, 5, 45, 6),  # B&S DEFERRED TAX ASSET
    ]
    
    # Apply merges
    for start_row, start_col, end_row, end_col in merge_ranges:
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    
    # Apply formatting
    # Font styles
    title_font = Font(name='Arial', size=12, bold=True)
    subtitle_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    data_font = Font(name='Arial', size=10)
    
    # Fill styles
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    # Format title and subtitle
    for row_idx in [1, 2, 3]:
        for col_idx in range(1, 8):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.font = title_font
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.font = subtitle_font
                cell.alignment = Alignment(horizontal='center')
    
    # Format table headers and section titles
    header_rows = [5, 15, 28, 37]
    section_title_rows = [13, 35]
    
    for row_idx in header_rows:
        for col_idx in range(1, 8):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
    
    for row_idx in section_title_rows:
        cell = sheet.cell(row=row_idx, column=2)  # Center column for section titles
        cell.font = subtitle_font
        cell.alignment = Alignment(horizontal='center')
    
    # Format all data cells with proper alignment and borders
    for row_idx in range(6, 46):
        for col_idx in range(1, 8):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Skip empty cells and headers
            if not cell.value or row_idx in header_rows or row_idx in section_title_rows:
                continue
            
            # Format numeric values
            if isinstance(cell.value, str) and (cell.value.replace('-', '').replace(',', '').replace('.', '').isdigit() or 
                                               cell.value.startswith('(') or cell.value == "-"):
                cell.alignment = Alignment(horizontal='right')
            elif col_idx in [1, 4]:  # Text columns (PARTICULARS and ASSET/LIABILITY)
                cell.alignment = Alignment(horizontal='left')
            
            # Add borders to cells in the table areas
            if ((row_idx >= 5 and row_idx <= 10) or 
                (row_idx >= 15 and row_idx <= 22) or
                (row_idx >= 28 and row_idx <= 33) or
                (row_idx >= 37 and row_idx <= 43)):
                cell.border = border
    
    # Format total rows
    total_rows = [10, 22, 33, 43]
    for row_idx in total_rows:
        for col_idx in range(1, 8):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.font = Font(bold=True)
                if "TOTAL" in str(cell.value) or col_idx in [2, 5]:  # Total label or amount columns
                    cell.fill = total_fill
    
    # Add summary labels formatting
    summary_rows = [24, 25, 44, 45]
    for row_idx in summary_rows:
        for col_idx in range(5, 7):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.alignment = Alignment(horizontal='right' if col_idx == 6 else 'left')
    
    # Set column widths
    column_widths = [20, 18, 15, 18, 20, 15, 15]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width
    
    # Add signature placeholder
    signature_row = 45
    signature_cell = sheet.cell(row=signature_row, column=2)
    signature_cell.value = "(Signature placeholder)"
    signature_cell.font = Font(italic=True, color="808080")
    signature_cell.alignment = Alignment(horizontal='center')

# Populate Annexure to Note 12 sheet
annexure_note12_sheet = wb['Annexure to Note 12']
populate_annexure_to_note12_sheet(annexure_note12_sheet, annexure_to_note12_data)

# Format subtotal in Note 15 (the bank balance subtotal)
def format_note15_subtotal(sheet):
    # Find the row with the bank subtotal (49,55,44,363.81)
    subtotal_row = 0
    for row_idx in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=2).value
        if isinstance(cell_value, str) and "49,55,44,363.81" in cell_value:
            subtotal_row = row_idx
            break
    
    # Apply formatting to subtotal if found
    if subtotal_row > 0:
        for col_idx in [2, 3]:
            subtotal_cell = sheet.cell(row=subtotal_row, column=col_idx)
            subtotal_cell.font = Font(bold=True)
            subtotal_cell.alignment = Alignment(horizontal='right')
            subtotal_cell.border = Border(bottom=Side(border_style="thin"))

# Populate annexure sheets
annexure_note11_sheet = wb['Annexure to Note 11']
annexure_note15_sheet = wb['Annexure to Note 15']

populate_annexure_to_note_sheet(annexure_note11_sheet, annexure_to_note11_data)
populate_annexure_to_note_sheet(annexure_note15_sheet, annexure_to_note15_data)

# Apply special formatting for Note 15 subtotal
format_note15_subtotal(annexure_note15_sheet)

# Annexure to Note 16 data - Short Term Loans & Advances
annexure_to_note16_data = [
    ["ANNEXURE TO NOTE 16", "", ""],
    ["SHORT TERM LOANS & ADVANCES", "", ""],
    ["ADVANCES TO VENDORS", "", ""],
    ["A & A Consulting Engineer", "4,00,000.00", "4,00,000.00"],
    ["Aaadnath Trade-Wings Private Limited", "7,69,494.24", "7,69,494.24"],
    ["Acc Limited", "58,854.00", "7,08,762.00"],
    ["Accurate Valuers And Engineers", "29,500.00", "29,500.00"],
    ["Amit Vikas Mahajan", "5,00,000.00", "5,00,000.00"],
    ["Ammann India Private Limited (Gujrat)", "17,263.54", "1,30,244.00"],
    ["Anshul & Associates", "11,41,686.00", "9,54,338.00"],
    ["Apple Chemie India Pvt. Ltd.", "22,390.00", "22,390.00"],
    ["Aqua- Tech Services", "80,000.00", "80,000.00"],
    ["Atrinandan Cement Products", "11,605.00", "11,605.00"],
    ["Avinash Ashok Aringale", "", "1,00,000.00"],
    ["Bharat Motors Co.", "16,124.00", "16,124.00"],
    ["Care Ratings Limited", "1,37,500.00", "1,10,500.00"],
    ["Countrywide Cargo Movers", "39,677.00", "39,677.00"],
    ["Dandade  Pritam Bajirao", "1,25,000.00", "1,25,000.00"],
    ["Dicksons Engineering Co. Pvt. Ltd.", "994.00", "994.00"],
    ["Dilip Chavan (Lc)", "2,83,149.98", "1,149.98"],
    ["Ganesh Shamrao Rede (Lc)", "10,000.00", "10,000.00"],
    ["FUTURE INC", "14,499.00", ""],
    ["Gurukrupa Cement Depot", "-", "4,00,000.00"],
    ["Gopal Govindrao Gunjikar", "13,880.00", ""],
    ["Hindustan Petroleum Corporation Limited", "3,55,140.85", "3,55,140.85"],
    ["Iron Trangle Ltd.", "5,24,808.00", "5,24,808.00"],
    ["Infomerics Valuation & Rating Pvt. Ltd.", "2,16,000.00", ""],
    ["Infraking Consulting Engineers", "10,42,360.00", ""],
    ["Jabir Jainal Shaikh", "", "2,80,000.00"],
    ["Jagdish Bhagwan Shivde", "", "4,00,000.00"],
    ["Jamnadas And Company", "", "30,025.00"],
    ["Jijau Shrusti", "2,00,000.00", "2,00,000.00"],
    ["Jitendra Kanhaiylalal Kalntri", "14,40,000.00", "70,000.00"],
    ["Jsk Corporation Pvt. Ltd.", "74747.00", "74,747.00"],
    ["Kamalkishor P. Bhattad", "0.00", "4,00,000.00"],
    ["Khalil Kalandar Khan (Lc)", "", "1,401.00"],
    ["Krd Borewells & Constructions", "20,000.00", "20,000.00"],
    ["Laxminarayan Cement Product", "", "2,00,000.00"],
    ["Lead Law", "1,44,000.00", ""],
    ["Manck Narang", "", "50,000.00"],
    ["Metso India Private Limited", "", "93,177.02"],
    ["Mukund Gangaram Belage", "55,577.00", ""],
    ["Messers Jagdamba Traders", "9,01,458.00", ""],
    ["Nandhra Engg. & Const. (India) Pvt.Ltd.(Adv.)", "11,59,22,844.00", "13,46,17,023.00"],
    ["Nandhra Engg. & Const. (India) Pvt.Ltd.(Bill)", "42,37,023.49", "41,94,474.44"],
    ["Naveen G.(L.C.)", "30,000.00", "30,000.00"],
    ["Nilesh Welding Works", "21,440.00", "21,440.00"],
    ["Nirman Infra Assocites", "9,66,213.00", "2,66,400.00"],
    ["Nitin Tukaram Jadhav", "", "4,00,000.00"],
    ["Nand Power Link", "64,30,068.00", ""],
    ["NAVNEET KAUR", "29,000.00", ""],
    ["Om Industries", "13,736.00", "13,736.00"],
    ["Pankaj Automobiles", "4,600.00", "4,600.00"],
    ["Parashuram Jayeppa Talakeri", "21,265.00", ""],
    ["Poonam Construction & Co. (Lc)", "8,40,90,474.98", "6,71,80,000.00"],
    ["Rahul Constructions", "-", "3,100.00"],
    ["Rajlaxmi Servo", "", "1,61,514.47"],
    ["Raj Water Suppliers", "23,673.00", ""],
    ["Raviraj Construction Co.", "4,34,265.00", "4,34,265.00"],
    ["Rustam Khan Taslim Khan", "", "50,000.00"],
    ["Resurgent India Limited", "15,000.00", ""],
    ["Saifee Iron & Steel", "26,816.00", "26,816.00"],
    ["Saltech Instruments Pvt. Ltd.", "", "26,200.00"],
    ["Sandip Suresh Shegakar (Lc)", "1,121.00", "1,121.00"],
    ["Sanjay Kumar", "2,06,220.00", "2,06,220.00"],
    ["Sanjeev Shibu Sardar ( Lc)", "1,34,324.00", "1,34,324.00"],
    ["Sarthi Enterprises", "6,00,000.00", ""],
    ["Shiv Ib Motors Pvt. Ltd.(Swalik Ford)", "", "3,064.00"],
    ["Shivkumar Prabhudas Atravalkar", "", "9,00,000.00"],
    ["Shravani Suppliers", "-", "3,00,000.00"],
    ["Shree Saaguru Kaka Stone Crusher", "", "2,33,880.00"],
    ["Sidma Infra Private Limited", "1,46,75,880.00", "75,48,520.00"],
    ["Sudhakar Tukaram Pundkar", "", "2,00,000.00"],
    ["Sukhija Real Estate Private Ltd.", "7,852.00", "7,852.00"],
    ["Shaurya Technosoft Pvt. Ltd.", "24,780.00", ""],
    ["Shivkumar Prabhudas Atravalkar", "9,00,000.00", ""],
    ["Shree Tading Company", "12,955.00", ""],
    ["Suspence Account", "1,30,000.00", "30,000.00"],
    ["Technogem Consultants Pvt. Ltd.", "6,00,000.00", ""],
    ["Udeep Systems", "54,918.00", "65,107.00"],
    ["Vaaraahee Structural Repair & Water Proofing Cont.", "", "1,35,700.00"],
    ["Vaibhav Zingaraji Parale (Lc)", "", "29,457.00"],
    ["Vision Infra", "2,75,000.00", "2,75,000.00"],
    ["World 13 Distributions Private Limited", "2,97,67,250.00", "1,63,50,810.00"],
    ["Yogesh Arvind Raut", "53,000.00", "53,000.00"],
    ["Subham Manohar Baharagi (LC)", "25,000.00", ""],
    ["Vinod Kishor Dhokane (Lc)", "50,00,000.00", ""],
    ["Vinod Sitaram Chavan (LC)", "50,00,000.00", ""],
    ["Abhajeet Harishchandra Patil (LC)", "50,00,000.00", ""],
    ["Abhaji Infrastructure", "66,47,872.00", ""],
    ["Amol Ramesh Gawai (Lc)", "1,03,021.00", ""],
    ["Bhagappa Ramchandra Reddy (Lc)", "50,00,000.00", ""],
    ["B. V. Parmar (LC)", "2,50,000.00", ""],
    ["Chandraprakash Dipchand Tailor", "31,80,098.00", ""],
    ["Ganesh Narayan Karale (LC)", "10,000.00", ""],
    ["Haribhau Vetai (LC)", "50,00,000.00", ""],
    ["Ismail Jamadar (LC)", "5,40,745.00", ""],
    ["Kawaljeet Singh Chhabra", "1,81,10,396.00", ""],
    ["Kawaljit Singh Chhabra (LC)", "18,72,965.00", ""],
    ["KNC Buildcon (LC)", "68,05,140.00", ""],
    ["Krushna Vishnu Kartade (Lc)", "50,00,000.00", ""],
    ["Malpshri Construction (LC)", "13,989.00", ""],
    ["Narendrasingh Munshilal Chabada (LC)", "47,09,980.00", ""],
    ["Om Sai Construction (Lc)", "9,07,082.00", ""],
    ["Poonam Subhash Sawale (LC)", "8,86,098.00", ""],
    ["Prakshvi Infraconstructors Pvt. Ltd.", "71,52,972.00", ""],
    ["Rajmudra Constructions (LC)", "2,00,000.00", ""],
    ["Sayaji Road Rollers", "2,50,000.00", ""],
    ["SES Engineering Builders & Civil Works Pvt.Ltd.Bill", "1,78,48,490.00", ""],
    ["Shyam Jadhav (Lc)", "50,00,000.00", ""],
    ["Bajaj Travells Ltd.", "12,59,430.00", ""],
    ["J V Engineering Works", "2,82,495.00", ""],
    ["Madhav Manikrao Deshmukh", "2,32,925.00", ""],
    ["Saffron Redymix Concrete", "1,90,87,878.20", ""],
    ["", "", ""],
    ["Total", "39,35,29,989.28", "24,10,12,701.00"],
    ["", "", ""],
    ["ADVANCES TO EMPLOYESS", "", ""],
    ["Rahat Virk (Salary)", "-", "7,700.00"],
    ["Kishor Ankush Mohite (Salary)", "20,03,200.00", "4,50,700.00"],
    ["Gurpalsingh Virk (Salary)", "", "29,350.00"],
    ["Gurtej Singh Nandhra (Salary)", "16,65,600.00", "6,13,100.00"],
    ["Santosh Ankush Mohite (Salary)", "40,83,200.00", "9,30,700.00"],
    ["Iqbal Singh  Nandhra (Salary)", "79,96,800.00", "37,44,300.00"],
    ["Rahat Virk (Adv.)", "5,500.00", "500.00"],
    ["Mohan Chandrabhan Singh (Adv.)", "3,000.00", "19,000.00"],
    ["Nanadhra Engineering & Construction (Solapur Road)", "50,000.00", "50,000.00"],
    ["Kishor Ankush Mohite (Adv.)", "(16,38,039.60)", "4,03,759.40"],
    ["Other Emplyee", "1,77,002.00", ""],
    ["", "", ""],
    ["Total", "1,43,46,262.40", "67,39,109.40"],
    ["", "", ""],
    ["OTHERS", "", ""],
    ["CGST Receivable against Reverse Charge", "", "5,78,945.85"],
    ["CGST (TDS)", "", "14,42,855.00"],
    ["GST Paid", "", "27,28,506.00"],
    ["GST Cess  20% On Car", "", "2,70,202.61"],
    ["IGST Claimed", "", "12,21,348.23"],
    ["IGST Receivable against Reverse Charge", "", "2,34,000.00"],
    ["SGST (TDS)", "", "14,42,855.00"],
    ["SGST Receivable against Reverse Charge", "", "5,43,346.57"],
    ["GST Input", "49,25,850.55", ""],
    ["TCS AY 2021-22", "90,007.92", "90,007.92"],
    ["TCS on Royalty Charges A.Y. 20-21", "21,092.00", "1,26,029.00"],
    ["TDS A.Y. 2020-21", "", "21,64,284.00"],
    ["TCS AY 2022-23", "1,20,424.83", ""],
    ["TDS A.Y. 2022-23", "52,29,276.00", ""],
    ["TCS on Royalty Charges A.Y. 22-23", "2,34,048.00", ""],
    ["", "", ""],
    ["Total", "1,06,20,699.30", "1,08,42,380.18"]
]

# Function to populate the Annexure to Note 16 sheet
def populate_annexure_to_note16_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Annexure title formatting
            if row_idx == 1 and col_idx == 1:
                cell.font = subheader_font
                cell.fill = subheader_fill
            
            # Section header formatting (SHORT TERM LOANS & ADVANCES)
            elif row_idx == 2 and col_idx == 1:
                cell.font = Font(bold=True)
            
            # Subsection header formatting (ADVANCES TO VENDORS, ADVANCES TO EMPLOYESS, OTHERS)
            elif col_idx == 1 and any(header in str(value) for header in ["ADVANCES TO VENDORS", "ADVANCES TO EMPLOYESS", "OTHERS"]):
                cell.font = Font(bold=True)
            
            # Format numeric values with right alignment
            elif col_idx in [2, 3] and value and isinstance(value, str) and (value[0].isdigit() or value == "-" or value.startswith("(")):
                cell.alignment = Alignment(horizontal='right')
            
            # Format total rows
            if isinstance(value, str) and value == "Total" and row_idx in [121, 134, 152]:
                # Format total label
                cell.font = Font(bold=True)
                
                # Format total amount cells
                total_row = row_idx
                for total_col in [2, 3]:
                    total_cell = sheet.cell(row=total_row, column=total_col)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = Border(bottom=Side(border_style="double"), top=Side(border_style="double"))
                
                # Add underline before total row
                for underline_col in [2, 3]:
                    underline_cell = sheet.cell(row=total_row-1, column=underline_col)
                    underline_cell.border = Border(bottom=Side(border_style="thin"))
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25

# Populate Annexure to Note 16 sheet
annexure_note16_sheet = wb['Annexure to Note 16']
populate_annexure_to_note16_sheet(annexure_note16_sheet, annexure_to_note16_data)

# Fixed Assets sheet data - Note 10 (both FY 2021 and FY 2022)
fixed_assets_data_2022 = [
    # FY 2021 data
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", "", "", "", "", "", "", "", "NOTE 10 FIXED ASSETS"],
    ["SCHEDULE OF FIXED ASSETS AS PER COMPANIES ACT, 2013 AS ON 31ST MARCH 2022", "", "", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["", "G R O S S  B L O C K", "", "", "", "D E P R E C I A T I O N", "", "", "", "N E T  B L O C K", ""],
    ["PARTICULARS", "", "", "", "", "", "", "ADJUSTMENTS", "", "", ""],
    ["", "COST AS ON 01.04.2020", "ADDITION", "DEDUCTION", "COST AS ON 31.03.2021", "UP TO 01.04.2020", "DURING THE YEAR", "ON A/C OF SALE or Other Adjustment", "TOTAL UP TO 31.03.2021", "W.D.V. AS ON 31.03.2021", "W.D.V. AS ON 31.03.2020"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["CWIP-TOOLS & EQUIPMENTS", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["TANGIBLE ASSETS", "", "", "", "", "", "", "", "", "", ""],
    ["BUILDING", "0.00", "", "-", "0.00", "", "", "-", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["PLANT & MACHINERY", "", "", "", "", "", "", "", "", "", ""],
    ["Plant & Machinery", "90,616.74", "2,88,86,381.19", "-", "2,89,76,997.93", "5,328.00", "42,98,200.00", "-", "43,03,528.00", "2,46,73,469.93", "85,288.74"],
    ["Generator", "0.00", "", "-", "0.00", "", "", "-", "-", "-", "-"],
    ["Lab Equipments", "21,25,200.60", "93,856.00", "", "22,19,056.60", "3,09,640.00", "4,79,284.00", "", "7,88,924.00", "14,30,132.60", "18,15,560.60"],
    ["Tools & Equipments", "1,08,958.00", "7,800.00", "-", "1,16,758.00", "3,298.00", "19,918.00", "-", "23,216.00", "93,542.00", "1,05,660.00"],
    ["Electrical Assets", "90,875.42", "1,18,875.41", "-", "2,09,750.83", "4,712.00", "30,775.00", "-", "35,487.00", "1,74,263.83", "86,163.42"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["OFFICE EQUIPMENTS", "", "", "", "", "", "", "", "", "", ""],
    ["Mobile and Mac Book", "2,37,712.45", "-", "-", "2,37,712.45", "53,197.00", "1,16,540.00", "-", "1,69,737.00", "67,975.00", "1,84,515.00"],
    ["Computer and Laptops", "9,83,277.22", "2,74,638.26", "-", "12,57,915.48", "3,79,340.00", "4,25,379.00", "-", "8,04,719.00", "4,53,196.00", "6,03,937.00"],
    ["Office Equipments", "2,62,486.40", "1,58,128.20", "-", "4,20,614.60", "64,353.00", "1,36,567.00", "-", "2,00,920.00", "2,19,695.00", "1,98,133.00"],
    ["Air Conditioner", "2,93,750.00", "1,23,671.88", "-", "4,17,421.88", "12,763.00", "89,310.00", "-", "1,02,073.00", "3,15,349.00", "2,80,987.00"],
    ["Refrigerator", "", "45,593.23", "", "45,593.23", "-", "9,504.00", "", "9,504.00", "36,089.00", "-"],
    ["Camera", "0.00", "1,47,485.00", "-", "1,47,485.00", "-", "41,158.00", "-", "41,158.00", "1,06,327.00", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["VEHICLES", "9,25,942.12", "18,29,700.07", "-", "27,55,642.19", "54,288.00", "2,86,308.00", "-", "3,40,596.00", "24,15,046.19", "8,71,654.12"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["FURNITURE & FIXTURES", "11,23,625.18", "78,882.23", "-", "12,02,507.41", "1,47,598.00", "2,66,057.00", "-", "4,13,655.00", "7,88,852.41", "9,76,027.18"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["INTANGIBLE ASSETS", "", "", "", "", "", "", "", "", "", ""],
    ["Computer Softwares", "0.00", "-", "-", "0.00", "", "", "-", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["TOTAL", "62,42,444.13", "3,17,65,011.47", "0.00", "3,80,07,455.60", "10,34,517.00", "61,99,000.00", "0.00", "72,33,517.00", "3,07,73,937.96", "52,07,926.06"],
    
    # FY 2022 data - Add a comma here to separate the two sets of data
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", "", "", "", "", "", "", "", "NOTE 10 FIXED ASSETS"],
    ["SCHEDULE OF FIXED ASSETS AS PER COMPANIES ACT, 2013 AS ON 31ST MARCH 2022", "", "", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["", "G R O S S  B L O C K", "", "", "", "D E P R E C I A T I O N", "", "", "", "N E T  B L O C K", ""],
    ["PARTICULARS", "", "", "", "", "", "", "ADJUSTMENTS", "", "", ""],
    ["", "COST AS ON 01.04.2021", "ADDITION", "DEDUCTION", "COST AS ON 31.03.2022", "UP TO 01.04.2021", "DURING THE YEAR", "ON A/C OF SALE or Other Adjustment", "TOTAL UP TO 31.03.2022", "W.D.V. AS ON 31.03.2022", "W.D.V. AS ON 31.03.2021"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["CWIP-TOOLS & EQUIPMENTS", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["TANGIBLE ASSETS", "", "", "", "", "", "", "", "", "", ""],
    ["BUILDING", "-", "", "-", "-", "", "", "-", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["PLANT & MACHINERY", "", "", "", "", "", "", "", "", "", ""],
    ["Plant & Machinery", "2,46,73,469.93", "-", "-", "2,46,73,469.93", "43,03,528.00", "44,65,898.06", "-", "87,69,426.06", "2,02,07,571.87", "2,46,73,469.93"],
    ["Generator", "0.00", "", "-", "0.00", "", "", "-", "", "", ""],
    ["Lab Equipments", "14,30,132.60", "45,140.00", "", "14,75,272.60", "7,88,924.00", "3,73,111.33", "", "11,62,035.33", "11,02,161.27", "14,30,132.60"],
    ["Tools & Equipments", "93,542.00", "-", "-", "93,542.00", "23,216.00", "16,931.10", "-", "40,147.10", "76,610.90", "93,542.00"],
    ["Electrical Assets", "1,74,263.83", "-", "-", "1,74,263.83", "35,487.00", "31,541.75", "-", "67,028.75", "1,42,722.08", "1,74,263.83"],
    ["Crusher", "", "5,55,84,745.00", "", "5,55,84,745.00", "", "", "", "", "5,55,84,745.00", ""],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["OFFICE EQUIPMENTS", "", "", "", "", "", "", "", "", "", ""],
    ["Mobile and Mac Book", "67,975.00", "24,500.00", "-", "92,475.00", "1,69,737.00", "54,094.01", "-", "2,23,831.01", "38,380.99", "67,975.00"],
    ["Computer and Laptops", "4,53,196.00", "2,62,457.48", "-", "7,15,653.48", "8,04,719.00", "3,93,187.30", "-", "11,97,906.30", "3,22,466.18", "4,53,196.00"],
    ["Office Equipments", "2,19,695.00", "-", "-", "2,19,695.00", "2,00,920.00", "99,016.54", "-", "2,99,936.54", "1,20,678.46", "2,19,695.00"],
    ["Air Conditioner", "3,15,349.00", "-", "-", "3,15,349.00", "1,02,073.00", "81,643.86", "-", "1,83,716.86", "2,33,705.14", "3,15,349.00"],
    ["Refrigerator", "36,089.00", "-", "", "36,089.00", "9,504.00", "6,532.11", "", "16,036.11", "29,556.89", "36,089.00"],
    ["Camera", "1,06,327.00", "-", "-", "1,06,327.00", "41,158.00", "47,921.58", "-", "89,079.58", "58,405.42", "1,06,327.00"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["VEHICLES", "24,15,046.19", "-", "-", "24,15,046.19", "3,40,596.00", "7,54,218.93", "-", "10,94,814.93", "16,60,827.26", "24,15,046.19"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["FURNITURE & FIXTURES", "7,88,852.41", "-", "-", "7,88,852.41", "4,13,655.00", "2,04,233.89", "-", "6,17,888.89", "5,84,618.52", "7,88,852.41"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["INTANGIBLE ASSETS", "", "", "", "", "", "", "", "", "", ""],
    ["Computer Softwares", "-", "-", "-", "-", "", "", "-", "-", "-", "-"],
    ["", "", "", "", "", "", "", "", "", "", ""],
    ["TOTAL", "3,07,73,937.96", "5,59,16,842.48", "0.00", "8,66,90,780.44", "72,33,517.00", "65,28,330.45", "0.00", "1,37,61,847.45", "8,01,62,449.98", "3,07,73,937.96"]
]

# Function to populate the Fixed Assets sheet
def populate_fixed_assets_sheet(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Set column widths
    column_widths = [25, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width
    
    # Apply formatting
    # Merge cells for title and subtitle
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    
    # Format title and subtitle
    for row_idx in [1, 2]:
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left')
    
    # Format section headers
    header_rows = [4, 5]
    for row_idx in header_rows:
        for col_idx in range(1, 12):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = subheader_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format GROSS BLOCK, DEPRECIATION, NET BLOCK merged cells
    gb_start, gb_end = 2, 5
    dep_start, dep_end = 6, 9
    nb_start, nb_end = 10, 11
    
    sheet.merge_cells(start_row=4, start_column=gb_start, end_row=4, end_column=gb_end)
    sheet.merge_cells(start_row=4, start_column=dep_start, end_row=4, end_column=dep_end)
    sheet.merge_cells(start_row=4, start_column=nb_start, end_row=4, end_column=nb_end)
    
    # Apply section headers formatting (TANGIBLE ASSETS, OFFICE EQUIPMENTS, etc.)
    section_headers = [8, 10, 13, 21, 27, 30, 32]
    for row_idx in section_headers:
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = Font(bold=True)
        if "TANGIBLE ASSETS" in str(cell.value) or "INTANGIBLE ASSETS" in str(cell.value):
            cell.fill = subheader_fill
    
    # Format all numeric cells
    for row_idx in range(7, 35):
        for col_idx in range(2, 12):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if isinstance(value, str) and (value.replace('-', '').replace('.', '').replace(',', '').isdigit() or value == "-"):
                cell.alignment = Alignment(horizontal='right')
    
    # Format the TOTAL row
    total_row = 34
    for col_idx in range(1, 12):
        cell = sheet.cell(row=total_row, column=col_idx)
        cell.font = Font(bold=True)
        if col_idx > 1:
            cell.fill = total_fill
    
    # Add borders to all cells in the table
    for row_idx in range(4, 35):
        for col_idx in range(1, 12):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = border
    
    # Apply special light blue background to headers
    for row_idx in range(4, 7):
        for col_idx in range(1, 12):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = light_blue_fill

# Populate Fixed Assets sheet
fixed_assets_sheet = wb['Fixed Assets']
populate_fixed_assets_sheet(fixed_assets_sheet, fixed_assets_data_2022)

# Annexure B data - Fixed Assets As Per Income Tax Act for both years
annexure_b_data = [
    # FY 2021-22 Data
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", "", "", "", "", "", ""],
    ["Schedule Of Fixed Assets As Per Income Tax Act, 1961 As on March 31, 2022", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", ""],
    ["ANNEXURE \"B\"", "", "", "", "", "", "", "", ""],
    ["", "", "ADDITION", "", "TOTAL", "", "", "", ""],
    ["PARTICULARS", "W.D.V. AS ON 01st April 2021", "UPTO 30-09-2021", "AFTER 30-09-2021", "SALE/ DEDUCTION", "AS ON 31.03.2022", "RATE", "DEPRECIATION", "W.D.V. AS ON 31st March 2022"],
    ["", "", "", "", "", "", "", "", ""],
    ["BLOCK 10%", "", "", "", "", "", "", "", ""],
    ["BUILDING", "-", "-", "", "-", "-", "10%", "-", "-"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "-", "-", "-", "-", "-", "", "-", "-"],
    ["BLOCK 10%", "", "", "", "", "", "", "", ""],
    ["FURNITURE & FITTINGS", "10,03,853.41", "-", "-", "-", "10,03,853.41", "10%", "1,00,385.00", "9,03,468.41"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "10,03,853.41", "-", "-", "-", "10,03,853.41", "", "1,00,385.00", "9,03,468.41"],
    ["BLOCK 15%", "", "", "", "", "", "", "", ""],
    ["PLANT & MACHINARY & TOOLS", "2,49,50,497.76", "-", "5,56,29,885.00", "-", "8,05,80,382.76", "15%", "79,14,816.00", "7,26,65,566.76"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "2,49,50,497.76", "-", "5,56,29,885.00", "-", "8,05,80,382.76", "", "79,14,816.00", "7,26,65,566.76"],
    ["BLOCK 15%", "", "", "", "", "", "", "", ""],
    ["OFFICE EQUIPMENTS", "25,27,511.31", "-", "-", "-", "25,27,511.31", "15%", "3,79,127.00", "21,48,384.31"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "25,27,511.31", "-", "-", "-", "25,27,511.31", "", "3,79,127.00", "21,48,384.31"],
    ["BLOCK 40%", "", "", "", "", "", "", "", ""],
    ["COMPUTER & SOFTWARE", "6,92,784.93", "2,86,957.48", "-", "", "9,79,742.41", "40%", "3,91,897.00", "5,87,845.41"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "6,92,784.93", "2,86,957.48", "-", "-", "9,79,742.41", "", "3,91,897.00", "5,87,845.41"],
    ["BLOCK 15%", "", "", "", "", "", "", "", ""],
    ["VEHICLES", "24,07,096.19", "-", "-", "-", "24,07,096.19", "15%", "3,61,064.00", "20,46,032.19"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "24,07,096.19", "-", "-", "-", "24,07,096.19", "", "3,61,064.00", "20,46,032.19"],
    ["BLOCK 80%", "", "", "", "", "", "", "", ""],
    ["SOLAR POWER PLANT", "-", "-", "-", "-", "-", "40%", "-", "-"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "-", "-", "-", "-", "-", "", "-", "-"],
    ["", "", "", "", "", "", "", "", ""],
    ["GRAND TOTAL", "3,15,81,743.60", "2,86,957.48", "5,56,29,885.00", "-", "8,74,98,586.08", "", "91,47,289.00", "7,83,51,297.08"],
    
    # FY 2020-21 Data
    ["NANDHRA ENGINEERING & CONSTRUCTION (SPV) PVT. LTD.", "", "", "", "", "", "", "", ""],
    ["Schedule Of Fixed Assets As Per Income Tax Act, 1961 As on March 31, 2021", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", ""],
    ["ANNEXURE \"B\"", "", "", "", "", "", "", "", ""],
    ["", "", "ADDITION", "", "TOTAL", "", "", "", ""],
    ["PARTICULARS", "W.D.V. AS ON 01st April 2020", "UPTO 30-09-2020", "AFTER 30-09-2020", "SALE/ DEDUCTION", "AS ON 31.03.2021", "RATE", "DEPRECIATION", "W.D.V. AS ON 31st March 2021"],
    ["", "", "", "", "", "", "", "", ""],
    ["BLOCK 10%", "", "", "", "", "", "", "", ""],
    ["BUILDING", "-", "-", "", "-", "-", "10%", "-", "-"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "-", "-", "-", "-", "-", "", "-", "-"],
    ["BLOCK 10%", "", "", "", "", "", "", "", ""],
    ["FURNITURE & FITTINGS", "10,36,510.18", "78,882.23", "-", "-", "11,15,392.41", "10%", "1,11,539.00", "10,03,853.41"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "10,36,510.18", "78,882.23", "-", "-", "11,15,392.41", "", "1,11,539.00", "10,03,853.41"],
    ["BLOCK 15%", "", "", "", "", "", "", "", ""],
    ["PLANT & MACHINARY & TOOLS", "2,66,001.16", "2,81,69,075.41", "8,43,981.19", "-", "2,92,79,057.76", "15%", "43,28,560.00", "2,49,50,497.76"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "2,66,001.16", "2,81,69,075.41", "8,43,981.19", "-", "2,92,79,057.76", "", "43,28,560.00", "2,49,50,497.76"],
    ["BLOCK 15%", "", "", "", "", "", "", "", ""],
    ["OFFICE EQUIPMENTS", "23,91,605.00", "4,19,095.60", "1,49,638.71", "-", "29,60,339.31", "15%", "4,32,828.00", "25,27,511.31"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "23,91,605.00", "4,19,095.60", "1,49,638.71", "-", "29,60,339.31", "", "4,32,828.00", "25,27,511.31"],
    ["BLOCK 40%", "", "", "", "", "", "", "", ""],
    ["COMPUTER & SOFTWARE", "8,08,655.67", "60,593.22", "2,14,045.04", "", "10,83,293.93", "40%", "3,90,509.00", "6,92,784.93"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "8,08,655.67", "60,593.22", "2,14,045.04", "-", "10,83,293.93", "", "3,90,509.00", "6,92,784.93"],
    ["BLOCK 15%", "", "", "", "", "", "", "", ""],
    ["VEHICLES", "8,40,733.12", "-", "18,29,700.07", "-", "26,70,433.19", "15%", "2,63,337.00", "24,07,096.19"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "8,40,733.12", "-", "18,29,700.07", "-", "26,70,433.19", "", "2,63,337.00", "24,07,096.19"],
    ["BLOCK 80%", "", "", "", "", "", "", "", ""],
    ["SOLAR POWER PLANT", "-", "-", "-", "-", "-", "40%", "-", "-"],
    ["", "", "", "", "", "", "", "", ""],
    ["TOTAL", "-", "-", "-", "-", "-", "", "-", "-"],
    ["", "", "", "", "", "", "", "", ""],
    ["GRAND TOTAL", "53,43,505.13", "2,87,27,646.46", "30,37,365.01", "-", "3,71,08,516.60", "", "55,26,773.00", "3,15,81,743.60"]
]

# Function to populate the Annexure B sheet
def populate_annexure_b_sheet_combined(sheet, data):
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Merge cells for titles and subtitles of both tables
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    sheet.merge_cells(start_row=39, start_column=1, end_row=39, end_column=9)
    sheet.merge_cells(start_row=40, start_column=1, end_row=40, end_column=9)
    
    # Merge cells for ADDITION and TOTAL headers in both tables
    sheet.merge_cells(start_row=5, start_column=3, end_row=5, end_column=4)
    sheet.merge_cells(start_row=5, start_column=5, end_row=5, end_column=6)
    sheet.merge_cells(start_row=43, start_column=3, end_row=43, end_column=4)
    sheet.merge_cells(start_row=43, start_column=5, end_row=43, end_column=6)
    
    # Apply formatting
    # Format title and subtitle
    for row_idx in [1, 2]:
        cell = sheet.cell(row=row_idx, column=1)
        if row_idx == 1:
            cell.font = Font(name='Arial', size=12, bold=True, color="000080")  # Dark blue
        else:
            cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Format "ANNEXURE B" text
    cell = sheet.cell(row=4, column=1)
    cell.font = Font(bold=True)
    
    # Format table headers
    for col_idx in range(1, 10):
        # Headers in row 5-6
        for row_idx in [5, 6]:
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = light_blue_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Format BLOCK headers and asset category rows
    block_rows = [8, 12, 16, 20, 24, 28, 32]
    asset_rows = [9, 13, 17, 21, 25, 29, 33]
    
    for row_idx in block_rows:
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = Font(bold=True)
        cell.fill = light_blue_fill
    
    for row_idx in asset_rows:
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = Font(bold=True)
    
    # Format total rows
    total_rows = [11, 15, 19, 23, 27, 31, 35, 37]
    for row_idx in total_rows:
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = Font(bold=True)
        if row_idx == 37:  # GRAND TOTAL
            for col_idx in range(1, 10):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = light_blue_fill
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                if col_idx > 1:
                    cell.alignment = Alignment(horizontal='right')
    
    # Format all numeric cells
    for row_idx in range(7, 38):
        for col_idx in range(2, 10):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if isinstance(value, str) and (value.replace('-', '').replace('.', '').replace(',', '').isdigit() or value == "-"):
                cell.alignment = Alignment(horizontal='right')
    
    # Add borders to all cells in the table
    for row_idx in range(4, 38):
        for col_idx in range(1, 10):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = border
    
    # Set column widths
    column_widths = [25, 18, 15, 15, 15, 18, 10, 15, 20]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width

# Populate Annexure B sheet
annexure_b_sheet = wb['Annexure B']
populate_annexure_b_sheet_combined(annexure_b_sheet, annexure_b_data)

# Function to safely populate a sheet with data and apply formatting
def populate_sheet(sheet, data, has_note_column=False):
    # Set the column offset based on note column presence
    column_offset = 1 if has_note_column else 0
    
    # First pass: Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Second pass: Apply merges
    for row_idx, row in enumerate(data, 1):
        if row_idx in [2, 3]:  # Company name and report title rows
            if len(row) > 1:
                sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(row))
    
    # Third pass: Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            try:
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Apply styles
                if row_idx in [2, 3]:  # Company name and report title
                    if col_idx == 1:  # Only format the first cell of merged cells
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                elif row_idx in [5, 6]:  # Column headers
                    cell.font = subheader_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                elif row_idx > 6 and col_idx > (1 + column_offset):  # Numeric values
                    # Right-align numeric values
                    cell.alignment = Alignment(horizontal='right')
                
                # Special formatting for section headers
                if col_idx == 1 and row_idx > 6 and isinstance(value, str) and value and not value.startswith(" "):
                    if value in ["EQUITY AND LIABILITIES", "ASSETS", "INCOME", "EXPENSES", 
                               "I) Cash Flows from Operating Activities", 
                               "II) Cash Flows from Investing Activities",
                               "III) Cash Flows From Financing Activities",
                               "TANGIBLE ASSETS", "INTANGIBLE ASSETS"]:
                        cell.font = subheader_font
                        cell.fill = subheader_fill
                
                # Special formatting for totals
                if isinstance(value, str) and value == "TOTAL" or (row_idx > 6 and isinstance(value, str) and "total" in value.lower()):
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
                
                # Add borders to all non-empty cells
                if value:
                    cell.border = border
            except:
                # Skip cells that can't be formatted (like merged cells)
                pass
    
    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max(max_length + 2, 12), 50)  # Min width 12, max 50
        sheet.column_dimensions[column_letter].width = adjusted_width

# Function to populate a note sheet
def populate_note_sheet(sheet, data):
    # Set all cell values
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            try:
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Header formatting for note title
                if row_idx == 1 and col_idx == 1:
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                
                # Format numeric values (columns 2 and 3)
                if col_idx in [2, 3] and isinstance(value, str) and value:
                    cell.alignment = Alignment(horizontal='right')
                
                # Format total row
                if isinstance(value, str) and "TOTAL" in value:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
                
                # Add borders to cells with content
                if value:
                    cell.border = border
            except:
                pass
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

# Modify the populate_annexure_sheet function to handle merged cells correctly
def populate_annexure_sheet(sheet, data):
    # FIRST: Define any merges we need to apply, if any
    # (Empty for now - add specific merges as needed)
    
    # SECOND: Set all cell values (avoiding merged cells)
    # For any row/column where merged cells exist, only set value on the top-left cell
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            # Check if this is a regular cell (not a MergedCell)
            if not isinstance(sheet.cell(row=row_idx, column=col_idx), openpyxl.cell.cell.MergedCell):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
    
    # THIRD: Apply formatting
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            try:
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Header formatting
                if row_idx <= 2:
                    cell.font = subheader_font
                    if row_idx == 1 and col_idx == 1:
                        cell.fill = subheader_fill
                
                # Column headers
                if row_idx == 3:
                    cell.font = subheader_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Format section headers
                if col_idx == 1 and row_idx > 3 and isinstance(value, str) and "BLOCK" in value:
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                
                # Format numeric values
                if col_idx > 1 and isinstance(value, str) and value:
                    cell.alignment = Alignment(horizontal='right')
                
                # Format total rows
                if isinstance(value, str) and "TOTAL" in value:
                    cell.font = Font(bold=True)
                    if "GRAND TOTAL" in value:
                        cell.fill = total_fill
                    
                # Add borders to cells with content
                if value:
                    cell.border = border
            except:
                # Skip cells that can't be formatted (like merged cells)
                pass
    
    # Adjust column widths based on content
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max(max_length + 2, 12), 50)  # Min width 12, max 50
        sheet.column_dimensions[column_letter].width = adjusted_width
        
# Sundry Creditors data
sundry_creditors_header = [
    ["Nandhra Engineering & Construction (SPV) Pvt. Ltd.", "", ""],
    ["2nd Floor, Nathkrupa Complex,", "", ""],
    ["Near Hinjwadi Police Chowki,", "", ""],
    ["Shivaji Chowk, Hinjawadi, Pune.", "", ""],
    ["CIN: U45203PN2018PTC180980", "", ""],
    ["Sundry Creditors", "", ""],
    ["Group Summary", "", ""],
    ["1-Apr-2021 to 31-Mar-2022", "", ""],
    ["", "Sundry Creditors", ""],
    ["", "Nandhra Engineering &", ""],
    ["Particulars", "1-Apr-2021 to 31-Mar-2022", ""],
    ["", "Closing Balance", ""],
    ["", "Debit", "Credit"]
]

# Sundry Creditors - Expenses
sundry_creditors_expenses = [
    ["Sundry Creditors - Expenses", "", ""],
    ["Aaadnath Trade-Wings Private Limited", "769494.24", ""],
    ["A & A Consulting Engineer", "400000.00", ""],
    ["Accurate Valuers And Engineers", "29500.00", ""],
    ["Akinrao Gulabrao Ambiore", "", "346217.00"],
    ["Akul Enterprises", "", "29441.00"],
    ["Amit Vikas Mahajan", "500000.00", ""],
    ["Ananta Namdeo Unde", "", "295644.00"],
    ["Anil Kisan Raut", "", "105458.00"],
    ["Anirudha Baban Dahake", "", "1146720.00"],
    ["Ankush Shriram Jagtap", "", "192555.00"],
    ["Arun Harikshan Bhagat", "", "3025.00"],
    ["Baban Marotirao Mohire", "", "1540.00"],
    ["Bhagwan Kisan Popalghat", "", "550097.00"],
    ["Bharat N Pandhare", "", "854333.00"],
    ["Bhimrao Laxman Kulsundar", "", "40000.00"],
    ["Care Ratings Limited", "137500.00", ""],
    ["Chaggan Budharam Chinchole", "", "9075.00"],
    ["Chhallani Offset", "", "77998.00"],
    ["Countrywide Cargo Movers", "39677.00", ""],
    ["Damodhar Sakharam Falke", "", "147361.00"],
    ["Dandade  Pritam Bajirao", "125000.00", ""],
    ["Danish Sayyed Earth Moving & Road Constructions", "", "360887.00"],
    ["Darubai Jairam Ghope", "", "65000.00"],
    ["DATTATRAY SAHEBRAO DEOKAR", "", "154440.00"],
    ["DATTRYA NAMDEO WAKODE", "", "599412.00"],
    ["Deepak Bhimrao Kulsundar", "", "140000.00"],
    ["Digambar Madhukar Gunikar", "", "102000.00"],
    ["Dinkar Namdev Raut", "", "10065.00"],
    ["Dipak Onkar Ambiore", "", "6875.00"],
    ["Dipak Wasudeo Gawande", "", "185994.00"],
    ["D-Max Traders", "578683.00", ""],
    ["Eagleway Infra", "71226.40", ""],
    ["FUTURE INC", "14499.00", ""],
    ["Gajanan Bhanudas Sonalkar", "", "40000.00"],
    ["Gajanan Dinkar Raut", "", "8250.00"],
    ["Gajanan Narayan Malokar", "", "104050.00"],
    ["Gajanan Rajaram Nakhod", "", "9424.00"],
    ["Ganesh Govindrao Gunjikar", "", "90900.00"],
    ["GANESH V. KHANDARE", "", "183047.00"],
    ["Gopal Govindrao Gunjikar", "13880.00", ""],
    ["Gupteshwar Aqua", "", "47685.00"],
    ["Hemant Ruprao Parkhede", "", "230000.00"],
    ["Hind Earthmovers", "", "371391.00"],
    ["Hindusthan Security Services", "", "281337.00"],
    ["Infomerics Valuation & Rating Pvt. Ltd.", "216000.00", ""],
    ["Infraking Consulting Engineers", "1042360.00", ""],
    ["Jai Gurudev Earthmovers (Urp)", "", "44902.00"],
    ["Janardhan Atmaram Falke", "", "190080.00"],
    ["Janardhan Keshav Ambiore", "", "3300.00"],
    ["Jitendra Kanhaiyalal Kalntri", "1440000.00", ""],
    ["Kalim Kha Kadir Kha", "", "31000.00"],
    ["KARAN CONSTRUCTION CO.", "", "15568.00"],
    ["Kisandev Earth Moving Services", "", "345026.00"],
    ["KSD Borewells & Constructions", "20000.00", ""],
    ["Lead Law", "144000.00", ""],
    ["Mahadev  Nathu Dhomne", "", "2585.00"],
    ["Manohar Atmaram Falke", "", "158524.00"],
    ["Meharunbi Mohamad Hanif", "", "15190.00"],
    ["Messers Jagdamba Traders", "901458.00", ""],
    ["Mohammad Qasim Faqir Mohammad", "", "62169.00"],
    ["M P Diesel Services", "", "767321.90"],
    ["Mrs. Sarita Manoj Kashelani", "", "120000.00"],
    ["Mukund Gangaram Belage", "55577.00", ""],
    ["Najera Bi Shekh Fakira", "", "62026.00"],
    ["Nandhra Engg. & Const. (India) Pvt.Ltd.(Bill)", "4237023.49", "0.00"],
    ["Nand Power Link", "6430068.00", ""],
    ["Navneet Kaur", "29000.00", ""],
    ["Nilesh Arun Lahudkar", "", "9185.00"],
    ["Omkar Digambar Gunjikar", "", "102000.00"],
    ["Omkar Tech Power Services", "", "1125183.00"],
    ["Om Sakshi Engineers Co.", "", "113591.88"],
    ["Parashuram Jayappa Talakeri", "21265.00", ""],
    ["Parmeshwar P.K.", "", "6875.00"],
    ["POGIRI ANITHA", "", "47008.00"],
    ["Poonam Construction & Co.", "", "13114069.30"],
    ["Poonam Construction & Co. (LC)", "", "70976405.68"],
    ["Pramila Prakash Raut", "", "69000.00"],
    ["Pramod Pandurang Korade", "", "51000.00"],
    ["Prashik Raj Construction", "", "1170725.00"],
    ["Pratik Nandkishor Purohit", "", "78000.00"],
    ["Procon Contractors Pvt. Ltd.", "", "0.00"],
    ["Purushottam Samadhan Raut", "", "2365.00"],
    ["Rahul Bhimrao Dhurandhar", "", "17750.00"],
    ["Rahul Constructions", "", "221600.00"],
    ["Rajesh Namdev Unde", "", "89100.00"],
    ["Rajkumar Ramkrushna Naswale", "", "90000.00"],
    ["Rajlaxmi Earth Movers", "", "1531372.00"],
    ["Rajlaxmi Pravin Ghadyale (Mess)", "", "1344154.00"],
    ["Raju Ragibji Hatkar", "", "75000.00"],
    ["Raj Water Suppliers", "23673.00", ""],
    ["Ramdas Santosh Gunjikar", "", "189888.00"],
    ["Ramdas Tukaram Dhole", "", "48000.00"],
    ["Ramesh Namdev Raut", "", "107500.00"],
    ["Rameshwar Uddhav Shelake", "", "313226.00"],
    ["Ramleela Aqwa Dasala", "", "24660.00"],
    ["Ranjit Balasaheb Jagtap", "", "620923.00"],
    ["Ravindra Bhimrao Kulsundar", "", "10000.00"],
    ["Raviraj Construction Co.", "434265.00", ""],
    ["Raviraj Shivhari Bhende", "", "631432.00"],
    ["Reliance Jio Infocomm Limited", "", "16882.00"],
    ["Resurgent India Limited", "15000.00", ""],
    ["Rishikesh Purushottam Tale", "", "150707.00"],
    ["R K Deepak & Co.", "", "398000.00"],
    ["Rohidas Mohan Chinchole", "", "25725.00"],
    ["Rooi Construction Service", "", "120929.00"],
    ["Sagar Crane Services", "", "387445.00"],
    ["Sai Gajanan Aqua", "", "38970.00"],
    ["Sai Shardha Land Developers", "", "638913.00"],
    ["SAIYAD IQBAL SAIYED SAMMAD", "", "183433.00"],
    ["Sakina Shaikh Gaffar", "", "22840.00"],
    ["Samadhan Daulat Ambiore", "", "2200.00"],
    ["Samadhan Haribhau Chengade", "", "3753.00"],
    ["Sandip Ashok Nikrad", "", "202353.00"],
    ["Sanjay Kumar", "206220.00", ""],
    ["Sanjay Manohar Bhuje", "", "3025.00"],
    ["Sankalp Provision", "", "11338.00"],
    ["Sarthi Enterprises", "600000.00", ""],
    ["Satish Motors Pvt. Ltd.", "", "0.00"],
    ["Satish Ramesh Shinde (URP)", "", "1112874.00"],
    ["Sayyad Ismail Sayyad Gaibu", "", "95700.00"],
    ["Sejol Vinod Shyamrao", "", "51480.00"],
    ["Shaikh Ejaj Shaikh Chand", "", "25675.00"],
    ["Shaikh Hamid Shaikh Majid", "", "26500.00"],
    ["Shaikh Harun Shaikh Hasan", "", "60000.00"],
    ["Shaikh Harun Shaikh Razzak", "", "27000.00"],
    ["Shaikh Jabir Shaikh Rasul", "", "55800.00"],
    ["Shaikh Shabbir Shaikh Mahebub", "", "2090.00"],
    ["Shaikh Sultan Shaikh Subhan", "", "31500.00"],
    ["Sharad Laxmikant Joshi", "", "39000.00"],
    ["Shaurya Technosoft Pvt. Ltd.", "24780.00", ""],
    ["Shivaji Bhimrao Kokane", "", "70000.00"],
    ["Shiv IB Motors Pvt.Ltd.(Swalik Ford)", "", "22645.00"],
    ["Shivkumar Prabhudas Atravalkar", "900000.00", ""],
    ["Shree Anil Water Supply", "", "87450.00"],
    ["Shree Tading Company", "12955.00", ""],
    ["Shridhar Digambar Gunjikar", "", "96000.00"],
    ["Shrikrishna Subhash Misal", "", "1431214.00"],
    ["Shri Laxmi Ganapati Traders & Services", "", "1490000.00"],
    ["Shripada Constructions", "", "2497368.00"],
    ["Shiram Tukaram Bhalerao", "", "28000.00"],
    ["Shrish Ramesh Kurhe", "", "1051731.00"],
    ["SOPAN GAJANAN APWADE", "", "278320.00"],
    ["Subhash Muralidhar Tere", "", "548139.00"],
    ["Sukhakarla Kirana", "", "29177.00"],
    ["Sukhija Real Estate Private Ltd.", "7852.00", ""],
    ["Suresh Ravaji Kuskar (Bolero)", "", "25784.00"],
    ["Suspence Account", "130000.00", ""],
    ["TCI Express Limited", "", "1658.00"],
    ["Technogem Consultants Pvt. Ltd.", "600000.00", ""],
    ["Thakre Digambar Namdeo", "", "117000.00"],
    ["Thakur Sales & Services", "", "696000.00"],
    ["Tirupati Enterprises", "", "663420.00"],
    ["Udhavrao Roduji Gawargur", "", "5280.00"],
    ["Vasudeo Janardhan Ambikar", "", "139048.00"],
    ["VISHNUDAS BHANUDAS JADHAV", "", "132932.00"],
    ["Vision Infra", "275000.00", ""],
    ["Vithoba Naithu Dhomne", "", "2585.00"],
    ["Yash Engineers Co.", "", "192020.00"],
    ["Yogesh Arvind Raut", "53000.00", ""],
    ["", "", ""],
    ["TOTAL", "10393522", "30687283"]
]

# Sundry Creditors - Purchase
sundry_creditors_purchase = [
    ["Sundry Creditors - Purchase", "", ""],
    ["Abhiraj Enterprises", "", "282905.00"],
    ["ACC Limited", "58854.00", ""],
    ["Ammann India Private Limited (Gujrat)", "17263.54", ""],
    ["Apple Chemie India Pvt. Ltd.", "22390.00", ""],
    ["Aqua- Tech Services", "80000.00", ""],
    ["Aradhya Enterpises", "", "3920631.00"],
    ["Arhant Agro Services", "", "25314.00"],
    ["Ashoka Enterprises", "", "53265.00"],
    ["Atrinandan Cement Products", "11605.00", ""],
    ["Bajaj Travels Ltd.", "1259430.00", ""],
    ["Bharat Motors Co.", "16124.00", ""],
    ["Dicksons Engineering Co. Pvt. Ltd.", "994.00", ""],
    ["Ganesh B. Pandhare", "", "767880.00"],
    ["Gurukrupa Cement Depot", "0.00", ""],
    ["Hindustan Petroleum Corporation Limited", "355140.85", ""],
    ["Iron Trangle Ltd.", "524808.00", ""],
    ["IWL India Limited", "", "902388.00"],
    ["Jamnadas Iron & Steel Company", "", "22888.00"],
    ["JSK Corporation Pvt. Ltd.", "74747.00", ""],
    ["J V Engineering Works", "282492.00", ""],
    ["Kalika Steel Alloys Pvt. Ltd.", "", "7509093.00"],
    ["Kamalkishor P. Bhattad", "0.00", ""],
    ["Laxminarayan Cement Product", "0.00", ""],
    ["Madhav Manikrao Deshmukh", "232925.00", ""],
    ["Mahalakshmi Energy Private Limited", "", "3557800.00"],
    ["Metso India Private Limited", "", "12653.14"],
    ["M M Deshmukh Building Material Suppliers", "", "100000.00"],
    ["Nilesh Welding Works", "21440.00", ""],
    ["Om Industries", "13736.00", ""],
    ["Palasidha Vibrotech Pipes", "", "1804620.00"],
    ["Pankaj Automobiles", "4600.00", ""],
    ["Paridar Boards Pvt. Ltd.", "", "505000.00"],
    ["Rajlaxmi Servo", "0.00", "138485.53"],
    ["Renuka Hardware & Electricals", "", "3060.00"],
    ["R R Bondre", "", "1000000.00"],
    ["Saffron Redymix Concrete", "19087878.20", ""],
    ["Saifee Iron & Steel", "26816.00", ""],
    ["Shivkrupa Petroleum", "", "4711051.73"],
    ["Shravani Suppliers", "0.00", ""],
    ["Shree Dattakrupa Stone Crusher", "", "30915.00"],
    ["Shree Sadguru Kaka Stone Crusher", "0.00", "66120"],
    ["Shree Traders-Khamgaon", "", "150414.00"],
    ["Shri Seva General and Electrical", "", "9175.00"],
    ["Udeep Systems", "54918.00", ""],
    ["Uma Maheshwar Agency", "", "302149.00"],
    ["Unique Petroleum Services", "", "301076.00"],
    ["Vijayshri Traders", "", "1120180.00"],
    ["", "", ""],
    ["TOTAL", "22146162", "27297063"]
]

# Sundry Creditors - Salary
sundry_creditors_salary = [
    ["Sundry Creditors ( Salary)", "", ""],
    ["Alka Shelar (Salary)", "", "220750.00"],
    ["Anand Sangamanath Sindhur (Salary)", "", "51370.00"],
    ["Baburao Shelar (Salary)", "", "326390.00"],
    ["Balaji Tonge ( Salary)", "", "18358.00"],
    ["Bhagappa Ramchandra Reddy (Salary)", "", "82027.00"],
    ["Datta Sathe (Salary)", "", "65368.00"],
    ["Devendra Kumar Khule (Salary)", "", "38981.00"],
    ["Devidas Ramchandra Shete (Salary)", "", "218160.00"],
    ["Dhanraj Namdev Munde (Salary)", "", "104192.00"],
    ["Dinesh Narayan Parmar (Salary)", "", "15364.00"],
    ["Ghanshyam Zinzuwadiya (Salary)", "", "7726.00"],
    ["Goroba Dagdu Thodsare (Salary)", "", "47481.00"],
    ["", "", ""],
    ["Gurpalsingh Virk (Salary)", "", "699990.00"],
    ["Gurtej Singh Nandhra (Salary)", "1665600.00", "0.00"],
    ["Harmeet Kaur (Salary)", "", "140050.00"],
    ["Imtiyaz Ahmed Siddiqui (Salary)", "", "81727.00"],
    ["Iqbal Singh  Nandhra (Salary)", "7996800.00", ""],
    ["Jaldeep Balvatrava Desai (Salary)", "", "185319.00"],
    ["Jasvinder Singh ( Salary)", "", "159892.00"],
    ["Jitendra Kumar (Salary)", "", "83780.00"],
    ["Kabir Nathu Ahmed (Salary)", "", "3600.00"],
    ["Kapil Ashok Nalawade (Salary)", "", "81732.00"],
    ["Kirankumar Kantayya Swami (Salary)", "", "4941.00"],
    ["Kishor Ankush Mohite (Salary)", "2003200.00", ""],
    ["Krishna Rajendra Karatade (Salary)", "", "89224.00"],
    ["Mahesh Basavaraj (Salary)", "", "16891.00"],
    ["Mahesh Nandkishor Gore (Salary)", "", "127313.00"],
    ["Mohan Chandrabhan Singh (Salary)", "", "210890.00"],
    ["Nitesh Dilip Jadhav (Salary)", "", "161136.00"],
    ["Nitin Mahadev Kambale (Salary)", "", "40009.00"],
    ["Prashant Uttam Kate (Salary)", "", "178792.00"],
    ["Prem Prakash Soni (Salary)", "", "20004.00"],
    ["Pulli Sambaiah (Salary)", "", "44448.00"],
    ["Rahat Virk (Salary)", "438000.00", ""],
    ["Rajesh Kumar Dhiman (Salary)", "", "167686.00"],
    ["Ram Jadhar ( Salary)", "", "111892.00"],
    ["Randeep Gautam (Salary)", "", "162145.00"],
    ["Sandip Suryawanshi (Salary)", "", "57696.00"],
    ["Sanjaykumar Yashwant Khadase (Salary)", "", "161892.00"],
    ["Sanjay Sharma (Salary)", "", "89852.00"],
    ["Santosh Ankush Mohite (Salary)", "4083200.00", ""],
    ["Satish Kumar Hiralal Katara (Salary)", "", "40785.00"],
    ["Shahid Qumar Abdul Qadar (Salary)", "", "213661.00"],
    ["Shaikh Juber Ahmed Abdul Karim (Salary)", "", "188048.00"],
    ["Shispal Singh (Salary)", "", "98102.00"],
    ["Shiva Purushottam Tale (Salary)", "", "45534.00"],
    ["Shrikrushna Vasudev Gavande (Salary)", "", "99392.00"],
    ["Suresh Ravaji Kuskar (Salary)", "", "645510.00"],
    ["Vinod Kishor Dhokane ( Salary)", "", "112570.00"],
    ["Vrandavan Dhakar (Salary)", "", "89572.00"],
    ["Yatlamelli Chittikiran Rambahu (Salary)", "", "40168.00"],
    ["", "", ""],
    ["TOTAL", "15748800.00", "6288410.00"]
]

# Combine all Sundry Creditors data
sundry_creditors_data = sundry_creditors_header + sundry_creditors_expenses + sundry_creditors_purchase + sundry_creditors_salary

# Function to populate the Sundry Creditors sheet
def populate_sundry_creditors_sheet(sheet, data):
    # First pass: Set all cell values
    row_index = 1
    for row in data:
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.value = value
        row_index += 1
    
    # Define formatting
    title_font = Font(bold=True, size=12)
    address_font = Font(size=10)
    header_font = Font(bold=True, size=10)
    section_font = Font(bold=True, italic=True, size=10)
    
    # Format the header section
    for row_idx in range(1, 6):
        cell = sheet.cell(row=row_idx, column=1)
        if row_idx == 1:
            cell.font = title_font
        else:
            cell.font = address_font
    
    # Format the Sundry Creditors title and Group Summary
    for row_idx in range(6, 9):
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = header_font
    
    # Format the column headers
    for row_idx in range(9, 14):
        for col_idx in range(1, 4):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if col_idx > 1 or row_idx == 11:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            if row_idx == 13 and col_idx > 1:
                cell.border = Border(bottom=Side(style='thin'))
    
    # Find and format section headers
    section_headers = ["Sundry Creditors - Expenses", "Sundry Creditors - Purchase", "Sundry Creditors ( Salary)"]
    row_idx = 14
    while row_idx <= row_index:
        cell = sheet.cell(row=row_idx, column=1)
        if cell.value in section_headers:
            cell.font = section_font
            
            # Add underline to section headers
            for col_idx in range(1, 4):
                section_cell = sheet.cell(row=row_idx, column=col_idx)
                section_cell.border = Border(bottom=Side(style='thin'))
        row_idx += 1
    
    # Find and format total rows
    total_rows = []
    row_idx = 14
    while row_idx <= row_index:
        cell = sheet.cell(row=row_idx, column=1)
        if cell.value == "TOTAL":
            total_rows.append(row_idx)
            
            # Format total row
            for col_idx in range(1, 4):
                total_cell = sheet.cell(row=row_idx, column=col_idx)
                total_cell.font = Font(bold=True)
                total_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
        row_idx += 1
    
    # Format amount values with right alignment
    row_idx = 14
    while row_idx <= row_index:
        for col_idx in [2, 3]:
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if isinstance(value, str) and (value.replace('.', '').isdigit() or value == "0.00"):
                cell.alignment = Alignment(horizontal='right')
        row_idx += 1
    
    # Set column widths
    sheet.column_dimensions['A'].width = 55
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    
    # Add borders to the main table
    for row_idx in range(9, row_index):
        for col_idx in range(1, 4):
            cell = sheet.cell(row=row_idx, column=col_idx)
            existing_border = cell.border
            
            # Skip cells that already have special borders
            if existing_border and (existing_border.bottom.style == 'double' or existing_border.bottom.style == 'thin'):
                continue
                
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                                left=Side(style='thin'), right=Side(style='thin'))

# Create and populate the Sundry Creditors sheet
sundry_creditors_sheet = wb.create_sheet("Sundry Creditors")
populate_sundry_creditors_sheet(sundry_creditors_sheet, sundry_creditors_data)

# Populate main sheets
populate_sheet(balance_sheet, balance_sheet_data, True)
populate_sheet(profit_loss, profit_loss_data, True)

# Populate annexure sheets
populate_annexure_sheet(annexure_b, annexure_b_data)

# Get current directory
current_dir = os.getcwd()
file_path = os.path.join(current_dir, 'NANDHRA_FINANCIALS_WITH_ANNEXURES.xlsx')

# Save the workbook with the full path
try:
    wb.save(file_path)
    print(f"Excel file with annexures created successfully at: {file_path}")
except Exception as e:
    print(f"Error saving Excel file: {e}")
    print("Current directory:", current_dir)